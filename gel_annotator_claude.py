
"""
Gel & Blot Annotator (MVP v0.3)

A step-by-step GUI to:
- Load gel electrophoresis / SDS-PAGE / Western blot images (JPG/PNG/TIF)
- Perform basic image edits (brightness/contrast/gamma/invert, rotate, crop)
- Configure lanes, headers, group brackets, marker annotations
- Add marker/ladder annotations from a saved marker library (JSON) or create new calibrations
- Export a final annotated figure; optionally continue with next image while keeping settings

Dependencies: tkinter (built-in), opencv-python, pillow, numpy
"""

from __future__ import annotations

import json
import math
import os
import pathlib
import re
import sys
import uuid
import weakref
import webbrowser
import copy
import hashlib
import logging
import threading
from functools import lru_cache
from dataclasses import dataclass, asdict, field
from typing import List, Optional, Tuple, Dict, Any

logger = logging.getLogger(__name__)

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, colorchooser
from tkinter import font as tkfont

import numpy as np
import cv2
from PIL import Image, ImageTk, ImageDraw, ImageFont, ImageOps

try:
    import openpyxl
    from openpyxl.styles import Font as XLFont, PatternFill, Alignment, Border, Side
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


# ----------------------------
# Persistence locations
# ----------------------------
APP_DIR = pathlib.Path.home() / ".gel_annotator"
SETTINGS_PATH = APP_DIR / "settings.json"
MARKER_DIR = APP_DIR / "markers"

APP_DIR.mkdir(parents=True, exist_ok=True)
MARKER_DIR.mkdir(parents=True, exist_ok=True)

README_BASENAME = "README_Gel_Annotator.md"
APP_AUTHOR = "Felix Gerlach"
APP_BUG_EMAIL = "felixgerlach@yahoo.de"

# ----------------------------
# Layout constants
# ----------------------------
SIDEBAR_WIDTH_NARROW = 360
SIDEBAR_WIDTH_WIDE = 520
SIDEBAR_WIDTH_MEDIUM = 420
PANEL_SPACING_PX = 20


def _runtime_base_dir() -> pathlib.Path:
    # PyInstaller one-file extracts resources into _MEIPASS; in normal Python use script dir.
    try:
        meipass = getattr(sys, "_MEIPASS", None)
        if meipass:
            return pathlib.Path(str(meipass))
    except Exception:
        pass
    try:
        return pathlib.Path(__file__).resolve().parent
    except Exception:
        return pathlib.Path.cwd()


def find_packaged_readme() -> Optional[pathlib.Path]:
    candidates: List[pathlib.Path] = []
    try:
        candidates.append(_runtime_base_dir() / README_BASENAME)
    except Exception:
        pass
    try:
        candidates.append(pathlib.Path(sys.executable).resolve().parent / README_BASENAME)
    except Exception:
        pass
    try:
        candidates.append(pathlib.Path(__file__).resolve().parent / README_BASENAME)
    except Exception:
        pass
    seen: set[str] = set()
    for p in candidates:
        sp = str(p)
        if sp in seen:
            continue
        seen.add(sp)
        if p.exists():
            return p
    return None


def open_readme_in_system() -> bool:
    p = find_packaged_readme()
    if not p:
        return False
    try:
        if os.name == "nt":
            os.startfile(str(p))  # type: ignore[attr-defined]
            return True
    except Exception:
        pass
    try:
        return bool(webbrowser.open(p.resolve().as_uri()))
    except Exception:
        return False


# ----------------------------
# Data models
# ----------------------------
@dataclass
class AppSettings:
    gel_type: str = "agarose"  # agarose | sds_page | western_blot
    panel_mode: str = "single"  # single | multi_panel
    source_mode: str = "single_image"  # single_image | multi_image
    default_unit: str = "bp"  # bp | kDa
    keep_settings_next_image: bool = True
    prefer_invert: bool = False
    font_path: str = ""  # optional path to .ttf
    font_size: int = 20
    font_family: str = ""       # empty = use default/path font
    font_bold: bool = False
    font_italic: bool = False
    lane_label_angle_deg: int = 0
    theme_color: str = "#5B8DB8"

@dataclass
class HighlightRect:
    x0: int
    y0: int
    x1: int
    y1: int
    width: int = 3
    color: str = "#DC0000"
    kind: str = "box"  # box | arrow | asterisk

@dataclass
class GroupLabel:
    start_lane: int
    end_lane: int
    text: str
    bracket: bool = True
    id: str = ""
    height_group: str = ""

@dataclass
class HeaderRow:
    name: str
    values: List[str]  # per lane
    id: str = ""
    position: str = "top"  # top | bottom
    angle_deg: int = 0
    heading_font_size: Optional[int] = None
    value_font_size: Optional[int] = None

@dataclass
class MarkerDefinition:
    name: str
    unit: str  # bp | kDa
    sizes: List[float]  # descending or ascending is ok
    # Either provide explicit y pixel positions (same length as sizes) OR calibrate from clicks.
    y_positions: Optional[List[int]] = None

@dataclass
class MarkerCalibration:
    marker_name: str = ""
    ladder_lane: int = 1
    unit: str = "bp"
    # Picked points (y pixels) paired with sizes; used to fit y = a*log10(size)+b or inverse
    picked: List[Tuple[int, float]] = field(default_factory=list)
    fit_a: Optional[float] = None  # y = a*log10(size) + b
    fit_b: Optional[float] = None

@dataclass
class ImagePlacement:
    x: int = 0
    y: int = 0
    angle_deg: float = 0.0
    scale_x: float = 1.0
    scale_y: float = 1.0
    brightness: float = 0.0
    contrast: float = 1.0
    gamma: float = 1.0
    white_balance: bool = False
    invert: bool = False
    bw: bool = True

@dataclass
class PanelConfig:
    lanes: int = 10
    gel_left: Optional[int] = None
    gel_right: Optional[int] = None
    gel_regions: List[Tuple[int, int]] = field(default_factory=list)
    gel_region_lane_counts: List[int] = field(default_factory=list)

    include_marker: bool = True
    # Marker styling (per panel)
    marker_x_offset: int = 0                                  # shift ticks/labels left/right
    marker_y_offset: int = 0                                  # shift ticks/labels up/down
    marker_tick_length: int = 8                               # tick length in px
    marker_label_gap: int = 10                                # gap between tick end and label

    marker_font_size: Optional[int] = 15                          # font size for marker size+unit labels

    marker_tick_overrides: Dict[float, int] = field(default_factory=dict)  # per-size y override in panel pixels
    marker_tick_hidden: List[float] = field(default_factory=list)              # sizes hidden for this panel

    header_rows: List[HeaderRow] = field(default_factory=list)
    header_value_angle_deg: int = 0
    header_heading_x_offset: int = 0
    header_heading_y_offset: int = 0
    header_values_x_offset: int = 0
    header_values_y_offset: int = 0
    bracket_label_x_offset: int = 0
    bracket_label_y_offset: int = 0
    bracket_line_x_offset: int = 0
    bracket_line_y_offset: int = 0
    group_labels: List[GroupLabel] = field(default_factory=list)
    top_annotation_order: List[str] = field(default_factory=list)
    bracket_text_gap: int = 10
    highlight_enabled: bool = False
    highlights: List[HighlightRect] = field(default_factory=list)
    highlight_color: str = "#DC0000"
    highlight_width: int = 3
    highlight_shape: str = "box"  # box | arrow | asterisk
    run_band_analysis: bool = True
    analysis_peak_threshold: float = 20.0
    analysis_prominence: float = 7.0
    analysis_min_distance: int = 10
    analysis_smooth_window: int = 10
    analysis_lane_half_width: int = 2
    analysis_polarity: str = "dark"  # dark | light
    analysis_background_correction: bool = False
    analysis_background_lane: int = 1
    analysis_bg_mode: str = "none"  # none | lane | rolling_ball | local_adjacent
    analysis_rolling_ball_radius: int = 50
    analysis_fit_mode: str = "monotone_interp"  # monotone_interp | log_linear | log_quadratic
    show_band_sizes_on_final: bool = True
    band_size_excluded_lanes: List[int] = field(default_factory=list)
    analysis_band_sizes_by_lane: Dict[int, List[str]] = field(default_factory=dict)
    analysis_band_values_by_lane: Dict[int, List[Optional[float]]] = field(default_factory=dict)
    analysis_band_size_unit: str = ""
    band_size_included_bands_by_lane: Dict[int, List[int]] = field(default_factory=dict)
    final_band_label_font_size: Optional[int] = 15
    review_text_overrides: Dict[str, Dict[str, Any]] = field(default_factory=dict)
    marker_calibration: MarkerCalibration = field(default_factory=MarkerCalibration)

@dataclass
class ProjectState:
    settings: AppSettings = field(default_factory=AppSettings)
    # Images
    original_path: str = ""
    loaded_pil: Optional[Image.Image] = None  # original un-cropped image for reset
    original_pil: Optional[Image.Image] = None
    edited_pil: Optional[Image.Image] = None
    # Multi-image input set (before optional multi-panel extraction)
    multi_input_paths: List[str] = field(default_factory=list)
    multi_input_images: List[Image.Image] = field(default_factory=list)
    # Multi-image compose workflow
    multi_source_paths: List[str] = field(default_factory=list)
    multi_source_images: List[Image.Image] = field(default_factory=list)
    multi_source_placements: List[ImagePlacement] = field(default_factory=list)
    multi_source_origin_ids: List[int] = field(default_factory=list)   # source-image id for each composed panel
    multi_source_origin_names: List[str] = field(default_factory=list) # source-image name for each composed panel
    multi_source_lock_y: bool = False
    multi_source_final_crop_undo: Optional[Dict[str, Any]] = None
    # Multi-panel: list of panel images extracted from edited_pil or loaded from multiple images
    panels: List[Image.Image] = field(default_factory=list)
    panel_configs: List[PanelConfig] = field(default_factory=list)
    current_panel_index: int = 0
    # last export
    last_export_path: str = ""


# ----------------------------
# Helpers
# ----------------------------
def load_settings() -> AppSettings:
    if SETTINGS_PATH.exists():
        try:
            data = json.loads(SETTINGS_PATH.read_text(encoding="utf-8"))
            return AppSettings(**data)
        except Exception as exc:
            logger.warning("load_settings failed: %s", exc)
            return AppSettings()
    return AppSettings()

def save_settings(settings: AppSettings) -> None:
    try:
        SETTINGS_PATH.write_text(json.dumps(asdict(settings), indent=2), encoding="utf-8")
    except Exception as exc:
        logger.warning("save_settings failed: %s", exc)

def list_markers() -> List[str]:
    names = []
    for p in MARKER_DIR.glob("*.json"):
        names.append(p.stem)
    names.sort()
    return names

def load_marker(name: str) -> MarkerDefinition:
    return _load_marker_cached(name)

@lru_cache(maxsize=256)
def _load_marker_cached(name: str) -> MarkerDefinition:
    p = MARKER_DIR / f"{name}.json"
    try:
        data = json.loads(p.read_text(encoding="utf-8"))
        return MarkerDefinition(**data)
    except Exception as exc:
        logger.warning("Could not load marker '%s': %s", name, exc)
        raise

def save_marker(marker: MarkerDefinition) -> None:
    p = MARKER_DIR / f"{marker.name}.json"
    try:
        p.write_text(json.dumps(asdict(marker), indent=2), encoding="utf-8")
    except Exception as exc:
        logger.warning("Could not save marker '%s': %s", marker.name, exc)
        raise
    _load_marker_cached.cache_clear()


def _hex_to_rgb(color: str) -> Tuple[int, int, int]:
    s = str(color or "").strip().lstrip("#")
    if len(s) == 3:
        s = "".join(ch * 2 for ch in s)
    if len(s) != 6 or any(ch not in "0123456789abcdefABCDEF" for ch in s):
        raise ValueError(f"invalid color: {color}")
    return int(s[0:2], 16), int(s[2:4], 16), int(s[4:6], 16)


def _rgb_to_hex(rgb: Tuple[int, int, int]) -> str:
    r, g, b = [max(0, min(255, int(v))) for v in rgb]
    return f"#{r:02X}{g:02X}{b:02X}"


def _mix_hex(c1: str, c2: str, t: float) -> str:
    t = max(0.0, min(1.0, float(t)))
    r1, g1, b1 = _hex_to_rgb(c1)
    r2, g2, b2 = _hex_to_rgb(c2)
    return _rgb_to_hex((
        int(round(r1 * (1.0 - t) + r2 * t)),
        int(round(g1 * (1.0 - t) + g2 * t)),
        int(round(b1 * (1.0 - t) + b2 * t)),
    ))


def _safe_hex_color(color: str, fallback: str = "#5B8DB8") -> str:
    try:
        _hex_to_rgb(color)
        s = str(color).strip()
        return s if s.startswith("#") else f"#{s}"
    except Exception:
        return fallback


def _contrast_text_for_bg(color: str) -> str:
    try:
        r, g, b = _hex_to_rgb(color)
    except Exception:
        return "black"
    # Relative luminance approximation
    lum = (0.299 * r + 0.587 * g + 0.114 * b)
    return "black" if lum >= 150 else "white"


def ensure_default_markers() -> None:
    """Create a couple of common marker templates if the marker folder is empty."""
    if any(MARKER_DIR.glob("*.json")):
        return
    defaults = [
        MarkerDefinition(
            name="DNA_1kb_ladder",
            unit="bp",
            sizes=[10000, 8000, 6000, 5000, 4000, 3000, 2000, 1500, 1000, 750, 500, 250],
            y_positions=None,
        ),
        MarkerDefinition(
            name="DNA_100bp_ladder",
            unit="bp",
            sizes=[1500, 1200, 1000, 900, 800, 700, 600, 500, 400, 300, 200, 100],
            y_positions=None,
        ),
        MarkerDefinition(
            name="Protein_Prestained",
            unit="kDa",
            sizes=[250, 150, 100, 75, 50, 37, 25, 20, 15, 10],
            y_positions=None,
        ),
    ]
    for m in defaults:
        try:
            save_marker(m)
        except Exception:
            pass

def pil_from_path(path: str) -> Image.Image:
    img = Image.open(path)
    # keep EXIF orientation consistent
    try:
        img = ImageOps.exif_transpose(img)  # type: ignore
    except Exception:
        pass
    return img.convert("RGB")

def pil_to_cv(img: Image.Image) -> np.ndarray:
    arr = np.array(img)
    return cv2.cvtColor(arr, cv2.COLOR_RGB2BGR)

def cv_to_pil(arr_bgr: np.ndarray) -> Image.Image:
    arr_rgb = cv2.cvtColor(arr_bgr, cv2.COLOR_BGR2RGB)
    return Image.fromarray(arr_rgb)

def _apply_compose_adjustments(img: Image.Image, pl: ImagePlacement) -> Image.Image:
    bgr = pil_to_cv(img)

    if bool(getattr(pl, "white_balance", False)):
        bgr = gray_world_white_balance(bgr)

    alpha = float(getattr(pl, "contrast", 1.0))
    beta = float(getattr(pl, "brightness", 0.0))
    bgr = cv2.convertScaleAbs(bgr, alpha=alpha, beta=beta)

    bgr = apply_gamma(bgr, float(getattr(pl, "gamma", 1.0)))

    if bool(getattr(pl, "bw", False)):
        bgr = grayscale_with_red_as_black(bgr)

    if bool(getattr(pl, "invert", False)):
        bgr = 255 - bgr

    return cv_to_pil(bgr)

def _transform_for_compose(img: Image.Image, pl: ImagePlacement, bg=(255, 255, 255)) -> Image.Image:
    img = _apply_compose_adjustments(img, pl)
    sx = max(0.05, float(getattr(pl, "scale_x", 1.0)))
    sy = max(0.05, float(getattr(pl, "scale_y", 1.0)))
    nw = max(1, int(round(img.width * sx)))
    nh = max(1, int(round(img.height * sy)))
    if nw != img.width or nh != img.height:
        img = img.resize((nw, nh), Image.Resampling.BICUBIC)
    return img.rotate(float(pl.angle_deg), expand=True, resample=Image.Resampling.BICUBIC, fillcolor=bg)

def compose_images_with_meta(
    images: List[Image.Image], placements: List[ImagePlacement], bg=(255, 255, 255)
) -> Tuple[Image.Image, List[Dict[str, int]]]:
    """
    Compose transformed images and return:
    - composed image
    - per-image metadata with shifted bbox in composed-image coordinates:
      {index, x, y, w, h, x0, y0, x1, y1}
    """
    if not images:
        return Image.new("RGB", (800, 600), bg), []

    if len(placements) != len(images):
        placements = [ImagePlacement() for _ in images]

    transformed: List[Tuple[int, Image.Image, int, int]] = []
    min_x = 0
    min_y = 0
    max_x = 1
    max_y = 1

    for idx, (img, pl) in enumerate(zip(images, placements)):
        rot = _transform_for_compose(img, pl, bg=bg)
        x = int(pl.x)
        y = int(pl.y)
        transformed.append((idx, rot, x, y))
        min_x = min(min_x, x)
        min_y = min(min_y, y)
        max_x = max(max_x, x + rot.width)
        max_y = max(max_y, y + rot.height)

    out = Image.new("RGB", (max_x - min_x, max_y - min_y), bg)
    ox = -min_x
    oy = -min_y
    meta: List[Dict[str, int]] = []
    for idx, rot, x, y in transformed:
        out.paste(rot, (x + ox, y + oy))
        x0 = int(x + ox)
        y0 = int(y + oy)
        meta.append(
            dict(
                index=int(idx),
                x=int(x),
                y=int(y),
                w=int(rot.width),
                h=int(rot.height),
                x0=x0,
                y0=y0,
                x1=int(x0 + rot.width),
                y1=int(y0 + rot.height),
            )
        )
    return out, meta

def compose_images(images: List[Image.Image], placements: List[ImagePlacement], bg=(255, 255, 255)) -> Image.Image:
    """
    Compose transformed images on a common canvas using per-image x/y offsets and rotation.
    """
    out, _meta = compose_images_with_meta(images, placements, bg=bg)
    return out

def apply_gamma(img_bgr: np.ndarray, gamma: float) -> np.ndarray:
    gamma = max(0.05, float(gamma))
    table = _gamma_lut(round(gamma, 3))
    return cv2.LUT(img_bgr, table)

@lru_cache(maxsize=256)
def _gamma_lut(gamma: float) -> np.ndarray:
    inv = 1.0 / gamma
    return np.array([(i / 255.0) ** inv * 255 for i in range(256)], dtype=np.uint8)

def gray_world_white_balance(img_bgr: np.ndarray) -> np.ndarray:
    # Simple gray-world WB; safe default for gels/blots.
    b, g, r = cv2.split(img_bgr)
    b_avg, g_avg, r_avg = np.mean(b), np.mean(g), np.mean(r)
    k = (b_avg + g_avg + r_avg) / 3.0
    b = np.clip(b * (k / (b_avg + 1e-6)), 0, 255)
    g = np.clip(g * (k / (g_avg + 1e-6)), 0, 255)
    r = np.clip(r * (k / (r_avg + 1e-6)), 0, 255)
    return cv2.merge([b.astype(np.uint8), g.astype(np.uint8), r.astype(np.uint8)])

def grayscale_with_red_as_black(img_bgr: np.ndarray) -> np.ndarray:
    """
    Convert to grayscale but force strongly red pixels to black.
    Useful when red fluorescence/overlays should appear as dark bands in B/W mode.
    """
    if img_bgr is None or img_bgr.size == 0:
        return img_bgr

    b, g, r = cv2.split(img_bgr)
    hsv = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2HSV)
    h = hsv[:, :, 0]
    s = hsv[:, :, 1]
    v = hsv[:, :, 2]

    # Red occupies low and high hue ranges in OpenCV HSV (0..179).
    red_hue = ((h <= 15) | (h >= 165))
    red_sat = (s >= 60) & (v >= 20)
    red_dom = (r.astype(np.int16) >= g.astype(np.int16) + 20) & (r.astype(np.int16) >= b.astype(np.int16) + 20)
    red_mask = red_hue & red_sat & red_dom

    gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
    gray[red_mask] = 0
    return cv2.cvtColor(gray, cv2.COLOR_GRAY2BGR)

@lru_cache(maxsize=256)
def _truetype_font_cached(path: str, size: int) -> ImageFont.FreeTypeFont:
    return ImageFont.truetype(path, size=size)


@lru_cache(maxsize=64)
def _dejavu_font_cached(size: int) -> ImageFont.FreeTypeFont:
    return ImageFont.truetype("DejaVuSans.ttf", size=size)


def get_font_with_size(settings: AppSettings, size: int) -> ImageFont.FreeTypeFont:
    """Get a font with an explicit size (used for marker labels)."""
    size = max(6, int(size))
    # 1. Explicit .ttf path takes priority
    if settings.font_path and pathlib.Path(settings.font_path).exists():
        try:
            return _truetype_font_cached(str(pathlib.Path(settings.font_path).resolve()), size=size)
        except Exception:
            pass
    # 2. Named font family (system fonts)
    family = str(getattr(settings, "font_family", "") or "").strip()
    bold = bool(getattr(settings, "font_bold", False))
    italic = bool(getattr(settings, "font_italic", False))
    if family:
        # Try common system font paths for the named family
        suffixes = []
        if bold and italic:
            suffixes = ["BoldItalic", "Bold-Italic", "bolditalic", "bi"]
        elif bold:
            suffixes = ["Bold", "bold", "b"]
        elif italic:
            suffixes = ["Italic", "italic", "i"]
        else:
            suffixes = ["Regular", "regular", ""]
        search_dirs = []
        if os.name == "nt":
            search_dirs = [pathlib.Path(os.environ.get("WINDIR", "C:/Windows")) / "Fonts"]
        else:
            search_dirs = [pathlib.Path("/usr/share/fonts"), pathlib.Path("/usr/local/share/fonts"),
                           pathlib.Path.home() / ".fonts"]
        for d in search_dirs:
            if not d.exists():
                continue
            for suf in suffixes:
                for ext in (".ttf", ".otf"):
                    candidates = [
                        d / f"{family}{suf}{ext}",
                        d / f"{family}-{suf}{ext}",
                        d / f"{family}_{suf}{ext}",
                        d / f"{family}{ext}",
                    ]
                    for p in candidates:
                        if p.exists():
                            try:
                                return _truetype_font_cached(str(p.resolve()), size=size)
                            except Exception:
                                pass
    # 3. Fallback default
    try:
        return _dejavu_font_cached(size=size)
    except Exception:
        return ImageFont.load_default()

def get_font(settings: AppSettings) -> ImageFont.FreeTypeFont:
    # Try user font path, else a reasonable default.
    return get_font_with_size(settings, max(8, int(settings.font_size)))

def clamp_int(x: float) -> int:
    return int(max(0, min(2**31 - 1, x)))

def safe_int(x: Any, default: int = 0) -> int:
    try:
        return int(x)
    except Exception:
        return default


def _preview_stroke_width(base_width: int, zoom: float) -> int:
    try:
        bw = max(1, int(base_width))
    except Exception:
        bw = 1
    try:
        z = float(zoom)
    except Exception:
        z = 1.0
    if not np.isfinite(z) or z <= 0:
        z = 1.0
    return max(1, int(round(float(bw) * float(z))))


def _preview_arrowshape(base_width: int, zoom: float) -> Tuple[int, int, int]:
    w = _preview_stroke_width(base_width, zoom)
    h1 = max(8, int(round(3.2 * float(w))))
    h2 = max(10, int(round(4.2 * float(w))))
    h3 = max(4, int(round(1.4 * float(w))))
    return (h1, h2, h3)


def _lane_circle_colors(settings: AppSettings) -> Tuple[str, str]:
    base = _safe_hex_color(getattr(settings, "theme_color", "#5B8DB8"))
    outline = _mix_hex(base, "#000000", 0.18)
    text = _mix_hex(base, "#000000", 0.35)
    return outline, text


def _asterisk_segments(cx: float, cy: float, r: float) -> List[Tuple[Tuple[float, float], Tuple[float, float]]]:
    # 3-line asterisk (6 arms) for clearer visual marking at small sizes.
    rr = max(3.0, float(r))
    segs: List[Tuple[Tuple[float, float], Tuple[float, float]]] = []
    for ang_deg in (0.0, 60.0, 120.0):
        a = math.radians(float(ang_deg))
        dx = rr * math.cos(a)
        dy = rr * math.sin(a)
        segs.append(((cx - dx, cy - dy), (cx + dx, cy + dy)))
    return segs


def _draw_canvas_asterisk(
    canvas: tk.Canvas,
    cx: float,
    cy: float,
    r: float,
    color: str,
    width: int,
    tags: Tuple[str, ...] = ("overlay",),
) -> List[int]:
    ids: List[int] = []
    for (x0, y0), (x1, y1) in _asterisk_segments(cx, cy, r):
        ids.append(canvas.create_line(float(x0), float(y0), float(x1), float(y1), fill=str(color), width=max(1, int(width)), tags=tags))
    return ids


def _draw_canvas_arrow(
    canvas: tk.Canvas,
    x0: float,
    y0: float,
    x1: float,
    y1: float,
    color: str,
    width: int,
    tags: Tuple[str, ...] = ("overlay",),
) -> List[int]:
    ids: List[int] = []
    w = max(1, int(width))
    ids.append(canvas.create_line(float(x0), float(y0), float(x1), float(y1), fill=str(color), width=w, tags=tags))
    dx = float(x1) - float(x0)
    dy = float(y1) - float(y0)
    if abs(dx) + abs(dy) < 1e-6:
        return ids
    ang = math.atan2(dy, dx)
    head_len = max(8.0, 3.4 * float(w))
    head_half = max(5.0, 2.3 * float(w))
    bx = float(x1) - head_len * math.cos(ang)
    by = float(y1) - head_len * math.sin(ang)
    lx = bx + head_half * math.cos(ang + math.pi / 2.0)
    ly = by + head_half * math.sin(ang + math.pi / 2.0)
    rx = bx + head_half * math.cos(ang - math.pi / 2.0)
    ry = by + head_half * math.sin(ang - math.pi / 2.0)
    ids.append(
        canvas.create_polygon(
            float(x1),
            float(y1),
            float(lx),
            float(ly),
            float(rx),
            float(ry),
            fill=str(color),
            outline=str(color),
            width=max(1, int(round(0.6 * float(w)))),
            tags=tags,
        )
    )
    return ids


def _new_annotation_id(prefix: str) -> str:
    return f"{prefix}{uuid.uuid4().hex[:10]}"


def _bracket_group_token(group_id: str) -> str:
    return f"BG:{group_id}"


def _normalize_top_annotations(pc: PanelConfig) -> None:
    """Ensure HeaderRow/GroupLabel IDs and top annotation order are valid."""
    used: set[str] = set()

    for hr in pc.header_rows:
        hid = str(getattr(hr, "id", "")).strip()
        if (not hid) or (hid in used):
            hid = _new_annotation_id("H")
            hr.id = hid
        used.add(hid)
        pos = str(getattr(hr, "position", "top")).strip().lower()
        hr.position = "bottom" if pos == "bottom" else "top"
        try:
            hr.angle_deg = max(-80, min(80, int(getattr(hr, "angle_deg", getattr(pc, "header_value_angle_deg", 0)))))
        except Exception:
            hr.angle_deg = int(getattr(pc, "header_value_angle_deg", 0))

    for g in pc.group_labels:
        gid = str(getattr(g, "id", "")).strip()
        if (not gid) or (gid in used):
            gid = _new_annotation_id("G")
            g.id = gid
        used.add(gid)
        hgid = str(getattr(g, "height_group", "")).strip()
        if not hgid:
            hgid = _new_annotation_id("BG")
            g.height_group = hgid

    header_ids = [hr.id for hr in pc.header_rows if str(getattr(hr, "position", "top")).strip().lower() != "bottom"]
    group_by_id = {g.id: g for g in pc.group_labels}
    height_groups: List[str] = []
    for g in pc.group_labels:
        hgid = str(g.height_group)
        if hgid not in height_groups:
            height_groups.append(hgid)
    bracket_tokens = [_bracket_group_token(hgid) for hgid in height_groups]
    valid_set = set(header_ids) | set(bracket_tokens)

    order = [str(x) for x in (getattr(pc, "top_annotation_order", []) or [])]
    normalized: List[str] = []
    for oid in order:
        tok = oid
        # Backward compatibility: old saves may store bracket IDs directly.
        if oid in group_by_id:
            tok = _bracket_group_token(str(group_by_id[oid].height_group))
        elif oid in height_groups:
            tok = _bracket_group_token(oid)
        if tok in valid_set and tok not in normalized:
            normalized.append(tok)

    # Keep user order stable; only append missing items.
    order = list(normalized)
    for oid in header_ids:
        if oid not in order:
            order.append(oid)
    for tok in bracket_tokens:
        if tok not in order:
            order.append(tok)
    pc.top_annotation_order = order



def get_canvas_dims(canvas: tk.Canvas, min_dim: int = 250) -> Tuple[int, int]:
    """Return a sane (width, height) for a canvas, even right after frame switches."""
    try:
        canvas.update_idletasks()
    except Exception:
        pass
    w = canvas.winfo_width()
    h = canvas.winfo_height()
    # On some platforms, winfo_* can be 1 right after tkraise(); clamp to a useful minimum.
    w = max(min_dim, int(w) if w else min_dim)
    h = max(min_dim, int(h) if h else min_dim)
    return w, h


def fit_toplevel_to_screen(win: tk.Toplevel, pref_w: int, pref_h: int, margin: int = 80) -> None:
    """Apply a centered geometry that always fits the current screen."""
    try:
        sw = int(win.winfo_screenwidth())
        sh = int(win.winfo_screenheight())
    except Exception:
        sw, sh = 1920, 1080
    tw = min(int(pref_w), max(360, int(sw - margin)))
    th = min(int(pref_h), max(280, int(sh - margin)))
    x = max(0, int((sw - tw) / 2))
    y = max(0, int((sh - th) / 2))
    try:
        win.geometry(f"{int(tw)}x{int(th)}+{int(x)}+{int(y)}")
    except Exception:
        pass


class ZoomableCanvas(ttk.Frame):
    """
    A scrollable, zoomable canvas for displaying a PIL image.

    - Mouse wheel: zoom (around cursor)
    - Right-drag (or middle-drag): pan
    - Buttons: Fit / +/- zoom
    """
    def __init__(self, parent, bg: str = "#222222", show_tools: bool = True):
        super().__init__(parent)

        self._base_pil: Optional[Image.Image] = None
        self._tk_img: Optional[ImageTk.PhotoImage] = None
        self._img_id: Optional[int] = None

        self.zoom: float = 1.0
        self.user_zoomed: bool = False  # if False, we auto-fit on resize

        self._rendering = False  # guard against recursive configure events
        self._interact_after_id: Optional[str] = None
        self._is_interacting: bool = False

        if show_tools:
            tools = ttk.Frame(self)
            tools.pack(fill="x", pady=(0, 4))
            ttk.Button(tools, text="Fit", command=self.fit_to_window).pack(side="left")
            ttk.Button(tools, text="-", command=lambda: self.zoom_by(1/1.2)).pack(side="left", padx=4)
            ttk.Button(tools, text="+", command=lambda: self.zoom_by(1.2)).pack(side="left")
            self._zoom_lbl = ttk.Label(tools, text="100%  (wheel=zoom, right-drag=pan)")
            self._zoom_lbl.pack(side="left", padx=10)

        body = ttk.Frame(self)
        body.pack(fill="both", expand=True)

        self.canvas = tk.Canvas(body, bg=bg, highlightthickness=0)
        vsb = ttk.Scrollbar(body, orient="vertical", command=self.canvas.yview)
        hsb = ttk.Scrollbar(body, orient="horizontal", command=self.canvas.xview)
        self.canvas.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        body.rowconfigure(0, weight=1)
        body.columnconfigure(0, weight=1)

        self.canvas.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        # Ensure the canvas receives wheel events when hovered
        self.canvas.bind("<Enter>", lambda _e=None: self.canvas.focus_set())

        # Zoom bindings
        self.canvas.bind("<MouseWheel>", self._on_mousewheel)
        self.canvas.bind("<Shift-MouseWheel>", self._on_shift_mousewheel)
        # Linux
        self.canvas.bind("<Button-4>", lambda e: self._on_mousewheel_linux(e, +1))
        self.canvas.bind("<Button-5>", lambda e: self._on_mousewheel_linux(e, -1))

        # Pan (right or middle)
        self.canvas.bind("<ButtonPress-3>", self._pan_start)
        self.canvas.bind("<B3-Motion>", self._pan_drag)
        self.canvas.bind("<ButtonPress-2>", self._pan_start)
        self.canvas.bind("<B2-Motion>", self._pan_drag)

        self.canvas.bind("<Configure>", self._on_configure)

    def set_image(self, pil: Image.Image, fit_if_needed: bool = True) -> None:
        self._base_pil = pil
        if fit_if_needed and not self.user_zoomed:
            # fit on next render (after geometry is known)
            self.fit_to_window()
        else:
            self.render()

    def fit_to_window(self) -> None:
        if self._base_pil is None:
            return
        cw, ch = get_canvas_dims(self.canvas)
        iw, ih = self._base_pil.size
        if iw <= 0 or ih <= 0:
            return
        # Fit so the whole image is visible
        z = min((cw) / iw, (ch) / ih)
        z = max(0.05, min(20.0, z))
        self.zoom = float(z)
        self.user_zoomed = False
        self.render(reset_view=True)

    def zoom_by(self, factor: float) -> None:
        if self._base_pil is None:
            return
        new_zoom = max(0.05, min(20.0, self.zoom * float(factor)))
        self.zoom = float(new_zoom)
        self.user_zoomed = True
        self.render()

    def canvas_to_image(self, x_win: float, y_win: float) -> Tuple[int, int]:
        """Convert window coords (event.x/y) to image pixel coords."""
        cx = self.canvas.canvasx(x_win)
        cy = self.canvas.canvasy(y_win)
        ix = cx / max(1e-9, self.zoom)
        iy = cy / max(1e-9, self.zoom)
        return clamp_int(ix), clamp_int(iy)

    def image_to_canvas(self, ix: float, iy: float) -> Tuple[float, float]:
        return (float(ix) * self.zoom, float(iy) * self.zoom)

    def _on_configure(self, _event=None):
        if self._base_pil is None or self._rendering:
            return
        if not self.user_zoomed:
            self.fit_to_window()
        else:
            self.render()

    def _on_mousewheel_linux(self, event, direction: int):
        # direction: +1 zoom in, -1 zoom out
        self._zoom_at(event.x, event.y, 1.15 if direction > 0 else 1/1.15)

    def _on_mousewheel(self, event):
        # Windows/macOS
        delta = getattr(event, "delta", 0)
        if delta == 0:
            return
        factor = 1.15 if delta > 0 else 1/1.15
        self._zoom_at(event.x, event.y, factor)

    def _on_shift_mousewheel(self, event):
        # horizontal scroll
        delta = getattr(event, "delta", 0)
        if delta == 0:
            return
        units = -1 if delta > 0 else 1
        try:
            self.canvas.xview_scroll(units * 3, "units")
        except Exception:
            pass

    def _zoom_at(self, x_win: float, y_win: float, factor: float):
        if self._base_pil is None:
            return
        old_zoom = float(self.zoom)
        new_zoom = max(0.05, min(20.0, old_zoom * float(factor)))
        if abs(new_zoom - old_zoom) < 1e-9:
            return

        # Preserve the point under the cursor
        cx = self.canvas.canvasx(x_win)
        cy = self.canvas.canvasy(y_win)
        ix = cx / old_zoom
        iy = cy / old_zoom

        self.zoom = new_zoom
        self.user_zoomed = True
        self._is_interacting = True
        if self._interact_after_id:
            try:
                self.after_cancel(self._interact_after_id)
            except Exception:
                pass
        def _finish_interact():
            self._is_interacting = False
            self.render()
        self._interact_after_id = self.after(180, _finish_interact)
        self.render()

        disp_w = max(1, int(self._base_pil.width * self.zoom))
        disp_h = max(1, int(self._base_pil.height * self.zoom))

        ncx = ix * self.zoom
        ncy = iy * self.zoom

        # Move viewport so that ncx/ncy lands under x_win/y_win
        fx = (ncx - x_win) / disp_w
        fy = (ncy - y_win) / disp_h
        fx = max(0.0, min(1.0, fx))
        fy = max(0.0, min(1.0, fy))
        try:
            self.canvas.xview_moveto(fx)
            self.canvas.yview_moveto(fy)
        except Exception:
            pass

    def _pan_start(self, event):
        self.canvas.scan_mark(event.x, event.y)

    def _pan_drag(self, event):
        self.canvas.scan_dragto(event.x, event.y, gain=1)

    def render(self, reset_view: bool = False):
        if self._base_pil is None or self._rendering:
            return
        self._rendering = True
        try:
            iw, ih = self._base_pil.size
            disp_w = max(1, int(round(iw * self.zoom)))
            disp_h = max(1, int(round(ih * self.zoom)))

            resample = Image.Resampling.BILINEAR if self._is_interacting else Image.Resampling.LANCZOS
            disp = self._base_pil.resize((disp_w, disp_h), resample)
            self._tk_img = ImageTk.PhotoImage(disp)

            self.canvas.delete("all")
            self._img_id = self.canvas.create_image(0, 0, image=self._tk_img, anchor="nw")
            self.canvas.configure(scrollregion=(0, 0, disp_w, disp_h))

            if show_tools := hasattr(self, "_zoom_lbl"):
                self._zoom_lbl.configure(text=f"{int(self.zoom*100)}%  (wheel=zoom, right-drag=pan)")

            if reset_view:
                try:
                    self.canvas.xview_moveto(0.0)
                    self.canvas.yview_moveto(0.0)
                except Exception:
                    pass
            try:
                self.canvas.event_generate("<<ZoomableCanvasRendered>>", when="tail")
            except Exception:
                pass
        finally:
            self._rendering = False

    def screen_to_image(self, sx: int, sy: int) -> Tuple[Optional[float], Optional[float]]:
        """Convert canvas screen coordinates to original image pixel coordinates."""
        try:
            if self._base_pil is None:
                return None, None
            # Account for canvas scroll position
            cx = self.canvas.canvasx(sx)
            cy = self.canvas.canvasy(sy)
            ix = float(cx) / float(self.zoom)
            iy = float(cy) / float(self.zoom)
            return ix, iy
        except Exception:
            return None, None

class VScrollPanel(ttk.Frame):
    """Simple vertically scrollable container for dense control sidebars."""
    _instances: List["weakref.ReferenceType[VScrollPanel]"] = []
    _global_bound: bool = False

    def __init__(self, parent, width: int = 420):
        super().__init__(parent)
        self.canvas = tk.Canvas(self, highlightthickness=0, borderwidth=0, width=int(width))
        self.vsb = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=self.vsb.set)

        self.canvas.pack(side="left", fill="both", expand=True)
        self.vsb.pack(side="right", fill="y")

        self.inner = ttk.Frame(self.canvas)
        self._inner_window = self.canvas.create_window((0, 0), window=self.inner, anchor="nw")

        self.inner.bind("<Configure>", self._on_inner_configure)
        self.canvas.bind("<Configure>", self._on_canvas_configure)
        self.canvas.bind("<MouseWheel>", self._on_mousewheel)
        self.canvas.bind("<Button-4>", lambda _e=None: self._scroll_units(-1))
        self.canvas.bind("<Button-5>", lambda _e=None: self._scroll_units(1))

        VScrollPanel._instances.append(weakref.ref(self))
        if not VScrollPanel._global_bound:
            self.bind_all("<MouseWheel>", VScrollPanel._on_mousewheel_global_dispatch, add="+")
            self.bind_all("<Button-4>", VScrollPanel._on_mousewheel_up_global_dispatch, add="+")
            self.bind_all("<Button-5>", VScrollPanel._on_mousewheel_down_global_dispatch, add="+")
            VScrollPanel._global_bound = True

    def _on_inner_configure(self, _event=None):
        try:
            self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        except Exception:
            pass

    def _on_canvas_configure(self, event):
        try:
            self.canvas.itemconfigure(self._inner_window, width=int(event.width))
        except Exception:
            pass

    def _on_mousewheel(self, event):
        delta = getattr(event, "delta", 0)
        if delta == 0:
            return
        units = -1 if delta > 0 else 1
        self._scroll_units(units)

    def _scroll_units(self, units: int):
        try:
            self.canvas.yview_scroll(int(units) * 3, "units")
        except Exception:
            pass

    def _contains_widget(self, widget: Optional[tk.Widget]) -> bool:
        w = widget
        while w is not None:
            if w is self or w is self.canvas or w is self.inner:
                return True
            try:
                w = w.master  # type: ignore[assignment]
            except Exception:
                break
        return False

    def _widget_under_pointer(self, event) -> Optional[tk.Widget]:
        try:
            return self.winfo_containing(event.x_root, event.y_root)
        except Exception:
            return None

    @classmethod
    def _alive_instances(cls) -> List["VScrollPanel"]:
        alive: List["VScrollPanel"] = []
        new_refs: List["weakref.ReferenceType[VScrollPanel]"] = []
        for r in list(cls._instances):
            inst = r()
            if inst is None:
                continue
            alive.append(inst)
            new_refs.append(r)
        cls._instances = new_refs
        return alive

    @classmethod
    def _on_mousewheel_global_dispatch(cls, event):
        for inst in cls._alive_instances():
            if inst._contains_widget(inst._widget_under_pointer(event)):
                inst._on_mousewheel(event)
                break

    @classmethod
    def _on_mousewheel_up_global_dispatch(cls, event):
        for inst in cls._alive_instances():
            if inst._contains_widget(inst._widget_under_pointer(event)):
                inst._scroll_units(-1)
                break

    @classmethod
    def _on_mousewheel_down_global_dispatch(cls, event):
        for inst in cls._alive_instances():
            if inst._contains_widget(inst._widget_under_pointer(event)):
                inst._scroll_units(1)
                break

class PanelAwareFrame(ttk.Frame):
    """Base class for frames that operate on the current panel."""

    def current_panel(self) -> "Image.Image":
        return self.app.state.panels[self.app.state.current_panel_index]

    def current_pc(self) -> "PanelConfig":
        return self.app.state.panel_configs[self.app.state.current_panel_index]

    def _has_current_panel(self) -> bool:
        panels = self.app.state.panels
        idx = int(self.app.state.current_panel_index)
        return bool(panels) and (0 <= idx < len(panels))

    def _guard_panel(self) -> bool:
        """Return True if a panel is available; navigate home if not."""
        if not self.app.state.panels:
            self.app.show_frame("SettingsFrame")
            return False
        return True


# ----------------------------
# Main App
# ----------------------------
class GelAnnotatorApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self._apply_ui_scale(scale_factor=self._adaptive_ui_scale())
        self.title("Gel & Blot Annotator (MVP)")

        # Fullscreen-friendly and responsive window behavior.
        # 1) Start maximized when possible.
        # 2) Ensure the stacked frames expand with the window.
        try:
            sw0 = int(self.winfo_screenwidth())
            sh0 = int(self.winfo_screenheight())
        except Exception:
            sw0, sh0 = 1920, 1080
        self.minsize(max(780, int(0.60 * sw0)), max(560, int(0.60 * sh0)))
        try:
            # Works on Windows; on some Linux WMs too.
            tk.Tk.state(self, "zoomed")
        except Exception:
            try:
                # Some Tk builds support this attribute (Linux).
                self.attributes("-zoomed", True)
            except Exception:
                # Fallback: use the full screen geometry.
                sw = self.winfo_screenwidth()
                sh = self.winfo_screenheight()
                self.geometry(f"{sw}x{sh}+0+0")

        self.state = ProjectState(settings=load_settings())
        ensure_default_markers()

        self._steps = [
            ("Settings", "SettingsFrame"),
            ("Load", "LoadImageFrame"),
            ("Compose", "ComposeFrame"),
            ("Edit", "EditFrame"),
            ("Panels", "PanelSelectFrame"),
            ("Layout", "LayoutFrame"),
            ("Highlight", "HighlightFrame"),
            ("Annotate", "AnnotationFrame"),
            ("Marker", "MarkerFrame"),
            ("Analysis", "AnalysisFrame"),
            ("Review", "ReviewFrame"),
        ]
        crumb_bar = tk.Frame(self, bg="#dde3ea", height=36)
        crumb_bar.pack(fill="x", side="top")
        crumb_bar.pack_propagate(False)
        # Thin bottom border line
        tk.Frame(crumb_bar, bg="#b0bcc8", height=1).pack(side="bottom", fill="x")
        self._crumb_labels: Dict[str, tk.Label] = {}
        for i, (step_name, frame_name) in enumerate(self._steps):
            # Separator chevron between steps
            if i > 0:
                tk.Label(crumb_bar, text="›", font=("Arial", 10),
                         bg="#dde3ea", fg="#a0aab5", pady=0).pack(side="left", padx=0)
            lbl = tk.Label(
                crumb_bar,
                text=f" {step_name} ",
                font=("Arial", 9),
                bg="#dde3ea",
                fg="#7a8a9a",
                padx=4,
                pady=6,
                cursor="hand2",
            )
            lbl.pack(side="left")
            lbl.bind("<Button-1>", lambda _e=None, fn=frame_name: self._crumb_click(fn))
            self._crumb_labels[frame_name] = lbl

        container = ttk.Frame(self)
        container.pack(fill="both", expand=True)
        container.rowconfigure(0, weight=1)
        container.columnconfigure(0, weight=1)

        # Optional fullscreen toggle (F11)
        self._is_fullscreen = False
        self.bind('<F11>', self._toggle_fullscreen)
        self.bind('<Escape>', self._exit_fullscreen)
        self.bind('<Control-s>', lambda _e=None: self._kb_save())
        self.bind('<Control-z>', lambda _e=None: self._kb_undo())
        self.bind('<Return>', lambda _e=None: self._kb_next())
        self.bind('<Control-Return>', lambda _e=None: self._kb_next())

        self.frames: Dict[str, ttk.Frame] = {}
        for F in (SettingsFrame, MultiSourcePanelSelectFrame, ComposeFrame, EditFrame, PanelSelectFrame,
                  LayoutFrame, HighlightFrame, AnnotationFrame, MarkerFrame, AnalysisFrame, ReviewFrame):
            frame = F(container, self)
            self.frames[F.__name__] = frame
            frame.grid(row=0, column=0, sticky="nsew")

        self._theme_applied_color = ""
        self.apply_color_theme()

        self._bug_footer = tk.Label(
            self,
            text=APP_BUG_EMAIL,
            font=("Arial", 9),
            fg="#666666",
            bg="#f5f5f5",
            padx=2,
            pady=1,
        )
        self._place_bug_footer()
        self._status_bar = tk.Label(
            self,
            text="",
            font=("Arial", 9, "italic"),
            fg="#2E6EA0",
            bg="#eaf2fb",
            anchor="w",
            padx=8,
            pady=2,
            relief="flat",
        )
        self._place_status_bar()
        self.bind("<Configure>", lambda _e=None: (self._place_bug_footer(), self._place_status_bar()))
        self._spinner_chars = ["|", "/", "—", "\\"]
        self._spinner_idx = 0
        self._spinner_after_id: Optional[str] = None

        self._current_frame_name: Optional[str] = None
        self._show_frame_nonce: int = 0
        self.show_frame("SettingsFrame")

    def _adaptive_ui_scale(self) -> float:
        """Choose a UI scale that stays usable across small and large displays."""
        try:
            sw = float(self.winfo_screenwidth())
            sh = float(self.winfo_screenheight())
        except Exception:
            sw, sh = 1920.0, 1080.0
        rel = min(sw / 1920.0, sh / 1080.0)
        # Keep controls readable but avoid oversized UI on small screens.
        return max(0.90, min(1.20, float(rel)))

    def _scaled_sidebar_width(self, base_width: int) -> int:
        """Scale right-side control panel widths with screen size."""
        try:
            sw = float(self.winfo_screenwidth())
        except Exception:
            sw = 1920.0
        rel = max(0.78, min(1.25, sw / 1920.0))
        return int(max(300, min(760, round(float(base_width) * rel))))

    def place_floating_next(self, btn: tk.Widget, x_pad: int = 14, y_pad: int = 12) -> None:
        """Place floating next-buttons above the bug footer and inside visible bounds."""
        try:
            footer_h = int(self._bug_footer.winfo_reqheight()) + 4
        except Exception:
            footer_h = 14
        try:
            btn.place(relx=1.0, rely=1.0, anchor="se", x=-int(x_pad), y=-(int(y_pad) + int(footer_h)))
            btn.lift()
        except Exception:
            pass
    def _apply_ui_scale(self, scale_factor: float = 1.2):
        """Increase overall UI size (fonts + ttk paddings) consistently."""
        sf = max(0.8, min(2.0, float(scale_factor)))
        try:
            cur = float(self.tk.call("tk", "scaling"))
            self.tk.call("tk", "scaling", cur * sf)
        except Exception:
            pass

        # Use "clam" theme — it exposes relief/borderwidth on buttons
        # and gives us much more control over colour and 3-D appearance.
        try:
            style = ttk.Style(self)
            if "clam" in style.theme_names():
                style.theme_use("clam")
        except Exception:
            pass

        # Scale named Tk fonts used by Tk and ttk widgets.
        for fname in ("TkDefaultFont", "TkTextFont", "TkMenuFont", "TkHeadingFont", "TkCaptionFont", "TkSmallCaptionFont"):
            try:
                f = tkfont.nametofont(fname)
                size = int(f.cget("size"))
                if size > 0:
                    f.configure(size=max(9, int(round(size * sf))))
            except Exception:
                pass

        # Generous paddings — bigger hit targets, plumper controls.
        try:
            style = ttk.Style(self)
            style.configure(".", padding=int(round(3 * sf)))
            style.configure("TButton",
                            padding=(int(round(12 * sf)), int(round(7 * sf))),
                            relief="raised", borderwidth=2)
            style.configure("TEntry",   padding=int(round(4 * sf)))
            style.configure("TCombobox", padding=int(round(4 * sf)))
            style.configure("TSpinbox",  padding=int(round(4 * sf)))
            style.configure("TNotebook.Tab",
                            padding=(int(round(14 * sf)), int(round(7 * sf))))
        except Exception:
            pass

    def show_frame(self, name: str):
        self._show_frame_nonce = int(getattr(self, "_show_frame_nonce", 0)) + 1
        nonce = int(self._show_frame_nonce)
        prev_name = str(getattr(self, "_current_frame_name", "") or "")
        if prev_name and prev_name in self.frames and prev_name != str(name):
            prev = self.frames[prev_name]
            if hasattr(prev, "on_hide"):
                try:
                    prev.on_hide()
                except Exception:
                    pass
        frame = self.frames[name]
        if hasattr(frame, "on_show"):
            frame.on_show()
        # If on_show() triggered another show_frame() call, abort this stale activation.
        if nonce != int(getattr(self, "_show_frame_nonce", 0)):
            return
        self.apply_color_theme()
        frame.tkraise()
        self._current_frame_name = str(name)
        self._place_bug_footer()
        try:
            self._bug_footer.lift()
        except Exception:
            pass
        try:
            self._update_breadcrumb(str(name))
        except Exception:
            pass

    def _place_bug_footer(self):
        try:
            self._bug_footer.place(relx=1.0, rely=1.0, anchor="se", x=-6, y=-4)
        except Exception:
            pass

    def _crumb_click(self, frame_name: str) -> None:
        # Only allow navigating to frames that have been visited (have a panel loaded).
        if frame_name in ("SettingsFrame", "LoadImageFrame"):
            self.show_frame(frame_name)
        elif self.state.panels:
            self.show_frame(frame_name)

    def _update_breadcrumb(self, active_frame_name: str) -> None:
        theme = _safe_hex_color(getattr(self.state.settings, "theme_color", "#5B8DB8"))
        active_fg = _contrast_text_for_bg(theme)
        visited_bg = _mix_hex(theme, "#dde3ea", 0.78)
        step_names = [fn for _, fn in self._steps]
        active_idx = step_names.index(active_frame_name) if active_frame_name in step_names else -1
        for i, (_, fn) in enumerate(self._steps):
            lbl = self._crumb_labels.get(fn)
            if lbl is None:
                continue
            if fn == active_frame_name:
                # Active step: solid theme colour, bold, slightly larger
                lbl.configure(bg=theme, fg=active_fg, font=("Arial", 10, "bold"),
                              relief="flat", padx=6, pady=6)
            elif i < active_idx:
                # Visited step: tinted background, regular weight
                lbl.configure(bg=visited_bg, fg=_mix_hex(theme, "#111111", 0.35),
                              font=("Arial", 9), relief="flat", padx=4, pady=6)
            else:
                # Future step: muted
                lbl.configure(bg="#dde3ea", fg="#9aA8b5",
                              font=("Arial", 9), relief="flat", padx=4, pady=6)

    def set_status(self, msg: str) -> None:
        try:
            self._status_bar.configure(text=str(msg))
            if msg:
                self._status_bar.lift()
        except Exception:
            pass

    def clear_status(self) -> None:
        self.set_status("")

    def _place_status_bar(self) -> None:
        try:
            footer_h = int(self._bug_footer.winfo_reqheight()) + 4
            self._status_bar.place(relx=0.0, rely=1.0, anchor="sw", x=6, y=-(int(footer_h)))
            self._status_bar.lift()
        except Exception:
            pass

    def start_spinner(self, base_msg: str = "Working") -> None:
        self._stop_spinner_flag = False
        def _tick():
            if getattr(self, "_stop_spinner_flag", True):
                return
            ch = self._spinner_chars[self._spinner_idx % len(self._spinner_chars)]
            self._spinner_idx += 1
            self.set_status(f"{base_msg} {ch}")
            self._spinner_after_id = self.after(150, _tick)
        _tick()

    def stop_spinner(self, final_msg: str = "") -> None:
        self._stop_spinner_flag = True
        if self._spinner_after_id:
            try:
                self.after_cancel(self._spinner_after_id)
            except Exception:
                pass
        self.set_status(final_msg)

    def _walk_widgets(self, root: tk.Misc):
        yield root
        try:
            kids = list(root.winfo_children())
        except Exception:
            kids = []
        for child in kids:
            yield from self._walk_widgets(child)

    def _apply_theme_to_tk_widgets(self, root: tk.Misc, palette: Dict[str, str]) -> None:
        for w in self._walk_widgets(root):
            try:
                if isinstance(w, tk.Canvas):
                    # Keep viewer canvases and scroll canvases on their own colors.
                    continue
                if isinstance(w, tk.Listbox):
                    w.configure(
                        bg=palette["field"],
                        fg=palette["fg"],
                        selectbackground=palette["accent"],
                        selectforeground=palette["accent_fg"],
                        highlightbackground=palette["border"],
                    )
                    continue
                if isinstance(w, (tk.Entry, tk.Spinbox, tk.Text)):
                    cfg = dict(bg=palette["field"], fg=palette["fg"], insertbackground=palette["fg"])
                    try:
                        cfg["disabledbackground"] = palette["panel"]
                    except Exception:
                        pass
                    w.configure(**cfg)
                    continue
                if isinstance(w, tk.Button):
                    w.configure(
                        bg=palette["accent_soft"],
                        fg=palette["fg"],
                        activebackground=palette["accent"],
                        activeforeground=palette["accent_fg"],
                        highlightbackground=palette["border"],
                    )
                    continue
                if isinstance(w, (tk.Checkbutton, tk.Radiobutton)):
                    w.configure(
                        bg=palette["panel"],
                        fg=palette["fg"],
                        activebackground=palette["panel2"],
                        activeforeground=palette["fg"],
                        selectcolor=palette["field"],
                        highlightbackground=palette["border"],
                    )
                    continue
                if isinstance(w, (tk.Frame, tk.LabelFrame, tk.Toplevel)):
                    w.configure(bg=palette["panel"])
                    continue
                if isinstance(w, tk.Label):
                    w.configure(bg=palette["panel"], fg=palette["fg"])
                    continue
            except Exception:
                continue

    def apply_color_theme(self) -> None:
        base = _safe_hex_color(getattr(self.state.settings, "theme_color", "#5B8DB8"))
        if base == getattr(self, "_theme_applied_color", ""):
            # Theme already applied; skip the expensive widget walk.
            return
        accent = _mix_hex(base, "#000000", 0.08)
        accent_soft = _mix_hex(base, "#FFFFFF", 0.55)
        bg = _mix_hex(base, "#FFFFFF", 0.88)
        panel = _mix_hex(base, "#FFFFFF", 0.82)
        panel2 = _mix_hex(base, "#FFFFFF", 0.74)
        field = _mix_hex(base, "#FFFFFF", 0.94)
        border = _mix_hex(base, "#000000", 0.35)
        fg = "#111111"
        accent_fg = _contrast_text_for_bg(accent)
        palette = {
            "base": base,
            "accent": accent,
            "accent_soft": accent_soft,
            "bg": bg,
            "panel": panel,
            "panel2": panel2,
            "field": field,
            "border": border,
            "fg": fg,
            "accent_fg": accent_fg,
        }
        try:
            self.configure(bg=bg)
        except Exception:
            pass
        try:
            style = ttk.Style(self)
            style.configure(".", background=panel, foreground=fg)
            style.configure("TFrame", background=panel)
            style.configure("TLabelframe", background=panel, bordercolor=border,
                            relief="groove", borderwidth=2)
            style.configure("TLabelframe.Label", background=panel, foreground=base,
                            font=("Arial", 9, "bold"))
            style.configure("TLabel", background=panel, foreground=fg)

            # Regular button — raised relief for 3-D look, accent-soft fill
            btn_fg = _contrast_text_for_bg(accent_soft) if _contrast_text_for_bg(accent_soft) != fg else fg
            style.configure("TButton",
                            background=accent_soft, foreground=fg,
                            relief="raised", borderwidth=2,
                            font=("Arial", 9),
                            padding=(12, 7))
            style.map("TButton",
                      background=[("active",  accent),
                                  ("pressed", _mix_hex(accent, "#000000", 0.15))],
                      foreground=[("active",  _contrast_text_for_bg(accent)),
                                  ("pressed", _contrast_text_for_bg(accent))],
                      relief=[("pressed", "sunken")])

            style.configure("TCheckbutton", background=panel, foreground=fg)
            style.configure("TRadiobutton", background=panel, foreground=fg)
            style.configure("TMenubutton",  background=accent_soft, foreground=fg,
                            relief="raised", borderwidth=2)

            style.configure("TEntry",    fieldbackground=field, foreground=fg,
                            bordercolor=border, lightcolor=field, darkcolor=border)
            style.configure("TSpinbox",  fieldbackground=field, foreground=fg,
                            bordercolor=border)
            style.configure("TCombobox", fieldbackground=field, foreground=fg,
                            bordercolor=border)
            style.map("TCombobox",
                      fieldbackground=[("readonly", field)],
                      background=[("readonly", panel2)])

            style.configure("Treeview", background=field, fieldbackground=field, foreground=fg,
                            rowheight=24)
            style.configure("Treeview.Heading", background=panel2, foreground=base,
                            font=("Arial", 9, "bold"), relief="raised")
            style.map("Treeview.Heading",
                      background=[("active", accent_soft)])

            style.configure("TNotebook", background=panel, tabmargins=[2, 5, 2, 0])
            style.configure("TNotebook.Tab", background=panel2, foreground=fg,
                            padding=(14, 7), font=("Arial", 9))
            style.map("TNotebook.Tab",
                      background=[("selected", base), ("active", accent_soft)],
                      foreground=[("selected", _contrast_text_for_bg(base))],
                      font=[("selected", ("Arial", 9, "bold"))])

            style.configure("TSeparator", background=border)

            # ── Primary action button (filled solid accent) ──────────────
            primary_fg = _contrast_text_for_bg(base)
            style.configure("Primary.TButton",
                            background=base, foreground=primary_fg,
                            font=("Arial", 10, "bold"),
                            padding=(16, 9),
                            relief="raised", borderwidth=2)
            style.map("Primary.TButton",
                      background=[("active",  accent),
                                  ("pressed", _mix_hex(accent, "#000000", 0.18))],
                      foreground=[("active",  _contrast_text_for_bg(accent)),
                                  ("pressed", _contrast_text_for_bg(accent))],
                      relief=[("pressed", "sunken")])

            # ── Success button (green, for save/export actions) ──────────
            success = "#2E7D55"
            success_hover = _mix_hex(success, "#000000", 0.14)
            style.configure("Success.TButton",
                            background=success, foreground="#ffffff",
                            font=("Arial", 9, "bold"),
                            padding=(12, 7),
                            relief="raised", borderwidth=2)
            style.map("Success.TButton",
                      background=[("active",  success_hover),
                                  ("pressed", _mix_hex(success, "#000000", 0.25))],
                      foreground=[("active",  "#ffffff"), ("pressed", "#ffffff")],
                      relief=[("pressed", "sunken")])

            # ── Danger button (red, for destructive actions) ─────────────
            danger = "#C0392B"
            style.configure("Danger.TButton",
                            background=danger, foreground="#ffffff",
                            padding=(12, 7),
                            relief="raised", borderwidth=2)
            style.map("Danger.TButton",
                      background=[("active",  _mix_hex(danger, "#000000", 0.15)),
                                  ("pressed", _mix_hex(danger, "#000000", 0.25))],
                      foreground=[("active",  "#ffffff"), ("pressed", "#ffffff")],
                      relief=[("pressed", "sunken")])

            # ── Scrollbar ────────────────────────────────────────────────
            style.configure("TScrollbar", background=panel2, troughcolor=field,
                            bordercolor=border, arrowcolor=base)
        except Exception:
            pass
        try:
            self.option_add("*Listbox.Background", field)
            self.option_add("*Listbox.Foreground", fg)
            self.option_add("*Listbox.selectBackground", accent)
            self.option_add("*Listbox.selectForeground", accent_fg)
        except Exception:
            pass
        try:
            self._apply_theme_to_tk_widgets(self, palette)
        except Exception:
            pass
        # Re-skin non-ttk chrome widgets that aren't in the widget tree walk
        try:
            self._bug_footer.configure(fg=_mix_hex(base, "#666666", 0.5), bg=bg)
        except Exception:
            pass
        try:
            self._status_bar.configure(fg=base, bg=_mix_hex(base, "#ffffff", 0.88))
        except Exception:
            pass
        # Re-skin breadcrumb bar background
        try:
            crumb_bg = _mix_hex(base, "#ffffff", 0.84)
            for widget in self.winfo_children():
                if isinstance(widget, tk.Frame) and int(widget.cget("height") or 0) in (36, 28):
                    widget.configure(bg=crumb_bg)
                    for child in widget.winfo_children():
                        if isinstance(child, tk.Label):
                            if child not in self._crumb_labels.values():
                                child.configure(bg=crumb_bg)
                        elif isinstance(child, tk.Frame):
                            child.configure(bg=_mix_hex(base, "#aaaaaa", 0.35))
        except Exception:
            pass
        self._theme_applied_color = base

    def _toggle_fullscreen(self, event=None):
        """Toggle true fullscreen (F11). Esc exits fullscreen."""
        self._is_fullscreen = not getattr(self, '_is_fullscreen', False)
        try:
            self.attributes('-fullscreen', self._is_fullscreen)
        except Exception:
            # Fallback: try maximize
            if self._is_fullscreen:
                try:
                    tk.Tk.state(self, "zoomed")
                except Exception:
                    pass
        return 'break'

    def _exit_fullscreen(self, event=None):
        """Exit fullscreen (Esc)."""
        self._is_fullscreen = False
        try:
            self.attributes('-fullscreen', False)
        except Exception:
            pass
        return 'break'

    def _kb_save(self) -> None:
        frame = self.frames.get(str(self._current_frame_name or ""), None)
        if frame is not None and hasattr(frame, "save"):
            try:
                frame.save()
            except Exception as exc:
                logger.warning("Ctrl+S save failed: %s", exc)

    def _kb_undo(self) -> None:
        frame = self.frames.get(str(self._current_frame_name or ""), None)
        if frame is not None and hasattr(frame, "undo"):
            try:
                frame.undo()
            except Exception as exc:
                logger.warning("Ctrl+Z undo failed: %s", exc)

    def _kb_next(self) -> None:
        frame = self.frames.get(str(self._current_frame_name or ""), None)
        if frame is not None and hasattr(frame, "_next"):
            try:
                frame._next()
            except Exception as exc:
                logger.warning("Return _next failed: %s", exc)
        elif frame is not None and hasattr(frame, "on_next"):
            try:
                frame.on_next()
            except Exception as exc:
                logger.warning("Return on_next failed: %s", exc)

    def reset_for_next_image(self):
        # Keep settings, reset all image-dependent state.
        keep = self.state.settings.keep_settings_next_image
        settings = self.state.settings if keep else AppSettings()
        self.state = ProjectState(settings=settings)
        if keep:
            save_settings(settings)

    def ensure_panel_configs(self):
        n = len(self.state.panels)
        if len(self.state.panel_configs) != n:
            self.state.panel_configs = [PanelConfig(lanes=10) for _ in range(n)]
        # Ensure header row lengths match lanes
        for pc in self.state.panel_configs:
            for hr in pc.header_rows:
                if len(hr.values) != pc.lanes:
                    hr.values = (hr.values + [""] * pc.lanes)[:pc.lanes]
            _normalize_top_annotations(pc)

    def save_project(self) -> None:
        path = filedialog.asksaveasfilename(
            title="Save project",
            defaultextension=".gelproj",
            filetypes=[("Gel Annotator Project", "*.gelproj")],
        )
        if not path:
            return
        try:
            data: Dict[str, Any] = {
                "version": 1,
                "original_path": str(self.state.original_path or ""),
                "multi_input_paths": list(self.state.multi_input_paths or []),
                "settings": asdict(self.state.settings),
                "panel_configs": [asdict(pc) for pc in self.state.panel_configs],
                "current_panel_index": int(self.state.current_panel_index),
                "last_export_path": str(self.state.last_export_path or ""),
            }
            pathlib.Path(path).write_text(json.dumps(data, indent=2), encoding="utf-8")
            messagebox.showinfo("Project saved", f"Project saved:\n{path}")
        except Exception as exc:
            logger.warning("Project save failed: %s", exc)
            messagebox.showerror("Save failed", str(exc))

    def load_project(self) -> None:
        path = filedialog.askopenfilename(
            title="Load project",
            filetypes=[("Gel Annotator Project", "*.gelproj"), ("All files", "*.*")],
        )
        if not path:
            return
        try:
            data = json.loads(pathlib.Path(path).read_text(encoding="utf-8"))
            settings_data = dict(data.get("settings", {}))
            self.state.settings = AppSettings(**{k: v for k, v in settings_data.items() if k in AppSettings.__dataclass_fields__})
            save_settings(self.state.settings)
            img_path = str(data.get("original_path", "") or "")
            if img_path and pathlib.Path(img_path).exists():
                self.state.original_path = img_path
                self.state.original_pil = pil_from_path(img_path)
                self.state.edited_pil = self.state.original_pil.copy()
                self.state.panels = [self.state.edited_pil.copy()]
            pc_list = list(data.get("panel_configs", []) or [])
            self.state.panel_configs = []
            for pc_data in pc_list:
                try:
                    pc = PanelConfig(**{k: v for k, v in pc_data.items() if k in PanelConfig.__dataclass_fields__})
                    # Reconstruct nested dataclasses
                    if "marker_calibration" in pc_data:
                        mc_data = dict(pc_data["marker_calibration"] or {})
                        pc.marker_calibration = MarkerCalibration(**{k: v for k, v in mc_data.items() if k in MarkerCalibration.__dataclass_fields__})
                    if "header_rows" in pc_data:
                        pc.header_rows = [HeaderRow(**{k: v for k, v in hr.items() if k in HeaderRow.__dataclass_fields__}) for hr in (pc_data["header_rows"] or [])]
                    if "group_labels" in pc_data:
                        pc.group_labels = [GroupLabel(**{k: v for k, v in gl.items() if k in GroupLabel.__dataclass_fields__}) for gl in (pc_data["group_labels"] or [])]
                    if "highlights" in pc_data:
                        pc.highlights = [HighlightRect(**{k: v for k, v in h.items() if k in HighlightRect.__dataclass_fields__}) for h in (pc_data["highlights"] or [])]
                    self.state.panel_configs.append(pc)
                except Exception as exc:
                    logger.warning("Could not restore panel config: %s", exc)
            if not self.state.panel_configs and self.state.panels:
                self.state.panel_configs = [PanelConfig() for _ in self.state.panels]
            self.state.current_panel_index = int(data.get("current_panel_index", 0))
            messagebox.showinfo("Project loaded", f"Project loaded:\n{path}")
            if self.state.panels:
                self.show_frame("ReviewFrame")
            else:
                self.show_frame("SettingsFrame")
        except Exception as exc:
            logger.warning("Project load failed: %s", exc)
            messagebox.showerror("Load failed", str(exc))


# ----------------------------
# Step 1: Settings
# ----------------------------
class SettingsFrame(ttk.Frame):
    def __init__(self, parent, app: GelAnnotatorApp):
        super().__init__(parent)
        self.app = app

        ttk.Label(self, text="Step 1 - Session settings", font=("Arial", 18, "bold")).pack(pady=10)
        ttk.Label(
            self,
            text=f"Created by {APP_AUTHOR}",
            justify="center",
            font=("Arial", 15, "bold"),
        ).pack(pady=(0, 2))
        ttk.Label(
            self,
            text=f"Bug reports: {APP_BUG_EMAIL}",
            justify="center",
            font=("Arial", 12),
        ).pack(pady=(0, 8))

        form = ttk.Frame(self)
        form.pack(pady=10)

        self.gel_type = tk.StringVar(value=self.app.state.settings.gel_type)
        self.panel_mode = tk.StringVar(value=self.app.state.settings.panel_mode)
        self.source_mode = tk.StringVar(value=self.app.state.settings.source_mode)
        self.unit = tk.StringVar(value=self.app.state.settings.default_unit)
        self.keep = tk.BooleanVar(value=self.app.state.settings.keep_settings_next_image)
        self.invert = tk.BooleanVar(value=self.app.state.settings.prefer_invert)
        self.theme_color = tk.StringVar(value=getattr(self.app.state.settings, "theme_color", "#5B8DB8"))
        self._selected_paths: List[str] = []
        self.image_info = tk.StringVar(value="No image selected")

        row = 0
        ttk.Label(form, text="Experiment type:").grid(row=row, column=0, sticky="e", padx=8, pady=6)
        ttk.Combobox(form, textvariable=self.gel_type, values=["agarose", "sds_page", "western_blot"], width=20, state="readonly")\
            .grid(row=row, column=1, sticky="w", padx=8)
        row += 1

        ttk.Label(form, text="Work mode:").grid(row=row, column=0, sticky="e", padx=8, pady=6)
        ttk.Combobox(form, textvariable=self.panel_mode, values=["single", "multi_panel"], width=20, state="readonly")\
            .grid(row=row, column=1, sticky="w", padx=8)
        row += 1

        ttk.Label(form, text="Image source:").grid(row=row, column=0, sticky="e", padx=8, pady=6)
        self.source_combo = ttk.Combobox(
            form,
            textvariable=self.source_mode,
            values=["single_image", "multi_image"],
            width=20,
            state="readonly",
        )
        self.source_combo.grid(row=row, column=1, sticky="w", padx=8)
        self.source_combo.bind("<<ComboboxSelected>>", lambda _e=None: self._on_source_mode_changed())
        row += 1

        ttk.Label(form, text="Default marker unit:").grid(row=row, column=0, sticky="e", padx=8, pady=6)
        ttk.Combobox(form, textvariable=self.unit, values=["bp", "kDa"], width=20, state="readonly")\
            .grid(row=row, column=1, sticky="w", padx=8)
        row += 1

        ttk.Checkbutton(form, text="Prefer inverted images (dark bands on light background)", variable=self.invert)\
            .grid(row=row, column=1, sticky="w", padx=8, pady=6)
        row += 1

        ttk.Checkbutton(form, text="Keep these settings for the next image (batch work)", variable=self.keep)\
            .grid(row=row, column=1, sticky="w", padx=8, pady=6)
        row += 1

        ttk.Label(form, text="UI theme color:").grid(row=row, column=0, sticky="e", padx=8, pady=6)
        theme_row = ttk.Frame(form)
        theme_row.grid(row=row, column=1, sticky="w", padx=8, pady=6)
        self._theme_chip = tk.Label(theme_row, width=3, relief="solid", bg=self.theme_color.get())
        self._theme_chip.pack(side="left")
        ttk.Button(theme_row, text="Choose...", command=self._choose_theme_color).pack(side="left", padx=6)
        ttk.Label(theme_row, textvariable=self.theme_color).pack(side="left", padx=(4, 0))
        row += 1

        ttk.Separator(form).grid(row=row, column=0, columnspan=2, sticky="ew", padx=8, pady=8)
        row += 1

        ttk.Label(form, text="Font family:").grid(row=row, column=0, sticky="e", padx=8, pady=6)
        fam_row = ttk.Frame(form)
        fam_row.grid(row=row, column=1, sticky="w", padx=8, pady=6)
        _FONT_FAMILIES = [
            "(default)",
            "Arial", "Arial Narrow",
            "Calibri", "Cambria",
            "Century Gothic",
            "Consolas", "Courier New",
            "DejaVu Sans", "DejaVu Serif", "DejaVu Sans Mono",
            "Franklin Gothic Medium",
            "Garamond",
            "Georgia",
            "Helvetica",
            "Impact",
            "Palatino Linotype",
            "Segoe UI",
            "Tahoma",
            "Times New Roman",
            "Trebuchet MS",
            "Verdana",
        ]
        current_fam = str(getattr(self.app.state.settings, "font_family", "") or "")
        self.font_family = tk.StringVar(value=current_fam if current_fam in _FONT_FAMILIES else "(default)")
        ttk.Combobox(fam_row, textvariable=self.font_family, values=_FONT_FAMILIES,
                     state="readonly", width=22).pack(side="left")
        row += 1

        ttk.Label(form, text="Font style:").grid(row=row, column=0, sticky="e", padx=8, pady=6)
        style_row = ttk.Frame(form)
        style_row.grid(row=row, column=1, sticky="w", padx=8, pady=6)
        self.font_bold_var = tk.BooleanVar(value=bool(getattr(self.app.state.settings, "font_bold", False)))
        self.font_italic_var = tk.BooleanVar(value=bool(getattr(self.app.state.settings, "font_italic", False)))
        ttk.Checkbutton(style_row, text="Bold", variable=self.font_bold_var).pack(side="left", padx=4)
        ttk.Checkbutton(style_row, text="Italic", variable=self.font_italic_var).pack(side="left", padx=4)
        row += 1

        ttk.Separator(form).grid(row=row, column=0, columnspan=2, sticky="ew", padx=8, pady=8)
        row += 1

        ttk.Label(form, text="Image(s):").grid(row=row, column=0, sticky="ne", padx=8, pady=6)
        image_box = ttk.Frame(form)
        image_box.grid(row=row, column=1, sticky="w", padx=8, pady=6)
        ttk.Button(image_box, text="Select image(s)...", command=self._select_images).pack(anchor="w")
        ttk.Button(image_box, text="Add more images...", command=self._add_images).pack(anchor="w", pady=(4, 0))
        ttk.Button(image_box, text="Clear selection", command=self._clear_images).pack(anchor="w", pady=(4, 0))
        ttk.Label(image_box, textvariable=self.image_info, justify="left").pack(anchor="w", pady=(4, 0))
        row += 1

        foot = ttk.Frame(self)
        foot.pack(side="bottom", fill="x", pady=20, padx=20)

        ttk.Button(foot, text="Readme / Help", command=self._open_readme).pack(side="left")
        ttk.Button(foot, text="Next", style="Primary.TButton", command=self.on_next).pack(side="right")


        note = (
            "Notes:\n"
            "- 'multi_panel' lets you crop multiple gel/blot regions from one image OR assemble several images.\n"
            "- You can refine fonts, sizes, and marker styles later; this MVP uses safe defaults."
        )
        ttk.Label(self, text=note, justify="left").pack(padx=20, pady=10, anchor="w")

    def _open_readme(self):
        if open_readme_in_system():
            return
        messagebox.showinfo(
            "Readme not found",
            f"Could not open '{README_BASENAME}'.\n\n"
            "Place the readme file next to the script/executable and try again.",
        )

    def on_show(self):
        s = self.app.state.settings
        self.gel_type.set(s.gel_type)
        self.panel_mode.set(s.panel_mode)
        self.source_mode.set(s.source_mode)
        self.unit.set(s.default_unit)
        self.keep.set(s.keep_settings_next_image)
        self.invert.set(s.prefer_invert)
        self.theme_color.set(_safe_hex_color(getattr(s, "theme_color", "#5B8DB8")))
        try:
            self._theme_chip.configure(bg=self.theme_color.get())
        except Exception:
            pass
        self._selected_paths = [p for p in self.app.state.original_path.split(";") if p]
        self._update_image_info()

    def _update_image_info(self):
        if not self._selected_paths:
            self.image_info.set("No image selected")
        elif len(self._selected_paths) == 1:
            self.image_info.set(pathlib.Path(self._selected_paths[0]).name)
        else:
            self.image_info.set(f"{len(self._selected_paths)} images selected")

    def _select_images(self):
        mode = self.source_mode.get()
        if mode == "multi_image":
            paths = filedialog.askopenfilenames(
                title="Select images",
                filetypes=[("Images", "*.jpg *.jpeg *.png *.tif *.tiff"), ("All files", "*.*")]
            )
            if not paths:
                return
            self._selected_paths = list(paths)
            self._update_image_info()
            return

        path = filedialog.askopenfilename(
            title="Select an image",
            filetypes=[("Images", "*.jpg *.jpeg *.png *.tif *.tiff"), ("All files", "*.*")]
        )
        if not path:
            return
        self._selected_paths = [path]
        self._update_image_info()

    def _add_images(self):
        paths = filedialog.askopenfilenames(
            title="Add images",
            filetypes=[("Images", "*.jpg *.jpeg *.png *.tif *.tiff"), ("All files", "*.*")]
        )
        if not paths:
            return
        existing = set(self._selected_paths)
        for p in paths:
            if p not in existing:
                self._selected_paths.append(p)
                existing.add(p)
        if len(self._selected_paths) > 1 and self.source_mode.get() != "multi_image":
            self.source_mode.set("multi_image")
        self._update_image_info()

    def _clear_images(self):
        self._selected_paths = []
        self._update_image_info()

    def _on_source_mode_changed(self):
        # Keep selections stable when switching modes:
        # single_image keeps the first path; multi_image keeps all selected paths.
        if self.source_mode.get() == "single_image" and len(self._selected_paths) > 1:
            self._selected_paths = [self._selected_paths[0]]
        self._update_image_info()

    def _choose_theme_color(self):
        picked = colorchooser.askcolor(color=self.theme_color.get(), title="Choose UI theme color")
        if not picked or not picked[1]:
            return
        c = _safe_hex_color(str(picked[1]))
        self.theme_color.set(c)
        try:
            self._theme_chip.configure(bg=c)
        except Exception:
            pass
        self.app.state.settings.theme_color = c
        self.app.apply_color_theme()

    def _load_selected_images(self) -> bool:
        mode = self.source_mode.get()
        if not self._selected_paths:
            self._select_images()
            if not self._selected_paths:
                return False
        if mode == "single_image":
            self._selected_paths = [self._selected_paths[0]]
            self._update_image_info()

        try:
            if mode == "multi_image":
                paths = self._selected_paths
                self.app.state.original_path = ";".join(paths)
                imgs = [pil_from_path(p) for p in paths]
                self.app.state.multi_input_paths = list(paths)
                self.app.state.multi_input_images = imgs
                self.app.state.multi_source_paths = []
                self.app.state.multi_source_images = []
                self.app.state.multi_source_placements = []
                self.app.state.multi_source_origin_ids = []
                self.app.state.multi_source_origin_names = []
                self.app.state.multi_source_lock_y = False
                self.app.state.panels = []
                self.app.state.panel_configs = []
                if self.app.state.settings.panel_mode == "multi_panel":
                    self.app.show_frame("MultiSourcePanelSelectFrame")
                else:
                    self.app.state.multi_source_paths = list(paths)
                    self.app.state.multi_source_images = imgs
                    self.app.state.multi_source_origin_ids = [int(i) for i in range(len(imgs))]
                    self.app.state.multi_source_origin_names = [pathlib.Path(p).stem for p in paths]
                    self.app.show_frame("ComposeFrame")
                return True

            path = self._selected_paths[0]
            self.app.state.original_path = path
            self.app.state.multi_input_paths = []
            self.app.state.multi_input_images = []
            self.app.state.multi_source_paths = []
            self.app.state.multi_source_images = []
            self.app.state.multi_source_placements = []
            self.app.state.multi_source_origin_ids = []
            self.app.state.multi_source_origin_names = []
            self.app.state.multi_source_lock_y = False
            img = pil_from_path(path)
            self.app.state.loaded_pil = img
            self.app.state.original_pil = img.copy()
            self.app.state.edited_pil = img.copy()
            self.app.show_frame("EditFrame")
            return True
        except Exception as e:
            messagebox.showerror("Image load", f"Could not load selected image(s): {e}")
            return False

    def on_next(self):
        s = self.app.state.settings
        s.gel_type = self.gel_type.get()
        s.panel_mode = self.panel_mode.get()
        s.source_mode = self.source_mode.get()
        s.default_unit = self.unit.get()
        s.keep_settings_next_image = bool(self.keep.get())
        s.prefer_invert = bool(self.invert.get())
        s.theme_color = _safe_hex_color(self.theme_color.get())
        _fam = str(self.font_family.get()).strip()
        self.app.state.settings.font_family = "" if _fam in ("(default)", "") else _fam
        self.app.state.settings.font_bold = bool(self.font_bold_var.get())
        self.app.state.settings.font_italic = bool(self.font_italic_var.get())
        self.app.apply_color_theme()

        save_settings(s)
        self._load_selected_images()


# ----------------------------
# Step 2: Load image(s)
# ----------------------------
class LoadImageFrame(ttk.Frame):
    def __init__(self, parent, app: GelAnnotatorApp):
        super().__init__(parent)
        self.app = app

        ttk.Label(self, text="Step 2 - Open image(s)", font=("Arial", 18, "bold")).pack(pady=10)

        self.msg = ttk.Label(self, text="", justify="left")
        self.msg.pack(pady=10)

        btns = ttk.Frame(self)
        btns.pack(pady=10)
        ttk.Button(btns, text="Open image(s)...", command=self.open_images).pack(side="left", padx=10)
        ttk.Button(btns, text="Back", command=lambda: self.app.show_frame("SettingsFrame")).pack(side="left", padx=10)

        self.preview_label = ttk.Label(self)
        self.preview_label.pack(pady=10)

    def on_show(self):
        if self.app.state.settings.source_mode == "multi_image":
            self.msg.configure(text="Select one or multiple images (each becomes a panel).")
        else:
            self.msg.configure(text="Select one image to edit/annotate.")
        self.preview_label.configure(image="", text="")

    def open_images(self):
        if self.app.state.settings.source_mode == "multi_image":
            paths = filedialog.askopenfilenames(
                title="Select images",
                filetypes=[("Images", "*.jpg *.jpeg *.png *.tif *.tiff"), ("All files", "*.*")]
            )
            if not paths:
                return
            self.app.state.original_path = ";".join(paths)
            # Load first as "original" for editing controls; for multi-image we skip editing-unify for MVP
            imgs = [Image.open(p).convert("RGB") for p in paths]
            self.app.state.loaded_pil = imgs[0]
            self.app.state.original_pil = imgs[0].copy()
            self.app.state.edited_pil = imgs[0].copy()
            self.app.state.panels = imgs  # each is a panel
            self.app.ensure_panel_configs()
            self.app.show_frame("LayoutFrame")
            return

        path = filedialog.askopenfilename(
            title="Select an image",
            filetypes=[("Images", "*.jpg *.jpeg *.png *.tif *.tiff"), ("All files", "*.*")]
        )
        if not path:
            return
        self.app.state.original_path = path
        img = Image.open(path).convert("RGB")
        self.app.state.loaded_pil = img
        self.app.state.original_pil = img.copy()
        self.app.state.edited_pil = img.copy()

        # quick preview
        prev = img.copy()
        prev.thumbnail((700, 450))
        self._preview_tk = ImageTk.PhotoImage(prev)
        self.preview_label.configure(image=self._preview_tk, text="")

        self.app.show_frame("EditFrame")


# ----------------------------
# Step 2a: Select panels from multiple source images
# ----------------------------
class MultiSourcePanelSelectFrame(ttk.Frame):
    def __init__(self, parent, app: GelAnnotatorApp):
        super().__init__(parent)
        self.app = app

        ttk.Label(self, text="Step 2a - Select panels from source images", font=("Arial", 18, "bold")).pack(pady=8)

        top = ttk.Frame(self)
        top.pack(fill="both", expand=True)

        self.viewport = ZoomableCanvas(top, bg="#222222", show_tools=True)
        self.viewport.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        self.canvas = self.viewport.canvas

        right_panel = VScrollPanel(top, width=self.app._scaled_sidebar_width(SIDEBAR_WIDTH_WIDE))
        right_panel.pack(side="left", fill="y", padx=10, pady=10)
        right = right_panel.inner

        ttk.Label(right, text="Source images").pack(anchor="w")
        self.source_list = tk.Listbox(right, height=8, exportselection=False)
        self.source_list.pack(fill="x", pady=4)
        self.source_list.bind("<<ListboxSelect>>", lambda _e=None: self._on_source_select())

        ttk.Separator(right).pack(fill="x", pady=8)
        ttk.Label(right, text="Pre-rotate selected source (before panel picks)").pack(anchor="w")
        self.source_angle = tk.DoubleVar(value=0.0)
        rot_row = ttk.Frame(right)
        rot_row.pack(fill="x", pady=2)
        self._src_rot_entry = ttk.Entry(rot_row, textvariable=self.source_angle, width=7)
        self._src_rot_entry.pack(side="left")
        self._src_rot_entry.bind("<Return>", lambda _e=None: self._apply_selected_source_rotation())
        self._src_rot_entry.bind("<FocusOut>", lambda _e=None: self._apply_selected_source_rotation())
        ttk.Button(rot_row, text="Apply", command=self._apply_selected_source_rotation).pack(side="left", padx=4)
        rot_btn_row = ttk.Frame(right)
        rot_btn_row.pack(fill="x", pady=(0, 2))
        for txt, delta in (("-90", -90.0), ("+90", 90.0), ("-1", -1.0), ("+1", 1.0)):
            ttk.Button(
                rot_btn_row,
                text=txt,
                width=6,
                command=(lambda d=delta: self._rotate_selected_source(d)),
            ).pack(side="left", padx=2, pady=1)
        ttk.Label(
            right,
            text="Note: rotating clears panel rectangles for this source.",
            justify="left",
        ).pack(anchor="w", pady=(2, 0))

        ttk.Separator(right).pack(fill="x", pady=8)
        ttk.Label(right, text="Panels for selected source").pack(anchor="w")
        self.panel_list = tk.Listbox(right, height=10, exportselection=False)
        self.panel_list.pack(fill="x", pady=4)
        self.panel_list.bind("<<ListboxSelect>>", lambda _e=None: self._on_panel_select())

        ttk.Label(right, text="Selected panel metadata").pack(anchor="w", pady=(4, 0))
        self.panel_is_marker_var = tk.BooleanVar(value=False)
        self.panel_scale_group_var = tk.IntVar(value=1)
        meta_row1 = ttk.Frame(right)
        meta_row1.pack(fill="x", pady=2)
        ttk.Checkbutton(
            meta_row1,
            text="This panel contains a marker / ladder",
            variable=self.panel_is_marker_var,
            command=self._apply_selected_panel_meta,
        ).pack(side="left")
        meta_row2 = ttk.Frame(right)
        meta_row2.pack(fill="x", pady=2)
        ttk.Label(meta_row2, text="Scaling group").pack(side="left")
        self._panel_group_entry = ttk.Entry(meta_row2, textvariable=self.panel_scale_group_var, width=6)
        self._panel_group_entry.pack(side="left", padx=6)
        self._panel_group_entry.bind("<Return>", lambda _e=None: self._apply_selected_panel_meta())
        self._panel_group_entry.bind("<FocusOut>", lambda _e=None: self._apply_selected_panel_meta())
        ttk.Label(
            right,
            text="Panels with the same scaling group will be scaled/aligned together using that group's marker panel.",
            justify="left",
        ).pack(anchor="w", pady=(0, 4))

        ttk.Button(right, text="Remove selected panel", command=self._remove_selected_panel).pack(fill="x", pady=2)
        ttk.Button(right, text="Clear selected source panels", command=self._clear_selected_source).pack(fill="x", pady=2)
        ttk.Button(right, text="Clear all panels", command=self._clear_all).pack(fill="x", pady=2)

        ttk.Separator(right).pack(fill="x", pady=8)
        self.summary_var = tk.StringVar(value="0 panels selected")
        ttk.Label(right, textvariable=self.summary_var).pack(anchor="w", pady=2)
        ttk.Label(
            right,
            text="Draw: left-click + drag rectangle.\nYou can add multiple panels per image.",
            justify="left",
        ).pack(anchor="w", pady=4)

        nav = ttk.Frame(right)
        nav.pack(fill="x", pady=10)
        ttk.Button(nav, text="Back", command=lambda: self.app.show_frame("SettingsFrame")).pack(side="left")
        ttk.Button(nav, text="Next -> Compose", style="Primary.TButton", command=self._next).pack(side="right")

        self._source_images: List[Image.Image] = []
        self._source_paths: List[str] = []
        self._rects_by_source: Dict[int, List[Tuple[int, int, int, int]]] = {}
        self._panel_meta_by_source: Dict[int, List[Dict[str, Any]]] = {}
        self._angle_by_source: Dict[int, float] = {}
        self._start_canvas_xy: Optional[Tuple[float, float]] = None
        self._rect_id: Optional[int] = None

        self.canvas.bind("<ButtonPress-1>", self._start)
        self.canvas.bind("<B1-Motion>", self._drag)
        self.canvas.bind("<ButtonRelease-1>", self._end)
        self.canvas.bind("<Button-3>", self._right_click_next, add="+")
        self.canvas.bind("<Button-2>", self._right_click_next, add="+")

    def _fit_listbox_height(self, lb: tk.Listbox, min_rows: int = 4, max_rows: int = 16):
        try:
            n = int(lb.size())
            h = max(int(min_rows), min(int(max_rows), max(1, n)))
            lb.configure(height=h)
        except Exception:
            pass

    def on_show(self):
        self._source_images = list(self.app.state.multi_input_images)
        self._source_paths = list(self.app.state.multi_input_paths)
        if not self._source_images:
            # Fallback: recover from selected paths when returning into this step.
            paths = [p for p in self.app.state.original_path.split(";") if p]
            try:
                self._source_images = [pil_from_path(p) for p in paths]
                self._source_paths = list(paths)
            except Exception:
                self._source_images = []
                self._source_paths = []

        if not self._source_images:
            self.app.show_frame("SettingsFrame")
            return

        valid_keys = set(range(len(self._source_images)))
        self._rects_by_source = {k: v for k, v in self._rects_by_source.items() if k in valid_keys}
        self._panel_meta_by_source = {k: v for k, v in self._panel_meta_by_source.items() if k in valid_keys}
        self._angle_by_source = {k: v for k, v in self._angle_by_source.items() if k in valid_keys}
        for k in valid_keys:
            self._rects_by_source.setdefault(k, [])
            self._panel_meta_by_source.setdefault(k, [])
            self._sync_panel_meta_lengths_for_source(k)
            self._angle_by_source.setdefault(k, 0.0)

        self._refresh_source_list()
        if self.source_list.size() > 0 and not self.source_list.curselection():
            self.source_list.selection_set(0)
        self._on_source_select()

    def _default_group_for_source(self, src_idx: int) -> int:
        return int(src_idx) + 1

    def _height_group_tolerance(self) -> float:
        # Normalized center-y tolerance (fraction of source image height).
        return 0.08

    def _rect_center_y_norm(self, src_idx: int, rect: Tuple[int, int, int, int]) -> float:
        try:
            img = self._source_images[int(src_idx)]
            h = max(1, int(getattr(img, "height", 1)))
            _x0, y0, _x1, y1 = rect
            cy = (float(y0) + float(y1)) / 2.0
            return float(cy / float(h))
        except Exception:
            return 0.0

    def _rect_group_anchor_y_norm(self, src_idx: int, rect: Tuple[int, int, int, int]) -> float:
        """
        Height-group anchor for auto grouping.
        Uses a top-biased anchor instead of pure center so differing panel heights
        do not cause wrong row matches.
        """
        try:
            img = self._source_images[int(src_idx)]
            h = max(1, int(getattr(img, "height", 1)))
            _x0, y0, _x1, y1 = rect
            hh = max(1.0, float(y1) - float(y0))
            anchor_y = float(y0) + 0.15 * hh
            return float(anchor_y / float(h))
        except Exception:
            return self._rect_center_y_norm(int(src_idx), rect)

    def _all_existing_panel_entries(self) -> List[Tuple[int, int, Tuple[int, int, int, int], Dict[str, Any]]]:
        out: List[Tuple[int, int, Tuple[int, int, int, int], Dict[str, Any]]] = []
        for src_idx in sorted(self._rects_by_source.keys()):
            self._sync_panel_meta_lengths_for_source(int(src_idx))
            rects = list(self._rects_by_source.get(int(src_idx), []) or [])
            meta = list(self._panel_meta_by_source.get(int(src_idx), []) or [])
            for j, rect in enumerate(rects):
                mm = meta[j] if j < len(meta) else self._default_panel_meta(int(src_idx))
                out.append((int(src_idx), int(j), tuple(rect), dict(mm)))
        return out

    def _group_has_marker_in_selection(self, group_id: int, exclude: Optional[Tuple[int, int]] = None) -> bool:
        gid = int(group_id)
        for src_idx, panel_idx, _rect, mm in self._all_existing_panel_entries():
            if exclude is not None and src_idx == int(exclude[0]) and panel_idx == int(exclude[1]):
                continue
            try:
                if int(mm.get("scale_group", -1)) == gid and bool(mm.get("is_marker", False)):
                    return True
            except Exception:
                continue
        return False

    def _group_used_by_other_source(self, group_id: int, src_idx: int, exclude: Optional[Tuple[int, int]] = None) -> bool:
        gid = int(group_id)
        s0 = int(src_idx)
        for s_idx, panel_idx, _rect, mm in self._all_existing_panel_entries():
            if exclude is not None and s_idx == int(exclude[0]) and panel_idx == int(exclude[1]):
                continue
            if int(s_idx) == s0:
                continue
            try:
                if int(mm.get("scale_group", -1)) == gid:
                    return True
            except Exception:
                continue
        return False

    def _next_unused_group_id(self) -> int:
        max_gid = 0
        for _s_idx, _panel_idx, _rect, mm in self._all_existing_panel_entries():
            try:
                max_gid = max(max_gid, int(mm.get("scale_group", 0)))
            except Exception:
                continue
        return max(1, int(max_gid) + 1)

    def _auto_group_for_new_panel(self, src_idx: int, rect: Tuple[int, int, int, int]) -> int:
        """
        Choose a scaling group based on approximate vertical position (normalized center y).
        Panels from different source images must never share a scaling group, so
        matching is restricted to panels from the same source only.
        """
        cy_norm = self._rect_center_y_norm(int(src_idx), rect)
        tol = float(self._height_group_tolerance())
        best_gid: Optional[int] = None
        best_delta = float("inf")
        max_gid = 0
        for s_idx, p_idx, r, mm in self._all_existing_panel_entries():
            try:
                gid = int(mm.get("scale_group", self._default_group_for_source(int(s_idx))))
            except Exception:
                gid = self._default_group_for_source(int(s_idx))
            max_gid = max(max_gid, int(gid))
            if int(s_idx) != int(src_idx):
                continue
            other_cy = self._rect_center_y_norm(int(s_idx), r)
            d = abs(float(other_cy) - float(cy_norm))
            if d <= tol and d < best_delta:
                best_delta = d
                best_gid = int(gid)

        if best_gid is not None:
            return int(best_gid)
        return max(max_gid + 1, 1)

    def _default_panel_meta(self, src_idx: int) -> Dict[str, Any]:
        return {"is_marker": False, "scale_group": int(self._default_group_for_source(src_idx))}

    def _sync_panel_meta_lengths_for_source(self, src_idx: int):
        rects = list(self._rects_by_source.get(src_idx, []))
        meta = list(self._panel_meta_by_source.get(src_idx, []))
        while len(meta) < len(rects):
            meta.append(self._default_panel_meta(int(src_idx)))
        if len(meta) > len(rects):
            meta = meta[: len(rects)]
        # Normalize values
        out_meta: List[Dict[str, Any]] = []
        for m in meta:
            try:
                is_marker = bool((m or {}).get("is_marker", False))
            except Exception:
                is_marker = False
            try:
                group_id = int((m or {}).get("scale_group", self._default_group_for_source(src_idx)))
            except Exception:
                group_id = self._default_group_for_source(src_idx)
            group_id = max(1, int(group_id))
            out_meta.append({"is_marker": bool(is_marker), "scale_group": int(group_id)})
        self._panel_meta_by_source[int(src_idx)] = out_meta

    def _refresh_source_list(self):
        self.source_list.delete(0, "end")
        if self._source_paths:
            names = [pathlib.Path(p).name for p in self._source_paths]
        else:
            names = [f"Image {i+1}" for i in range(len(self._source_images))]
        for i, n in enumerate(names, start=1):
            self.source_list.insert("end", f"{i}. {n}")
        self._fit_listbox_height(self.source_list, min_rows=3, max_rows=10)

    def _selected_source_index(self) -> Optional[int]:
        sel = self.source_list.curselection()
        if not sel:
            return None
        idx = int(sel[0])
        if 0 <= idx < len(self._source_images):
            return idx
        return None

    def _on_source_select(self):
        idx = self._selected_source_index()
        if idx is None:
            return
        self._start_canvas_xy = None
        self._rect_id = None
        self.source_angle.set(float(self._angle_by_source.get(idx, 0.0)))
        self._render_current_source()
        self._refresh_panel_list()
        self._on_panel_select()

    def _render_current_source(self):
        idx = self._selected_source_index()
        if idx is None:
            return
        img = self._source_images[idx]
        self.viewport.set_image(img, fit_if_needed=True)
        try:
            self.canvas.delete("panel_overlay")
        except Exception:
            pass
        z = float(self.viewport.zoom)
        for j, (x0, y0, x1, y1) in enumerate(self._rects_by_source.get(idx, []), start=1):
            cx0, cy0 = x0 * z, y0 * z
            cx1, cy1 = x1 * z, y1 * z
            self.canvas.create_rectangle(cx0, cy0, cx1, cy1, outline="cyan", width=2, tags=("panel_overlay",))
            self.canvas.create_text(cx0 + 6, cy0 + 6, text=str(j), anchor="nw", fill="cyan", tags=("panel_overlay",))

    def _refresh_panel_list(self):
        idx = self._selected_source_index()
        self.panel_list.delete(0, "end")
        if idx is not None:
            self._sync_panel_meta_lengths_for_source(int(idx))
            meta = self._panel_meta_by_source.get(int(idx), [])
            for j, (x0, y0, x1, y1) in enumerate(self._rects_by_source.get(idx, []), start=1):
                mm = meta[j - 1] if (j - 1) < len(meta) else self._default_panel_meta(int(idx))
                tags: List[str] = []
                if bool(mm.get("is_marker", False)):
                    tags.append("M")
                tags.append(f"G{int(mm.get('scale_group', self._default_group_for_source(int(idx))))}")
                tag_txt = f" [{' | '.join(tags)}]" if tags else ""
                self.panel_list.insert("end", f"Panel {j}: ({x0},{y0})({x1},{y1}){tag_txt}")
        total = sum(len(v) for v in self._rects_by_source.values())
        self.summary_var.set(f"{total} panels selected")
        self._fit_listbox_height(self.panel_list, min_rows=4, max_rows=18)

    def _selected_panel_index_for_selected_source(self) -> Optional[Tuple[int, int]]:
        src_idx = self._selected_source_index()
        if src_idx is None:
            return None
        sel = self.panel_list.curselection()
        if not sel:
            return None
        j = int(sel[0])
        rects = self._rects_by_source.get(int(src_idx), [])
        if 0 <= j < len(rects):
            return int(src_idx), int(j)
        return None

    def _on_panel_select(self):
        got = self._selected_panel_index_for_selected_source()
        if got is None:
            self.panel_is_marker_var.set(False)
            src_idx = self._selected_source_index()
            self.panel_scale_group_var.set(self._default_group_for_source(int(src_idx)) if src_idx is not None else 1)
            return
        src_idx, panel_idx = got
        self._sync_panel_meta_lengths_for_source(int(src_idx))
        mm = self._panel_meta_by_source[int(src_idx)][int(panel_idx)]
        self.panel_is_marker_var.set(bool(mm.get("is_marker", False)))
        self.panel_scale_group_var.set(max(1, int(mm.get("scale_group", self._default_group_for_source(int(src_idx))))))

    def _apply_selected_panel_meta(self):
        got = self._selected_panel_index_for_selected_source()
        if got is None:
            return
        src_idx, panel_idx = got
        self._sync_panel_meta_lengths_for_source(int(src_idx))
        try:
            gid = max(1, int(self.panel_scale_group_var.get()))
        except Exception:
            gid = self._default_group_for_source(int(src_idx))
        if self._group_used_by_other_source(int(gid), int(src_idx), exclude=(int(src_idx), int(panel_idx))):
            gid = self._next_unused_group_id()
        self.panel_scale_group_var.set(int(gid))
        is_marker = bool(self.panel_is_marker_var.get())
        # Helpful default: if user moves a panel into a group that currently has no marker, promote it.
        if (not is_marker) and (not self._group_has_marker_in_selection(int(gid), exclude=(int(src_idx), int(panel_idx)))):
            is_marker = True
            self.panel_is_marker_var.set(True)
        self._panel_meta_by_source[int(src_idx)][int(panel_idx)] = {
            "is_marker": bool(is_marker),
            "scale_group": int(gid),
        }
        sel = self.panel_list.curselection()
        self._refresh_panel_list()
        if sel and 0 <= int(sel[0]) < self.panel_list.size():
            self.panel_list.selection_set(int(sel[0]))

    def _remove_selected_panel(self):
        idx = self._selected_source_index()
        if idx is None:
            return
        sel = self.panel_list.curselection()
        if not sel:
            return
        j = int(sel[0])
        rects = self._rects_by_source.get(idx, [])
        self._sync_panel_meta_lengths_for_source(int(idx))
        meta = self._panel_meta_by_source.get(int(idx), [])
        if 0 <= j < len(rects):
            rects.pop(j)
            if 0 <= j < len(meta):
                meta.pop(j)
        self._render_current_source()
        self._refresh_panel_list()
        self._on_panel_select()

    def _clear_selected_source(self):
        idx = self._selected_source_index()
        if idx is None:
            return
        self._rects_by_source[idx] = []
        self._panel_meta_by_source[idx] = []
        self._render_current_source()
        self._refresh_panel_list()
        self._on_panel_select()

    def _clear_all(self):
        for k in list(self._rects_by_source.keys()):
            self._rects_by_source[k] = []
            self._panel_meta_by_source[k] = []
        self._render_current_source()
        self._refresh_panel_list()
        self._on_panel_select()

    def _persist_source_images(self):
        # Persist current pre-rotated sources so back/forward keeps orientation.
        self.app.state.multi_input_images = list(self._source_images)
        self.app.state.multi_input_paths = list(self._source_paths)

    def _rotate_selected_source(self, delta_deg: float):
        idx = self._selected_source_index()
        if idx is None:
            return
        delta = float(delta_deg)
        if abs(delta) < 1e-9:
            return
        img = self._source_images[idx]
        rot = img.rotate(delta, expand=True, resample=Image.Resampling.BICUBIC, fillcolor=(255, 255, 255))
        self._source_images[idx] = rot
        self._angle_by_source[idx] = float(self._angle_by_source.get(idx, 0.0)) + delta
        self.source_angle.set(float(self._angle_by_source[idx]))
        # Existing rectangles no longer match after rotation.
        self._rects_by_source[idx] = []
        self._panel_meta_by_source[idx] = []
        self._persist_source_images()
        self._render_current_source()
        self._refresh_panel_list()
        self._on_panel_select()

    def _apply_selected_source_rotation(self):
        idx = self._selected_source_index()
        if idx is None:
            return
        try:
            target = float(self.source_angle.get())
        except Exception:
            return
        cur = float(self._angle_by_source.get(idx, 0.0))
        self._rotate_selected_source(target - cur)

    def _canvas_to_img(self, x_canvas: float, y_canvas: float) -> Tuple[int, int]:
        z = max(1e-9, float(self.viewport.zoom))
        return clamp_int(x_canvas / z), clamp_int(y_canvas / z)

    def _start(self, event):
        if self._selected_source_index() is None:
            return
        self._start_canvas_xy = (self.canvas.canvasx(event.x), self.canvas.canvasy(event.y))
        if self._rect_id:
            try:
                self.canvas.delete(self._rect_id)
            except Exception:
                pass
            self._rect_id = None

    def _drag(self, event):
        if not self._start_canvas_xy:
            return
        x0, y0 = self._start_canvas_xy
        x1, y1 = self.canvas.canvasx(event.x), self.canvas.canvasy(event.y)
        if self._rect_id:
            self.canvas.coords(self._rect_id, x0, y0, x1, y1)
        else:
            self._rect_id = self.canvas.create_rectangle(x0, y0, x1, y1, outline="yellow", width=2, tags=("panel_overlay",))

    def _end(self, event):
        idx = self._selected_source_index()
        if idx is None or not self._start_canvas_xy:
            return
        x0c, y0c = self._start_canvas_xy
        x1c, y1c = self.canvas.canvasx(event.x), self.canvas.canvasy(event.y)
        self._start_canvas_xy = None
        if self._rect_id:
            try:
                self.canvas.delete(self._rect_id)
            except Exception:
                pass
            self._rect_id = None

        x0, y0 = self._canvas_to_img(min(x0c, x1c), min(y0c, y1c))
        x1, y1 = self._canvas_to_img(max(x0c, x1c), max(y0c, y1c))
        img = self._source_images[idx]
        x0 = max(0, min(img.width - 1, int(x0)))
        y0 = max(0, min(img.height - 1, int(y0)))
        x1 = max(1, min(img.width, int(x1)))
        y1 = max(1, min(img.height, int(y1)))
        if abs(x1 - x0) < 10 or abs(y1 - y0) < 10:
            return

        new_rect = (x0, y0, x1, y1)
        auto_gid = self._auto_group_for_new_panel(int(idx), new_rect)
        auto_marker = not self._group_has_marker_in_selection(int(auto_gid))
        self._rects_by_source.setdefault(idx, []).append(new_rect)
        self._panel_meta_by_source.setdefault(idx, []).append({"is_marker": bool(auto_marker), "scale_group": int(auto_gid)})
        self._render_current_source()
        self._refresh_panel_list()
        try:
            self.panel_list.selection_clear(0, "end")
            self.panel_list.selection_set(self.panel_list.size() - 1)
        except Exception:
            pass
        self._on_panel_select()

    def _next(self):
        crops: List[Image.Image] = []
        labels: List[str] = []
        origin_ids: List[int] = []
        origin_names: List[str] = []
        source_rects: List[Tuple[int, int, int, int]] = []
        panel_marker_flags: List[bool] = []
        panel_group_ids: List[int] = []

        for src_idx, img in enumerate(self._source_images):
            rects = self._rects_by_source.get(src_idx, [])
            self._sync_panel_meta_lengths_for_source(int(src_idx))
            meta = self._panel_meta_by_source.get(int(src_idx), [])
            if self._source_paths and src_idx < len(self._source_paths):
                src_name = pathlib.Path(self._source_paths[src_idx]).stem
            else:
                src_name = f"Image{src_idx + 1}"
            for j, (x0, y0, x1, y1) in enumerate(rects, start=1):
                crops.append(img.crop((x0, y0, x1, y1)))
                labels.append(f"{src_name}_panel{j}")
                origin_ids.append(int(src_idx))
                origin_names.append(str(src_name))
                source_rects.append((int(x0), int(y0), int(x1), int(y1)))
                mm = meta[j - 1] if (j - 1) < len(meta) else self._default_panel_meta(int(src_idx))
                panel_marker_flags.append(bool(mm.get("is_marker", False)))
                panel_group_ids.append(max(1, int(mm.get("scale_group", self._default_group_for_source(int(src_idx))))))

        if not crops:
            messagebox.showinfo("Panels", "No panels defined. Draw at least one rectangle.")
            return

        self.app.state.multi_source_paths = labels
        self.app.state.multi_source_images = crops
        self.app.state.multi_source_placements = []
        self.app.state.multi_source_origin_ids = origin_ids
        self.app.state.multi_source_origin_names = origin_names
        self.app.state.multi_source_panel_source_rects = [tuple(map(int, r)) for r in source_rects]
        self.app.state.multi_source_lock_y = False
        # Marker-scaling assignments are now defined in panel selection (Step 2a).
        self.app.state.multi_source_marker_flags = [bool(v) for v in panel_marker_flags]
        self.app.state.multi_source_marker_group_ids = [int(v) for v in panel_group_ids]
        self.app.state.multi_source_marker_picks_by_panel = {}
        self.app.state.multi_source_marker_ref_panel = -1

        # Derive follower mapping from group IDs + marker flags.
        n = len(crops)
        group_to_marker: Dict[int, int] = {}
        dup_groups: List[int] = []
        for i in range(n):
            if not bool(panel_marker_flags[i]):
                continue
            gid = int(panel_group_ids[i])
            if gid in group_to_marker and gid not in dup_groups:
                dup_groups.append(gid)
            group_to_marker.setdefault(gid, int(i))
        follows: List[int] = [-1 for _ in range(n)]
        for i in range(n):
            if bool(panel_marker_flags[i]):
                follows[i] = int(i)
                continue
            gid = int(panel_group_ids[i])
            follows[i] = int(group_to_marker.get(gid, -1))
        self.app.state.multi_source_marker_follow_panel = follows

        self._persist_source_images()
        if dup_groups:
            messagebox.showwarning(
                "Marker groups",
                "Some scaling groups contain more than one marker panel (e.g. "
                + ", ".join(f"G{int(g)}" for g in dup_groups[:10])
                + ").\nOnly the first marker panel in each group will be used as the controller for follower panels.",
            )
        self.app.show_frame("ComposeFrame")

    def _right_click_next(self, _event=None):
        self._next()
        return "break"


# ----------------------------
# Step 2b: Compose multi-image canvas
# ----------------------------
class ComposeFrame(ttk.Frame):
    def __init__(self, parent, app: GelAnnotatorApp):
        super().__init__(parent)
        self.app = app

        ttk.Label(self, text="Step 2b - Compose multiple images", font=("Arial", 18, "bold")).pack(pady=8)

        main = ttk.Frame(self)
        main.pack(fill="both", expand=True)

        self.viewport = ZoomableCanvas(main, bg="#222222", show_tools=True)
        self.viewport.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        self.canvas = self.viewport.canvas

        right_panel = VScrollPanel(main, width=self.app._scaled_sidebar_width(560))
        right_panel.pack(side="left", fill="y", padx=10, pady=10)
        right = right_panel.inner

        ttk.Label(right, text="Images").pack(anchor="w")
        self.img_list = tk.Listbox(right, height=10)
        self.img_list.pack(fill="x", pady=4)
        self.img_list.bind("<<ListboxSelect>>", lambda _e=None: self._on_select())

        ttk.Separator(right).pack(fill="x", pady=8)
        ttk.Label(right, text="Selected image transform:").pack(anchor="w")

        self.sel_x = tk.IntVar(value=0)
        self.sel_y = tk.IntVar(value=0)
        self.sel_angle = tk.DoubleVar(value=0.0)
        self.sel_scale_x = tk.DoubleVar(value=1.0)
        self.sel_scale_y = tk.DoubleVar(value=1.0)
        self.sel_brightness = tk.DoubleVar(value=0.0)
        self.sel_contrast = tk.DoubleVar(value=1.0)
        self.sel_gamma = tk.DoubleVar(value=1.0)
        self.sel_wb = tk.BooleanVar(value=False)
        self.sel_invert = tk.BooleanVar(value=False)
        self.sel_bw = tk.BooleanVar(value=True)

        row = ttk.Frame(right)
        row.pack(fill="x", pady=2)
        ttk.Label(row, text="X").pack(side="left")
        ex = ttk.Entry(row, textvariable=self.sel_x, width=7)
        ex.pack(side="left", padx=4)
        ttk.Label(row, text="Y").pack(side="left", padx=(8, 0))
        ey = ttk.Entry(row, textvariable=self.sel_y, width=7)
        ey.pack(side="left", padx=4)
        self._sel_y_entry = ey

        row2 = ttk.Frame(right)
        row2.pack(fill="x", pady=2)
        ttk.Label(row2, text="Angle").pack(side="left")
        ea = ttk.Entry(row2, textvariable=self.sel_angle, width=7)
        ea.pack(side="left", padx=4)
        row2a = ttk.Frame(right)
        row2a.pack(fill="x", pady=(0, 2))
        for txt, delta in (("-90", -90.0), ("+90", 90.0), ("-1", -1.0), ("+1", 1.0)):
            ttk.Button(
                row2a,
                text=txt,
                width=6,
                command=(lambda d=delta: self._rotate_selected(d)),
            ).pack(side="left", padx=2, pady=1)

        row3 = ttk.Frame(right)
        row3.pack(fill="x", pady=2)
        ttk.Label(row3, text="Scale X").pack(side="left")
        esx = ttk.Entry(row3, textvariable=self.sel_scale_x, width=7)
        esx.pack(side="left", padx=4)
        ttk.Label(row3, text="Scale Y").pack(side="left", padx=(8, 0))
        esy = ttk.Entry(row3, textvariable=self.sel_scale_y, width=7)
        esy.pack(side="left", padx=4)
        ttk.Button(row3, text="-10%", command=lambda: self._scale_selected(0.9, 0.9)).pack(side="left", padx=2)
        ttk.Button(row3, text="+10%", command=lambda: self._scale_selected(1.1, 1.1)).pack(side="left", padx=2)

        ttk.Separator(right).pack(fill="x", pady=8)
        ttk.Label(right, text="Selected image color / B-W:").pack(anchor="w")

        row4 = ttk.Frame(right)
        row4.pack(fill="x", pady=2)
        ttk.Label(row4, text="B").pack(side="left")
        eb = ttk.Entry(row4, textvariable=self.sel_brightness, width=6)
        eb.pack(side="left", padx=4)
        ttk.Label(row4, text="C").pack(side="left", padx=(8, 0))
        ec = ttk.Entry(row4, textvariable=self.sel_contrast, width=6)
        ec.pack(side="left", padx=4)
        ttk.Label(row4, text="G").pack(side="left", padx=(8, 0))
        eg = ttk.Entry(row4, textvariable=self.sel_gamma, width=6)
        eg.pack(side="left", padx=4)

        row5 = ttk.Frame(right)
        row5.pack(fill="x", pady=2)
        ttk.Checkbutton(row5, text="Gray-world WB", variable=self.sel_wb, command=self._apply_selected).pack(side="left")
        ttk.Checkbutton(row5, text="Invert", variable=self.sel_invert, command=self._apply_selected).pack(side="left", padx=(8, 0))
        ttk.Checkbutton(row5, text="B/W", variable=self.sel_bw, command=self._apply_selected).pack(side="left", padx=(8, 0))
        ttk.Button(right, text="Reset color", command=self._reset_selected_color).pack(fill="x", pady=2)

        ex.bind("<KeyRelease>", lambda _e=None: self._apply_selected())
        ey.bind("<KeyRelease>", lambda _e=None: self._apply_selected())
        ea.bind("<KeyRelease>", lambda _e=None: self._apply_selected())
        esx.bind("<KeyRelease>", lambda _e=None: self._apply_selected())
        esy.bind("<KeyRelease>", lambda _e=None: self._apply_selected())
        eb.bind("<KeyRelease>", lambda _e=None: self._apply_selected())
        ec.bind("<KeyRelease>", lambda _e=None: self._apply_selected())
        eg.bind("<KeyRelease>", lambda _e=None: self._apply_selected())

        ttk.Button(right, text="Apply transform", command=self._apply_selected).pack(fill="x", pady=4)
        ttk.Button(right, text="Crop selected image...", command=self._crop_selected_dialog).pack(fill="x", pady=4)
        ttk.Button(right, text="Auto-arrange horizontally", command=self._auto_arrange).pack(fill="x", pady=4)
        ttk.Label(right, text="Tip: drag selected image in preview to move it.", justify="left").pack(anchor="w", pady=(4, 0))

        ttk.Separator(right).pack(fill="x", pady=10)
        self._y_lock_text = tk.StringVar(value="Y alignment lock: OFF")
        ttk.Label(right, textvariable=self._y_lock_text).pack(anchor="w", pady=(0, 4))
        self._temp_unlock_y_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            right,
            text="Temporarily disable Y lock (manual correction)",
            variable=self._temp_unlock_y_var,
            command=self._update_y_lock_ui,
        ).pack(anchor="w", pady=(0, 4))
        self._source_label_to_origin: Dict[str, int] = {}
        ttk.Label(right, text="Marker scaling (assignments come from Step 2a panel selection)").pack(anchor="w")
        self.global_ref_marker_panel_var = tk.StringVar(value="")
        self._marker_panel_label_to_idx: Dict[str, int] = {}
        row_mm = ttk.Frame(right)
        row_mm.pack(fill="x", pady=2)
        ttk.Button(row_mm, text="Pick 3 markers for selected panel", command=self._pick_markers_for_selected_panel).pack(side="left")
        ttk.Button(row_mm, text="Clear picks", command=self._clear_markers_for_selected_panel).pack(side="left", padx=6)
        ttk.Label(
            right,
            text="Marker panels and scaling groups are defined in Step 2a.\nHere you only pick marker bands; scaling/alignment runs automatically after the last marker panel is picked.",
            justify="left",
        ).pack(anchor="w", pady=(0, 2))

        ttk.Separator(right).pack(fill="x", pady=10)
        ttk.Button(right, text="Final crop composed result...", command=self._crop_composed_dialog).pack(fill="x", pady=4)
        ttk.Button(right, text="Undo final crop", command=self._undo_final_crop).pack(fill="x", pady=2)
        ttk.Label(
            right,
            text="Use this after arranging panels to crop all panels together so edges align.",
            justify="left",
        ).pack(anchor="w", pady=(0, 2))

        ttk.Separator(right).pack(fill="x", pady=10)

        nav = ttk.Frame(right)
        nav.pack(fill="x", pady=8)
        ttk.Button(nav, text="Back", command=self._back).pack(side="left")
        ttk.Button(nav, text="Next -> Layout", style="Primary.TButton", command=self._next).pack(side="right")

        self._layout_meta: List[Dict[str, int]] = []
        self._preview_initialized = False
        self._drag_idx: Optional[int] = None
        self._drag_anchor_img: Optional[Tuple[int, int]] = None
        self._drag_start_xy: Optional[Tuple[int, int]] = None
        self._transform_cache: Dict[Tuple[Any, ...], Image.Image] = {}
        self._render_after_id: Optional[str] = None

        self.canvas.bind("<ButtonPress-1>", self._drag_start)
        self.canvas.bind("<B1-Motion>", self._drag_move)
        self.canvas.bind("<ButtonRelease-1>", self._drag_end)
        # Keep selection overlay visible after canvas redraw/zoom events.
        self.canvas.bind("<Configure>", lambda _e=None: self.after(0, self._draw_overlay), add="+")
        self.canvas.bind("<MouseWheel>", lambda _e=None: self.after(0, self._draw_overlay), add="+")
        self.canvas.bind("<Button-4>", lambda _e=None: self.after(0, self._draw_overlay), add="+")
        self.canvas.bind("<Button-5>", lambda _e=None: self.after(0, self._draw_overlay), add="+")

    def _fit_listbox_height(self, lb: tk.Listbox, min_rows: int = 4, max_rows: int = 16):
        try:
            n = int(lb.size())
            h = max(int(min_rows), min(int(max_rows), max(1, n)))
            lb.configure(height=h)
        except Exception:
            pass

    def _snapshot_compose_state_for_final_crop_undo(self) -> Dict[str, Any]:
        imgs = [img.copy() for img in list(self.app.state.multi_source_images or [])]
        pls = [
            ImagePlacement(
                x=int(getattr(pl, "x", 0)),
                y=int(getattr(pl, "y", 0)),
                angle_deg=float(getattr(pl, "angle_deg", 0.0)),
                scale_x=float(getattr(pl, "scale_x", 1.0)),
                scale_y=float(getattr(pl, "scale_y", 1.0)),
                brightness=float(getattr(pl, "brightness", 0.0)),
                contrast=float(getattr(pl, "contrast", 1.0)),
                gamma=float(getattr(pl, "gamma", 1.0)),
                white_balance=bool(getattr(pl, "white_balance", False)),
                invert=bool(getattr(pl, "invert", False)),
                bw=bool(getattr(pl, "bw", True)),
            )
            for pl in list(self.app.state.multi_source_placements or [])
        ]
        picks_raw = dict(getattr(self.app.state, "multi_source_marker_picks_by_panel", {}) or {})
        picks = {int(k): [float(v) for v in list(vals or [])] for k, vals in picks_raw.items()}
        snap = {
            "images": imgs,
            "paths": list(self.app.state.multi_source_paths or []),
            "placements": pls,
            "origin_ids": [int(v) for v in list(self.app.state.multi_source_origin_ids or [])],
            "origin_names": [str(v) for v in list(self.app.state.multi_source_origin_names or [])],
            "panel_source_rects": [tuple(map(int, r)) for r in list(getattr(self.app.state, "multi_source_panel_source_rects", []) or [])],
            "lock_y": bool(getattr(self.app.state, "multi_source_lock_y", False)),
            "marker_flags": [bool(v) for v in list(getattr(self.app.state, "multi_source_marker_flags", []) or [])],
            "marker_groups": [int(v) for v in list(getattr(self.app.state, "multi_source_marker_group_ids", []) or [])],
            "marker_follows": [int(v) for v in list(getattr(self.app.state, "multi_source_marker_follow_panel", []) or [])],
            "marker_picks": picks,
            "marker_ref": int(getattr(self.app.state, "multi_source_marker_ref_panel", -1) or -1),
        }
        return snap

    def _restore_compose_state_from_snapshot(self, snap: Dict[str, Any]) -> None:
        self.app.state.multi_source_images = [img.copy() for img in list(snap.get("images", []) or [])]
        self.app.state.multi_source_paths = list(snap.get("paths", []) or [])
        self.app.state.multi_source_placements = list(snap.get("placements", []) or [])
        self.app.state.multi_source_origin_ids = [int(v) for v in list(snap.get("origin_ids", []) or [])]
        self.app.state.multi_source_origin_names = [str(v) for v in list(snap.get("origin_names", []) or [])]
        self.app.state.multi_source_panel_source_rects = [tuple(map(int, r)) for r in list(snap.get("panel_source_rects", []) or [])]
        self.app.state.multi_source_lock_y = bool(snap.get("lock_y", False))
        self.app.state.multi_source_marker_flags = [bool(v) for v in list(snap.get("marker_flags", []) or [])]
        self.app.state.multi_source_marker_group_ids = [int(v) for v in list(snap.get("marker_groups", []) or [])]
        self.app.state.multi_source_marker_follow_panel = [int(v) for v in list(snap.get("marker_follows", []) or [])]
        self.app.state.multi_source_marker_picks_by_panel = {
            int(k): [float(v) for v in list(vals or [])]
            for k, vals in dict(snap.get("marker_picks", {}) or {}).items()
        }
        self.app.state.multi_source_marker_ref_panel = int(snap.get("marker_ref", -1) or -1)

    def _undo_final_crop(self):
        snap = getattr(self.app.state, "multi_source_final_crop_undo", None)
        if not snap:
            messagebox.showinfo("Undo final crop", "No final crop to undo.")
            return
        try:
            self._restore_compose_state_from_snapshot(dict(snap))
            self.app.state.multi_source_final_crop_undo = None
            self._transform_cache.clear()
            self._preview_initialized = False
            self._ensure_placements()
            self._refresh_list()
            self._refresh_multi_marker_controls()
            self._update_y_lock_ui()
            if self.img_list.size() > 0:
                self.img_list.selection_clear(0, "end")
                self.img_list.selection_set(0)
                self._on_select()
            self._render_preview()
        except Exception as e:
            messagebox.showerror("Undo final crop", f"Could not restore previous composition: {e}")

    def _is_y_locked(self) -> bool:
        hard_lock = bool(getattr(self.app.state, "multi_source_lock_y", False))
        temp_unlock = bool(getattr(self, "_temp_unlock_y_var", tk.BooleanVar(value=False)).get()) if hasattr(self, "_temp_unlock_y_var") else False
        return bool(hard_lock and (not temp_unlock))

    def _update_y_lock_ui(self):
        hard_lock = bool(getattr(self.app.state, "multi_source_lock_y", False))
        temp_unlock = bool(self._temp_unlock_y_var.get()) if hasattr(self, "_temp_unlock_y_var") else False
        locked = self._is_y_locked()
        if hard_lock and temp_unlock:
            self._y_lock_text.set("Y alignment lock: ON (temporarily unlocked)")
        elif locked:
            self._y_lock_text.set("Y alignment lock: ON (Y fixed)")
        else:
            self._y_lock_text.set("Y alignment lock: OFF")
        try:
            self._sel_y_entry.configure(state=("disabled" if locked else "normal"))
        except Exception:
            pass

    def on_show(self):
        if not self.app.state.multi_source_images:
            self.app.show_frame("SettingsFrame")
            return
        self._cancel_scheduled_render()
        self._preview_initialized = False
        self._transform_cache.clear()
        if hasattr(self, "_temp_unlock_y_var"):
            self._temp_unlock_y_var.set(False)
        self._ensure_placements()
        self._ensure_multi_marker_state()
        self._refresh_list()
        self._refresh_multi_marker_controls()
        self._update_y_lock_ui()
        if self.img_list.size() > 0 and not self.img_list.curselection():
            self.img_list.selection_set(0)
            self._on_select()
        self._render_preview()

    def on_hide(self):
        # Prevent queued heavy preview renders from running after frame switches.
        self._cancel_scheduled_render()

    def _cancel_scheduled_render(self):
        if self._render_after_id is not None:
            try:
                self.after_cancel(self._render_after_id)
            except Exception:
                pass
            self._render_after_id = None

    def _schedule_render_preview(self, delay_ms: int = 16):
        self._cancel_scheduled_render()
        self._render_after_id = self.after(max(1, int(delay_ms)), self._render_preview)

    def _ensure_placements(self):
        imgs = self.app.state.multi_source_images
        pls = self.app.state.multi_source_placements
        if len(pls) == len(imgs) and len(pls) > 0:
            for pl in pls:
                if not hasattr(pl, "scale_x"):
                    pl.scale_x = 1.0
                if not hasattr(pl, "scale_y"):
                    pl.scale_y = 1.0
                if not hasattr(pl, "brightness"):
                    pl.brightness = 0.0
                if not hasattr(pl, "contrast"):
                    pl.contrast = 1.0
                if not hasattr(pl, "gamma"):
                    pl.gamma = 1.0
                if not hasattr(pl, "white_balance"):
                    pl.white_balance = False
                if not hasattr(pl, "invert"):
                    pl.invert = False
                if not hasattr(pl, "bw"):
                    pl.bw = True
            self._ensure_source_origin_meta()
            self._ensure_multi_marker_state()
            return
        new_pls: List[ImagePlacement] = []
        x = 0
        for img in imgs:
            new_pls.append(ImagePlacement(x=x, y=0, angle_deg=0.0, scale_x=1.0, scale_y=1.0))
            x += img.width + 20
        self.app.state.multi_source_placements = new_pls
        self._ensure_source_origin_meta()
        self._ensure_multi_marker_state()

    def _ensure_source_origin_meta(self):
        n = len(self.app.state.multi_source_images)
        ids = list(self.app.state.multi_source_origin_ids)
        names = list(self.app.state.multi_source_origin_names)
        paths = list(self.app.state.multi_source_paths)
        if len(ids) != n:
            ids = [int(i) for i in range(n)]
        if len(names) != n:
            if len(paths) == n:
                names = [pathlib.Path(p).stem for p in paths]
            else:
                names = [f"Source {int(oid) + 1}" for oid in ids]
        self.app.state.multi_source_origin_ids = ids
        self.app.state.multi_source_origin_names = names

    def _ensure_panel_source_rect_meta(self):
        n = len(self.app.state.multi_source_images)
        imgs = list(self.app.state.multi_source_images)
        raw = list(getattr(self.app.state, "multi_source_panel_source_rects", []) or [])
        out: List[Tuple[int, int, int, int]] = []
        for i in range(n):
            if i < len(raw):
                try:
                    x0, y0, x1, y1 = raw[i]
                    out.append((int(x0), int(y0), int(x1), int(y1)))
                    continue
                except Exception:
                    pass
            try:
                w = int(getattr(imgs[i], "width", 1))
                h = int(getattr(imgs[i], "height", 1))
            except Exception:
                w, h = 1, 1
            out.append((0, 0, max(1, w), max(1, h)))
        self.app.state.multi_source_panel_source_rects = out

    def _panel_source_rects(self) -> List[Tuple[int, int, int, int]]:
        self._ensure_panel_source_rect_meta()
        return list(getattr(self.app.state, "multi_source_panel_source_rects", []) or [])

    def _ensure_multi_marker_state(self):
        n = len(self.app.state.multi_source_images)
        flags = list(getattr(self.app.state, "multi_source_marker_flags", []) or [])
        follows = list(getattr(self.app.state, "multi_source_marker_follow_panel", []) or [])
        picks_any = getattr(self.app.state, "multi_source_marker_picks_by_panel", {}) or {}
        ref_idx = int(getattr(self.app.state, "multi_source_marker_ref_panel", -1) or -1)

        if len(flags) != n:
            flags = (flags + [False] * n)[:n]
        if len(follows) != n:
            follows = (follows + [-1] * n)[:n]

        # Normalize picks map to in-range integer keys with 3 floats per entry.
        picks: Dict[int, List[float]] = {}
        try:
            items = list(picks_any.items())
        except Exception:
            items = []
        for k, v in items:
            try:
                idx = int(k)
            except Exception:
                continue
            if not (0 <= idx < n):
                continue
            vals: List[float] = []
            for x in list(v or []):
                try:
                    fx = float(x)
                except Exception:
                    continue
                if np.isfinite(fx):
                    vals.append(float(fx))
            if len(vals) == 3:
                picks[idx] = sorted(vals)

        for i in range(n):
            fi = int(follows[i]) if i < len(follows) else -1
            if fi < 0 or fi >= n:
                follows[i] = -1
            else:
                follows[i] = fi

        if ref_idx < 0 or ref_idx >= n:
            ref_idx = -1

        self.app.state.multi_source_marker_flags = [bool(x) for x in flags]
        self.app.state.multi_source_marker_follow_panel = [int(x) for x in follows]
        self.app.state.multi_source_marker_picks_by_panel = picks
        self.app.state.multi_source_marker_ref_panel = int(ref_idx)

    def _marker_flags(self) -> List[bool]:
        self._ensure_multi_marker_state()
        return self.app.state.multi_source_marker_flags

    def _marker_follows(self) -> List[int]:
        self._ensure_multi_marker_state()
        return self.app.state.multi_source_marker_follow_panel

    def _marker_picks(self) -> Dict[int, List[float]]:
        self._ensure_multi_marker_state()
        return self.app.state.multi_source_marker_picks_by_panel

    def _marker_label_for_panel_idx(self, idx: int) -> str:
        self._ensure_source_origin_meta()
        self._ensure_multi_marker_state()
        n = len(self.app.state.multi_source_images)
        if idx < 0 or idx >= n:
            return "(invalid)"
        src_id = int(self.app.state.multi_source_origin_ids[idx]) if idx < len(self.app.state.multi_source_origin_ids) else idx
        src_name = self._source_name_for_id(src_id)
        groups = self._source_groups()
        src_panels = groups.get(int(src_id), [])
        panel_no = (src_panels.index(idx) + 1) if idx in src_panels else (idx + 1)
        picks = self._marker_picks().get(int(idx))
        picks_txt = " (3 picks)" if picks and len(picks) == 3 else ""
        return f"P{idx + 1} | S{src_id + 1}:{src_name} | src panel {panel_no}{picks_txt}"

    def _available_marker_panel_indices(self) -> List[int]:
        self._ensure_multi_marker_state()
        flags = self._marker_flags()
        return [i for i, is_marker in enumerate(flags) if bool(is_marker)]

    def _marker_panel_control_labels(self, include_none: bool = True) -> List[str]:
        labels: List[str] = []
        self._marker_panel_label_to_idx = {}
        if include_none:
            labels.append("(none)")
            self._marker_panel_label_to_idx["(none)"] = -1
        for idx in self._available_marker_panel_indices():
            lbl = self._marker_label_for_panel_idx(int(idx))
            labels.append(lbl)
            self._marker_panel_label_to_idx[lbl] = int(idx)
        return labels

    def _default_marker_follow_for_panel(self, panel_idx: int) -> int:
        self._ensure_source_origin_meta()
        self._ensure_multi_marker_state()
        if panel_idx < 0 or panel_idx >= len(self.app.state.multi_source_images):
            return -1
        flags = self._marker_flags()
        if bool(flags[panel_idx]):
            return int(panel_idx)
        group_ids = [int(v) for v in list(getattr(self.app.state, "multi_source_marker_group_ids", []) or [])]
        gid = int(group_ids[panel_idx]) if 0 <= panel_idx < len(group_ids) else -1
        candidates = [
            i for i in self._available_marker_panel_indices()
            if (0 <= i < len(group_ids)) and int(group_ids[i]) == gid
        ]
        # Fallback to same source if group mapping is unavailable/incomplete.
        if not candidates:
            try:
                src_id = int(self.app.state.multi_source_origin_ids[panel_idx])
            except Exception:
                src_id = -1
            if src_id >= 0:
                candidates = [
                    i for i in self._available_marker_panel_indices()
                    if i < len(self.app.state.multi_source_origin_ids) and int(self.app.state.multi_source_origin_ids[i]) == src_id
                ]
        if candidates:
            return int(candidates[0])
        return -1

    def _refresh_multi_marker_controls(self):
        self._ensure_multi_marker_state()
        labels_with_none = self._marker_panel_control_labels(include_none=True)
        labels_markers_only = [lbl for lbl in labels_with_none if self._marker_panel_label_to_idx.get(lbl, -1) >= 0]
        if hasattr(self, "sel_follow_marker_combo"):
            try:
                self.sel_follow_marker_combo["values"] = labels_with_none
            except Exception:
                pass
        if hasattr(self, "global_ref_marker_combo"):
            try:
                self.global_ref_marker_combo["values"] = labels_markers_only
            except Exception:
                pass

        ref_idx = int(getattr(self.app.state, "multi_source_marker_ref_panel", -1) or -1)
        if ref_idx >= 0:
            ref_label = next((lbl for lbl, idx in self._marker_panel_label_to_idx.items() if idx == ref_idx), "")
            if ref_label:
                self.global_ref_marker_panel_var.set(ref_label)
        elif labels_markers_only and (self.global_ref_marker_panel_var.get() not in labels_markers_only):
            self.global_ref_marker_panel_var.set(labels_markers_only[0])
            self._on_global_marker_ref_changed()
        elif not labels_markers_only:
            self.global_ref_marker_panel_var.set("")

    def _sync_selected_panel_marker_controls(self):
        if not hasattr(self, "sel_has_marker_var"):
            return
        i = self._selected_index()
        if i is None:
            self.sel_has_marker_var.set(False)
            self.sel_follow_marker_var.set("(none)")
            return
        self._ensure_multi_marker_state()
        self._refresh_multi_marker_controls()
        flags = self._marker_flags()
        follows = self._marker_follows()
        self.sel_has_marker_var.set(bool(flags[i]))
        fi = int(follows[i]) if 0 <= i < len(follows) else -1
        if fi < 0:
            fi = self._default_marker_follow_for_panel(i)
            follows[i] = int(fi)
        lbl = next((k for k, v in self._marker_panel_label_to_idx.items() if int(v) == int(fi)), "(none)")
        self.sel_follow_marker_var.set(lbl if lbl else "(none)")

    def _on_selected_marker_config_changed(self):
        if not hasattr(self, "sel_has_marker_var"):
            return
        i = self._selected_index()
        if i is None:
            return
        self._ensure_multi_marker_state()
        flags = self._marker_flags()
        follows = self._marker_follows()
        picks = self._marker_picks()

        is_marker = bool(self.sel_has_marker_var.get())
        flags[i] = bool(is_marker)
        if is_marker:
            # Marker panels default to following themselves.
            follows[i] = int(i)
            if int(getattr(self.app.state, "multi_source_marker_ref_panel", -1)) < 0:
                self.app.state.multi_source_marker_ref_panel = int(i)
        else:
            # Keep picks only for active marker panels; clear if marker flag turned off.
            picks.pop(int(i), None)
            if follows[i] == int(i):
                follows[i] = self._default_marker_follow_for_panel(i)
            if int(getattr(self.app.state, "multi_source_marker_ref_panel", -1)) == int(i):
                self.app.state.multi_source_marker_ref_panel = -1

        sel_lbl = str(self.sel_follow_marker_var.get()).strip()
        if sel_lbl:
            fi = int(self._marker_panel_label_to_idx.get(sel_lbl, follows[i]))
            if fi >= 0 and fi < len(self.app.state.multi_source_images):
                follows[i] = int(fi)
        if bool(self.sel_has_marker_var.get()):
            follows[i] = int(i)

        self._refresh_multi_marker_controls()
        self._sync_selected_panel_marker_controls()
        self._refresh_list()

    def _on_global_marker_ref_changed(self):
        self._ensure_multi_marker_state()
        lbl = str(self.global_ref_marker_panel_var.get()).strip()
        idx = int(self._marker_panel_label_to_idx.get(lbl, -1)) if lbl else -1
        self.app.state.multi_source_marker_ref_panel = int(idx)
        self._refresh_list()

    def _set_selected_as_global_reference_marker(self):
        i = self._selected_index()
        if i is None:
            messagebox.showinfo("Marker scaling", "Select a panel first.")
            return
        self._ensure_multi_marker_state()
        flags = self._marker_flags()
        if not (0 <= i < len(flags)) or not bool(flags[i]):
            messagebox.showinfo("Marker scaling", "Selected panel is not designated as a marker panel (set this in Step 2a).")
            return
        self.app.state.multi_source_marker_ref_panel = int(i)
        self._refresh_multi_marker_controls()
        self._refresh_list()

    def _pick_markers_for_selected_panel(self):
        i = self._selected_index()
        if i is None:
            messagebox.showinfo("Markers", "Select a panel first.")
            return
        self._ensure_multi_marker_state()
        flags = self._marker_flags()
        follows = self._marker_follows()
        if not (0 <= i < len(flags)) or not bool(flags[i]):
            messagebox.showinfo("Markers", "Selected panel is not designated as a marker panel (set this in Step 2a).")
            return

        follows[i] = int(i)
        marker_panels = [idx for idx, is_marker in enumerate(flags) if bool(is_marker)]
        if not marker_panels:
            messagebox.showinfo("Markers", "No marker panels are designated. Set them in Step 2a first.")
            return
        picks_map = self._marker_picks()
        had_all_marker_picks_before = all(len(list(picks_map.get(int(idx), []) or [])) == 3 for idx in marker_panels)

        # Start with the selected marker panel, then automatically continue with other marker panels
        # that do not yet have 3 picks (wrap-around order).
        start_pos = marker_panels.index(int(i)) if int(i) in marker_panels else 0
        ordered = marker_panels[start_pos:] + marker_panels[:start_pos]
        pending_after_first = [idx for idx in ordered[1:] if len(list(picks_map.get(int(idx), []) or [])) != 3]
        cycle_order = [int(i)] + pending_after_first
        cancelled_cycle = False

        for k, panel_idx in enumerate(cycle_order):
            try:
                self.img_list.selection_clear(0, "end")
                self.img_list.selection_set(int(panel_idx))
                self.img_list.activate(int(panel_idx))
                self._on_select()
            except Exception:
                pass
            title = f"Pick 3 markers - Panel {int(panel_idx) + 1}"
            ys = self._pick_three_marker_ys(int(panel_idx), title)
            if not ys:
                # User cancelled mid-cycle: keep already stored picks.
                cancelled_cycle = True
                break
            picks_map[int(panel_idx)] = [float(v) for v in sorted(ys)]
            follows[int(panel_idx)] = int(panel_idx)
            if int(getattr(self.app.state, "multi_source_marker_ref_panel", -1)) < 0:
                self.app.state.multi_source_marker_ref_panel = int(panel_idx)

            # Continue automatically only after the first panel as requested.
            if k == 0 and pending_after_first:
                continue

        self._refresh_multi_marker_controls()
        self._sync_selected_panel_marker_controls()
        self._refresh_list()

        # Auto-apply scaling once the last designated marker panel has been completed.
        # Only trigger on the transition from "incomplete" to "all complete" and not on cancel.
        if not cancelled_cycle:
            has_all_marker_picks_after = all(len(list(picks_map.get(int(idx), []) or [])) == 3 for idx in marker_panels)
            if (not had_all_marker_picks_before) and has_all_marker_picks_after:
                self._apply_all_marker_group_scaling()

    def _clear_markers_for_selected_panel(self):
        i = self._selected_index()
        if i is None:
            return
        self._ensure_multi_marker_state()
        self._marker_picks().pop(int(i), None)
        self._refresh_multi_marker_controls()
        self._sync_selected_panel_marker_controls()
        self._refresh_list()

    def _source_groups(self) -> Dict[int, List[int]]:
        self._ensure_source_origin_meta()
        groups: Dict[int, List[int]] = {}
        for idx, oid in enumerate(self.app.state.multi_source_origin_ids):
            key = int(oid)
            groups.setdefault(key, []).append(int(idx))
        return groups

    def _source_name_for_id(self, source_id: int) -> str:
        sid = int(source_id)
        ids = self.app.state.multi_source_origin_ids
        names = self.app.state.multi_source_origin_names
        for i, oid in enumerate(ids):
            if int(oid) == sid and i < len(names) and names[i]:
                return str(names[i])
        return f"Source {sid + 1}"

    def _refresh_source_scale_controls(self):
        if (not hasattr(self, "ref_source_combo")) or (not hasattr(self, "tgt_source_combo")):
            return
        groups = self._source_groups()
        ordered_ids = sorted(groups.keys())
        self._source_label_to_origin = {}
        labels: List[str] = []
        for sid in ordered_ids:
            name = self._source_name_for_id(sid)
            label = f"{sid + 1}: {name}"
            labels.append(label)
            self._source_label_to_origin[label] = int(sid)

        self.ref_source_combo["values"] = labels
        self.tgt_source_combo["values"] = labels

        if labels:
            if self.ref_source_var.get() not in self._source_label_to_origin:
                self.ref_source_var.set(labels[0])
            if self.tgt_source_var.get() not in self._source_label_to_origin:
                self.tgt_source_var.set(labels[1] if len(labels) > 1 else labels[0])
        else:
            self.ref_source_var.set("")
            self.tgt_source_var.set("")
        self._update_source_panel_spins()

    def _update_source_panel_spins(self):
        if (not hasattr(self, "ref_panel_spin")) or (not hasattr(self, "tgt_panel_spin")):
            return
        groups = self._source_groups()
        ref_sid = self._source_label_to_origin.get(self.ref_source_var.get())
        tgt_sid = self._source_label_to_origin.get(self.tgt_source_var.get())

        ref_n = len(groups.get(int(ref_sid), [])) if ref_sid is not None else 1
        tgt_n = len(groups.get(int(tgt_sid), [])) if tgt_sid is not None else 1
        ref_n = max(1, int(ref_n))
        tgt_n = max(1, int(tgt_n))

        self.ref_panel_spin.configure(from_=1, to=ref_n)
        self.tgt_panel_spin.configure(from_=1, to=tgt_n)
        self.ref_panel_num.set(max(1, min(ref_n, int(self.ref_panel_num.get()))))
        self.tgt_panel_num.set(max(1, min(tgt_n, int(self.tgt_panel_num.get()))))

    def _transform_cache_key(self, idx: int, img: Image.Image, pl: ImagePlacement) -> Tuple[Any, ...]:
        return (
            int(idx),
            int(id(img)),
            int(img.width),
            int(img.height),
            round(float(getattr(pl, "angle_deg", 0.0)), 4),
            round(float(getattr(pl, "scale_x", 1.0)), 4),
            round(float(getattr(pl, "scale_y", 1.0)), 4),
            round(float(getattr(pl, "brightness", 0.0)), 3),
            round(float(getattr(pl, "contrast", 1.0)), 3),
            round(float(getattr(pl, "gamma", 1.0)), 3),
            bool(getattr(pl, "white_balance", False)),
            bool(getattr(pl, "invert", False)),
            bool(getattr(pl, "bw", True)),
        )

    def _get_transformed_cached(self, idx: int, img: Image.Image, pl: ImagePlacement) -> Image.Image:
        key = self._transform_cache_key(idx, img, pl)
        got = self._transform_cache.get(key)
        if got is not None:
            return got
        out = _transform_for_compose(img, pl, bg=(255, 255, 255))
        self._transform_cache[key] = out
        # Simple cap to avoid unbounded growth when many edits happen.
        if len(self._transform_cache) > 256:
            self._transform_cache.clear()
            self._transform_cache[key] = out
        return out

    def _compose_preview_with_cache(self) -> Tuple[Image.Image, List[Dict[str, int]]]:
        images = self.app.state.multi_source_images
        placements = self.app.state.multi_source_placements
        bg = (255, 255, 255)
        if not images:
            return Image.new("RGB", (800, 600), bg), []
        if len(placements) != len(images):
            placements = [ImagePlacement() for _ in images]

        transformed: List[Tuple[int, Image.Image, int, int]] = []
        min_x = 0
        min_y = 0
        max_x = 1
        max_y = 1

        for idx, (img, pl) in enumerate(zip(images, placements)):
            rot = self._get_transformed_cached(idx, img, pl)
            x = int(pl.x)
            y = int(pl.y)
            transformed.append((idx, rot, x, y))
            min_x = min(min_x, x)
            min_y = min(min_y, y)
            max_x = max(max_x, x + rot.width)
            max_y = max(max_y, y + rot.height)

        out = Image.new("RGB", (max_x - min_x, max_y - min_y), bg)
        ox = -min_x
        oy = -min_y
        meta: List[Dict[str, int]] = []
        for idx, rot, x, y in transformed:
            out.paste(rot, (x + ox, y + oy))
            x0 = int(x + ox)
            y0 = int(y + oy)
            meta.append(
                dict(
                    index=int(idx),
                    x=int(x),
                    y=int(y),
                    w=int(rot.width),
                    h=int(rot.height),
                    x0=x0,
                    y0=y0,
                    x1=int(x0 + rot.width),
                    y1=int(y0 + rot.height),
                )
            )
        return out, meta

    def _refresh_list(self):
        self._ensure_multi_marker_state()
        self.img_list.delete(0, "end")
        names = [pathlib.Path(p).name for p in self.app.state.multi_source_paths]
        if not names:
            names = [f"Image {i+1}" for i in range(len(self.app.state.multi_source_images))]
        flags = self._marker_flags()
        follows = self._marker_follows()
        picks = self._marker_picks()
        group_ids = [int(v) for v in list(getattr(self.app.state, "multi_source_marker_group_ids", []) or [])]
        ref_idx = int(getattr(self.app.state, "multi_source_marker_ref_panel", -1) or -1)
        for i, n in enumerate(names, start=1):
            idx = i - 1
            tags: List[str] = []
            if idx < len(group_ids) and int(group_ids[idx]) > 0:
                tags.append(f"G{int(group_ids[idx])}")
            if idx < len(flags) and bool(flags[idx]):
                tags.append("M")
            if idx in picks and len(picks[idx]) == 3:
                tags.append("3p")
            if idx == ref_idx:
                tags.append("REF")
            fi = int(follows[idx]) if idx < len(follows) else -1
            if fi >= 0 and fi != idx:
                tags.append(f"->P{fi + 1}")
            tag_txt = f" [{' | '.join(tags)}]" if tags else ""
            self.img_list.insert("end", f"{i}. {n}{tag_txt}")
        self._fit_listbox_height(self.img_list, min_rows=4, max_rows=18)

    def _selected_index(self) -> Optional[int]:
        sel = self.img_list.curselection()
        if not sel:
            return None
        i = int(sel[0])
        if 0 <= i < len(self.app.state.multi_source_images):
            return i
        return None

    def _on_select(self):
        i = self._selected_index()
        if i is None:
            return
        self._update_y_lock_ui()
        pl = self.app.state.multi_source_placements[i]
        self.sel_x.set(int(pl.x))
        self.sel_y.set(int(pl.y))
        self.sel_angle.set(float(pl.angle_deg))
        self.sel_scale_x.set(float(getattr(pl, "scale_x", 1.0)))
        self.sel_scale_y.set(float(getattr(pl, "scale_y", 1.0)))
        self.sel_brightness.set(float(getattr(pl, "brightness", 0.0)))
        self.sel_contrast.set(float(getattr(pl, "contrast", 1.0)))
        self.sel_gamma.set(float(getattr(pl, "gamma", 1.0)))
        self.sel_wb.set(bool(getattr(pl, "white_balance", False)))
        self.sel_invert.set(bool(getattr(pl, "invert", False)))
        self.sel_bw.set(bool(getattr(pl, "bw", True)))
        self._sync_selected_panel_marker_controls()
        self._draw_overlay()

    def _apply_selected(self):
        i = self._selected_index()
        if i is None:
            return
        pl = self.app.state.multi_source_placements[i]
        locked_y = self._is_y_locked()
        try:
            pl.x = int(self.sel_x.get())
            if not locked_y:
                pl.y = int(self.sel_y.get())
            pl.angle_deg = float(self.sel_angle.get())
            pl.scale_x = max(0.05, min(20.0, float(self.sel_scale_x.get())))
            pl.scale_y = max(0.05, min(20.0, float(self.sel_scale_y.get())))
            pl.brightness = max(-80.0, min(80.0, float(self.sel_brightness.get())))
            pl.contrast = max(0.4, min(2.5, float(self.sel_contrast.get())))
            pl.gamma = max(0.4, min(2.5, float(self.sel_gamma.get())))
            pl.white_balance = bool(self.sel_wb.get())
            pl.invert = bool(self.sel_invert.get())
            pl.bw = bool(self.sel_bw.get())
        except Exception:
            return
        self.sel_scale_x.set(float(pl.scale_x))
        self.sel_scale_y.set(float(pl.scale_y))
        self.sel_y.set(int(pl.y))
        self.sel_brightness.set(float(pl.brightness))
        self.sel_contrast.set(float(pl.contrast))
        self.sel_gamma.set(float(pl.gamma))
        self._render_preview()

    def _reset_selected_color(self):
        i = self._selected_index()
        if i is None:
            return
        pl = self.app.state.multi_source_placements[i]
        pl.brightness = 0.0
        pl.contrast = 1.0
        pl.gamma = 1.0
        pl.white_balance = False
        pl.invert = False
        pl.bw = True
        self.sel_brightness.set(0.0)
        self.sel_contrast.set(1.0)
        self.sel_gamma.set(1.0)
        self.sel_wb.set(False)
        self.sel_invert.set(False)
        self.sel_bw.set(True)
        self._render_preview()

    def _rotate_selected(self, delta: float):
        i = self._selected_index()
        if i is None:
            return
        pl = self.app.state.multi_source_placements[i]
        pl.angle_deg = float(pl.angle_deg) + float(delta)
        self.sel_angle.set(float(pl.angle_deg))
        self._render_preview()

    def _scale_selected(self, mx: float, my: float):
        i = self._selected_index()
        if i is None:
            return
        pl = self.app.state.multi_source_placements[i]
        pl.scale_x = max(0.05, min(20.0, float(getattr(pl, "scale_x", 1.0)) * float(mx)))
        pl.scale_y = max(0.05, min(20.0, float(getattr(pl, "scale_y", 1.0)) * float(my)))
        self.sel_scale_x.set(float(pl.scale_x))
        self.sel_scale_y.set(float(pl.scale_y))
        self._render_preview()

    def _auto_arrange(self):
        imgs = self.app.state.multi_source_images
        pls = self.app.state.multi_source_placements
        locked_y = self._is_y_locked()
        x = 0
        order = sorted(
            range(min(len(imgs), len(pls))),
            key=lambda i: (float(getattr(pls[i], "x", 0)), float(getattr(pls[i], "y", 0)), int(i)),
        )
        for i in order:
            img = imgs[i]
            pl = pls[i]
            rot = self._get_transformed_cached(i, img, pl)
            pls[i].x = x
            if not locked_y:
                pls[i].y = 0
            x += rot.width
        self._on_select()
        self._render_preview()

    def _pick_three_marker_ys(self, panel_idx: int, title: str) -> Optional[List[float]]:
        if panel_idx < 0 or panel_idx >= len(self.app.state.multi_source_images):
            return None
        img = _transform_for_compose(
            self.app.state.multi_source_images[panel_idx],
            self.app.state.multi_source_placements[panel_idx],
            bg=(255, 255, 255),
        )

        win = tk.Toplevel(self)
        win.title(title)
        try:
            win.attributes("-fullscreen", True)
        except Exception:
            try:
                win.state("zoomed")
            except Exception:
                fit_toplevel_to_screen(win, 1200, 900)

        body = ttk.Frame(win)
        body.pack(fill="both", expand=True, padx=8, pady=8)
        vp = ZoomableCanvas(body, bg="#222222", show_tools=False)
        vp.pack(fill="both", expand=True)
        vp.set_image(img, fit_if_needed=True)
        c = vp.canvas
        # Disable zoom/pan in this picker to keep click positions and overlays stable.
        for seq in (
            "<MouseWheel>",
            "<Shift-MouseWheel>",
            "<Button-4>",
            "<Button-5>",
            "<ButtonPress-3>",
            "<B3-Motion>",
            "<ButtonPress-2>",
            "<B2-Motion>",
        ):
            try:
                c.unbind(seq)
            except Exception:
                pass

        status = tk.StringVar(value="Click marker band 1 / 3")
        picks: List[float] = []
        star_items: List[int] = []
        cancelled = {"v": False}

        info = ttk.Frame(win)
        info.pack(fill="x", padx=8, pady=(0, 8))
        ttk.Label(info, textvariable=status).pack(side="left")

        def on_click(event):
            if len(picks) >= 3:
                return
            _ix, iy = vp.canvas_to_image(event.x, event.y)
            y = float(max(0, min(img.height - 1, int(iy))))
            picks.append(y)
            try:
                cx = float(event.x)
                cy = float(event.y)
                r = 6.0
                star_items.extend(
                    [
                        c.create_line(cx - r, cy, cx + r, cy, fill="#D80000", width=2),
                        c.create_line(cx, cy - r, cx, cy + r, fill="#D80000", width=2),
                        c.create_line(cx - r * 0.75, cy - r * 0.75, cx + r * 0.75, cy + r * 0.75, fill="#D80000", width=2),
                        c.create_line(cx - r * 0.75, cy + r * 0.75, cx + r * 0.75, cy - r * 0.75, fill="#D80000", width=2),
                        c.create_text(cx + 8, cy - 8, text=str(len(picks)), anchor="sw", fill="#B00000"),
                    ]
                )
            except Exception:
                pass
            status.set(f"Picked {len(picks)} / 3")
            if len(picks) == 3:
                status.set("Picked 3 / 3. Click Done.")

        c.bind("<ButtonPress-1>", on_click)

        footer = ttk.Frame(win)
        footer.pack(fill="x", padx=8, pady=(0, 8))

        def reset():
            picks.clear()
            try:
                for it in list(star_items):
                    c.delete(it)
            except Exception:
                pass
            star_items.clear()
            status.set("Click marker band 1 / 3")

        def cancel():
            cancelled["v"] = True
            win.destroy()

        def done():
            if len(picks) != 3:
                messagebox.showinfo("Markers", "Please click exactly 3 marker bands.", parent=win)
                return
            win.destroy()

        ttk.Button(footer, text="Reset picks", command=reset).pack(side="left")
        ttk.Button(footer, text="Cancel", command=cancel).pack(side="left", padx=6)
        ttk.Button(footer, text="Done", command=done).pack(side="right")
        try:
            win.bind("<Escape>", lambda _e=None: cancel())
            win.bind("<F11>", lambda _e=None: None)
        except Exception:
            pass

        win.transient(self.winfo_toplevel())
        win.grab_set()
        win.wait_window()

        if cancelled["v"] or len(picks) != 3:
            return None
        out = sorted([float(v) for v in picks])
        return out

    def _marker_scale_factor_from_triplets(self, ref_ys: List[float], tgt_ys: List[float]) -> Optional[float]:
        try:
            r = [float(v) for v in list(ref_ys or [])]
            t = [float(v) for v in list(tgt_ys or [])]
        except Exception:
            return None
        if len(r) != 3 or len(t) != 3:
            return None
        r = sorted(r)
        t = sorted(t)
        pairs = [(0, 1), (1, 2), (0, 2)]
        ratios: List[float] = []
        for a, b in pairs:
            dr = float(r[b] - r[a])
            dt = float(t[b] - t[a])
            if abs(dt) > 1e-6 and abs(dr) > 1e-6:
                ratios.append(dr / dt)
        if not ratios:
            return None
        factor = float(np.median(np.asarray(ratios, dtype=float)))
        if not np.isfinite(factor) or factor <= 0:
            return None
        return max(0.1, min(10.0, factor))

    def _apply_all_marker_group_scaling(self):
        self._ensure_multi_marker_state()
        self._ensure_source_origin_meta()
        self._ensure_panel_source_rect_meta()
        n = len(self.app.state.multi_source_images)
        if n <= 0:
            return

        flags = self._marker_flags()
        follows = self._marker_follows()
        picks = self._marker_picks()
        placements = self.app.state.multi_source_placements
        origin_ids = [int(v) for v in list(getattr(self.app.state, "multi_source_origin_ids", []) or [])]
        source_rects = self._panel_source_rects()

        marker_panels = [i for i in range(n) if bool(flags[i])]
        if not marker_panels:
            messagebox.showinfo("Marker scaling", "Mark at least one panel as a marker panel.")
            return

        ref_idx = int(getattr(self.app.state, "multi_source_marker_ref_panel", -1) or -1)
        if ref_idx < 0 or ref_idx >= n or not bool(flags[ref_idx]):
            ref_with_picks = [i for i in marker_panels if i in picks and len(picks[i]) == 3]
            if not ref_with_picks:
                messagebox.showinfo("Marker scaling", "Pick 3 marker bands for at least one marker panel first.")
                return
            ref_idx = int(ref_with_picks[0])
            self.app.state.multi_source_marker_ref_panel = int(ref_idx)

        if ref_idx not in picks or len(picks.get(ref_idx, [])) != 3:
            messagebox.showinfo("Marker scaling", "The global reference marker panel needs 3 picked marker bands.")
            return

        # Normalize follows: non-marker panels without explicit assignment default to same-source marker if available.
        for i in range(n):
            fi = int(follows[i]) if i < len(follows) else -1
            if fi < 0 or fi >= n or not bool(flags[fi]):
                follows[i] = int(self._default_marker_follow_for_panel(i))
            if bool(flags[i]) and (follows[i] < 0 or not bool(flags[follows[i]])):
                follows[i] = int(i)

        ref_ys = [float(v) for v in picks[int(ref_idx)]]
        ref_pl = placements[int(ref_idx)]

        # Group panels by controlling marker panel.
        groups_by_marker: Dict[int, List[int]] = {}
        for panel_idx, marker_idx in enumerate(follows):
            mi = int(marker_idx)
            if 0 <= mi < n and bool(flags[mi]):
                groups_by_marker.setdefault(mi, []).append(int(panel_idx))

        applied_msgs: List[str] = []
        skipped_msgs: List[str] = []

        for marker_idx in sorted(groups_by_marker.keys()):
            if marker_idx == int(ref_idx):
                marker_picks = list(ref_ys)
                factor = 1.0
            else:
                marker_picks = picks.get(int(marker_idx))
                if not marker_picks or len(marker_picks) != 3:
                    skipped_msgs.append(f"Panel {marker_idx + 1}: missing 3 marker picks")
                    continue

                factor = self._marker_scale_factor_from_triplets(ref_ys, marker_picks)
                if factor is None:
                    skipped_msgs.append(f"Panel {marker_idx + 1}: invalid scale factor")
                    continue

            panel_group = list(groups_by_marker.get(int(marker_idx), []))
            if int(marker_idx) not in panel_group:
                panel_group.append(int(marker_idx))

            # Apply Y scaling to all followers of this marker panel.
            for pi in panel_group:
                pl = placements[int(pi)]
                pl.scale_y = max(0.05, min(20.0, float(getattr(pl, "scale_y", 1.0)) * float(factor)))
                # Keep stored local marker picks in sync with local y scaling if this follower also has picks.
                if int(pi) in picks and len(picks[int(pi)]) == 3:
                    picks[int(pi)] = [float(v) * float(factor) for v in list(picks[int(pi)])]

            # Align this marker group's Y by matching the middle picked marker to the reference marker panel.
            tgt_mid_before = float(sorted([float(v) for v in marker_picks])[1])
            tgt_mid_after = float(tgt_mid_before) * float(factor)
            # Prefer actual post-scale stored picks (absolute current panel-local coords) over
            # the incremental estimate to avoid drift when scaling is applied repeatedly.
            try:
                tgt_now = [float(v) for v in list(picks.get(int(marker_idx), []) or [])]
                if len(tgt_now) == 3:
                    tgt_mid_after = float(sorted(tgt_now)[1])
            except Exception:
                pass
            ref_mid = float(sorted(ref_ys)[1])
            try:
                ref_now = [float(v) for v in list(picks.get(int(ref_idx), []) or [])]
                if len(ref_now) == 3:
                    ref_mid = float(sorted(ref_now)[1])
            except Exception:
                pass
            tgt_pl = placements[int(marker_idx)]
            delta_y = (float(ref_pl.y) + ref_mid) - (float(tgt_pl.y) + tgt_mid_after)
            # First move the controlling marker panel to the matched Y.
            tgt_pl.y = int(round(float(tgt_pl.y) + float(delta_y)))

            # Propagate to follower panels. Panels from the same original source as the controller
            # get a source-relative Y target based on their original crop-top position so identical
            # bands line up across side-by-side panels after scaling.
            marker_src_id = int(origin_ids[int(marker_idx)]) if int(marker_idx) < len(origin_ids) else -1
            marker_src_y0 = 0
            if int(marker_idx) < len(source_rects):
                try:
                    marker_src_y0 = int(source_rects[int(marker_idx)][1])
                except Exception:
                    marker_src_y0 = 0
            marker_y_after = int(tgt_pl.y)
            marker_scale_y_after = float(getattr(tgt_pl, "scale_y", 1.0))

            for pi in panel_group:
                pi = int(pi)
                if pi == int(marker_idx):
                    continue
                if pi < 0 or pi >= len(placements):
                    continue
                same_source = (pi < len(origin_ids)) and (int(origin_ids[pi]) == int(marker_src_id))
                if same_source and (pi < len(source_rects)):
                    try:
                        panel_src_y0 = int(source_rects[pi][1])
                        target_y = float(marker_y_after) + float(marker_scale_y_after) * float(panel_src_y0 - marker_src_y0)
                        placements[pi].y = int(round(target_y))
                        continue
                    except Exception:
                        pass
                placements[pi].y = int(round(float(placements[pi].y) + float(delta_y)))

            applied_msgs.append(
                f"Marker panel {marker_idx + 1}: factor {factor:.4f} applied to {len(panel_group)} panel(s)"
            )

        if not applied_msgs:
            msg = "No marker groups were applied."
            if skipped_msgs:
                msg += "\n" + "\n".join(skipped_msgs[:8])
            messagebox.showinfo("Marker scaling", msg)
            self._refresh_multi_marker_controls()
            self._sync_selected_panel_marker_controls()
            self._refresh_list()
            return

        self.app.state.multi_source_lock_y = True
        self._update_y_lock_ui()
        self._on_select()
        self._refresh_multi_marker_controls()
        self._sync_selected_panel_marker_controls()
        self._refresh_list()
        self._render_preview()

        summary = "\n".join(applied_msgs[:10])
        if skipped_msgs:
            summary += "\n\nSkipped:\n" + "\n".join(skipped_msgs[:6])
        messagebox.showinfo("Marker scaling", summary)

    def _auto_scale_target_source_from_markers(self):
        messagebox.showinfo(
            "Auto-scale",
            "Legacy source-level auto-scale has been removed.\nUse Step 2a panel marker/group assignments and the compose-step marker buttons instead.",
        )

    def _crop_composed_dialog(self):
        composite = compose_images(self.app.state.multi_source_images, self.app.state.multi_source_placements)
        if composite is None:
            return

        win = tk.Toplevel(self)
        win.title("Final crop composed result")
        fit_toplevel_to_screen(win, 1000, 700)

        body = ttk.Frame(win)
        body.pack(fill="both", expand=True, padx=8, pady=8)

        vp = ZoomableCanvas(body, bg="#222222", show_tools=True)
        vp.pack(fill="both", expand=True)
        vp.set_image(composite, fit_if_needed=True)
        c = vp.canvas

        state: Dict[str, Any] = {"start": None, "rect_id": None, "crop": None}

        def on_start(event):
            state["start"] = (c.canvasx(event.x), c.canvasy(event.y))
            if state["rect_id"] is not None:
                try:
                    c.delete(state["rect_id"])
                except Exception:
                    pass
                state["rect_id"] = None

        def on_drag(event):
            if state["start"] is None:
                return
            x0, y0 = state["start"]
            x1, y1 = c.canvasx(event.x), c.canvasy(event.y)
            if state["rect_id"] is None:
                state["rect_id"] = c.create_rectangle(x0, y0, x1, y1, outline="yellow", width=2)
            else:
                c.coords(state["rect_id"], x0, y0, x1, y1)

        def on_end(event):
            if state["start"] is None:
                return
            x0, y0 = state["start"]
            x1, y1 = c.canvasx(event.x), c.canvasy(event.y)
            ix0, iy0 = vp.canvas_to_image(min(x0, x1), min(y0, y1))
            ix1, iy1 = vp.canvas_to_image(max(x0, x1), max(y0, y1))
            ix0 = max(0, min(composite.width - 1, int(ix0)))
            iy0 = max(0, min(composite.height - 1, int(iy0)))
            ix1 = max(0, min(composite.width, int(ix1)))
            iy1 = max(0, min(composite.height, int(iy1)))
            if abs(ix1 - ix0) < 6 or abs(iy1 - iy0) < 6:
                state["crop"] = None
            else:
                state["crop"] = (ix0, iy0, ix1, iy1)

        c.bind("<ButtonPress-1>", on_start)
        c.bind("<B1-Motion>", on_drag)
        c.bind("<ButtonRelease-1>", on_end)

        footer = ttk.Frame(win)
        footer.pack(fill="x", padx=8, pady=(0, 8))

        def apply_crop():
            if not state["crop"]:
                messagebox.showinfo("Crop", "Draw a crop rectangle first.", parent=win)
                return
            x0, y0, x1, y1 = state["crop"]
            cropped = composite.crop((x0, y0, x1, y1))
            self.app.state.multi_source_final_crop_undo = self._snapshot_compose_state_for_final_crop_undo()
            self.app.state.multi_source_images = [cropped]
            self.app.state.multi_source_paths = ["Composed crop"]
            self.app.state.multi_source_placements = [ImagePlacement()]
            self.app.state.multi_source_origin_ids = [0]
            self.app.state.multi_source_origin_names = ["Composed"]
            self.app.state.multi_source_panel_source_rects = [(0, 0, int(cropped.width), int(cropped.height))]
            self._transform_cache.clear()
            self._preview_initialized = False
            self._refresh_list()
            self._refresh_multi_marker_controls()
            if self.img_list.size() > 0:
                self.img_list.selection_clear(0, "end")
                self.img_list.selection_set(0)
                self._on_select()
            self._render_preview()
            win.destroy()

        ttk.Button(footer, text="Cancel", command=win.destroy).pack(side="left")
        ttk.Button(footer, text="Apply crop", command=apply_crop).pack(side="right")

    def _crop_selected_dialog(self):
        i = self._selected_index()
        if i is None:
            messagebox.showinfo("Crop", "Select an image first.")
            return

        pl = self.app.state.multi_source_placements[i]
        base = _transform_for_compose(self.app.state.multi_source_images[i], pl, bg=(255, 255, 255))
        win = tk.Toplevel(self)
        win.title("Crop selected image")
        fit_toplevel_to_screen(win, 1000, 700)

        body = ttk.Frame(win)
        body.pack(fill="both", expand=True, padx=8, pady=8)

        vp = ZoomableCanvas(body, bg="#222222", show_tools=True)
        vp.pack(fill="both", expand=True)
        vp.set_image(base, fit_if_needed=True)
        c = vp.canvas

        state: Dict[str, Any] = {"start": None, "rect_id": None, "crop": None}

        def on_start(event):
            state["start"] = (c.canvasx(event.x), c.canvasy(event.y))
            if state["rect_id"] is not None:
                try:
                    c.delete(state["rect_id"])
                except Exception:
                    pass
                state["rect_id"] = None

        def on_drag(event):
            if state["start"] is None:
                return
            x0, y0 = state["start"]
            x1, y1 = c.canvasx(event.x), c.canvasy(event.y)
            if state["rect_id"] is None:
                state["rect_id"] = c.create_rectangle(x0, y0, x1, y1, outline="yellow", width=2)
            else:
                c.coords(state["rect_id"], x0, y0, x1, y1)

        def on_end(event):
            if state["start"] is None:
                return
            x0, y0 = state["start"]
            x1, y1 = c.canvasx(event.x), c.canvasy(event.y)
            ix0, iy0 = vp.canvas_to_image(min(x0, x1), min(y0, y1))
            ix1, iy1 = vp.canvas_to_image(max(x0, x1), max(y0, y1))
            ix0 = max(0, min(base.width - 1, int(ix0)))
            iy0 = max(0, min(base.height - 1, int(iy0)))
            ix1 = max(0, min(base.width, int(ix1)))
            iy1 = max(0, min(base.height, int(iy1)))
            if abs(ix1 - ix0) < 6 or abs(iy1 - iy0) < 6:
                state["crop"] = None
            else:
                state["crop"] = (ix0, iy0, ix1, iy1)

        c.bind("<ButtonPress-1>", on_start)
        c.bind("<B1-Motion>", on_drag)
        c.bind("<ButtonRelease-1>", on_end)

        footer = ttk.Frame(win)
        footer.pack(fill="x", padx=8, pady=(0, 8))

        def apply_crop():
            if not state["crop"]:
                messagebox.showinfo("Crop", "Draw a crop rectangle first.", parent=win)
                return
            x0, y0, x1, y1 = state["crop"]
            cropped = base.crop((x0, y0, x1, y1))
            try:
                self._ensure_panel_source_rect_meta()
                rects = list(getattr(self.app.state, "multi_source_panel_source_rects", []) or [])
                if 0 <= int(i) < len(rects):
                    old_x0, old_y0, old_x1, old_y1 = [int(v) for v in rects[int(i)]]
                    if (
                        abs(float(getattr(pl, "angle_deg", 0.0))) < 1e-6
                        and abs(float(getattr(pl, "scale_x", 1.0)) - 1.0) < 1e-6
                        and abs(float(getattr(pl, "scale_y", 1.0)) - 1.0) < 1e-6
                    ):
                        rects[int(i)] = (old_x0 + int(x0), old_y0 + int(y0), old_x0 + int(x1), old_y0 + int(y1))
                    self.app.state.multi_source_panel_source_rects = rects
            except Exception:
                pass
            self.app.state.multi_source_images[i] = cropped
            pl.x = int(pl.x) + int(x0)
            pl.y = int(pl.y) + int(y0)
            pl.angle_deg = 0.0
            pl.scale_x = 1.0
            pl.scale_y = 1.0
            pl.brightness = 0.0
            pl.contrast = 1.0
            pl.gamma = 1.0
            pl.white_balance = False
            pl.invert = False
            pl.bw = True
            self._transform_cache.clear()
            self._on_select()
            self._render_preview()
            win.destroy()

        ttk.Button(footer, text="Cancel", command=win.destroy).pack(side="left")
        ttk.Button(footer, text="Apply crop", command=apply_crop).pack(side="right")

    def _drag_start(self, event):
        ix, iy = self.viewport.canvas_to_image(event.x, event.y)
        # hit-test topmost image first (later images are pasted later)
        hit_idx = None
        for m in reversed(self._layout_meta):
            if int(m["x0"]) <= ix <= int(m["x1"]) and int(m["y0"]) <= iy <= int(m["y1"]):
                hit_idx = int(m["index"])
                break
        if hit_idx is None:
            self._drag_idx = None
            self._drag_anchor_img = None
            self._drag_start_xy = None
            return

        self.img_list.selection_clear(0, "end")
        self.img_list.selection_set(hit_idx)
        self._on_select()

        pl = self.app.state.multi_source_placements[hit_idx]
        self._drag_idx = hit_idx
        self._drag_anchor_img = (int(ix), int(iy))
        self._drag_start_xy = (int(pl.x), int(pl.y))

    def _drag_move(self, event):
        if self._drag_idx is None or self._drag_anchor_img is None or self._drag_start_xy is None:
            return
        ix, iy = self.viewport.canvas_to_image(event.x, event.y)
        ax, ay = self._drag_anchor_img
        sx, sy = self._drag_start_xy
        pl = self.app.state.multi_source_placements[self._drag_idx]
        pl.x = int(sx + (ix - ax))
        if not self._is_y_locked():
            pl.y = int(sy + (iy - ay))
        if self._selected_index() == self._drag_idx:
            self.sel_x.set(int(pl.x))
            self.sel_y.set(int(pl.y))
        self._schedule_render_preview(delay_ms=16)

    def _drag_end(self, _event):
        self._drag_idx = None
        self._drag_anchor_img = None
        self._drag_start_xy = None
        self._cancel_scheduled_render()
        self._render_preview()

    def _draw_overlay(self):
        try:
            self.canvas.delete("compose_overlay")
        except Exception:
            pass
        if not self._layout_meta:
            return
        z = float(self.viewport.zoom)
        sel = self._selected_index()
        for m in self._layout_meta:
            x0 = float(m["x0"]) * z
            y0 = float(m["y0"]) * z
            x1 = float(m["x1"]) * z
            y1 = float(m["y1"]) * z
            idx = int(m["index"])
            if sel is not None and idx == sel:
                outline = "yellow"
                w = 3
            else:
                outline = "cyan"
                w = 1
            self.canvas.create_rectangle(x0, y0, x1, y1, outline=outline, width=w, tags=("compose_overlay",))
            self.canvas.create_text(x0 + 6, y0 + 6, text=str(idx + 1), anchor="nw", fill=outline, tags=("compose_overlay",))

    def _render_preview(self):
        self._render_after_id = None
        img, meta = self._compose_preview_with_cache()
        self._layout_meta = meta
        self.viewport.set_image(img, fit_if_needed=not self._preview_initialized)
        self._preview_initialized = True
        self._draw_overlay()

    def _next(self):
        self._cancel_scheduled_render()
        composite = compose_images(self.app.state.multi_source_images, self.app.state.multi_source_placements)
        self.app.state.loaded_pil = composite.copy()
        self.app.state.original_pil = composite.copy()
        self.app.state.edited_pil = composite.copy()
        self.app.state.panels = [composite]
        self.app.state.panel_configs = [PanelConfig()]
        self.app.state.current_panel_index = 0
        self.app.show_frame("LayoutFrame")

    def _back(self):
        self._cancel_scheduled_render()
        if (
            self.app.state.settings.source_mode == "multi_image"
            and self.app.state.settings.panel_mode == "multi_panel"
            and self.app.state.multi_input_images
        ):
            self.app.show_frame("MultiSourcePanelSelectFrame")
            return
        self.app.show_frame("SettingsFrame")


# ----------------------------
# Step 3: Edit image
# ----------------------------
class EditFrame(ttk.Frame):
    def __init__(self, parent, app: GelAnnotatorApp):
        super().__init__(parent)
        self.app = app

        ttk.Label(self, text="Step 3 - Edit (contrast, color, crop, rotate)", font=("Arial", 18, "bold")).pack(pady=8)

        main = ttk.Frame(self)
        main.pack(fill="both", expand=True)

        # Left: controls
        controls = ttk.Frame(main)
        controls.pack(side="left", fill="y", padx=10, pady=10)

        self.brightness = tk.DoubleVar(value=0.0)
        self.contrast = tk.DoubleVar(value=1.0)
        self.gamma = tk.DoubleVar(value=1.0)
        self.wb = tk.BooleanVar(value=False)
        self.invert = tk.BooleanVar(value=self.app.state.settings.prefer_invert)
        self.bw = tk.BooleanVar(value=True)
        self.rotate_deg = tk.DoubleVar(value=0.0)
        self.zoom = tk.DoubleVar(value=1.0)
        self._refresh_after_id: Optional[str] = None
        self._undo_stack: List[Image.Image] = []

        def add_slider(label, var, from_, to, resolution=0.1, fmt=".2f"):
            ttk.Label(controls, text=label).pack(anchor="w", pady=(8, 2))
            s = ttk.Scale(controls, variable=var, from_=from_, to=to, orient="horizontal", command=lambda _e=None: self._schedule_preview())
            s.pack(fill="x")
            disp_var = tk.StringVar(value=f"{var.get():{fmt}}")
            def _update_label(*_):
                try:
                    disp_var.set(f"{var.get():{fmt}}")
                except Exception:
                    pass
            var.trace_add("write", _update_label)
            ttk.Label(controls, textvariable=disp_var).pack(anchor="w")

        add_slider("Brightness (beta)", self.brightness, -80, 80, 1.0, fmt=".0f")
        add_slider("Contrast (alpha)", self.contrast, 0.4, 2.5, 0.05, fmt=".2f")
        add_slider("Gamma", self.gamma, 0.4, 2.5, 0.05, fmt=".2f")

        ttk.Checkbutton(controls, text="Gray-world white balance", variable=self.wb, command=self._schedule_preview).pack(anchor="w", pady=6)
        ttk.Checkbutton(controls, text="Invert", variable=self.invert, command=self._schedule_preview).pack(anchor="w", pady=6)
        ttk.Checkbutton(controls, text="B/W (grayscale)", variable=self.bw, command=self._schedule_preview).pack(anchor="w", pady=6)

        ttk.Label(controls, text="Rotate (degrees):").pack(anchor="w", pady=(10, 2))
        rot_row = ttk.Frame(controls)
        rot_row.pack(fill="x")
        ttk.Entry(rot_row, textvariable=self.rotate_deg, width=8).pack(side="left")
        ttk.Button(rot_row, text="Apply", command=self.refresh_preview).pack(side="left", padx=6)
        rot_btn_row = ttk.Frame(controls)
        rot_btn_row.pack(fill="x", pady=(2, 0))
        for txt, delta in (("-90", -90.0), ("+90", 90.0), ("-1", -1.0), ("+1", 1.0)):
            ttk.Button(
                rot_btn_row,
                text=txt,
                width=6,
                command=(lambda d=delta: self._rotate90(d)),
            ).pack(side="left", padx=2, pady=1)

        ttk.Separator(controls).pack(fill="x", pady=10)

        ttk.Label(controls, text="Crop tool: drag a rectangle on the image, then click 'Crop'.").pack(anchor="w", pady=2)
        ttk.Button(controls, text="Crop", command=self.apply_crop).pack(fill="x", pady=4)
        ttk.Button(controls, text="Reset edits", command=self.reset_edits).pack(fill="x", pady=4)

        ttk.Separator(controls).pack(fill="x", pady=10)

        nav = ttk.Frame(controls)
        nav.pack(fill="x", pady=6)
        ttk.Button(nav, text="Back", command=lambda: self.app.show_frame("SettingsFrame")).pack(side="left")
        ttk.Button(nav, text="Next -> Panels / Layout", style="Primary.TButton", command=self.on_next).pack(side="right")

        nav2 = ttk.Frame(controls)
        nav2.pack(fill="x", pady=(0, 10))
        ttk.Button(nav2, text="Select multiple panels", command=self.on_next_panels).pack(fill="x")

        # Right: image canvas
        viewer = ttk.Frame(main)
        viewer.pack(side="left", fill="both", expand=True, padx=10, pady=10)

        self.viewport = ZoomableCanvas(viewer, bg="#222222", show_tools=True)
        self.viewport.pack(fill="both", expand=True)
        self.canvas = self.viewport.canvas

        self.canvas.bind("<ButtonPress-1>", self._crop_start)
        self.canvas.bind("<B1-Motion>", self._crop_drag)
        self.canvas.bind("<ButtonRelease-1>", self._crop_end)
        self.canvas.bind("<Button-3>", self._right_click_next, add="+")
        self.canvas.bind("<Button-2>", self._right_click_next, add="+")

        self._crop_rect_id = None
        self._crop_start_xy = None
        self._crop_box_img = None  # (x0,y0,x1,y1) in image pixels

        self._tk_img = None
        self._preview_pil = None  # current adjusted image for preview
        self._preview_initialized = False

    def on_show(self):
        if self.app.state.original_pil is None:
            self.app.show_frame("SettingsFrame")
            return

        self.brightness.set(0.0)
        self.contrast.set(1.0)
        self.gamma.set(1.0)
        self.wb.set(False)
        self.invert.set(self.app.state.settings.prefer_invert)
        self.bw.set(True)
        self.rotate_deg.set(0.0)
        self._crop_box_img = None
        self._preview_initialized = False
        self.refresh_preview()

    def _schedule_preview(self, delay_ms: int = 120):
        if self._refresh_after_id:
            try:
                self.after_cancel(self._refresh_after_id)
            except Exception:
                pass
        self._refresh_after_id = self.after(delay_ms, self.refresh_preview)

    def _push_undo(self) -> None:
        img = self.app.state.original_pil
        if img is not None:
            self._undo_stack.append(img.copy())
            if len(self._undo_stack) > 10:
                self._undo_stack.pop(0)

    def undo(self) -> None:
        if not self._undo_stack:
            return
        self.app.state.original_pil = self._undo_stack.pop()
        self.refresh_preview()

    def _rotate90(self, deg: int):
        self._push_undo()
        self.rotate_deg.set(self.rotate_deg.get() + deg)
        self.refresh_preview()

    def reset_edits(self):
        # Restore original (un-cropped) image and reset sliders / crop.
        if self.app.state.loaded_pil is not None:
            self.app.state.original_pil = self.app.state.loaded_pil.copy()
        self._crop_box_img = None
        if self._crop_rect_id:
            try:
                self.canvas.delete(self._crop_rect_id)
            except Exception:
                pass
            self._crop_rect_id = None
        self.on_show()

    def refresh_preview(self):
        img = self.app.state.original_pil
        if img is None:
            return

        bgr = pil_to_cv(img)

        # WB
        if self.wb.get():
            bgr = gray_world_white_balance(bgr)

        # contrast/brightness
        alpha = float(self.contrast.get())
        beta = float(self.brightness.get())
        bgr = cv2.convertScaleAbs(bgr, alpha=alpha, beta=beta)

        # gamma
        bgr = apply_gamma(bgr, float(self.gamma.get()))
        # B/W
        if self.bw.get():
            bgr = grayscale_with_red_as_black(bgr)

        # invert
        if self.invert.get():
            bgr = 255 - bgr

        # rotate
        angle = float(self.rotate_deg.get())
        if abs(angle) > 1e-6:
            h, w = bgr.shape[:2]
            M = cv2.getRotationMatrix2D((w / 2, h / 2), angle, 1.0)
            cos = abs(M[0, 0])
            sin = abs(M[0, 1])
            nW = int((h * sin) + (w * cos))
            nH = int((h * cos) + (w * sin))
            M[0, 2] += (nW / 2) - w / 2
            M[1, 2] += (nH / 2) - h / 2
            bgr = cv2.warpAffine(bgr, M, (nW, nH), borderValue=(255, 255, 255))

        pil = cv_to_pil(bgr)
        self._preview_pil = pil

        self.viewport.set_image(pil, fit_if_needed=not self._preview_initialized)
        self._preview_initialized = True

        # reset crop rectangle overlay (preview redraw clears canvas)
        self._crop_rect_id = None

    def _canvas_to_img_xy(self, x_canvas: int, y_canvas: int) -> Tuple[int, int]:
        if self._preview_pil is None:
            return (0, 0)
        # x_canvas/y_canvas are canvas coordinates (already account for scrolling)
        x = x_canvas / max(1e-9, self.viewport.zoom)
        y = y_canvas / max(1e-9, self.viewport.zoom)
        return (clamp_int(x), clamp_int(y))

    def _crop_start(self, event):
        self._crop_start_xy = (self.canvas.canvasx(event.x), self.canvas.canvasy(event.y))
        if self._crop_rect_id:
            self.canvas.delete(self._crop_rect_id)
            self._crop_rect_id = None

    def _crop_drag(self, event):
        if not self._crop_start_xy:
            return
        x0, y0 = self._crop_start_xy
        x1, y1 = self.canvas.canvasx(event.x), self.canvas.canvasy(event.y)
        if self._crop_rect_id:
            self.canvas.coords(self._crop_rect_id, x0, y0, x1, y1)
        else:
            self._crop_rect_id = self.canvas.create_rectangle(x0, y0, x1, y1, outline="yellow", width=2)

    def _crop_end(self, event):
        if not self._crop_start_xy or self._preview_pil is None:
            return
        x0, y0 = self._crop_start_xy
        x1, y1 = self.canvas.canvasx(event.x), self.canvas.canvasy(event.y)
        ix0, iy0 = self._canvas_to_img_xy(min(x0, x1), min(y0, y1))
        ix1, iy1 = self._canvas_to_img_xy(max(x0, x1), max(y0, y1))
        if abs(ix1 - ix0) < 5 or abs(iy1 - iy0) < 5:
            self._crop_box_img = None
        else:
            self._crop_box_img = (ix0, iy0, ix1, iy1)

    def apply_crop(self):
        if self._preview_pil is None:
            return
        if not self._crop_box_img:
            messagebox.showinfo("Crop", "No crop rectangle selected.")
            return
        self._push_undo()
        x0, y0, x1, y1 = self._crop_box_img
        cropped = self._preview_pil.crop((x0, y0, x1, y1))
        # Commit the current transformed image as new baseline and reset controls
        # so transforms are not applied a second time after crop.
        self.app.state.original_pil = cropped
        self.brightness.set(0.0)
        self.contrast.set(1.0)
        self.gamma.set(1.0)
        self.wb.set(False)
        self.invert.set(False)
        self.bw.set(True)
        self.rotate_deg.set(0.0)
        self._preview_initialized = False
        self.refresh_preview()
        self._crop_box_img = None

    def on_next(self):
        if self._preview_pil is None:
            return
        self.app.state.edited_pil = self._preview_pil.copy()

        if self.app.state.settings.panel_mode == "multi_panel" and self.app.state.settings.source_mode == "single_image":
            self.app.show_frame("PanelSelectFrame")
        else:
            # Single panel: panels = [edited]
            self.app.state.panels = [self.app.state.edited_pil]
            self.app.ensure_panel_configs()
            self.app.state.current_panel_index = 0
            self.app.show_frame("LayoutFrame")

    def _right_click_next(self, _event=None):
        self.on_next()
        return "break"

    def on_next_panels(self):
        """Always go to panel selection to extract multiple blocks from the same image."""
        if self._preview_pil is None:
            return
        self.app.state.edited_pil = self._preview_pil.copy()
        if self.app.state.settings.source_mode != "single_image":
            messagebox.showinfo("Panels", "Panel selection requires 'Single image' source mode.")
            return
        self.app.show_frame("PanelSelectFrame")


# ----------------------------
# Step 3b: Select multiple panels from one edited image
# ----------------------------

class PanelSelectFrame(ttk.Frame):
    def __init__(self, parent, app: GelAnnotatorApp):
        super().__init__(parent)
        self.app = app

        ttk.Label(self, text="Step 3b - Define panels (multi-panel mode)", font=("Arial", 18, "bold")).pack(pady=8)

        top = ttk.Frame(self)
        top.pack(fill="both", expand=True)

        # left: zoomable viewer
        self.viewport = ZoomableCanvas(top, bg="#222222", show_tools=True)
        self.viewport.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        self.canvas = self.viewport.canvas

        right = ttk.Frame(top)
        right.pack(side="left", fill="y", padx=10, pady=10)

        ttk.Label(right, text="Panels").pack(anchor="w")
        self.listbox = tk.Listbox(right, height=12)
        self.listbox.pack(fill="x", pady=6)

        ttk.Button(right, text="Remove selected", command=self.remove_selected).pack(fill="x", pady=4)
        ttk.Button(right, text="Clear all", command=self.clear_all).pack(fill="x", pady=4)

        ttk.Separator(right).pack(fill="x", pady=10)

        nav = ttk.Frame(right)
        nav.pack(fill="x", pady=10)
        ttk.Button(nav, text="Back", command=lambda: self.app.show_frame("EditFrame")).pack(side="left")
        ttk.Button(nav, text="Next -> Layout", style="Primary.TButton", command=self.on_next).pack(side="right")

        ttk.Label(
            right,
            text="How to:\nDrag (left mouse) to draw a rectangle.\nRelease to add a panel crop.\n(You can add multiple panels.)",
            justify="left",
        ).pack(anchor="w", pady=8)

        self._rect_id = None
        self._start_canvas_xy = None  # in canvas coords
        self._rects_img: List[Tuple[int, int, int, int]] = []

        self.canvas.bind("<ButtonPress-1>", self._start)
        self.canvas.bind("<B1-Motion>", self._drag)
        self.canvas.bind("<ButtonRelease-1>", self._end)

    def on_show(self):
        if self.app.state.edited_pil is None:
            self.app.show_frame("EditFrame")
            return
        self._rects_img = []
        self.listbox.delete(0, "end")
        self._render()

    def _render(self):
        pil = self.app.state.edited_pil
        if pil is None:
            return
        self.viewport.set_image(pil, fit_if_needed=True)

        # draw stored rectangles (overlay)
        z = self.viewport.zoom
        for (x0, y0, x1, y1) in self._rects_img:
            cx0, cy0 = x0 * z, y0 * z
            cx1, cy1 = x1 * z, y1 * z
            self.canvas.create_rectangle(cx0, cy0, cx1, cy1, outline="cyan", width=2, tags=("overlay",))

    def _canvas_to_img(self, x_canvas: float, y_canvas: float) -> Tuple[int, int]:
        z = max(1e-9, float(self.viewport.zoom))
        return clamp_int(x_canvas / z), clamp_int(y_canvas / z)

    def _start(self, event):
        self._start_canvas_xy = (self.canvas.canvasx(event.x), self.canvas.canvasy(event.y))
        if self._rect_id:
            try:
                self.canvas.delete(self._rect_id)
            except Exception:
                pass
            self._rect_id = None

    def _drag(self, event):
        if not self._start_canvas_xy:
            return
        x0, y0 = self._start_canvas_xy
        x1, y1 = self.canvas.canvasx(event.x), self.canvas.canvasy(event.y)
        if self._rect_id:
            self.canvas.coords(self._rect_id, x0, y0, x1, y1)
        else:
            self._rect_id = self.canvas.create_rectangle(x0, y0, x1, y1, outline="yellow", width=2, tags=("overlay",))

    def _end(self, event):
        if not self._start_canvas_xy:
            return
        x0c, y0c = self._start_canvas_xy
        x1c, y1c = self.canvas.canvasx(event.x), self.canvas.canvasy(event.y)

        # convert to image pixels
        x0, y0 = self._canvas_to_img(min(x0c, x1c), min(y0c, y1c))
        x1, y1 = self._canvas_to_img(max(x0c, x1c), max(y0c, y1c))

        if abs(x1 - x0) < 10 or abs(y1 - y0) < 10:
            # too small; ignore
            self._start_canvas_xy = None
            if self._rect_id:
                self.canvas.delete(self._rect_id)
                self._rect_id = None
            return

        self._rects_img.append((int(x0), int(y0), int(x1), int(y1)))
        self.listbox.insert("end", f"Panel {len(self._rects_img)}: ({x0},{y0})({x1},{y1})")
        self._start_canvas_xy = None
        if self._rect_id:
            self.canvas.delete(self._rect_id)
            self._rect_id = None
        self._render()

    def remove_selected(self):
        sel = list(self.listbox.curselection())
        if not sel:
            return
        idx = sel[0]
        if 0 <= idx < len(self._rects_img):
            self._rects_img.pop(idx)
        self.listbox.delete(0, "end")
        for i, r in enumerate(self._rects_img, start=1):
            self.listbox.insert("end", f"Panel {i}: ({r[0]},{r[1]})({r[2]},{r[3]})")
        self._render()

    def clear_all(self):
        self._rects_img = []
        self.listbox.delete(0, "end")
        self._render()

    def on_next(self):
        pil = self.app.state.edited_pil
        if pil is None:
            self.app.show_frame("EditFrame")
            return
        if not self._rects_img:
            messagebox.showinfo("Panels", "No panels defined. Please draw at least one rectangle.")
            return

        crops: List[Image.Image] = []
        for r in self._rects_img:
            x0, y0, x1, y1 = r
            crop = pil.crop((x0, y0, x1, y1))
            crops.append(crop)

        if len(crops) >= 2:
            self.app.state.multi_source_paths = [f"Panel {i+1}" for i in range(len(crops))]
            self.app.state.multi_source_images = crops
            self.app.state.multi_source_placements = []
            src_name = pathlib.Path(self.app.state.original_path).stem if self.app.state.original_path else "Image1"
            self.app.state.multi_source_origin_ids = [0 for _ in crops]
            self.app.state.multi_source_origin_names = [src_name for _ in crops]
            self.app.state.multi_source_lock_y = False
            self.app.show_frame("ComposeFrame")
            return

        self.app.state.panels = crops
        self.app.state.panel_configs = [PanelConfig() for _ in crops]
        self.app.state.current_panel_index = 0
        self.app.show_frame("LayoutFrame")

    def _right_click_next(self, _event=None):
        self.on_next()
        return "break"


class LayoutFrame(PanelAwareFrame):
    def __init__(self, parent, app: GelAnnotatorApp):
        super().__init__(parent)
        self.app = app

        self.title = ttk.Label(self, text="Step 4 - Panel layout", font=("Arial", 18, "bold"))
        self.title.pack(pady=8)

        main = ttk.Frame(self)
        main.pack(fill="both", expand=True)

        # left viewer (zoom/pan)
        self.viewport = ZoomableCanvas(main, bg="#222222", show_tools=True)
        self.viewport.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        self.canvas = self.viewport.canvas

        # right config
        right = ttk.Frame(main)
        right.pack(side="left", fill="y", padx=10, pady=10)

        self.lanes = tk.IntVar(value=10)
        ttk.Label(right, text="Number of lanes/columns:").pack(anchor="w")
        ttk.Spinbox(right, from_=1, to=60, textvariable=self.lanes, width=10, command=self._apply_lanes).pack(anchor="w", pady=4)

        ttk.Separator(right).pack(fill="x", pady=8)

        self.include_marker = tk.BooleanVar(value=True)
        self.ladder_lane = tk.IntVar(value=1)
        self.enable_highlight_step = tk.BooleanVar(value=False)
        ttk.Checkbutton(right, text="Include marker (ladder)", variable=self.include_marker).pack(anchor="w", pady=4)
        ladder_row = ttk.Frame(right)
        ladder_row.pack(fill="x", pady=2)
        ttk.Label(ladder_row, text="Ladder lane index:").pack(side="left")
        ttk.Spinbox(ladder_row, from_=1, to=60, textvariable=self.ladder_lane, width=6).pack(side="left", padx=6)

        ttk.Separator(right).pack(fill="x", pady=8)

        ttk.Label(
            right,
            text="Gel regions:\nDrag across the gel width (left-to-right).\nClick 'Set gel region' again to add another region.\nDrag cyan handles to resize.",
            justify="left",
        ).pack(anchor="w", pady=4)
        self._gel_btn = tk.Button(right, text="Set gel region", command=self._set_gel_region_mode)
        self._gel_btn.pack(fill="x", pady=2)
        self._gel_btn_default_bg = str(self._gel_btn.cget("bg"))
        ttk.Button(right, text="Clear gel region", command=self._clear_gel_region).pack(fill="x", pady=2)
        ttk.Label(right, text="Region lane counts (per region):").pack(anchor="w", pady=(6, 2))
        self.region_lane_list = tk.Listbox(right, height=4, exportselection=False)
        self.region_lane_list.pack(fill="x", pady=2)
        self.region_lane_list.bind("<<ListboxSelect>>", lambda _e=None: self._on_region_lane_select())
        rl_row = ttk.Frame(right)
        rl_row.pack(fill="x", pady=2)
        ttk.Label(rl_row, text="Selected region lanes:").pack(side="left")
        self.region_lane_var = tk.IntVar(value=1)
        self.region_lane_spin = ttk.Spinbox(rl_row, from_=1, to=60, textvariable=self.region_lane_var, width=6)
        self.region_lane_spin.pack(side="left", padx=6)
        ttk.Button(rl_row, text="Apply", command=self._apply_selected_region_lanes).pack(side="left")

        ttk.Separator(right).pack(fill="x", pady=8)
        ttk.Checkbutton(
            right,
            text="Open optional band highlight step before annotations",
            variable=self.enable_highlight_step,
        ).pack(anchor="w", pady=4)
        ttk.Label(
            right,
            text="If enabled, a separate step opens next where you can add box or arrow highlights.\nDefault OFF = no highlights are rendered.",
            justify="left",
        ).pack(anchor="w", pady=(0, 2))

        ttk.Separator(right).pack(fill="x", pady=10)

        nav = ttk.Frame(right)
        nav.pack(fill="x", pady=10)
        ttk.Button(nav, text="Back", command=self._back).pack(side="left")
        ttk.Button(nav, text="Next -> Annotations", style="Primary.TButton", command=self._next).pack(side="right")
        self._next_float_btn = tk.Button(self, text="Next -> Annotations", command=self._next, bg="#6ea84f", fg="white")
        self.bind("<Configure>", lambda _e=None: self._place_float_next())
        self._place_float_next()

        # internal
        self._mode = "gel_region"  # none | gel_region
        self._start_xy: Optional[Tuple[float, float]] = None  # in CANVAS coords (supports scroll)
        self._rect_id: Optional[int] = None
        self._drag_edge: Optional[str] = None  # left | right while resizing gel region handles
        self._drag_region_idx: Optional[int] = None

        self.canvas.bind("<ButtonPress-1>", self._start)
        self.canvas.bind("<B1-Motion>", self._drag)
        self.canvas.bind("<ButtonRelease-1>", self._end)
        self.canvas.bind("<Button-3>", self._right_click_next, add="+")
        self.canvas.bind("<Button-2>", self._right_click_next, add="+")
        self.canvas.bind("<<ZoomableCanvasRendered>>", lambda _e=None: self.after(0, self._draw_overlay_only), add="+")

    def on_show(self):
        if not self.app.state.panels:
            self.app.show_frame("SettingsFrame")
            return
        i = self.app.state.current_panel_index + 1
        n = len(self.app.state.panels)
        self.title.configure(text=f"Step 4 - Panel layout ({i}/{n})")

        pc = self.current_pc()
        self.lanes.set(pc.lanes)
        self.include_marker.set(pc.include_marker)
        self.ladder_lane.set(pc.marker_calibration.ladder_lane if pc.marker_calibration.ladder_lane else 1)
        self.enable_highlight_step.set(bool(getattr(pc, "highlight_enabled", False)))

        has_custom = bool(pc.gel_regions) or ((pc.gel_left is not None) and (pc.gel_right is not None))
        self._mode = "gel_region" if not has_custom else "none"
        self._start_xy = None
        self._rect_id = None
        self._drag_region_idx = None
        self._update_mode_ui()
        self._place_float_next()
        self._render()

    def _place_float_next(self):
        try:
            self.app.place_floating_next(self._next_float_btn)
        except Exception:
            pass

    def _get_custom_regions(self) -> List[Tuple[int, int]]:
        pc = self.current_pc()
        W = self.current_panel().width
        out: List[Tuple[int, int]] = []
        for reg in list(getattr(pc, "gel_regions", []) or []):
            try:
                a, b = int(reg[0]), int(reg[1])
            except Exception:
                continue
            if b < a:
                a, b = b, a
            a = max(0, min(W - 1, a))
            b = max(1, min(W, b))
            if b - a >= 2:
                out.append((a, b))
        if not out and (pc.gel_left is not None) and (pc.gel_right is not None):
            a = int(pc.gel_left)
            b = int(pc.gel_right)
            if b < a:
                a, b = b, a
            a = max(0, min(W - 1, a))
            b = max(1, min(W, b))
            if b - a >= 2:
                out.append((a, b))
        out.sort(key=lambda t: (t[0], t[1]))
        return out

    def _set_custom_regions(self, regions: List[Tuple[int, int]]) -> None:
        pc = self.current_pc()
        W = self.current_panel().width
        old_regs = self._get_custom_regions()
        old_counts_raw = list(getattr(pc, "gel_region_lane_counts", []) or [])
        old_counts: List[int] = []
        for i in range(len(old_regs)):
            try:
                c = int(old_counts_raw[i])
            except Exception:
                c = 1
            old_counts.append(max(1, c))

        cleaned: List[Tuple[int, int]] = []
        for a, b in regions:
            x0 = int(a)
            x1 = int(b)
            if x1 < x0:
                x0, x1 = x1, x0
            x0 = max(0, min(W - 1, x0))
            x1 = max(1, min(W, x1))
            if x1 - x0 >= 2:
                cleaned.append((x0, x1))
        cleaned.sort(key=lambda t: (t[0], t[1]))
        merged: List[Tuple[int, int]] = []
        for a, b in cleaned:
            if not merged:
                merged.append((a, b))
                continue
            pa, pb = merged[-1]
            if a <= pb:
                merged[-1] = (pa, max(pb, b))
            else:
                merged.append((a, b))
        pc.gel_regions = [(int(a), int(b)) for (a, b) in merged]
        if pc.gel_regions:
            new_counts: List[int] = []
            for a, b in pc.gel_regions:
                best_idx = -1
                best_ov = -1
                for i, (oa, ob) in enumerate(old_regs):
                    ov = max(0, min(b, ob) - max(a, oa))
                    if ov > best_ov:
                        best_ov = ov
                        best_idx = i
                if best_idx >= 0 and best_idx < len(old_counts):
                    new_counts.append(max(1, int(old_counts[best_idx])))
                else:
                    new_counts.append(1)
            pc.gel_region_lane_counts = new_counts
            pc.lanes = max(1, int(sum(pc.gel_region_lane_counts)))
            pc.gel_left = int(pc.gel_regions[0][0])
            pc.gel_right = int(pc.gel_regions[0][1])
        else:
            pc.gel_region_lane_counts = []
            pc.gel_left = None
            pc.gel_right = None
        self.lanes.set(int(pc.lanes))

    def _region_lane_counts_for_regions(self, regions: List[Tuple[int, int]]) -> List[int]:
        pc = self.current_pc()
        has_custom = bool(pc.gel_regions) or ((pc.gel_left is not None) and (pc.gel_right is not None))
        raw = list(getattr(pc, "gel_region_lane_counts", []) or [])
        counts: List[int] = []
        if has_custom and len(raw) == len(regions):
            ok = True
            for v in raw:
                try:
                    iv = int(v)
                except Exception:
                    ok = False
                    break
                if iv < 1:
                    ok = False
                    break
                counts.append(iv)
            if ok:
                return counts

        if not regions:
            return []
        widths = [max(1, b - a) for a, b in regions]
        total = max(len(regions), int(pc.lanes))
        counts = _allocate_lanes_to_regions(total, widths)
        pc.gel_region_lane_counts = [int(c) for c in counts]
        pc.lanes = max(1, int(sum(counts)))
        self.lanes.set(int(pc.lanes))
        return counts

    def _refresh_region_lane_list(self):
        regions = self._get_custom_regions()
        counts = self._region_lane_counts_for_regions(regions)
        prev = self.region_lane_list.curselection()
        prev_idx = int(prev[0]) if prev else None
        self.region_lane_list.delete(0, "end")
        for i, ((a, b), c) in enumerate(zip(regions, counts), start=1):
            self.region_lane_list.insert("end", f"R{i}: x={a}-{b}  lanes={int(c)}")
        if prev_idx is not None and 0 <= prev_idx < self.region_lane_list.size():
            self.region_lane_list.selection_set(prev_idx)
        elif self.region_lane_list.size() > 0:
            self.region_lane_list.selection_set(0)
        self._on_region_lane_select()

    def _on_region_lane_select(self):
        regions = self._get_custom_regions()
        counts = self._region_lane_counts_for_regions(regions)
        sel = self.region_lane_list.curselection()
        if not sel:
            return
        idx = int(sel[0])
        if 0 <= idx < len(counts):
            self.region_lane_var.set(int(counts[idx]))

    def _apply_selected_region_lanes(self):
        regions = self._get_custom_regions()
        if not regions:
            return
        counts = self._region_lane_counts_for_regions(regions)
        sel = self.region_lane_list.curselection()
        if not sel:
            messagebox.showinfo("Regions", "Select a region first.")
            return
        idx = int(sel[0])
        if not (0 <= idx < len(counts)):
            return
        try:
            val = max(1, int(float(self.region_lane_var.get())))
        except Exception:
            messagebox.showerror("Regions", "Please enter a valid lane count.")
            return
        counts[idx] = int(val)
        pc = self.current_pc()
        pc.gel_region_lane_counts = [int(c) for c in counts]
        pc.lanes = max(1, int(sum(counts)))
        self.lanes.set(int(pc.lanes))
        self._render()

    def _render(self):
        pil = self.current_panel()
        self.viewport.set_image(pil, fit_if_needed=True)
        self._draw_overlay_only()

    def _draw_overlay_only(self):
        if not self._has_current_panel():
            return
        try:
            self.canvas.delete("overlay")
        except Exception:
            pass
        pil = self.current_panel()
        pc = self.current_pc()
        z = float(self.viewport.zoom)
        Hc = pil.height * z
        lane_bounds = compute_lane_bounds(pil, pc)
        regions = get_panel_gel_regions(pil, pc)
        region_edges = {int(a) for a, _b in regions} | {int(b) for _a, b in regions}

        # lane boundaries across all regions, numbered continuously from left to right.
        boundaries: Dict[int, str] = {}
        for x0, x1 in lane_bounds:
            boundaries[int(round(x0))] = "gray"
            boundaries[int(round(x1))] = "gray"
        for ex in region_edges:
            boundaries[int(round(ex))] = "orange"
        for bx in sorted(boundaries.keys()):
            cx = float(bx) * z
            self.canvas.create_line(cx, 0, cx, Hc, fill=boundaries[bx], width=1, tags=("overlay",))

        # Lane numbers in white circles centered beneath each lane.
        try:
            lane_centers = [0.5 * (float(a) + float(b)) for (a, b) in lane_bounds]
            radius = max(7.0, 10.0 * z)
            cy = float(Hc) + float(radius) + 5.0
            c_w = max(1, int(round(1.4 * z)))
            t_size = max(9, int(round(10 * z)))
            circ_outline, txt_color = _lane_circle_colors(self.app.state.settings)
            for i, cx_img in enumerate(lane_centers, start=1):
                cx = float(cx_img) * z
                self.canvas.create_oval(cx - radius, cy - radius, cx + radius, cy + radius, fill="white", outline=circ_outline, width=c_w, tags=("overlay",))
                self.canvas.create_text(cx, cy, text=str(i), fill=txt_color, font=("Arial", t_size, "bold"), tags=("overlay",))
        except Exception:
            pass

        # highlight annotations (shown only when enabled for this panel)
        if bool(getattr(pc, "highlight_enabled", False)):
            for h in pc.highlights:
                cx0, cy0 = h.x0 * z, h.y0 * z
                cx1, cy1 = h.x1 * z, h.y1 * z
                outline = h.color if getattr(h, "color", "") else "#DC0000"
                width_img = max(1, int(getattr(h, "width", 2)))
                width = _preview_stroke_width(width_img, z)
                kind = str(getattr(h, "kind", "box") or "box").lower()
                if kind == "arrow":
                    _draw_canvas_arrow(self.canvas, float(cx0), float(cy0), float(cx1), float(cy1), color=str(outline), width=int(width), tags=("overlay",))
                elif kind == "asterisk":
                    cx = 0.5 * (float(cx0) + float(cx1))
                    cy = 0.5 * (float(cy0) + float(cy1))
                    r = max(5.0, 0.5 * max(abs(float(cx1) - float(cx0)), abs(float(cy1) - float(cy0))))
                    _draw_canvas_asterisk(self.canvas, cx=float(cx), cy=float(cy), r=float(r), color=str(outline), width=int(width), tags=("overlay",))
                else:
                    self.canvas.create_rectangle(
                        cx0, cy0, cx1, cy1,
                        outline=outline,
                        width=width,
                        tags=("overlay",),
                    )

        # Custom gel region markers + handles.
        custom_regions = self._get_custom_regions()
        region_counts = self._region_lane_counts_for_regions(custom_regions)
        for idx, (rx0, rx1) in enumerate(custom_regions, start=1):
            cx0 = rx0 * z
            cx1 = rx1 * z
            self.canvas.create_rectangle(cx0, 0, cx1, Hc, outline="cyan", width=2, tags=("overlay",))
            self.canvas.create_rectangle(cx0 - 6, 0, cx0 + 6, 14, outline="cyan", fill="cyan", tags=("overlay",))
            self.canvas.create_rectangle(cx1 - 6, 0, cx1 + 6, 14, outline="cyan", fill="cyan", tags=("overlay",))
            lanes_txt = int(region_counts[idx - 1]) if idx - 1 < len(region_counts) else 0
            self.canvas.create_text(cx0 + 4, 18, text=f"R{idx} ({lanes_txt})", anchor="nw", fill="cyan", tags=("overlay",))

        # if actively drawing a rectangle, keep it (re-created by _drag)
        self._refresh_region_lane_list()

    def _img_xy(self, x_canvas: float, y_canvas: float) -> Tuple[int, int]:
        pil = self.current_panel()
        z = max(1e-9, float(self.viewport.zoom))
        ix = x_canvas / z
        iy = y_canvas / z
        ix = max(0, min(pil.width - 1, clamp_int(ix)))
        iy = max(0, min(pil.height - 1, clamp_int(iy)))
        return int(ix), int(iy)

    def _start(self, event):
        z = float(self.viewport.zoom)
        cx = self.canvas.canvasx(event.x)
        cy = self.canvas.canvasy(event.y)
        self._drag_edge = None
        self._drag_region_idx = None

        # Dragging existing region handles (any region).
        if self._mode == "none":
            regs = self._get_custom_regions()
            if regs and 0 <= cy <= 18:
                best: Optional[Tuple[float, int, str]] = None
                for i, (a, b) in enumerate(regs):
                    cx_left = a * z
                    cx_right = b * z
                    dl = abs(cx - cx_left)
                    dr = abs(cx - cx_right)
                    if dl <= 10:
                        cand = (dl, i, "left")
                        if best is None or cand[0] < best[0]:
                            best = cand
                    if dr <= 10:
                        cand = (dr, i, "right")
                        if best is None or cand[0] < best[0]:
                            best = cand
                if best is not None:
                    _d, idx, edge = best
                    self._start_xy = (cx, cy)
                    self._drag_region_idx = int(idx)
                    self._drag_edge = str(edge)
                    return

        self._start_xy = (cx, cy)
        if self._rect_id is not None:
            try:
                self.canvas.delete(self._rect_id)
            except Exception:
                pass
            self._rect_id = None

    def _drag(self, event):
        if not self._start_xy:
            return
        cx = self.canvas.canvasx(event.x)
        cy = self.canvas.canvasy(event.y)

        # Dragging gel region edge
        if self._drag_edge in ("left", "right") and self._drag_region_idx is not None:
            pil = self.current_panel()
            regs = self._get_custom_regions()
            idx = int(self._drag_region_idx)
            if not (0 <= idx < len(regs)):
                return
            ix, _ = self._img_xy(cx, cy)
            min_w = 10
            left_n = regs[idx - 1][1] + 1 if idx > 0 else 0
            right_n = regs[idx + 1][0] - 1 if idx + 1 < len(regs) else pil.width
            a, b = regs[idx]
            if self._drag_edge == "left":
                new_a = int(max(left_n, min(ix, b - min_w)))
                regs[idx] = (new_a, b)
            else:
                new_b = int(min(right_n, max(ix, a + min_w)))
                regs[idx] = (a, new_b)
            self._set_custom_regions(regs)
            self._render()
            return

        x0, y0 = self._start_xy
        x1, y1 = cx, cy

        if self._mode == "gel_region":
            if self._rect_id is None:
                self._rect_id = self.canvas.create_rectangle(x0, y0, x1, y1, outline="yellow", width=2, tags=("overlay",))
            else:
                self.canvas.coords(self._rect_id, x0, y0, x1, y1)

    def _end(self, event):
        if not self._start_xy:
            return
        if self._drag_edge:
            self._start_xy = None
            self._drag_edge = None
            self._drag_region_idx = None
            self._render()
            return

        x0c, y0c = self._start_xy
        x1c, y1c = self.canvas.canvasx(event.x), self.canvas.canvasy(event.y)

        if self._mode == "gel_region":
            ix0, _ = self._img_xy(min(x0c, x1c), min(y0c, y1c))
            ix1, _ = self._img_xy(max(x0c, x1c), max(y0c, y1c))
            if abs(ix1 - ix0) > 10:
                regs = self._get_custom_regions()
                regs.append((int(ix0), int(ix1)))
                self._set_custom_regions(regs)
            self._mode = "none"

        self._start_xy = None
        self._drag_region_idx = None
        if self._rect_id is not None:
            try:
                self.canvas.delete(self._rect_id)
            except Exception:
                pass
            self._rect_id = None

        self._update_mode_ui()
        self._render()

    def _apply_lanes(self):
        pc = self.current_pc()
        total = max(1, int(self.lanes.get()))
        regs = self._get_custom_regions()
        if regs:
            widths = [max(1, b - a) for a, b in regs]
            counts = _allocate_lanes_to_regions(max(len(regs), total), widths)
            pc.gel_region_lane_counts = [int(c) for c in counts]
            pc.lanes = max(1, int(sum(counts)))
            self.lanes.set(int(pc.lanes))
        else:
            pc.lanes = int(total)
        self._render()

    def _set_gel_region_mode(self):
        self._mode = "gel_region"
        self._drag_region_idx = None
        self._drag_edge = None
        self._update_mode_ui()

    def _clear_gel_region(self):
        self._set_custom_regions([])
        self._mode = "gel_region"
        self._update_mode_ui()
        self._render()

    def _update_mode_ui(self):
        try:
            if self._mode == "gel_region":
                self._gel_btn.configure(text="Set gel region (drag to add)")
                self._gel_btn.configure(bg="#6ea84f", fg="white", activebackground="#5f9443")
            else:
                self._gel_btn.configure(text="Set gel region")
                self._gel_btn.configure(bg=self._gel_btn_default_bg, fg="black")
        except Exception:
            pass

    def _sync_to_state(self):
        if not self._has_current_panel():
            return
        pc = self.current_pc()
        self._set_custom_regions(self._get_custom_regions())
        regs = self._get_custom_regions()
        if regs:
            counts = self._region_lane_counts_for_regions(regs)
            pc.gel_region_lane_counts = [int(c) for c in counts]
            pc.lanes = max(1, int(sum(counts)))
            self.lanes.set(int(pc.lanes))
        else:
            pc.gel_region_lane_counts = []
            pc.lanes = max(1, int(self.lanes.get()))
        pc.include_marker = bool(self.include_marker.get())
        pc.marker_calibration.ladder_lane = int(self.ladder_lane.get())
        pc.highlight_enabled = bool(self.enable_highlight_step.get())

    def _back(self):
        self._sync_to_state()
        if self.app.state.current_panel_index > 0:
            self.app.state.current_panel_index -= 1
            self.on_show()
        else:
            # back to Edit or panel select depending on mode
            if self.app.state.multi_source_images:
                self.app.show_frame("ComposeFrame")
            elif (
                self.app.state.settings.panel_mode == "multi_panel"
                and self.app.state.settings.source_mode == "single_image"
            ):
                self.app.show_frame("PanelSelectFrame")
            else:
                self.app.show_frame("EditFrame")

    def _next(self):
        self._sync_to_state()
        pc = self.current_pc()
        if bool(getattr(pc, "highlight_enabled", False)):
            self.app.show_frame("HighlightFrame")
        else:
            self.app.show_frame("AnnotationFrame")

    def _right_click_next(self, _event=None):
        self._next()
        return "break"


class HighlightFrame(ttk.Frame):
    def __init__(self, parent, app: GelAnnotatorApp):
        super().__init__(parent)
        self.app = app

        self.title = ttk.Label(self, text="Step 4b - Band highlights (optional)", font=("Arial", 18, "bold"))
        self.title.pack(pady=8)

        main = ttk.Frame(self)
        main.pack(fill="both", expand=True)

        self.viewport = ZoomableCanvas(main, bg="#222222", show_tools=True)
        self.viewport.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        self.canvas = self.viewport.canvas

        self.right_scroll = VScrollPanel(main, width=self.app._scaled_sidebar_width(SIDEBAR_WIDTH_NARROW))
        self.right_scroll.pack(side="left", fill="y", padx=10, pady=10)
        right = self.right_scroll.inner

        self.enabled_var = tk.BooleanVar(value=False)
        self.shape_var = tk.StringVar(value="box")
        self.highlight_color = tk.StringVar(value="#DC0000")
        self.highlight_width = tk.IntVar(value=3)
        self._mode = "none"  # none | draw
        self._start_xy: Optional[Tuple[float, float]] = None
        self._temp_items: List[int] = []

        ttk.Checkbutton(
            right,
            text="Render highlights for this panel",
            variable=self.enabled_var,
            command=self._on_enabled_changed,
        ).pack(anchor="w", pady=(0, 6))

        shape_row = ttk.Frame(right)
        shape_row.pack(fill="x", pady=2)
        ttk.Label(shape_row, text="Highlight style").pack(side="left")
        shape_combo = ttk.Combobox(shape_row, textvariable=self.shape_var, values=["box", "arrow", "asterisk"], width=10, state="readonly")
        shape_combo.pack(side="left", padx=6)
        shape_combo.bind("<<ComboboxSelected>>", lambda _e=None: self._on_style_changed())

        style_row = ttk.Frame(right)
        style_row.pack(fill="x", pady=2)
        ttk.Label(style_row, text="Color").pack(side="left")
        self._color_chip = tk.Label(style_row, width=2, relief="solid", bg=self.highlight_color.get())
        self._color_chip.pack(side="left", padx=6)
        ttk.Button(style_row, text="Choose...", command=self._choose_color).pack(side="left", padx=4)
        ttk.Label(style_row, text="Width").pack(side="left", padx=(8, 2))
        width_spin = ttk.Spinbox(style_row, from_=1, to=12, textvariable=self.highlight_width, width=4, command=self._on_style_changed)
        width_spin.pack(side="left")
        width_spin.bind("<KeyRelease>", lambda _e=None: self._on_style_changed())

        self._draw_btn = tk.Button(right, text="Draw mode: OFF", command=self._toggle_draw_mode)
        self._draw_btn.pack(fill="x", pady=4)
        ttk.Button(right, text="Apply style to existing highlights", command=self._apply_style_to_existing).pack(fill="x", pady=2)
        ttk.Button(right, text="Remove last highlight", command=self._remove_last).pack(fill="x", pady=2)
        ttk.Button(right, text="Clear all highlights", command=self._clear_all).pack(fill="x", pady=2)
        ttk.Label(
            right,
            text="Box: drag rectangle around a band.\nArrow: drag from tail to band.\nAsterisk: click/drag to set center and size.",
            justify="left",
        ).pack(anchor="w", pady=(6, 2))

        ttk.Separator(right).pack(fill="x", pady=10)
        nav = ttk.Frame(right)
        nav.pack(fill="x", pady=8)
        ttk.Button(nav, text="Back", command=self._back).pack(side="left")
        ttk.Button(nav, text="Next -> Annotations", style="Primary.TButton", command=self._next).pack(side="right")
        self._next_float_btn = tk.Button(self, text="Next -> Annotations", command=self._next, bg="#6ea84f", fg="white")
        self.bind("<Configure>", lambda _e=None: self._place_float_next())
        self._place_float_next()

        self.canvas.bind("<ButtonPress-1>", self._start)
        self.canvas.bind("<B1-Motion>", self._drag)
        self.canvas.bind("<ButtonRelease-1>", self._end)
        self.canvas.bind("<Button-3>", self._right_click_next, add="+")
        self.canvas.bind("<Button-2>", self._right_click_next, add="+")

    def _place_float_next(self):
        try:
            self.app.place_floating_next(self._next_float_btn)
        except Exception:
            pass

    def on_show(self):
        if not self.app.state.panels:
            self.app.show_frame("SettingsFrame")
            return
        i = self.app.state.current_panel_index + 1
        n = len(self.app.state.panels)
        self.title.configure(text=f"Step 4b - Band highlights (optional) ({i}/{n})")
        pc = self.current_pc()
        self.enabled_var.set(bool(getattr(pc, "highlight_enabled", False)))
        self.shape_var.set(str(getattr(pc, "highlight_shape", "box") or "box"))
        self.highlight_color.set(str(getattr(pc, "highlight_color", "#DC0000") or "#DC0000"))
        self.highlight_width.set(max(1, int(getattr(pc, "highlight_width", 3) or 3)))
        self._start_xy = None
        self._temp_items = []
        self._mode = "none"
        self._update_mode_ui()
        self._on_style_changed()
        self._place_float_next()
        self._render()

    def _img_xy(self, x_canvas: float, y_canvas: float) -> Tuple[int, int]:
        pil = self.current_panel()
        z = max(1e-9, float(self.viewport.zoom))
        ix = x_canvas / z
        iy = y_canvas / z
        ix = max(0, min(pil.width - 1, clamp_int(ix)))
        iy = max(0, min(pil.height - 1, clamp_int(iy)))
        return int(ix), int(iy)

    def _sync_to_state(self):
        if not self._has_current_panel():
            return
        pc = self.current_pc()
        pc.highlight_enabled = bool(self.enabled_var.get())
        shp = str(self.shape_var.get()).strip().lower()
        if shp not in {"box", "arrow", "asterisk"}:
            shp = "box"
        pc.highlight_shape = str(shp)
        pc.highlight_color = str(self.highlight_color.get()).strip() or "#DC0000"
        try:
            pc.highlight_width = max(1, int(self.highlight_width.get()))
        except Exception:
            pc.highlight_width = 3
            self.highlight_width.set(3)

    def _render(self):
        panel = self.current_panel()
        self.viewport.set_image(panel, fit_if_needed=True)
        self._draw_overlay_only()

    def _draw_overlay_only(self):
        if not self._has_current_panel():
            return
        try:
            self.canvas.delete("hl_overlay")
        except Exception:
            pass
        panel = self.current_panel()
        z = float(self.viewport.zoom)
        pc = self.current_pc()

        # Optional lane guides for easier placement.
        try:
            for x0, x1 in compute_lane_bounds(panel, pc):
                self.canvas.create_line(x0 * z, 0, x0 * z, panel.height * z, fill="gray", width=1, tags=("hl_overlay",))
                self.canvas.create_line(x1 * z, 0, x1 * z, panel.height * z, fill="gray", width=1, tags=("hl_overlay",))
        except Exception:
            pass

        if bool(getattr(pc, "highlight_enabled", False)):
            for h in pc.highlights:
                x0 = float(getattr(h, "x0", 0)) * z
                y0 = float(getattr(h, "y0", 0)) * z
                x1 = float(getattr(h, "x1", 0)) * z
                y1 = float(getattr(h, "y1", 0)) * z
                color = str(getattr(h, "color", "#DC0000") or "#DC0000")
                width_img = max(1, int(getattr(h, "width", 2)))
                width = _preview_stroke_width(width_img, z)
                kind = str(getattr(h, "kind", "box") or "box").lower()
                if kind == "arrow":
                    _draw_canvas_arrow(self.canvas, float(x0), float(y0), float(x1), float(y1), color=str(color), width=int(width), tags=("hl_overlay",))
                elif kind == "asterisk":
                    cx = 0.5 * (float(x0) + float(x1))
                    cy = 0.5 * (float(y0) + float(y1))
                    r = max(5.0, 0.5 * max(abs(float(x1) - float(x0)), abs(float(y1) - float(y0))))
                    _draw_canvas_asterisk(self.canvas, cx=float(cx), cy=float(cy), r=float(r), color=str(color), width=int(width), tags=("hl_overlay",))
                else:
                    self.canvas.create_rectangle(x0, y0, x1, y1, outline=color, width=width, tags=("hl_overlay",))

    def _on_enabled_changed(self):
        self._sync_to_state()
        self._update_mode_ui()
        self._render()

    def _on_style_changed(self):
        try:
            self._color_chip.configure(bg=self.highlight_color.get())
        except Exception:
            pass
        self._sync_to_state()
        self._render()

    def _choose_color(self):
        picked = colorchooser.askcolor(color=self.highlight_color.get(), title="Highlight color")
        if not picked or not picked[1]:
            return
        self.highlight_color.set(str(picked[1]))
        self._on_style_changed()

    def _toggle_draw_mode(self):
        if not bool(self.enabled_var.get()):
            self.enabled_var.set(True)
            self._sync_to_state()
        self._mode = "none" if self._mode == "draw" else "draw"
        self._update_mode_ui()

    def _update_mode_ui(self):
        try:
            enabled = bool(self.enabled_var.get())
            if self._mode == "draw" and enabled:
                self._draw_btn.configure(text="Draw mode: ON", bg="#6ea84f", fg="white", activebackground="#5f9443")
            elif not enabled:
                self._draw_btn.configure(text="Draw mode: OFF (highlights disabled)", bg="#d9d9d9", fg="black")
            else:
                self._draw_btn.configure(text="Draw mode: OFF", bg="#d9d9d9", fg="black")
        except Exception:
            pass

    def _apply_style_to_existing(self):
        pc = self.current_pc()
        self._sync_to_state()
        color = pc.highlight_color
        width = pc.highlight_width
        kind = pc.highlight_shape
        for h in pc.highlights:
            h.color = color
            h.width = int(width)
            h.kind = kind
        self._render()

    def _remove_last(self):
        pc = self.current_pc()
        if pc.highlights:
            pc.highlights.pop()
        self._render()

    def _clear_all(self):
        pc = self.current_pc()
        pc.highlights = []
        self._render()

    def _start(self, event):
        if self._mode != "draw" or not bool(self.enabled_var.get()):
            return
        self._start_xy = (self.canvas.canvasx(event.x), self.canvas.canvasy(event.y))
        for item_id in list(self._temp_items):
            try:
                self.canvas.delete(item_id)
            except Exception:
                pass
        self._temp_items = []

    def _drag(self, event):
        if self._mode != "draw" or not self._start_xy or not bool(self.enabled_var.get()):
            return
        x0, y0 = self._start_xy
        x1 = self.canvas.canvasx(event.x)
        y1 = self.canvas.canvasy(event.y)
        color = str(self.highlight_color.get()).strip() or "#DC0000"
        width_img = max(1, int(self.highlight_width.get()))
        width = _preview_stroke_width(width_img, float(self.viewport.zoom))
        shape = str(self.shape_var.get()).strip().lower()
        for item_id in list(self._temp_items):
            try:
                self.canvas.delete(item_id)
            except Exception:
                pass
        self._temp_items = []
        if shape == "arrow":
            self._temp_items = _draw_canvas_arrow(
                self.canvas,
                float(x0),
                float(y0),
                float(x1),
                float(y1),
                color=str(color),
                width=int(width),
                tags=("hl_overlay",),
            )
        elif shape == "asterisk":
            r = max(5.0, float(math.hypot(float(x1) - float(x0), float(y1) - float(y0))))
            self._temp_items = _draw_canvas_asterisk(
                self.canvas,
                cx=float(x0),
                cy=float(y0),
                r=float(r),
                color=str(color),
                width=int(width),
                tags=("hl_overlay",),
            )
        else:
            self._temp_items = [
                self.canvas.create_rectangle(
                    x0, y0, x1, y1,
                    outline=color,
                    width=width,
                    tags=("hl_overlay",),
                )
            ]

    def _end(self, event):
        if self._mode != "draw" or not self._start_xy or not bool(self.enabled_var.get()):
            return
        x0c, y0c = self._start_xy
        x1c = self.canvas.canvasx(event.x)
        y1c = self.canvas.canvasy(event.y)
        ix0, iy0 = self._img_xy(x0c, y0c)
        ix1, iy1 = self._img_xy(x1c, y1c)
        self._sync_to_state()
        pc = self.current_pc()
        kind = str(self.shape_var.get()).strip().lower()
        if kind not in {"box", "arrow", "asterisk"}:
            kind = "box"
        width = max(1, int(self.highlight_width.get()))
        color = str(self.highlight_color.get()).strip() or "#DC0000"
        if kind == "arrow":
            if abs(ix1 - ix0) + abs(iy1 - iy0) >= 6:
                pc.highlights.append(HighlightRect(ix0, iy0, ix1, iy1, width=width, color=color, kind="arrow"))
        elif kind == "asterisk":
            if abs(ix1 - ix0) + abs(iy1 - iy0) < 3:
                ix1 = int(ix0 + max(6, 3 * width))
                iy1 = int(iy0)
            pc.highlights.append(HighlightRect(int(ix0), int(iy0), int(ix1), int(iy1), width=width, color=color, kind="asterisk"))
        else:
            rx0, rx1 = sorted((int(ix0), int(ix1)))
            ry0, ry1 = sorted((int(iy0), int(iy1)))
            if abs(rx1 - rx0) > 6 and abs(ry1 - ry0) > 6:
                pc.highlights.append(HighlightRect(rx0, ry0, rx1, ry1, width=width, color=color, kind="box"))
        self._start_xy = None
        for item_id in list(self._temp_items):
            try:
                self.canvas.delete(item_id)
            except Exception:
                pass
        self._temp_items = []
        self._render()

    def _back(self):
        self._sync_to_state()
        self.app.show_frame("LayoutFrame")

    def _next(self):
        self._sync_to_state()
        self.app.show_frame("AnnotationFrame")

    def _right_click_next(self, _event=None):
        self._next()
        return "break"

class AnnotationFrame(PanelAwareFrame):
    def __init__(self, parent, app: GelAnnotatorApp):
        super().__init__(parent)
        self.app = app
        self._undo_stack: List[PanelConfig] = []

        self.title = ttk.Label(self, text="Step 5 - Header/column annotations", font=("Arial", 18, "bold"))
        self.title.pack(pady=8)

        main = ttk.Frame(self)
        main.pack(fill="both", expand=True)

        # left preview (zoom/pan)
        self.viewport = ZoomableCanvas(main, bg="#222222", show_tools=True)
        self.viewport.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        self.canvas = self.viewport.canvas

        # right: tables + groups + symbols
        self.right_scroll = VScrollPanel(main, width=self.app._scaled_sidebar_width(560))
        self.right_scroll.pack(side="left", fill="y", padx=10, pady=10)
        right = self.right_scroll.inner

        ttk.Label(right, text="Header rows (condition table):").pack(anchor="w")
        row_btns = ttk.Frame(right)
        row_btns.pack(fill="x", pady=4)
        ttk.Button(row_btns, text="Add row", command=self.add_row).pack(side="left", padx=4)
        ttk.Button(row_btns, text="Remove last", command=self.remove_row).pack(side="left", padx=4)
        self.header_value_angle = tk.IntVar(value=0)
        angle_row = ttk.Frame(right)
        angle_row.pack(fill="x", pady=2)
        ttk.Label(angle_row, text="Default row angle (deg, for new rows):").pack(side="left")
        angle_spin = ttk.Spinbox(
            angle_row,
            from_=-80,
            to=80,
            textvariable=self.header_value_angle,
            width=6,
            command=self._schedule_preview,
        )
        angle_spin.pack(side="left", padx=6)
        angle_spin.bind("<KeyRelease>", lambda _e=None: self._schedule_preview())

        ttk.Label(right, text="Position offsets (px):").pack(anchor="w", pady=(6, 2))
        self.header_heading_x_off = tk.IntVar(value=0)
        self.header_heading_y_off = tk.IntVar(value=0)
        self.header_values_x_off = tk.IntVar(value=0)
        self.header_values_y_off = tk.IntVar(value=0)
        self.bracket_label_x_off = tk.IntVar(value=0)
        self.bracket_label_y_off = tk.IntVar(value=0)
        self.bracket_line_x_off = tk.IntVar(value=0)
        self.bracket_line_y_off = tk.IntVar(value=0)

        def _offset_row(label: str, xvar: tk.IntVar, yvar: tk.IntVar):
            rr = ttk.Frame(right)
            rr.pack(fill="x", pady=1)
            ttk.Label(rr, text=label).pack(side="left")
            ttk.Label(rr, text="X").pack(side="left", padx=(8, 2))
            sx = ttk.Spinbox(rr, from_=-300, to=300, textvariable=xvar, width=5, command=self._schedule_preview)
            sx.pack(side="left")
            sx.bind("<KeyRelease>", lambda _e=None: self._schedule_preview())
            ttk.Label(rr, text="Y").pack(side="left", padx=(8, 2))
            sy = ttk.Spinbox(rr, from_=-300, to=300, textvariable=yvar, width=5, command=self._schedule_preview)
            sy.pack(side="left")
            sy.bind("<KeyRelease>", lambda _e=None: self._schedule_preview())

        _offset_row("Header headings", self.header_heading_x_off, self.header_heading_y_off)
        _offset_row("Header values", self.header_values_x_off, self.header_values_y_off)
        _offset_row("Bracket labels", self.bracket_label_x_off, self.bracket_label_y_off)
        _offset_row("Bracket lines", self.bracket_line_x_off, self.bracket_line_y_off)

        rows_wrap = ttk.Frame(right)
        rows_wrap.pack(fill="x", pady=6)
        self.rows_canvas = tk.Canvas(rows_wrap, height=190, highlightthickness=1, borderwidth=0)
        self.rows_canvas.pack(side="top", fill="x", expand=True)
        self.rows_xbar = ttk.Scrollbar(rows_wrap, orient="horizontal", command=self.rows_canvas.xview)
        self.rows_xbar.pack(side="top", fill="x")
        self.rows_canvas.configure(xscrollcommand=self.rows_xbar.set)
        self.rows_frame = ttk.Frame(self.rows_canvas)
        self._rows_window = self.rows_canvas.create_window((0, 0), window=self.rows_frame, anchor="nw")
        self.rows_frame.bind("<Configure>", self._on_rows_inner_configure)
        self.rows_canvas.bind("<Shift-MouseWheel>", self._on_rows_shift_wheel)
        self.rows_canvas.bind("<Button-4>", lambda _e=None: self.rows_canvas.xview_scroll(-3, "units"))
        self.rows_canvas.bind("<Button-5>", lambda _e=None: self.rows_canvas.xview_scroll(3, "units"))
        ttk.Label(right, text="Tip: Shift + mouse wheel scrolls left/right across lane text boxes.").pack(anchor="w", pady=(2, 0))

        ttk.Separator(right).pack(fill="x", pady=8)

        ttk.Label(right, text="Group labels (brackets spanning lanes):").pack(anchor="w")
        grp = ttk.Frame(right)
        grp.pack(fill="x", pady=4)
        self.grp_start = tk.IntVar(value=1)
        self.grp_end = tk.IntVar(value=3)
        self.grp_text = tk.StringVar(value="")
        self.grp_height_target = tk.StringVar(value="New height")
        self._height_choice_map: Dict[str, Optional[str]] = {}
        self.bracket_gap = tk.IntVar(value=10)
        ttk.Label(grp, text="From").grid(row=0, column=0, padx=2)
        ttk.Spinbox(grp, from_=1, to=60, textvariable=self.grp_start, width=5).grid(row=0, column=1, padx=2)
        ttk.Label(grp, text="to").grid(row=0, column=2, padx=2)
        ttk.Spinbox(grp, from_=1, to=60, textvariable=self.grp_end, width=5).grid(row=0, column=3, padx=2)
        ttk.Entry(grp, textvariable=self.grp_text, width=18).grid(row=0, column=4, padx=4)
        ttk.Button(grp, text="Add", command=self.add_group).grid(row=0, column=5, padx=2)
        grp2 = ttk.Frame(right)
        grp2.pack(fill="x", pady=(0, 4))
        ttk.Label(grp2, text="Bracket height:").pack(side="left")
        self.grp_height_combo = ttk.Combobox(grp2, textvariable=self.grp_height_target, state="readonly", width=34)
        self.grp_height_combo.pack(side="left", padx=6)

        gap_row = ttk.Frame(right)
        gap_row.pack(fill="x", pady=2)
        ttk.Label(gap_row, text="Bracket text distance:").pack(side="left")
        gap_spin = ttk.Spinbox(gap_row, from_=0, to=120, textvariable=self.bracket_gap, width=6, command=self._schedule_preview)
        gap_spin.pack(side="left", padx=6)
        gap_spin.bind("<KeyRelease>", lambda _e=None: self._schedule_preview())

        region_row = ttk.Frame(right)
        region_row.pack(fill="x", pady=(4, 2))
        self._bracket_select_btn = ttk.Button(region_row, text="Bracket select tool: OFF", command=self.toggle_bracket_select_mode)
        self._bracket_select_btn.pack(side="left", padx=2)
        ttk.Button(region_row, text="Add selected regions", command=self.add_groups_from_selected_regions).pack(side="left", padx=2)
        ttk.Button(region_row, text="Clear selected", command=self.clear_selected_regions).pack(side="left", padx=2)
        ttk.Label(
            right,
            text="Drag on preview to select bracket lane ranges.\nCtrl+drag adds more ranges. All selected ranges can be added with one label.",
            justify="left",
        ).pack(anchor="w", pady=(0, 2))

        self.group_list = tk.Listbox(right, height=6)
        self.group_list.pack(fill="x", pady=4)
        ttk.Button(right, text="Remove selected group", command=self.remove_group).pack(fill="x", pady=2)

        ttk.Separator(right).pack(fill="x", pady=8)
        ttk.Label(right, text="Top annotation order (headers + brackets):").pack(anchor="w")
        self.order_list = tk.Listbox(right, height=8)
        self.order_list.pack(fill="x", pady=4)
        order_btns = ttk.Frame(right)
        order_btns.pack(fill="x", pady=2)
        ttk.Button(order_btns, text="Move up", command=lambda: self.move_order_item(-1)).pack(side="left", padx=2)
        ttk.Button(order_btns, text="Move down", command=lambda: self.move_order_item(1)).pack(side="left", padx=2)

        ttk.Separator(right).pack(fill="x", pady=8)

        ttk.Label(right, text="Quick symbols:").pack(anchor="w")
        sym_frame = ttk.Frame(right)
        sym_frame.pack(fill="x", pady=4)
        symbols = ["+", "-", "+/-", "up", "down", "->", "Delta", "u", "beta", "alpha", "gamma", "kDa", "bp", "h", "min"]
        for i, sym in enumerate(symbols):
            ttk.Button(sym_frame, text=sym, width=4, command=lambda s=sym: self.insert_symbol(s)).grid(row=i // 6, column=i % 6, padx=2, pady=2)

        ttk.Separator(right).pack(fill="x", pady=10)

        nav = ttk.Frame(right)
        nav.pack(fill="x", pady=10)
        ttk.Button(nav, text="Back", command=self._back).pack(side="left")
        self._next_float_btn = tk.Button(self, text="Next -> Marker", command=self._next, bg="#6ea84f", fg="white")
        self.bind("<Configure>", lambda _e=None: self._place_float_next())
        self._place_float_next()

        # internal
        self._tk_img = None
        self._disp_scale = 1.0
        self._row_widgets: List[Tuple[str, tk.Entry, tk.StringVar, tk.IntVar, tk.IntVar, tk.IntVar, List[tk.Entry]]] = []
        self._focused_entry: Optional[tk.Entry] = None
        self._refresh_after_id: Optional[str] = None
        self._bracket_select_mode: bool = False
        self._drag_start_xy: Optional[Tuple[float, float]] = None
        self._drag_lane_range: Optional[Tuple[int, int]] = None
        self._selected_lane_ranges: List[Tuple[int, int]] = []

        self.canvas.bind("<ButtonPress-1>", self._on_canvas_press, add="+")
        self.canvas.bind("<B1-Motion>", self._on_canvas_drag, add="+")
        self.canvas.bind("<ButtonRelease-1>", self._on_canvas_release, add="+")
        self.canvas.bind("<Configure>", lambda _e=None: self.after(0, self._draw_temp_overlays), add="+")
        self.canvas.bind("<MouseWheel>", lambda _e=None: self.after(0, self._draw_temp_overlays), add="+")
        self.canvas.bind("<Button-4>", lambda _e=None: self.after(0, self._draw_temp_overlays), add="+")
        self.canvas.bind("<Button-5>", lambda _e=None: self.after(0, self._draw_temp_overlays), add="+")
        self.canvas.bind("<<ZoomableCanvasRendered>>", lambda _e=None: self.after(0, self._draw_temp_overlays), add="+")
        self.canvas.bind("<Button-3>", self._right_click_next, add="+")
        self.canvas.bind("<Button-2>", self._right_click_next, add="+")

    def on_show(self):
        if not self.app.state.panels:
            self.app.show_frame("SettingsFrame")
            return
        i = self.app.state.current_panel_index + 1
        n = len(self.app.state.panels)
        self.title.configure(text=f"Step 5 - Header/column annotations ({i}/{n})")
        pc = self.current_pc()
        _normalize_top_annotations(pc)
        self.bracket_gap.set(max(0, int(pc.bracket_text_gap)))
        self.header_value_angle.set(int(getattr(pc, "header_value_angle_deg", 0)))
        self.header_heading_x_off.set(int(getattr(pc, "header_heading_x_offset", 0)))
        self.header_heading_y_off.set(int(getattr(pc, "header_heading_y_offset", 0)))
        self.header_values_x_off.set(int(getattr(pc, "header_values_x_offset", 0)))
        self.header_values_y_off.set(int(getattr(pc, "header_values_y_offset", 0)))
        self.bracket_label_x_off.set(int(getattr(pc, "bracket_label_x_offset", 0)))
        self.bracket_label_y_off.set(int(getattr(pc, "bracket_label_y_offset", 0)))
        self.bracket_line_x_off.set(int(getattr(pc, "bracket_line_x_offset", 0)))
        self.bracket_line_y_off.set(int(getattr(pc, "bracket_line_y_offset", 0)))
        self._drag_start_xy = None
        self._drag_lane_range = None
        self._selected_lane_ranges = []
        self._update_bracket_select_button()
        self.build_row_grid()
        self._place_float_next()
        self.refresh_preview()

    def _place_float_next(self):
        try:
            self.app.place_floating_next(self._next_float_btn)
        except Exception:
            pass

    def _push_undo(self) -> None:
        try:
            pc_copy = copy.deepcopy(self.current_pc())
            self._undo_stack.append(pc_copy)
            if len(self._undo_stack) > 15:
                self._undo_stack.pop(0)
        except Exception:
            pass

    def undo(self) -> None:
        if not self._undo_stack:
            return
        try:
            old_pc = self._undo_stack.pop()
            idx = self.app.state.current_panel_index
            self.app.state.panel_configs[idx] = old_pc
            self.on_show()
        except Exception:
            pass

    def _on_rows_inner_configure(self, _event=None):
        try:
            self.rows_canvas.configure(scrollregion=self.rows_canvas.bbox("all"))
        except Exception:
            pass
        self._fit_rows_canvas_height()

    def _fit_listbox_height(self, lb: tk.Listbox, min_rows: int = 2, max_rows: int = 12):
        try:
            n = int(lb.size())
            h = max(int(min_rows), min(int(max_rows), max(1, n)))
            lb.configure(height=h)
        except Exception:
            pass

    def _fit_rows_canvas_height(self, min_px: int = 70, max_px: int = 260):
        try:
            self.rows_frame.update_idletasks()
            bbox = self.rows_canvas.bbox("all")
            if not bbox:
                return
            h = int(max(1, bbox[3] - bbox[1])) + 6
            h = max(int(min_px), min(int(max_px), int(h)))
            self.rows_canvas.configure(height=h)
        except Exception:
            pass

    def _on_rows_shift_wheel(self, event):
        delta = getattr(event, "delta", 0)
        if delta == 0:
            return
        units = -1 if delta > 0 else 1
        try:
            self.rows_canvas.xview_scroll(units * 3, "units")
        except Exception:
            pass

    def _schedule_rerender(self):
        if self._rerender_after_id is not None:
            try:
                self.after_cancel(self._rerender_after_id)
            except Exception:
                pass
        self._rerender_after_id = self.after(80, self.refresh_preview)

    def build_row_grid(self):
        # clear
        for w in self.rows_frame.winfo_children():
            w.destroy()
        self._row_widgets = []

        pc = self.current_pc()
        _normalize_top_annotations(pc)
        lanes = pc.lanes

        # Header line (lane indices)
        ttk.Label(self.rows_frame, text="Row name").grid(row=0, column=0, padx=2, pady=2, sticky="w")
        ttk.Label(self.rows_frame, text="Position").grid(row=0, column=1, padx=2, pady=2, sticky="w")
        ttk.Label(self.rows_frame, text="Angle").grid(row=0, column=2, padx=2, pady=2, sticky="w")
        ttk.Label(self.rows_frame, text="Heading size", wraplength=64, justify="center").grid(row=0, column=3, padx=2, pady=2, sticky="w")
        ttk.Label(self.rows_frame, text="Label size", wraplength=64, justify="center").grid(row=0, column=4, padx=2, pady=2, sticky="w")
        for li in range(lanes):
            ttk.Label(self.rows_frame, text=str(li + 1)).grid(row=0, column=li + 5, padx=2, pady=2)

        def bind_focus(e: tk.Entry):
            e.bind("<FocusIn>", lambda _ev, ent=e: self._set_focus(ent))
            e.bind("<KeyRelease>", lambda _ev: self._schedule_preview())

        # Populate existing rows
        for r, hr in enumerate(pc.header_rows, start=1):
            name_ent = ttk.Entry(self.rows_frame, width=12)
            name_ent.insert(0, hr.name)
            name_ent.grid(row=r, column=0, padx=2, pady=2)
            bind_focus(name_ent)
            pos_var = tk.StringVar(value="Bottom" if str(getattr(hr, "position", "top")).strip().lower() == "bottom" else "Top")
            pos_box = ttk.Combobox(
                self.rows_frame,
                textvariable=pos_var,
                values=["Top", "Bottom"],
                state="readonly",
                width=8,
            )
            pos_box.grid(row=r, column=1, padx=2, pady=2)
            pos_box.bind("<<ComboboxSelected>>", lambda _e=None: self._schedule_preview())
            angle_var = tk.IntVar(value=max(-80, min(80, int(getattr(hr, "angle_deg", getattr(pc, "header_value_angle_deg", 0))))))
            angle_spin = ttk.Spinbox(
                self.rows_frame,
                from_=-80,
                to=80,
                textvariable=angle_var,
                width=5,
                command=self._schedule_preview,
            )
            angle_spin.grid(row=r, column=2, padx=2, pady=2)
            angle_spin.bind("<KeyRelease>", lambda _e=None: self._schedule_preview())
            heading_fs_var = tk.IntVar(value=max(6, int(getattr(hr, "heading_font_size", 0) or self.app.state.settings.font_size)))
            heading_fs_spin = ttk.Spinbox(
                self.rows_frame,
                from_=6,
                to=96,
                textvariable=heading_fs_var,
                width=5,
                command=self._schedule_preview,
            )
            heading_fs_spin.grid(row=r, column=3, padx=2, pady=2)
            heading_fs_spin.bind("<KeyRelease>", lambda _e=None: self._schedule_preview())
            value_fs_var = tk.IntVar(value=max(6, int(getattr(hr, "value_font_size", 0) or self.app.state.settings.font_size)))
            value_fs_spin = ttk.Spinbox(
                self.rows_frame,
                from_=6,
                to=96,
                textvariable=value_fs_var,
                width=5,
                command=self._schedule_preview,
            )
            value_fs_spin.grid(row=r, column=4, padx=2, pady=2)
            value_fs_spin.bind("<KeyRelease>", lambda _e=None: self._schedule_preview())

            val_ents: List[tk.Entry] = []
            for li in range(lanes):
                ent = ttk.Entry(self.rows_frame, width=4)
                ent.insert(0, hr.values[li] if li < len(hr.values) else "")
                ent.grid(row=r, column=li + 5, padx=1, pady=1)
                bind_focus(ent)
                val_ents.append(ent)
            self._row_widgets.append((hr.id, name_ent, pos_var, angle_var, heading_fs_var, value_fs_var, val_ents))

        # Group list
        self.group_list.delete(0, "end")
        horder: List[str] = []
        for g in pc.group_labels:
            hgid = str(getattr(g, "height_group", "")).strip()
            if hgid and hgid not in horder:
                horder.append(hgid)
        for g in pc.group_labels:
            hgid = str(getattr(g, "height_group", "")).strip()
            row_no = (horder.index(hgid) + 1) if (hgid in horder) else 0
            self.group_list.insert("end", f"Row {row_no} | {g.start_lane}-{g.end_lane}: {g.text}")
        self._fit_listbox_height(self.group_list, min_rows=2, max_rows=10)
        self._refresh_order_list()
        self._fit_rows_canvas_height()

    def _set_focus(self, entry: tk.Entry):
        self._focused_entry = entry

    def _bracket_groups_by_height(self) -> Dict[str, List[GroupLabel]]:
        pc = self.current_pc()
        groups: Dict[str, List[GroupLabel]] = {}
        for g in pc.group_labels:
            hgid = str(getattr(g, "height_group", "")).strip()
            if not hgid:
                continue
            groups.setdefault(hgid, []).append(g)
        return groups

    def _refresh_height_target_options(self):
        pc = self.current_pc()
        _normalize_top_annotations(pc)
        groups = self._bracket_groups_by_height()
        choice_map: Dict[str, Optional[str]] = {"New height": None}
        labels = ["New height"]
        for hgid, glist in groups.items():
            if not glist:
                continue
            glist2 = sorted(glist, key=lambda g: (min(int(g.start_lane), int(g.end_lane)), max(int(g.start_lane), int(g.end_lane))))
            ex = glist2[0]
            lbl = f"Same as {ex.start_lane}-{ex.end_lane}: {str(ex.text).strip() or '(no text)'}"
            if lbl in choice_map:
                lbl = f"{lbl} [{hgid[:4]}]"
            choice_map[lbl] = hgid
            labels.append(lbl)
        self._height_choice_map = choice_map
        self.grp_height_combo["values"] = labels
        cur = self.grp_height_target.get().strip()
        if cur not in choice_map:
            self.grp_height_target.set("New height")

    def _selected_height_group_id(self) -> Optional[str]:
        key = self.grp_height_target.get().strip()
        return self._height_choice_map.get(key)

    def _order_item_label(self, oid: str) -> str:
        pc = self.current_pc()
        for i, hr in enumerate(pc.header_rows, start=1):
            if hr.id == oid:
                title = hr.name.strip() or f"Header row {i}"
                return f"Header: {title}"
        hgid = oid[3:] if oid.startswith("BG:") else oid
        groups = self._bracket_groups_by_height().get(hgid, [])
        if groups:
            groups2 = sorted(groups, key=lambda g: (min(int(g.start_lane), int(g.end_lane)), max(int(g.start_lane), int(g.end_lane))))
            first = groups2[0]
            txt = str(first.text).strip() or "(no text)"
            return f"Bracket row ({len(groups2)}): {first.start_lane}-{first.end_lane} {txt}"
        return f"(missing) {oid}"

    def _refresh_order_list(self):
        pc = self.current_pc()
        _normalize_top_annotations(pc)
        sel = self.order_list.curselection()
        sel_idx = int(sel[0]) if sel else None
        self.order_list.delete(0, "end")
        for oid in pc.top_annotation_order:
            self.order_list.insert("end", self._order_item_label(oid))
        if sel_idx is not None and 0 <= sel_idx < self.order_list.size():
            self.order_list.selection_set(sel_idx)
        self._fit_listbox_height(self.order_list, min_rows=2, max_rows=12)
        self._refresh_height_target_options()

    def move_order_item(self, delta: int):
        pc = self.current_pc()
        _normalize_top_annotations(pc)
        sel = self.order_list.curselection()
        if not sel:
            return
        i = int(sel[0])
        j = i + int(delta)
        if j < 0 or j >= len(pc.top_annotation_order):
            return
        order = list(pc.top_annotation_order)
        order[i], order[j] = order[j], order[i]
        pc.top_annotation_order = order
        self._refresh_order_list()
        self.order_list.selection_set(j)
        self._schedule_preview()

    def _update_bracket_select_button(self):
        txt = "Bracket select tool: ON" if self._bracket_select_mode else "Bracket select tool: OFF"
        try:
            self._bracket_select_btn.configure(text=txt)
        except Exception:
            pass

    def toggle_bracket_select_mode(self):
        self._bracket_select_mode = not self._bracket_select_mode
        self._drag_start_xy = None
        self._drag_lane_range = None
        self._update_bracket_select_button()
        self._draw_temp_overlays()

    def clear_selected_regions(self):
        self._selected_lane_ranges = []
        self._drag_lane_range = None
        self._draw_temp_overlays()

    def _xrange_to_lane_range(self, x0: float, x1: float) -> Optional[Tuple[int, int]]:
        if not self._has_current_panel():
            return None
        layout = compute_render_layout(self.current_panel(), self.current_pc(), self.app.state.settings)
        lane_lefts = list(layout["lane_lefts"])
        lane_rights = list(layout["lane_rights"])
        if not lane_lefts or not lane_rights:
            return None
        lo = min(float(x0), float(x1))
        hi = max(float(x0), float(x1))
        covered: List[int] = []
        for i, (lx, rx) in enumerate(zip(lane_lefts, lane_rights), start=1):
            if hi < float(lx) or lo > float(rx):
                continue
            covered.append(i)
        if not covered:
            return None
        return (int(min(covered)), int(max(covered)))

    def _event_to_img_xy(self, event) -> Tuple[int, int]:
        return self.viewport.canvas_to_image(event.x, event.y)

    def _on_canvas_press(self, event):
        if not self._has_current_panel():
            return
        if not self._bracket_select_mode:
            return
        ctrl_down = bool(int(getattr(event, "state", 0)) & 0x0004)
        if not ctrl_down:
            self._selected_lane_ranges = []
        x, y = self._event_to_img_xy(event)
        self._drag_start_xy = (float(x), float(y))
        self._drag_lane_range = None
        self._draw_temp_overlays()

    def _on_canvas_drag(self, event):
        if not self._has_current_panel():
            return
        if not self._bracket_select_mode or self._drag_start_xy is None:
            return
        x, y = self._event_to_img_xy(event)
        x0, _y0 = self._drag_start_xy
        self._drag_lane_range = self._xrange_to_lane_range(x0, float(x))
        self._draw_temp_overlays()

    def _on_canvas_release(self, event):
        if not self._has_current_panel():
            return
        if not self._bracket_select_mode or self._drag_start_xy is None:
            return
        x, y = self._event_to_img_xy(event)
        x0, _y0 = self._drag_start_xy
        rng = self._xrange_to_lane_range(x0, float(x))
        if rng is not None:
            if rng not in self._selected_lane_ranges:
                self._selected_lane_ranges.append(rng)
            self._selected_lane_ranges.sort(key=lambda t: (t[0], t[1]))
        self._drag_start_xy = None
        self._drag_lane_range = None
        self._draw_temp_overlays()

    def add_groups_from_selected_regions(self):
        pc = self.current_pc()
        text = self.grp_text.get().strip()
        if not text:
            messagebox.showinfo("Brackets", "Enter bracket text first.")
            return
        if not self._selected_lane_ranges:
            messagebox.showinfo("Brackets", "Select one or more regions first (drag on preview).")
            return
        _normalize_top_annotations(pc)
        hgid = self._selected_height_group_id()
        if not hgid:
            hgid = _new_annotation_id("BG")
        for s, e in self._selected_lane_ranges:
            g = GroupLabel(
                start_lane=int(s),
                end_lane=int(e),
                text=text,
                bracket=True,
                id=_new_annotation_id("G"),
                height_group=hgid,
            )
            pc.group_labels.append(g)
        tok = _bracket_group_token(hgid)
        if tok not in pc.top_annotation_order:
            pc.top_annotation_order.append(tok)
        self._selected_lane_ranges = []
        self.build_row_grid()
        self._schedule_preview()

    def _draw_temp_overlays(self):
        try:
            self.canvas.delete("annotation_ui")
        except Exception:
            pass
        if not self._has_current_panel():
            return

        panel = self.current_panel()
        pc = self.current_pc()
        layout = compute_render_layout(panel, pc, self.app.state.settings)
        z = float(self.viewport.zoom)

        # Temporary lane numbers for Step 5 only (not rendered/exported).
        y_lane = float(layout["panel_origin"][1] + panel.height + 10) * z
        circ_outline, txt_color = _lane_circle_colors(self.app.state.settings)
        for i, cx in enumerate(layout["lane_centers"], start=1):
            cxz = float(cx) * z
            radius = max(7.0, 8.0 * z)
            self.canvas.create_oval(
                cxz - radius,
                y_lane - radius,
                cxz + radius,
                y_lane + radius,
                fill="white",
                outline=circ_outline,
                width=max(1, int(round(1.2 * z))),
                tags=("annotation_ui",),
            )
            self.canvas.create_text(
                cxz,
                y_lane,
                text=str(i),
                fill=txt_color,
                font=("Arial", max(9, int(round(9 * z))), "bold"),
                tags=("annotation_ui",),
            )

        # Visualize selected bracket regions.
        lane_lefts = list(layout["lane_lefts"])
        lane_rights = list(layout["lane_rights"])
        y_base = float(layout["panel_origin"][1] + 8) * z
        ranges = list(self._selected_lane_ranges)
        if self._drag_lane_range is not None:
            ranges.append(self._drag_lane_range)
        for idx, (s, e) in enumerate(ranges):
            s2 = max(1, min(len(lane_lefts), int(s)))
            e2 = max(1, min(len(lane_rights), int(e)))
            if e2 < s2:
                s2, e2 = e2, s2
            x0 = float(lane_lefts[s2 - 1]) * z
            x1 = float(lane_rights[e2 - 1]) * z
            y = y_base + idx * 10.0
            color = "#ffc000" if (idx == len(ranges) - 1 and self._drag_lane_range is not None) else "#00d6a1"
            self.canvas.create_line(x0, y, x1, y, fill=color, width=2, tags=("annotation_ui",))
            self.canvas.create_line(x0, y, x0, y + 6, fill=color, width=2, tags=("annotation_ui",))
            self.canvas.create_line(x1, y, x1, y + 6, fill=color, width=2, tags=("annotation_ui",))

    def insert_symbol(self, sym: str):
        if not self._focused_entry:
            return
        self._focused_entry.insert("insert", sym)

    def _schedule_preview(self):
        # Debounce redraw while typing (real-time preview)
        if self._refresh_after_id is not None:
            try:
                self.after_cancel(self._refresh_after_id)
            except Exception:
                pass
        self._refresh_after_id = self.after(150, self.refresh_preview)

    def add_row(self):
        self._push_undo()
        pc = self.current_pc()
        _normalize_top_annotations(pc)
        hr = HeaderRow(
            name="",
            values=[""] * pc.lanes,
            id=_new_annotation_id("H"),
            position="top",
            angle_deg=max(-80, min(80, int(self.header_value_angle.get()))),
            heading_font_size=max(6, int(self.app.state.settings.font_size)),
            value_font_size=max(6, int(self.app.state.settings.font_size)),
        )
        pc.header_rows.append(hr)
        pc.top_annotation_order.append(hr.id)
        self.build_row_grid()
        self._schedule_preview()

    def remove_row(self):
        self._push_undo()
        pc = self.current_pc()
        if pc.header_rows:
            _normalize_top_annotations(pc)
            hr = pc.header_rows.pop()
            pc.top_annotation_order = [oid for oid in pc.top_annotation_order if oid != hr.id]
            self.build_row_grid()
            self._schedule_preview()

    def add_group(self):
        self._push_undo()
        pc = self.current_pc()
        s = max(1, int(self.grp_start.get()))
        e = max(1, int(self.grp_end.get()))
        if e < s:
            s, e = e, s
        text = self.grp_text.get().strip()
        if not text:
            return
        _normalize_top_annotations(pc)
        hgid = self._selected_height_group_id()
        if not hgid:
            hgid = _new_annotation_id("BG")
        g = GroupLabel(start_lane=s, end_lane=e, text=text, bracket=True, id=_new_annotation_id("G"), height_group=hgid)
        pc.group_labels.append(g)
        tok = _bracket_group_token(hgid)
        if tok not in pc.top_annotation_order:
            pc.top_annotation_order.append(tok)
        self.build_row_grid()
        self._schedule_preview()

    def remove_group(self):
        self._push_undo()
        pc = self.current_pc()
        sel = self.group_list.curselection()
        if not sel:
            return
        idx = sel[0]
        g = pc.group_labels[idx]
        hgid = str(getattr(g, "height_group", "")).strip()
        del pc.group_labels[idx]
        if hgid and not any(str(getattr(gg, "height_group", "")).strip() == hgid for gg in pc.group_labels):
            tok = _bracket_group_token(hgid)
            pc.top_annotation_order = [oid for oid in pc.top_annotation_order if oid != tok]
        self.build_row_grid()
        self._schedule_preview()

    def _sync_to_state(self):
        pc = self.current_pc()
        pc.bracket_text_gap = max(0, int(self.bracket_gap.get()))
        pc.header_value_angle_deg = max(-80, min(80, int(self.header_value_angle.get())))
        pc.header_heading_x_offset = int(self.header_heading_x_off.get())
        pc.header_heading_y_offset = int(self.header_heading_y_off.get())
        pc.header_values_x_offset = int(self.header_values_x_off.get())
        pc.header_values_y_offset = int(self.header_values_y_off.get())
        pc.bracket_label_x_offset = int(self.bracket_label_x_off.get())
        pc.bracket_label_y_offset = int(self.bracket_label_y_off.get())
        pc.bracket_line_x_offset = int(self.bracket_line_x_off.get())
        pc.bracket_line_y_offset = int(self.bracket_line_y_off.get())
        # read back row grid
        new_rows: List[HeaderRow] = []
        for row_id, name_ent, pos_var, angle_var, heading_fs_var, value_fs_var, val_ents in self._row_widgets:
            name = name_ent.get().strip()
            vals = [v.get().strip() for v in val_ents]
            pos = "bottom" if str(pos_var.get()).strip().lower() == "bottom" else "top"
            ang = max(-80, min(80, int(angle_var.get())))
            heading_fs = max(6, safe_int(heading_fs_var.get(), int(self.app.state.settings.font_size)))
            value_fs = max(6, safe_int(value_fs_var.get(), int(self.app.state.settings.font_size)))
            new_rows.append(
                HeaderRow(
                    name=name,
                    values=vals,
                    id=row_id,
                    position=pos,
                    angle_deg=ang,
                    heading_font_size=heading_fs,
                    value_font_size=value_fs,
                )
            )
        pc.header_rows = new_rows
        _normalize_top_annotations(pc)
        self._refresh_order_list()

    def _render_preview_image(self) -> Image.Image:
        panel = self.current_panel().copy()
        pc = self.current_pc()
        return render_panel(panel, pc, self.app.state.settings, preview=True)

    def refresh_preview(self):
        if not self._has_current_panel():
            try:
                self.canvas.delete("annotation_ui")
            except Exception:
                pass
            return
        self._sync_to_state()
        img = self._render_preview_image()
        self.viewport.set_image(img, fit_if_needed=True)
        self._draw_temp_overlays()

    def _back(self):
        self._sync_to_state()
        pc = self.current_pc()
        if bool(getattr(pc, "highlight_enabled", False)):
            self.app.show_frame("HighlightFrame")
        else:
            self.app.show_frame("LayoutFrame")

    def _next(self):
        self._sync_to_state()
        self.app.ensure_panel_configs()
        self.app.show_frame("MarkerFrame")

    def _right_click_next(self, _event=None):
        self._next()
        return "break"


# ----------------------------
# Step 6: Marker selection
# ----------------------------
class MarkerFrame(PanelAwareFrame):
    def __init__(self, parent, app: GelAnnotatorApp):
        super().__init__(parent)
        self.app = app

        self.title = ttk.Label(self, text="Step 6 - Marker / ladder", font=("Arial", 18, "bold"))
        self.title.pack(pady=8)

        main = ttk.Frame(self)
        main.pack(fill="both", expand=True)

        # left preview (zoom/pan)
        self.viewport = ZoomableCanvas(main, bg="#222222", show_tools=True)
        self.viewport.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        self.canvas = self.viewport.canvas

        # right controls
        self.right_scroll = VScrollPanel(main, width=self.app._scaled_sidebar_width(620))
        self.right_scroll.pack(side="left", fill="y", padx=10, pady=10)
        right = self.right_scroll.inner

        ttk.Label(right, text="Marker library:").pack(anchor="w")
        self.marker_name = tk.StringVar(value="")
        self.marker_combo = ttk.Combobox(
            right, textvariable=self.marker_name, values=list_markers(), width=28, state="readonly"
        )
        self.marker_combo.pack(anchor="w", pady=4)
        self.marker_combo.bind("<<ComboboxSelected>>", lambda _e=None: self._on_marker_selected())

        btns = ttk.Frame(right)
        btns.pack(fill="x", pady=4)
        ttk.Button(btns, text="Refresh", command=self._refresh_markers).pack(side="left", padx=4)
        ttk.Button(btns, text="New", command=self._new_marker_dialog).pack(side="left", padx=4)
        ttk.Button(btns, text="Edit", command=self._edit_marker_dialog).pack(side="left", padx=4)

        tick_style_name = "MarkerTickSmall.Treeview"
        try:
            base = tkfont.nametofont("TkDefaultFont")
            base_size = int(base.cget("size"))
            small_size = max(9, base_size - 2) if base_size > 0 else 10
            self._tick_font = tkfont.Font(self, family=str(base.cget("family")), size=small_size)
            self._tick_heading_font = tkfont.Font(self, family=str(base.cget("family")), size=max(8, small_size), weight="bold")
            st = ttk.Style(self)
            st.configure(tick_style_name, font=self._tick_font, rowheight=max(24, small_size + 12))
            st.configure(f"{tick_style_name}.Heading", font=self._tick_heading_font)
        except Exception:
            pass

        ttk.Separator(right).pack(fill="x", pady=8)
        ttk.Label(right, text="Marker tick list (size -> y in panel px):").pack(anchor="w")
        self.tick_tree = ttk.Treeview(right, columns=("size", "y", "mode"), show="headings", height=14, style=tick_style_name)
        self.tick_tree.heading("size", text="Size")
        self.tick_tree.heading("y", text="Y(px)")
        self.tick_tree.heading("mode", text="Mode")
        self.tick_tree.column("size", width=130, anchor="center")
        self.tick_tree.column("y", width=130, anchor="center")
        self.tick_tree.column("mode", width=250, anchor="w")
        self.tick_tree.pack(fill="x", pady=4)
        self.tick_tree.bind("<<TreeviewSelect>>", lambda _e=None: self._on_tick_selected())

        tick_edit = ttk.Frame(right)
        tick_edit.pack(fill="x", pady=2)
        ttk.Label(tick_edit, text="Set Y:").pack(side="left")
        self.tick_y_var = tk.StringVar(value="")
        self.tick_y_entry = ttk.Entry(tick_edit, textvariable=self.tick_y_var, width=8)
        self.tick_y_entry.pack(side="left", padx=4)
        self.tick_y_entry.bind("<Return>", lambda _e=None: self.apply_tick_y())
        ttk.Button(tick_edit, text="Apply", command=self.apply_tick_y).pack(side="left", padx=4)
        ttk.Button(tick_edit, text="Clear override", command=self.clear_tick_override).pack(side="left", padx=4)

        tick_btns = ttk.Frame(right)
        tick_btns.pack(fill="x", pady=2)
        ttk.Button(tick_btns, text="Hide selected", command=self.hide_selected_tick).pack(side="left", padx=4)
        ttk.Button(tick_btns, text="Show all", command=self.show_all_ticks).pack(side="left", padx=4)
        self.assign_hint = ttk.Label(right, text="Tip: select a size in the list, then click the image to assign that one size.", justify="left")
        self.assign_hint.pack(anchor="w", pady=(2, 0))

        ttk.Separator(right).pack(fill="x", pady=8)

        ttk.Label(right, text="Marker positioning (this panel):").pack(anchor="w")
        self.marker_x_off = tk.IntVar(value=0)
        self.marker_y_off = tk.IntVar(value=0)
        self.marker_tick_len = tk.IntVar(value=8)
        self.marker_label_gap = tk.IntVar(value=10)
        self.marker_font_size = tk.IntVar(value=0)

        def spin_row(label: str, var: tk.IntVar, from_: int, to: int):
            row = ttk.Frame(right)
            row.pack(fill="x", pady=2)
            ttk.Label(row, text=label).pack(side="left")
            sp = ttk.Spinbox(row, from_=from_, to=to, textvariable=var, width=6, command=self._on_style_changed)
            sp.pack(side="right")
            sp.bind("<KeyRelease>", lambda _e=None: self._on_style_changed())
            return sp

        spin_row("X offset (px)", self.marker_x_off, -200, 200)
        spin_row("Y offset (px)", self.marker_y_off, -200, 200)
        spin_row("Tick length", self.marker_tick_len, 2, 30)
        spin_row("Label gap", self.marker_label_gap, 0, 50)
        spin_row("Marker label font size", self.marker_font_size, 6, 96)

        ttk.Separator(right).pack(fill="x", pady=8)
        ttk.Label(
            right,
            text="Assign marker bands from the size table above:\nselect a marker size, then click the matching band in the preview.",
            justify="left",
        ).pack(anchor="w", pady=(0, 2))
        self.calib_info = None
        self.picks_box = None

        ttk.Separator(right).pack(fill="x", pady=10)
        self.run_analysis_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            right,
            text="Run lane intensity/band analysis before review",
            variable=self.run_analysis_var,
        ).pack(anchor="w", pady=(2, 8))

        nav = ttk.Frame(right)
        nav.pack(fill="x", pady=10)
        ttk.Button(nav, text="Back", command=self._back).pack(side="left")
        ttk.Button(nav, text="Next -> Review", style="Primary.TButton", command=self._next).pack(side="right")

        # internal state
        self._tk_img = None
        self._disp_scale = 1.0
        self._rerender_after_id = None
        self._text_after_id = None
        self._picking = False
        self._reposition_pick_idx: Optional[int] = None
        self._assign_tick_size: Optional[float] = None
        self._picked_click_points: List[Tuple[int, int]] = []
        self._tick_assign_click_points: List[Tuple[int, int]] = []


        self.canvas.bind("<ButtonPress-1>", self._on_press)
        self.canvas.bind("<Button-3>", self._right_click_next, add="+")
        self.canvas.bind("<Button-2>", self._right_click_next, add="+")
        self.canvas.bind("<<ZoomableCanvasRendered>>", lambda _e=None: self.after(0, self._draw_overlays), add="+")

    def on_show(self):
        if not self.app.state.panels:
            self.app.show_frame("SettingsFrame")
            return

        i = self.app.state.current_panel_index + 1
        n = len(self.app.state.panels)
        self.title.configure(text=f"Step 6 - Marker / ladder ({i}/{n})")

        pc = self.current_pc()
        self.marker_combo["values"] = list_markers()

        if pc.marker_calibration.marker_name:
            self.marker_name.set(pc.marker_calibration.marker_name)
        elif list_markers():
            self.marker_name.set(list_markers()[0])
            pc.marker_calibration.marker_name = self.marker_name.get()

        # load per-panel style vars
        self.marker_x_off.set(int(pc.marker_x_offset))
        self.marker_y_off.set(int(pc.marker_y_offset))
        self.marker_tick_len.set(int(pc.marker_tick_length))
        self.marker_label_gap.set(int(pc.marker_label_gap))
        self.marker_font_size.set(int(pc.marker_font_size) if pc.marker_font_size else 15)
        self.run_analysis_var.set(bool(pc.run_band_analysis))
        self._fit_tick_tree_height()


        self._picking = False
        self._reposition_pick_idx = None
        self._assign_tick_size = None
        self._picked_click_points = []
        self._tick_assign_click_points = []
        self._update_pick_list()
        self.refresh_preview()

    def _schedule_rerender(self):
        if self._rerender_after_id is not None:
            try:
                self.after_cancel(self._rerender_after_id)
            except Exception:
                pass
        self._rerender_after_id = self.after(80, self.refresh_preview)

    def _sync_style_to_state(self):
        pc = self.current_pc()
        pc.marker_x_offset = int(self.marker_x_off.get())
        pc.marker_y_offset = int(self.marker_y_off.get())
        pc.marker_tick_length = int(self.marker_tick_len.get())
        pc.marker_label_gap = int(self.marker_label_gap.get())
        pc.marker_font_size = int(self.marker_font_size.get())

    def _on_style_changed(self):
        self._sync_style_to_state()
        self.refresh_preview()

    def _refresh_markers(self):
        self.marker_combo["values"] = list_markers()

    def _fit_tick_tree_height(self):
        if not hasattr(self, "tick_tree"):
            return
        n = 12
        try:
            name = self.marker_name.get().strip()
            if name:
                marker = load_marker(name)
                n = max(1, int(len(marker.sizes)))
        except Exception:
            pass
        h = max(8, min(40, n + 1))
        try:
            self.tick_tree.configure(height=h)
        except Exception:
            pass

    def _on_marker_selected(self):
        pc = self.current_pc()
        pc.marker_calibration.marker_name = self.marker_name.get()
        pc.marker_calibration.picked = []
        pc.marker_calibration.fit_a = None
        pc.marker_calibration.fit_b = None
        pc.marker_tick_overrides = {}
        pc.marker_tick_hidden = []
        self._reposition_pick_idx = None
        self._assign_tick_size = None
        self._fit_tick_tree_height()
        self._update_pick_list()
        self.refresh_preview()

    def _edit_marker_dialog(self):
        name = self.marker_name.get().strip()
        if not name:
            messagebox.showinfo("Marker", "Select a marker first.")
            return

        try:
            marker = load_marker(name)
        except Exception as e:
            messagebox.showerror("Marker", f"Could not load marker: {e}")
            return

        win = tk.Toplevel(self)
        win.title(f"Edit marker  {marker.name}")
        fit_toplevel_to_screen(win, 460, 460)

        name_var = tk.StringVar(value=marker.name)
        unit_var = tk.StringVar(value=marker.unit)
        sizes_text = tk.Text(win, width=52, height=16)
        sizes_text.insert("1.0", "\n".join(str(s) for s in marker.sizes))

        ttk.Label(win, text="Marker name:").pack(anchor="w", padx=10, pady=(10, 2))
        ttk.Entry(win, textvariable=name_var).pack(fill="x", padx=10)

        ttk.Label(win, text="Unit:").pack(anchor="w", padx=10, pady=(10, 2))
        ttk.Combobox(win, textvariable=unit_var, values=["bp", "kDa"], state="readonly").pack(fill="x", padx=10)

        ttk.Label(win, text="Sizes (one per line):").pack(anchor="w", padx=10, pady=(10, 2))
        sizes_text.pack(fill="both", expand=True, padx=10, pady=6)

        def save_edit():
            new_name = name_var.get().strip()
            if not new_name:
                messagebox.showerror("Marker", "Please provide a name.", parent=win)
                return
            try:
                sizes: List[float] = []
                for line in sizes_text.get("1.0", "end").splitlines():
                    line = line.strip()
                    if not line:
                        continue
                    sizes.append(float(line))
                if len(sizes) < 2:
                    raise ValueError("Need at least 2 sizes.")
            except Exception as e:
                messagebox.showerror("Marker", f"Invalid sizes: {e}", parent=win)
                return

            edited = MarkerDefinition(name=new_name, unit=unit_var.get(), sizes=sizes, y_positions=marker.y_positions)

            # If name changed, keep old file (safe); user can delete manually.
            save_marker(edited)
            self._refresh_markers()
            self.marker_name.set(new_name)
            self._on_marker_selected()
            win.destroy()

        ttk.Button(win, text="Save", command=save_edit).pack(pady=10)

    def _new_marker_dialog(self):
        win = tk.Toplevel(self)
        win.title("New marker")
        fit_toplevel_to_screen(win, 420, 400)

        name_var = tk.StringVar(value="NewMarker")
        unit_var = tk.StringVar(value=self.app.state.settings.default_unit)
        sizes_text = tk.Text(win, width=46, height=12)

        ttk.Label(win, text="Marker name:").pack(anchor="w", padx=10, pady=(10, 2))
        ttk.Entry(win, textvariable=name_var).pack(fill="x", padx=10)
        ttk.Label(win, text="Unit:").pack(anchor="w", padx=10, pady=(10, 2))
        ttk.Combobox(win, textvariable=unit_var, values=["bp", "kDa"], state="readonly").pack(fill="x", padx=10)
        ttk.Label(win, text="Sizes (one per line, e.g. 1000):").pack(anchor="w", padx=10, pady=(10, 2))
        sizes_text.pack(fill="both", expand=True, padx=10, pady=6)

        def save_new():
            name = name_var.get().strip()
            if not name:
                messagebox.showerror("Marker", "Please provide a marker name.", parent=win)
                return
            try:
                sizes = []
                for line in sizes_text.get("1.0", "end").splitlines():
                    line = line.strip()
                    if not line:
                        continue
                    sizes.append(float(line))
                if len(sizes) < 2:
                    raise ValueError("Need at least 2 sizes.")
            except Exception as e:
                messagebox.showerror("Marker", f"Invalid sizes: {e}", parent=win)
                return

            marker = MarkerDefinition(name=name, unit=unit_var.get(), sizes=sizes, y_positions=None)
            save_marker(marker)
            self._refresh_markers()
            self.marker_name.set(name)
            self._on_marker_selected()
            win.destroy()

        ttk.Button(win, text="Save marker", command=save_new).pack(pady=10)

    def start_picking(self):
        pc = self.current_pc()
        if not pc.include_marker:
            messagebox.showinfo("Marker", "Marker is disabled for this panel. Enable it in Step 4.")
            return
        if not self.marker_name.get():
            messagebox.showinfo("Marker", "Select or create a marker first.")
            return
        self._picking = True
        self._reposition_pick_idx = None
        self._assign_tick_size = None
        marker = load_marker(self.marker_name.get())
        pc.marker_calibration.marker_name = marker.name
        pc.marker_calibration.unit = marker.unit
        pc.marker_calibration.picked = []
        pc.marker_calibration.fit_a = None
        pc.marker_calibration.fit_b = None
        self._picked_click_points = []
        self._update_pick_list()
        self.refresh_preview()

    def undo_pick(self):
        pc = self.current_pc()
        if pc.marker_calibration.picked:
            pc.marker_calibration.picked.pop()
            if self._picked_click_points:
                self._picked_click_points.pop()
            self._recompute_fit()
            self._update_pick_list()
            self.refresh_preview()

    def clear_picks(self):
        pc = self.current_pc()
        pc.marker_calibration.picked = []
        pc.marker_calibration.fit_a = None
        pc.marker_calibration.fit_b = None
        self._reposition_pick_idx = None
        self._assign_tick_size = None
        self._picked_click_points = []
        self._tick_assign_click_points = []
        self._update_pick_list()
        self.refresh_preview()

    def reposition_selected(self):
        if not hasattr(self, "picks_box") or self.picks_box is None:
            messagebox.showinfo("Reposition", "Marker calibration picking is disabled in this version.")
            return
        pc = self.current_pc()
        sel = self.picks_box.curselection()
        if not sel:
            messagebox.showinfo("Reposition", "Select a picked point first.")
            return
        self._reposition_pick_idx = int(sel[0])
        self._picking = False
        self._assign_tick_size = None
        messagebox.showinfo("Reposition", "Click on the new band position in the preview.")

    def _update_pick_list(self):
        pc = self.current_pc()
        if hasattr(self, "picks_box") and self.picks_box is not None:
            self.picks_box.delete(0, "end")
            for y, s in pc.marker_calibration.picked:
                self.picks_box.insert("end", f"y={y}px  size={s:g}")
        if hasattr(self, "calib_info") and self.calib_info is not None:
            if pc.marker_calibration.marker_name:
                self.calib_info.configure(text=f"Marker: {pc.marker_calibration.marker_name} ({pc.marker_calibration.unit})")
            else:
                self.calib_info.configure(text="")

    def _set_visible_ticks_to_assigned(self, include_size: Optional[float] = None):
        """Hide all non-assigned marker sizes so only explicitly positioned sizes are shown."""
        pc = self.current_pc()
        try:
            marker = load_marker(pc.marker_calibration.marker_name)
        except Exception:
            return
        keep = {float(k) for k in pc.marker_tick_overrides.keys()}
        if include_size is not None:
            keep.add(float(include_size))
        pc.marker_tick_hidden = [float(s) for s in marker.sizes if float(s) not in keep]

    def _next_smaller_tick_size(self, current_size: float) -> Optional[float]:
        pc = self.current_pc()
        try:
            marker = load_marker(pc.marker_calibration.marker_name)
        except Exception:
            return None
        cur = float(current_size)
        smaller = []
        for s in marker.sizes:
            try:
                fs = float(s)
            except Exception:
                continue
            if fs < cur - 1e-12:
                smaller.append(fs)
        if not smaller:
            return None
        return float(max(smaller))

    def _click_to_panel_y(self, x_win: int, y_win: int) -> int:
        # Convert click to panel-relative Y in render_panel() coordinates.
        _ix, iy_out = self._canvas_to_img_xy(x_win, y_win)
        layout = compute_render_layout(self.current_panel(), self.current_pc(), self.app.state.settings)
        panel_origin_y = int(layout["panel_origin"][1])
        panel_h = int(self.current_panel().height)
        y_panel = int(round(float(iy_out) - float(panel_origin_y)))
        # Store override in raw panel coordinates so rendered tick (which adds marker_y_offset)
        # stays exactly at the click position.
        try:
            y_panel -= int(getattr(self.current_pc(), "marker_y_offset", 0))
        except Exception:
            pass
        return max(0, min(panel_h - 1, y_panel))

    def _click_to_panel_xy(self, x_win: int, y_win: int) -> Tuple[int, int]:
        # Convert click to panel-relative X/Y in render_panel() coordinates.
        ix_out, iy_out = self._canvas_to_img_xy(x_win, y_win)
        panel = self.current_panel()
        layout = compute_render_layout(panel, self.current_pc(), self.app.state.settings)
        panel_origin_x = int(layout["panel_origin"][0])
        panel_origin_y = int(layout["panel_origin"][1])
        x_panel = int(round(float(ix_out) - float(panel_origin_x)))
        y_panel = int(round(float(iy_out) - float(panel_origin_y)))
        x_panel = max(0, min(int(panel.width) - 1, x_panel))
        y_panel = max(0, min(int(panel.height) - 1, y_panel))
        return x_panel, y_panel

    
    # --- Marker tick list editing (per panel) ---
    def _populate_tick_tree(self):
        if not hasattr(self, "tick_tree"):
            return
        selected_size = self._assign_tick_size
        try:
            for it in self.tick_tree.get_children():
                self.tick_tree.delete(it)
        except Exception:
            return

        pc = self.current_pc()
        if not pc.include_marker or not pc.marker_calibration.marker_name:
            return
        try:
            marker = load_marker(pc.marker_calibration.marker_name)
        except Exception:
            return

        base_map = marker_base_positions(marker, pc.marker_calibration)
        overrides = pc.marker_tick_overrides or {}

        hidden = set(pc.marker_tick_hidden or [])

        for s in marker.sizes:
            fs = float(s)
            if fs in overrides:
                y: Any = int(overrides[fs])
            elif fs in base_map:
                y = int(base_map[fs])
            else:
                y = ""
            if fs in hidden:
                mode = "hidden"
            elif fs in overrides:
                mode = "override"
            elif fs in base_map:
                mode = "auto"
            else:
                mode = "unassigned"
            try:
                self.tick_tree.insert("", "end", iid=str(fs), values=(_format_marker_size(fs), y, mode))
            except Exception:
                # fallback if iid collisions occur
                self.tick_tree.insert("", "end", values=(_format_marker_size(fs), y, mode))

        if selected_size is not None:
            iid = str(float(selected_size))
            if self.tick_tree.exists(iid):
                try:
                    self.tick_tree.selection_set(iid)
                    self.tick_tree.focus(iid)
                    self.tick_tree.see(iid)
                except Exception:
                    pass

    def _selected_tick_size(self) -> Optional[float]:
        if not hasattr(self, "tick_tree"):
            return None
        sel = self.tick_tree.selection()
        if not sel:
            return None
        try:
            return float(sel[0])
        except Exception:
            # try reading size from values
            vals = self.tick_tree.item(sel[0]).get("values", [])
            if vals:
                try:
                    return float(str(vals[0]).replace("kDa", "").replace("bp", "").strip())
                except Exception:
                    return None
        return None

    def _on_tick_selected(self):
        s = self._selected_tick_size()
        if s is None:
            self._assign_tick_size = None
            return
        self._assign_tick_size = float(s)
        pc = self.current_pc()
        try:
            marker = load_marker(pc.marker_calibration.marker_name)
        except Exception:
            return
        base_map = marker_base_positions(marker, pc.marker_calibration)
        if s in pc.marker_tick_overrides:
            y = int(pc.marker_tick_overrides[s])
            self.tick_y_var.set(str(y))
        elif s in base_map:
            y = int(base_map[s])
            self.tick_y_var.set(str(y))
        else:
            self.tick_y_var.set("")

    def apply_tick_y(self):
        s = self._selected_tick_size()
        if s is None:
            messagebox.showinfo("Marker ticks", "Select a size first.")
            return
        try:
            y = int(float(self.tick_y_var.get().strip()))
        except Exception:
            messagebox.showerror("Marker ticks", "Please enter a valid integer Y position (panel pixels).")
            return

        panel_h = self.current_panel().height
        y = max(0, min(panel_h - 1, y))
        pc = self.current_pc()
        pc.marker_tick_overrides[float(s)] = int(y)
        try:
            pc.marker_tick_hidden = [v for v in pc.marker_tick_hidden if float(v) != float(s)]
        except Exception:
            pass
        self.refresh_preview()

    def clear_tick_override(self):
        s = self._selected_tick_size()
        if s is None:
            return
        pc = self.current_pc()
        try:
            pc.marker_tick_overrides.pop(float(s), None)
        except Exception:
            pass
        self.refresh_preview()

    def hide_selected_tick(self):
        s = self._selected_tick_size()
        if s is None:
            return
        pc = self.current_pc()
        fs = float(s)
        if fs not in pc.marker_tick_hidden:
            pc.marker_tick_hidden.append(fs)
        self.refresh_preview()

    def show_all_ticks(self):
        pc = self.current_pc()
        pc.marker_tick_hidden = []
        self.refresh_preview()

    # --- UI overlays (move handles + picked band highlights) ---
    def _draw_overlays(self):
        try:
            self.canvas.delete("overlay_ui")
        except Exception:
            pass

        pc = self.current_pc()
        panel = self.current_panel()
        layout = compute_render_layout(panel, pc, self.app.state.settings)
        z = float(self.viewport.zoom)

        def _draw_star(cx: float, cy: float, color: str = "#D80000", r: float = 5.0):
            _draw_canvas_asterisk(
                self.canvas,
                cx=float(cx),
                cy=float(cy),
                r=max(3.0, float(r)),
                color=str(color),
                width=2,
                tags=("overlay_ui",),
            )

        # Show picked points as horizontal guides (panel-relative y)
        try:
            if pc.marker_calibration.picked:
                lanes = max(1, pc.lanes)
                ladder_lane = max(1, min(lanes, int(pc.marker_calibration.ladder_lane) if pc.marker_calibration.ladder_lane else 1))
                x_tick = layout["lane_lefts"][ladder_lane - 1] - 8 + int(pc.marker_x_offset)
                out_w = int(layout["out_w"])
                py0 = int(layout["panel_origin"][1])
                for idx, (y_panel, _s) in enumerate(pc.marker_calibration.picked):
                    y_out = py0 + int(y_panel) + int(pc.marker_y_offset)
                    cy = y_out * z
                    self.canvas.create_line(0, cy, out_w * z, cy, fill="lime", width=1, tags=("overlay_ui",))
                    cx = x_tick * z
                    self.canvas.create_oval(cx - 3, cy - 3, cx + 3, cy + 3, outline="lime", fill="lime", tags=("overlay_ui",))
        except Exception:
            pass

        # Exact clicked points for current picking session (temporary overlay).
        try:
            if len(self._picked_click_points) != len(pc.marker_calibration.picked):
                lanes = max(1, pc.lanes)
                ladder_lane = max(1, min(lanes, int(pc.marker_calibration.ladder_lane) if pc.marker_calibration.ladder_lane else 1))
                lx0, lx1 = lane_x_bounds(panel, pc, ladder_lane)
                x_default = int(round((float(lx0) + float(lx1)) / 2.0))
                self._picked_click_points = [(x_default, int(y)) for (y, _s) in pc.marker_calibration.picked]
            if self._picked_click_points:
                px0, py0 = layout["panel_origin"]
                for idx, (xp, yp) in enumerate(self._picked_click_points, start=1):
                    x_out = (int(px0) + int(xp)) * z
                    y_out = (int(py0) + int(yp)) * z
                    _draw_star(float(x_out), float(y_out), color="#D80000", r=5.0)
                    self.canvas.create_text(
                        x_out + 8, y_out - 8, text=str(int(idx)),
                        fill="#B00000", anchor="sw", tags=("overlay_ui",)
                    )
        except Exception:
            pass

        try:
            if self._tick_assign_click_points:
                px0, py0 = layout["panel_origin"]
                for xp, yp in self._tick_assign_click_points[-24:]:
                    x_out = (int(px0) + int(xp)) * z
                    y_out = (int(py0) + int(yp)) * z
                    _draw_star(float(x_out), float(y_out), color="#D80000", r=4.0)
        except Exception:
            pass

    def _recompute_fit(self):
        pc = self.current_pc()
        if len(pc.marker_calibration.picked) >= 2:
            ys = np.array([p[0] for p in pc.marker_calibration.picked], dtype=float)
            sizes = np.array([p[1] for p in pc.marker_calibration.picked], dtype=float)
            xs = np.log10(sizes)
            a, b = np.polyfit(xs, ys, 1)
            pc.marker_calibration.fit_a = float(a)
            pc.marker_calibration.fit_b = float(b)
        else:
            pc.marker_calibration.fit_a = None
            pc.marker_calibration.fit_b = None

    def _canvas_to_img_xy(self, x: int, y: int) -> Tuple[int, int]:
        # rendered-image coordinates (not the raw panel) since preview uses render_panel()
        return self.viewport.canvas_to_image(x, y)

    def _sync_text_to_state(self):
        pc = self.current_pc()
        name = self.marker_name.get().strip()
        if name:
            pc.marker_calibration.marker_name = name

    def _render_preview_image(self) -> Image.Image:
        # Sync both style and text so the preview updates immediately while typing/toggling.
        self._sync_style_to_state()
        self._sync_text_to_state()
        panel = self.current_panel().copy()
        pc = self.current_pc()
        return render_panel(panel, pc, self.app.state.settings, preview=True)

    def refresh_preview(self):
        img = self._render_preview_image()
        self.viewport.set_image(img, fit_if_needed=True)
        self._draw_overlays()
        self._populate_tick_tree()

    def _on_press(self, event):
        pc = self.current_pc()
        if not pc.include_marker and self._picking:
            self._picking = False
        if not pc.include_marker:
            self._assign_tick_size = None

        # Reposition mode takes precedence
        if self._reposition_pick_idx is not None:
            # next click sets y of selected pick
            x_panel, y_panel = self._click_to_panel_xy(event.x, event.y)
            i = self._reposition_pick_idx
            if 0 <= i < len(pc.marker_calibration.picked):
                old_size = pc.marker_calibration.picked[i][1]
                pc.marker_calibration.picked[i] = (int(y_panel), float(old_size))
                if 0 <= i < len(self._picked_click_points):
                    self._picked_click_points[i] = (int(x_panel), int(y_panel))
                self._recompute_fit()
            self._reposition_pick_idx = None
            self._update_pick_list()
            self.refresh_preview()
            return

        # Picking mode (ladder calibration)
        if self._picking:
            if not self.marker_name.get():
                return
            marker = load_marker(self.marker_name.get())
            idx = len(pc.marker_calibration.picked)
            if idx >= len(marker.sizes):
                messagebox.showinfo("Marker", "You have already picked all sizes in this marker.")
                self._picking = False
                return
            size = float(marker.sizes[idx])
            x_in_panel, y_in_panel = self._click_to_panel_xy(event.x, event.y)
            pc.marker_calibration.picked.append((y_in_panel, size))
            self._picked_click_points.append((x_in_panel, y_in_panel))
            self._recompute_fit()
            self._update_pick_list()
            self.refresh_preview()
            return

        # Direct single-size assignment from selected tick row.
        if self._assign_tick_size is not None:
            s = float(self._assign_tick_size)
            x_in_panel, y_click_panel = self._click_to_panel_xy(event.x, event.y)
            y_in_panel = self._click_to_panel_y(event.x, event.y)
            pc.marker_tick_overrides[s] = int(y_in_panel)
            self._tick_assign_click_points.append((int(x_in_panel), int(y_click_panel)))
            if len(self._tick_assign_click_points) > 64:
                self._tick_assign_click_points = self._tick_assign_click_points[-64:]
            next_s = self._next_smaller_tick_size(s)
            self._set_visible_ticks_to_assigned(include_size=next_s if next_s is not None else s)
            self._assign_tick_size = next_s
            self.refresh_preview()
            return

    def _back(self):
        self._picking = False
        self._reposition_pick_idx = None
        self._assign_tick_size = None
        self.app.show_frame("AnnotationFrame")

    def _next(self):
        self._picking = False
        self._reposition_pick_idx = None
        self._assign_tick_size = None
        pc = self.current_pc()
        pc.run_band_analysis = bool(self.run_analysis_var.get())
        if pc.run_band_analysis:
            self.app.show_frame("AnalysisFrame")
        else:
            self.app.show_frame("ReviewFrame")

    def _right_click_next(self, _event=None):
        self._next()
        return "break"


# ----------------------------
# Step 6b: Optional lane/band analysis
# ----------------------------
class AnalysisFrame(PanelAwareFrame):
    _FIT_MODE_LABELS: List[Tuple[str, str]] = [
        ("monotone_interp", "Monotone interpolation"),
        ("log_linear", "Log-linear regression"),
        ("log_quadratic", "Log-quadratic regression"),
        ("log_cubic", "Log-cubic regression (deg-3)"),
    ]

    def __init__(self, parent, app: GelAnnotatorApp):
        super().__init__(parent)
        self.app = app

        self.title = ttk.Label(self, text="Step 6b - Lane intensity & band analysis", font=("Arial", 18, "bold"))
        self.title.pack(pady=8)

        main = ttk.Frame(self)
        main.pack(fill="both", expand=True)

        # left: preview tabs (lane histogram + panel overlay)
        left = ttk.Frame(main)
        left.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        self.left_tabs = ttk.Notebook(left)
        self.left_tabs.pack(fill="both", expand=True)
        tab_hist = ttk.Frame(self.left_tabs)
        tab_overlay = ttk.Frame(self.left_tabs)
        self.left_tabs.add(tab_hist, text="Lane Histogram")
        self.left_tabs.add(tab_overlay, text="Lane Overlay + Fit")
        tab_pixhist = ttk.Frame(self.left_tabs)
        self.left_tabs.add(tab_pixhist, text="Pixel Histogram")
        self.hist_view = ZoomableCanvas(tab_hist, bg="#222222", show_tools=True)
        self.hist_view.pack(fill="both", expand=True)
        self.overlay_view = ZoomableCanvas(tab_overlay, bg="#222222", show_tools=True)
        self.overlay_view.pack(fill="both", expand=True)
        self.overlay_view.bind("<Button-3>", self._on_overlay_right_click)
        self.pixhist_view = ZoomableCanvas(tab_pixhist, bg="#222222", show_tools=False)
        self.pixhist_view.pack(fill="both", expand=True)
        self.left_tabs.bind("<<NotebookTabChanged>>", lambda _e=None: self._on_tab_changed())

        # right: controls + report
        self.right_scroll = VScrollPanel(main, width=self.app._scaled_sidebar_width(SIDEBAR_WIDTH_WIDE))
        self.right_scroll.pack(side="left", fill="y", padx=10, pady=10)
        right = self.right_scroll.inner

        self.lane_var = tk.IntVar(value=1)
        self.threshold_var = tk.DoubleVar(value=20.0)
        self.prom_var = tk.DoubleVar(value=7.0)
        self.min_dist_var = tk.IntVar(value=10)
        self.smooth_var = tk.IntVar(value=10)
        self.polarity_var = tk.StringVar(value="dark")
        self.bg_correct_var = tk.BooleanVar(value=False)
        self.bg_lane_var = tk.IntVar(value=1)
        self.fit_mode_var = tk.StringVar(value="Monotone interpolation")
        self.show_sizes_final_var = tk.BooleanVar(value=True)
        self.final_label_font_var = tk.IntVar(value=15)
        self.eq_var = tk.StringVar(value="")
        self.band_pick_summary_var = tk.StringVar(value="")
        self._lane_exclude_vars: Dict[int, tk.BooleanVar] = {}
        self._detected_band_values_by_lane: Dict[int, List[Optional[float]]] = {}
        self._band_tracks: List[Dict[int, int]] = []
        self._analysis_after_id: Optional[str] = None

        # ── Histogram lane selector ──────────────────────────────────────
        hist_lf = ttk.LabelFrame(right, text="Histogram lane")
        hist_lf.pack(fill="x", pady=(4, 6), padx=2)
        lane_row = ttk.Frame(hist_lf)
        lane_row.pack(fill="x", pady=4, padx=4)
        ttk.Label(lane_row, text="Lane:").pack(side="left")
        self.lane_spin = ttk.Spinbox(lane_row, from_=1, to=60, textvariable=self.lane_var, width=6, command=self._refresh_hist)
        self.lane_spin.pack(side="left", padx=6)
        self.lane_spin.bind("<KeyRelease>", lambda _e=None: self._refresh_hist())
        self.lane_indicator = ttk.Label(lane_row, text="", font=("Arial", 9, "italic"))
        self.lane_indicator.pack(side="left", padx=4)

        # ── Band detection parameters ─────────────────────────────────────
        det_lf = ttk.LabelFrame(right, text="Band detection")
        det_lf.pack(fill="x", pady=(0, 6), padx=2)

        def row_spin(label: str, var: tk.Variable, frm: int | float, to: int | float, parent=det_lf):
            rr = ttk.Frame(parent)
            rr.pack(fill="x", pady=2, padx=4)
            ttk.Label(rr, text=label).pack(side="left")
            sp = ttk.Spinbox(rr, from_=frm, to=to, textvariable=var, width=7, command=self._schedule_run_analysis)
            sp.pack(side="right")
            sp.bind("<KeyRelease>", lambda _e=None: self._schedule_run_analysis())
            return sp

        row_spin("Peak threshold", self.threshold_var, 0, 255)
        row_spin("Peak prominence", self.prom_var, 0, 255)
        row_spin("Min distance (px)", self.min_dist_var, 1, 200)
        row_spin("Smoothing window", self.smooth_var, 1, 51)

        fit_row = ttk.Frame(det_lf)
        fit_row.pack(fill="x", pady=2, padx=4)
        ttk.Label(fit_row, text="Marker fit model").pack(side="left")
        fit_cb = ttk.Combobox(
            fit_row,
            textvariable=self.fit_mode_var,
            values=[label for _k, label in self._FIT_MODE_LABELS],
            state="readonly",
            width=25,
        )
        fit_cb.pack(side="right")
        fit_cb.bind("<<ComboboxSelected>>", lambda _e=None: self._schedule_run_analysis())

        pol_row = ttk.Frame(det_lf)
        pol_row.pack(fill="x", pady=2, padx=4)
        ttk.Label(pol_row, text="Band polarity").pack(side="left")
        pol_cb = ttk.Combobox(pol_row, textvariable=self.polarity_var, values=["dark", "light"], state="readonly", width=8)
        pol_cb.pack(side="right")
        pol_cb.bind("<<ComboboxSelected>>", lambda _e=None: self._schedule_run_analysis())

        # Background subtraction mode
        bg_lf = ttk.LabelFrame(det_lf, text="Background subtraction")
        bg_lf.pack(fill="x", pady=(4, 4), padx=4)

        self.bg_mode_var = tk.StringVar(value="none")  # none | lane | rolling_ball | local_adjacent
        bg_modes = [
            ("None", "none"),
            ("Reference lane", "lane"),
            ("Rolling ball", "rolling_ball"),
            ("Local adjacent", "local_adjacent"),
        ]
        for label, val in bg_modes:
            ttk.Radiobutton(bg_lf, text=label, variable=self.bg_mode_var, value=val,
                            command=self._schedule_run_analysis).pack(anchor="w", padx=4, pady=1)

        bg_opts = ttk.Frame(bg_lf)
        bg_opts.pack(fill="x", padx=4, pady=(2, 4))
        ttk.Label(bg_opts, text="Ref lane:").pack(side="left")
        bg_spin = ttk.Spinbox(bg_opts, from_=1, to=60, textvariable=self.bg_lane_var, width=5,
                              command=self._schedule_run_analysis)
        bg_spin.pack(side="left", padx=4)
        bg_spin.bind("<KeyRelease>", lambda _e=None: self._schedule_run_analysis())

        self.rolling_ball_radius_var = tk.IntVar(value=50)
        ttk.Label(bg_opts, text="  Ball radius:").pack(side="left")
        rb_spin = ttk.Spinbox(bg_opts, from_=5, to=500, textvariable=self.rolling_ball_radius_var, width=6,
                              command=self._schedule_run_analysis)
        rb_spin.pack(side="left", padx=4)
        rb_spin.bind("<KeyRelease>", lambda _e=None: self._schedule_run_analysis())

        # ── Export / size labels ──────────────────────────────────────────
        exp_lf = ttk.LabelFrame(right, text="Band size labels (final image)")
        exp_lf.pack(fill="x", pady=(0, 6), padx=2)
        ttk.Checkbutton(
            exp_lf,
            text="Add calculated band sizes below image",
            variable=self.show_sizes_final_var,
            command=self._sync_final_size_options,
        ).pack(anchor="w", pady=(4, 2), padx=4)
        font_row = ttk.Frame(exp_lf)
        font_row.pack(fill="x", pady=2, padx=4)
        ttk.Label(font_row, text="Font size").pack(side="left")
        self.final_label_font_spin = ttk.Spinbox(
            font_row,
            from_=6,
            to=96,
            textvariable=self.final_label_font_var,
            width=6,
            command=self._sync_final_size_options,
        )
        self.final_label_font_spin.pack(side="right")
        self.final_label_font_spin.bind("<KeyRelease>", lambda _e=None: self._sync_final_size_options())
        ttk.Button(exp_lf, text="Select included bands per lane...", command=self._open_band_include_dialog).pack(fill="x", pady=(2, 2), padx=4)
        ttk.Label(exp_lf, textvariable=self.band_pick_summary_var, justify="left", wraplength=360).pack(anchor="w", padx=4, pady=(0, 2))
        ttk.Label(exp_lf, text="Exclude lanes:").pack(anchor="w", padx=4)
        self.exclude_wrap = ttk.Frame(exp_lf)
        self.exclude_wrap.pack(fill="x", pady=(0, 4), padx=4)

        # ── Run button ────────────────────────────────────────────────────
        ttk.Button(right, text="Run / Update analysis", style="Primary.TButton", command=self.run_analysis).pack(fill="x", pady=6, padx=2)
        ttk.Button(right, text="Save analysis report...", style="Success.TButton", command=self.save_report).pack(fill="x", pady=2, padx=2)

        # ── Marker fit result ─────────────────────────────────────────────
        fit_lf = ttk.LabelFrame(right, text="Marker fit result")
        fit_lf.pack(fill="x", pady=(6, 4), padx=2)
        ttk.Label(fit_lf, textvariable=self.eq_var, justify="left", wraplength=360).pack(anchor="w", padx=4, pady=4)
        self.report_box: Optional[tk.Text] = None

        ttk.Separator(right).pack(fill="x", pady=8)
        nav = ttk.Frame(right)
        nav.pack(fill="x")
        ttk.Button(nav, text="Back", command=lambda: self.app.show_frame("MarkerFrame")).pack(side="left")
        ttk.Button(nav, text="Next -> Review", style="Primary.TButton", command=lambda: self.app.show_frame("ReviewFrame")).pack(side="right")

        self._profiles: Dict[int, np.ndarray] = {}
        self._peaks: Dict[int, List[int]] = {}
        self._fit: Optional[Dict[str, Any]] = None
        self._fit_points: List[Tuple[float, float]] = []
        self._report_text: str = ""
        self._lane_meta_text: Dict[int, str] = {}
        self._lane_meta_lines: Dict[int, List[str]] = {}
        self._lane_bounds: List[Tuple[int, int]] = []
        self._global_hist_x_max: float = 1.0

    def _fit_mode_key_from_label(self, label: str) -> str:
        s = str(label or "").strip()
        for key, lab in self._FIT_MODE_LABELS:
            if s == lab:
                return key
        for key, _lab in self._FIT_MODE_LABELS:
            if s == key:
                return key
        return "monotone_interp"

    def _fit_mode_label_from_key(self, key: str) -> str:
        s = str(key or "").strip()
        for k, lab in self._FIT_MODE_LABELS:
            if s == k:
                return lab
        return self._FIT_MODE_LABELS[0][1]

    def on_show(self):
        if not self.app.state.panels:
            self.app.show_frame("SettingsFrame")
            return
        i = self.app.state.current_panel_index + 1
        n = len(self.app.state.panels)
        self.title.configure(text=f"Step 6b - Lane intensity & band analysis ({i}/{n})")

        pc = self.current_pc()
        self.threshold_var.set(float(pc.analysis_peak_threshold))
        self.prom_var.set(float(pc.analysis_prominence))
        self.min_dist_var.set(int(pc.analysis_min_distance))
        self.smooth_var.set(int(pc.analysis_smooth_window))
        self.polarity_var.set(str(pc.analysis_polarity))
        self.bg_correct_var.set(bool(getattr(pc, "analysis_background_correction", False)))
        self.bg_lane_var.set(max(1, int(getattr(pc, "analysis_background_lane", 1) or 1)))
        self.bg_mode_var.set(str(getattr(pc, "analysis_bg_mode", "none") or "none"))
        self.rolling_ball_radius_var.set(int(getattr(pc, "analysis_rolling_ball_radius", 50) or 50))
        self.fit_mode_var.set(self._fit_mode_label_from_key(str(getattr(pc, "analysis_fit_mode", "monotone_interp"))))
        self.show_sizes_final_var.set(bool(getattr(pc, "show_band_sizes_on_final", True)))
        self.final_label_font_var.set(int(getattr(pc, "final_band_label_font_size", 15) or 15))
        self.lane_var.set(max(1, min(int(pc.lanes), int(self.lane_var.get()))))
        try:
            self.lane_spin.configure(to=max(1, int(pc.lanes)))
        except Exception:
            pass
        self._build_lane_exclude_ui()
        self._update_band_pick_summary()
        self.run_analysis()

    def on_hide(self):
        # Cancel delayed analysis when navigating away from this step.
        if self._analysis_after_id is not None:
            try:
                self.after_cancel(self._analysis_after_id)
            except Exception:
                pass
            self._analysis_after_id = None

    def _sync_to_state(self):
        pc = self.current_pc()
        pc.analysis_peak_threshold = float(self.threshold_var.get())
        pc.analysis_prominence = float(self.prom_var.get())
        pc.analysis_min_distance = max(1, int(self.min_dist_var.get()))
        pc.analysis_smooth_window = max(1, int(self.smooth_var.get()))
        pc.analysis_polarity = str(self.polarity_var.get())
        pc.analysis_background_correction = (self.bg_mode_var.get() != "none")
        pc.analysis_background_lane = max(1, int(self.bg_lane_var.get()))
        pc.analysis_bg_mode = str(self.bg_mode_var.get())
        pc.analysis_rolling_ball_radius = max(5, int(self.rolling_ball_radius_var.get()))
        pc.analysis_fit_mode = self._fit_mode_key_from_label(self.fit_mode_var.get())
        self._sync_final_size_options()

    def _schedule_run_analysis(self):
        if self._analysis_after_id is not None:
            try:
                self.after_cancel(self._analysis_after_id)
            except Exception:
                pass
        self._analysis_after_id = self.after(180, self.run_analysis)

    def _on_bg_correction_toggled(self):
        # Mode-specific defaults requested by user.
        if bool(self.bg_correct_var.get()):
            try:
                self.threshold_var.set(3.0)
            except Exception:
                pass
            try:
                self.smooth_var.set(5)
            except Exception:
                pass
        self._schedule_run_analysis()

    def _build_lane_exclude_ui(self):
        if not self._has_current_panel():
            return
        for w in self.exclude_wrap.winfo_children():
            w.destroy()
        self._lane_exclude_vars = {}
        pc = self.current_pc()
        excluded = set(int(x) for x in (getattr(pc, "band_size_excluded_lanes", []) or []) if isinstance(x, int) or str(x).isdigit())
        lanes = max(1, int(pc.lanes))
        cols = 8
        for lane in range(1, lanes + 1):
            v = tk.BooleanVar(value=(lane in excluded))
            self._lane_exclude_vars[lane] = v
            cb = ttk.Checkbutton(
                self.exclude_wrap,
                text=str(lane),
                variable=v,
                command=self._sync_final_size_options,
            )
            cb.grid(row=(lane - 1) // cols, column=(lane - 1) % cols, sticky="w", padx=4, pady=2)

    def _sync_final_size_options(self):
        if not self._has_current_panel():
            return
        pc = self.current_pc()
        pc.show_band_sizes_on_final = bool(self.show_sizes_final_var.get())
        try:
            fs = int(self.final_label_font_var.get())
        except Exception:
            fs = 15
        pc.final_band_label_font_size = max(6, int(fs))
        excluded: List[int] = []
        for lane, var in self._lane_exclude_vars.items():
            try:
                if bool(var.get()):
                    excluded.append(int(lane))
            except Exception:
                continue
        excluded.sort()
        pc.band_size_excluded_lanes = excluded
        self._normalize_band_include_selection()
        self._update_band_pick_summary()

    def _normalize_band_include_selection(self):
        if not self._has_current_panel():
            return
        if not self._detected_band_values_by_lane:
            return
        pc = self.current_pc()
        raw = getattr(pc, "band_size_included_bands_by_lane", {}) or {}
        normalized: Dict[int, List[int]] = {}
        for lane, vals in self._detected_band_values_by_lane.items():
            try:
                li = int(lane)
            except Exception:
                continue
            count = len(vals or [])
            if count <= 0:
                continue
            src_present = (li in raw) or (str(li) in raw)
            src = raw.get(li, raw.get(str(li), []))
            picks: List[int] = []
            for v in list(src or []):
                try:
                    bi = int(v)
                except Exception:
                    continue
                if 1 <= bi <= count:
                    picks.append(bi)
            picks = sorted(set(picks))
            if not picks and not src_present:
                picks = list(range(1, count + 1))
            normalized[li] = picks
        pc.band_size_included_bands_by_lane = normalized

    def _update_band_pick_summary(self):
        if not self._has_current_panel():
            self.band_pick_summary_var.set("")
            return
        pc = self.current_pc()
        detected = self._detected_band_values_by_lane
        if not detected:
            self.band_pick_summary_var.set("Included bands per lane: pending (run analysis).")
            return
        sel = getattr(pc, "band_size_included_bands_by_lane", {}) or {}
        customized = 0
        lanes_with_bands = 0
        for lane, vals in detected.items():
            n = len(vals or [])
            if n <= 0:
                continue
            lanes_with_bands += 1
            picks = sel.get(int(lane), sel.get(str(int(lane)), list(range(1, n + 1))))
            norm = sorted(set(int(x) for x in list(picks or []) if str(x).isdigit()))
            if norm != list(range(1, n + 1)):
                customized += 1
        if lanes_with_bands <= 0:
            self.band_pick_summary_var.set("Included bands per lane: no detected bands.")
        elif customized <= 0:
            self.band_pick_summary_var.set("Included bands per lane: all detected bands selected.")
        else:
            self.band_pick_summary_var.set(f"Included bands per lane: custom selection in {customized} lane(s).")

    def _open_band_include_dialog(self):
        if not self._has_current_panel():
            return
        if not self._detected_band_values_by_lane:
            messagebox.showinfo("Band selection", "Run analysis first to detect bands.")
            return
        pc = self.current_pc()
        unit = str(getattr(pc, "analysis_band_size_unit", "") or "").strip()
        selected = getattr(pc, "band_size_included_bands_by_lane", {}) or {}

        win = tk.Toplevel(self)
        win.title("Select included bands per lane")
        fit_toplevel_to_screen(win, 540, 620)

        body = ttk.Frame(win)
        body.pack(fill="both", expand=True, padx=8, pady=8)
        sc = VScrollPanel(body, width=self.app._scaled_sidebar_width(500))
        sc.pack(fill="both", expand=True)
        inner = sc.inner

        vars_by_lane: Dict[int, Dict[int, tk.BooleanVar]] = {}
        lanes = sorted(int(k) for k in self._detected_band_values_by_lane.keys())
        for lane in lanes:
            vals = list(self._detected_band_values_by_lane.get(lane, []) or [])
            if not vals:
                continue
            frm = ttk.LabelFrame(inner, text=f"Lane {lane}")
            frm.pack(fill="x", padx=4, pady=4)
            sel_lane = selected.get(lane, selected.get(str(lane), list(range(1, len(vals) + 1))))
            sel_set = set()
            for v in list(sel_lane or []):
                try:
                    sel_set.add(int(v))
                except Exception:
                    continue
            vars_by_lane[lane] = {}
            for idx, est in enumerate(vals, start=1):
                if est is None or (not np.isfinite(float(est))) or float(est) <= 0:
                    est_txt = "n/a"
                else:
                    est_txt = f"{int(round(float(est)))}"
                txt = f"Band {idx}: {est_txt}"
                if unit and est_txt != "n/a":
                    txt += f" {unit}"
                v = tk.BooleanVar(value=(idx in sel_set))
                vars_by_lane[lane][idx] = v
                ttk.Checkbutton(frm, text=txt, variable=v).pack(anchor="w", padx=6, pady=1)

        footer = ttk.Frame(win)
        footer.pack(fill="x", padx=8, pady=(0, 8))

        def _select_all():
            for lane_vars in vars_by_lane.values():
                for v in lane_vars.values():
                    v.set(True)

        def _clear_all():
            for lane_vars in vars_by_lane.values():
                for v in lane_vars.values():
                    v.set(False)

        def _apply():
            new_map: Dict[int, List[int]] = {}
            for lane, lane_vars in vars_by_lane.items():
                picks: List[int] = []
                for idx, v in lane_vars.items():
                    try:
                        if bool(v.get()):
                            picks.append(int(idx))
                    except Exception:
                        continue
                picks.sort()
                new_map[int(lane)] = picks
            pc.band_size_included_bands_by_lane = new_map
            self._sync_final_size_options()
            win.destroy()

        ttk.Button(footer, text="Select all", command=_select_all).pack(side="left")
        ttk.Button(footer, text="Clear all", command=_clear_all).pack(side="left", padx=6)
        ttk.Button(footer, text="Cancel", command=win.destroy).pack(side="right")
        ttk.Button(footer, text="Apply", command=_apply).pack(side="right", padx=6)

    def _set_report_text(self, text: str):
        self._report_text = str(text)
        if self.report_box is None:
            return
        self.report_box.configure(state="normal")
        self.report_box.delete("1.0", "end")
        self.report_box.insert("1.0", self._report_text)
        self.report_box.configure(state="disabled")

    def _lane_context(self, lane: int) -> Tuple[str, List[str]]:
        """Return per-lane metadata string and display lines (headers + brackets)."""
        pc = self.current_pc()
        idx = int(lane) - 1
        header_parts: List[str] = []
        for hr in pc.header_rows:
            name = str(hr.name).strip()
            val = ""
            if 0 <= idx < len(hr.values):
                val = str(hr.values[idx]).strip()
            if name and val:
                header_parts.append(f"{name}={val}")
            elif val:
                header_parts.append(val)
            elif name:
                header_parts.append(f"{name}=-")

        bracket_parts: List[str] = []
        for g in pc.group_labels:
            s = int(g.start_lane)
            e = int(g.end_lane)
            if e < s:
                s, e = e, s
            if s <= int(lane) <= e:
                txt = str(g.text).strip()
                if txt and txt not in bracket_parts:
                    bracket_parts.append(txt)

        parts: List[str] = []
        if header_parts:
            parts.append("; ".join(header_parts))
        if bracket_parts:
            parts.append("Brackets: " + ", ".join(bracket_parts))
        context = " | ".join(parts) if parts else "-"

        lines: List[str] = [f"Lane {int(lane)}"]
        if header_parts:
            lines.append("Headers: " + "; ".join(header_parts))
        if bracket_parts:
            lines.append("Brackets: " + ", ".join(bracket_parts))
        return context, lines

    def run_analysis(self):
        self._analysis_after_id = None
        if not self._has_current_panel():
            return
        self._sync_to_state()
        panel = self.current_panel()
        pc = self.current_pc()
        lanes = max(1, int(pc.lanes))
        marker_lane = int(pc.marker_calibration.ladder_lane) if pc.marker_calibration else 1
        gray_panel = np.asarray(panel.convert("L"), dtype=float)
        lane_bounds = compute_lane_bounds(panel, pc)
        self._lane_bounds = list(lane_bounds)
        self._profiles = {}
        self._peaks = {}
        self._fit_points = []
        self._global_hist_x_max = 1.0
        lane_sizes_map: Dict[int, List[str]] = {}
        lane_values_map: Dict[int, List[Optional[float]]] = {}
        self._lane_meta_text = {}
        self._lane_meta_lines = {}

        try:
            marker_pos = marker_positions_for_panel(panel, pc, self.app.state.settings, include_hidden=True)
            self._fit_points = [(float(s), float(y) + float(pc.marker_y_offset)) for s, y in marker_pos if float(s) > 0]
        except Exception:
            marker_pos = []
            self._fit_points = []
        self._fit = fit_marker_curve_for_panel(panel, pc, self.app.state.settings, marker_positions_hidden=marker_pos)

        lines: List[str] = []
        bg_signal: Optional[np.ndarray] = None
        bg_lane = max(1, min(lanes, int(getattr(pc, "analysis_background_lane", 1) or 1)))
        _bg_mode = str(getattr(pc, "analysis_bg_mode", "none") or "none")
        _rolling_ball_radius = int(getattr(pc, "analysis_rolling_ball_radius", 50) or 50)
        if _bg_mode == "lane":
            try:
                bg_mean = _lane_mean_gray_profile(panel, pc, bg_lane, gray=gray_panel, lane_bounds=lane_bounds)
                bg_signal = (255.0 - bg_mean) if str(pc.analysis_polarity).lower() == "dark" else bg_mean
                bg_signal = np.asarray(np.maximum(bg_signal, 0.0), dtype=float)
                lines.append(f"Background correction: ON (lane {bg_lane})")
            except Exception as e:
                bg_signal = None
                lines.append(f"Background correction: OFF (failed: {e})")
        elif _bg_mode == "rolling_ball":
            lines.append(f"Background correction: Rolling ball (radius={_rolling_ball_radius})")
        elif _bg_mode == "local_adjacent":
            lines.append("Background correction: Local adjacent lanes")
        else:
            lines.append("Background correction: OFF")
        lines.append("")
        if self._fit:
            eq = _fit_equation_text(self._fit)
            self.eq_var.set(eq)
            lines.append(f"Marker fit: {eq}")
        else:
            self.eq_var.set("No marker fit available (need at least 2 marker points).")
            lines.append("Marker fit: unavailable")
        lines.append("")

        for lane in range(1, lanes + 1):
            lane_meta_txt, lane_meta_lines = self._lane_context(lane)
            self._lane_meta_text[int(lane)] = lane_meta_txt
            self._lane_meta_lines[int(lane)] = lane_meta_lines

            sig = lane_intensity_profile(
                panel,
                pc,
                lane_idx_1based=lane,
                polarity=str(pc.analysis_polarity),
                smooth_window=int(pc.analysis_smooth_window),
                background_signal=bg_signal if _bg_mode == "lane" else None,
                gray=gray_panel,
                lane_bounds=lane_bounds,
                bg_mode=_bg_mode,
                rolling_ball_radius=_rolling_ball_radius,
            )
            peaks = detect_bands_from_profile(
                sig,
                threshold=float(pc.analysis_peak_threshold),
                prominence=float(pc.analysis_prominence),
                min_distance=int(pc.analysis_min_distance),
            )
            self._profiles[lane] = sig
            self._peaks[lane] = peaks
            try:
                smax = float(np.nanmax(np.asarray(sig, dtype=float)))
                if np.isfinite(smax):
                    self._global_hist_x_max = max(float(self._global_hist_x_max), float(smax))
            except Exception:
                pass

            lines.append(f"Lane {lane}: {lane_meta_txt}")
            if not peaks:
                lines.append("  no bands detected")
                continue
            lane_sizes_map[int(lane)] = []
            lane_values_map[int(lane)] = []
            for band_idx, y in enumerate(peaks, start=1):
                intensity = float(sig[y]) if 0 <= y < len(sig) else float("nan")
                est = estimate_size_from_y(float(y), self._fit)
                unit = str(self._fit.get("unit", "")) if self._fit else ""
                est_txt = format_size_pretty(est, unit=unit)
                lane_sizes_map[int(lane)].append(str(est_txt))
                if est is None:
                    lane_values_map[int(lane)].append(None)
                else:
                    try:
                        estf = float(est)
                        if np.isfinite(estf) and estf > 0:
                            lane_values_map[int(lane)].append(estf)
                        else:
                            lane_values_map[int(lane)].append(None)
                    except Exception:
                        lane_values_map[int(lane)].append(None)
                lines.append(f"  y={int(y)} px, signal={intensity:.1f}, approx={est_txt}")
            lines.append("")

        pc.analysis_band_sizes_by_lane = lane_sizes_map
        pc.analysis_band_values_by_lane = lane_values_map
        if self._fit:
            pc.analysis_band_size_unit = str(self._fit.get("unit", "")).strip()
        else:
            pc.analysis_band_size_unit = str(getattr(pc.marker_calibration, "unit", "")).strip()
        self._detected_band_values_by_lane = dict(lane_values_map)
        self._normalize_band_include_selection()
        self._update_band_pick_summary()
        self._band_tracks = self._compute_band_tracks()
        report_text = "\n".join(lines).strip() + "\n"
        self._set_report_text(report_text)
        self._refresh_hist()
        self._refresh_overlay()
        self._refresh_pixel_histogram()

    def _render_overlay_image(self) -> Image.Image:
        panel = self.current_panel()
        pc = self.current_pc()
        lanes = max(1, int(pc.lanes))
        plot_w = 360
        gap = 14
        lane_num_strip = 30  # extra pixels at bottom for lane number circles
        out = Image.new("RGB", (panel.width + gap + plot_w, panel.height + lane_num_strip), (255, 255, 255))
        out.paste(panel, (0, 0))
        draw = ImageDraw.Draw(out)
        lane_bounds = list(self._lane_bounds or compute_lane_bounds(panel, pc))

        # Collect lane centers once for reuse
        lane_centers: List[float] = []
        if lane_bounds:
            for x0, x1 in lane_bounds:
                lane_centers.append(0.5 * (float(x0) + float(x1)))
        else:
            for lane in range(1, lanes + 1):
                lane_centers.append(float(lane_center_x(panel, pc, lane)))

        # Lane center lines + detected band marks
        for idx, cx in enumerate(lane_centers):
            lane = idx + 1
            draw.line((cx, 0, cx, panel.height), fill=(60, 180, 255), width=1)
            for y in self._peaks.get(lane, []):
                yy = int(y)
                draw.line((cx - 8, yy, cx + 8, yy), fill=(220, 0, 0), width=2)
                draw.ellipse((cx - 2, yy - 2, cx + 2, yy + 2), fill=(220, 0, 0), outline=(220, 0, 0))

        # Lane numbers in circles in the bottom strip
        try:
            num_font = get_font_with_size(self.app.state.settings, 13)
            r = 11.0
            cy = float(panel.height) + lane_num_strip / 2.0
            circ_outline_hex, txt_hex = _lane_circle_colors(self.app.state.settings)
            circ_outline = _hex_to_rgb(circ_outline_hex)
            txt_color = _hex_to_rgb(txt_hex)
            for i, cx in enumerate(lane_centers, start=1):
                draw.ellipse((cx - r, cy - r, cx + r, cy + r), fill=(255, 255, 255), outline=circ_outline, width=2)
                txt = str(int(i))
                bb = draw.textbbox((0, 0), txt, font=num_font)
                tw = float(bb[2] - bb[0]); th = float(bb[3] - bb[1])
                draw.text((float(cx) - tw / 2.0, float(cy) - th / 2.0 - float(bb[1])), txt, fill=txt_color, font=num_font)
        except Exception:
            pass

        # Fitted marker curve on right-side plot: y vs size (linear x-axis, no log plotting)
        plot_left = panel.width + gap
        px0 = plot_left + 58
        px1 = plot_left + plot_w - 14
        py0 = 30
        py1 = panel.height - 58
        draw.rectangle((px0, py0, px1, py1), outline=(0, 0, 0), width=1)
        unit = str(self._fit.get("unit", "")) if self._fit else ""
        draw.text((plot_left + 8, 6), "Marker fit (linear size axis)", fill=(0, 0, 0))
        draw.text((px0, panel.height - 16), f"Size ({unit})", fill=(0, 0, 0))
        draw.text((plot_left + 6, py0 - 2), "Y(px)", fill=(0, 0, 0))

        if self._fit and self._fit_points:
            xs = np.array([float(s) for s, _ in self._fit_points], dtype=float)
            x_min = float(np.min(xs))
            x_max = float(np.max(xs))
            if x_max - x_min < 1e-9:
                x_max = x_min + 1.0

            def x_to_px(xv: float) -> float:
                return px0 + (float(xv) - x_min) / (x_max - x_min) * (px1 - px0)

            # Grid/ticks first so labels can be placed outside the plot.
            y_tick_vals = sorted(set([0, int((panel.height - 1) * 0.25), int((panel.height - 1) * 0.5), int((panel.height - 1) * 0.75), max(0, panel.height - 1)]))
            for yv in y_tick_vals:
                yy = py0 + (float(yv) / max(1, panel.height - 1)) * (py1 - py0)
                draw.line((px0 - 4, yy, px0, yy), fill=(0, 0, 0), width=1)
                draw.line((px0, yy, px1, yy), fill=(232, 232, 232), width=1)
                txt = str(int(yv))
                bb = draw.textbbox((0, 0), txt)
                tw = bb[2] - bb[0]
                th = bb[3] - bb[1]
                draw.text((px0 - 8 - tw, yy - th / 2.0), txt, fill=(0, 0, 0))

            x_tick_vals = [x_min, x_min + 0.25 * (x_max - x_min), x_min + 0.5 * (x_max - x_min), x_min + 0.75 * (x_max - x_min), x_max]
            row_last_right = [-1e9, -1e9]
            for idx_tick, xv in enumerate(x_tick_vals):
                xx = x_to_px(float(xv))
                draw.line((xx, py1, xx, py1 + 4), fill=(0, 0, 0), width=1)
                draw.line((xx, py0, xx, py1), fill=(240, 240, 240), width=1)
                txt = format_size_pretty(float(xv), "")
                bb = draw.textbbox((0, 0), txt)
                tw = bb[2] - bb[0]
                tx = float(xx) - tw / 2.0
                tx = max(float(px0), min(float(px1 - tw), tx))
                row_idx = idx_tick % 2
                if tx <= row_last_right[row_idx] + 4:
                    row_idx = 1 - row_idx
                if tx <= row_last_right[row_idx] + 4:
                    tx = row_last_right[row_idx] + 6
                    tx = max(float(px0), min(float(px1 - tw), tx))
                ty = float(py1 + 7 + (12 * row_idx))
                draw.text((tx, ty), txt, fill=(0, 0, 0))
                row_last_right[row_idx] = max(row_last_right[row_idx], tx + tw)

            # Draw fitted curve by sampling across the x-axis (size domain). This ensures
            # regression curves reach the largest/smallest marker sizes without artificial clipping.
            pts: List[Tuple[float, float]] = []
            for i in range(700):
                xv = x_min + (x_max - x_min) * (i / 699.0)
                yv = _fit_eval_y_from_size(float(xv), self._fit)
                if yv is None:
                    continue
                try:
                    yvf = float(yv)
                except Exception:
                    continue
                if not np.isfinite(yvf):
                    continue
                if (yvf < 0.0) or (yvf > float(panel.height - 1)):
                    continue
                yy = py0 + (yvf / max(1, panel.height - 1)) * (py1 - py0)
                pts.append((x_to_px(xv), yy))
            if len(pts) >= 2:
                draw.line(pts, fill=(30, 90, 180), width=2)

            # Draw marker points
            for s, y in self._fit_points:
                xv = float(s)
                xx = x_to_px(xv)
                yy_panel = max(0.0, min(float(panel.height - 1), float(y)))
                yy = py0 + (yy_panel / max(1, panel.height - 1)) * (py1 - py0)
                draw.ellipse((xx - 3, yy - 3, xx + 3, yy + 3), fill=(0, 160, 0), outline=(0, 100, 0))
        return out

    def _refresh_overlay(self):
        if not self._has_current_panel():
            return
        img = self._render_overlay_image()
        self.overlay_view.set_image(img, fit_if_needed=True)

    def _render_pixel_histogram(self) -> Image.Image:
        """Render a pixel intensity histogram of the current panel."""
        try:
            panel = self.current_panel()
            gray = np.asarray(panel.convert("L"), dtype=float)
            counts, _ = np.histogram(gray.ravel(), bins=256, range=(0, 256))
            W, H = 640, 300
            img = Image.new("RGB", (W, H), (30, 30, 35))
            draw = ImageDraw.Draw(img)
            pad_l, pad_r, pad_t, pad_b = 48, 16, 16, 36
            plot_w = W - pad_l - pad_r
            plot_h = H - pad_t - pad_b
            max_count = max(1, int(np.max(counts)))
            bar_w = max(1.0, float(plot_w) / 256.0)
            # Draw bars
            for i in range(256):
                bh = int(float(counts[i]) / max_count * plot_h)
                x0 = pad_l + int(i * bar_w)
                x1 = pad_l + int((i + 1) * bar_w)
                y0 = pad_t + plot_h - bh
                y1 = pad_t + plot_h
                gray_val = i
                draw.rectangle((x0, y0, x1, y1), fill=(gray_val, gray_val, gray_val))
            # Axes
            draw.rectangle((pad_l, pad_t, pad_l + plot_w, pad_t + plot_h), outline=(180, 180, 180), width=1)
            # X-axis labels
            try:
                ax_font = get_font_with_size(self.app.state.settings, 10)
            except Exception:
                ax_font = None
            for xv in [0, 64, 128, 192, 255]:
                xx = pad_l + int(xv / 255.0 * plot_w)
                draw.line((xx, pad_t + plot_h, xx, pad_t + plot_h + 4), fill=(180, 180, 180), width=1)
                lbl = str(xv)
                bb = draw.textbbox((0, 0), lbl, font=ax_font)
                tw = bb[2] - bb[0]
                draw.text((xx - tw // 2, pad_t + plot_h + 6), lbl, fill=(200, 200, 200), font=ax_font)
            # Y-axis label
            draw.text((4, pad_t + plot_h // 2), "Count", fill=(180, 180, 180), font=ax_font)
            draw.text((pad_l + plot_w // 2 - 30, H - 14), "Pixel intensity (0-255)", fill=(180, 180, 180), font=ax_font)
            # Title
            draw.text((pad_l, 2), "Panel pixel intensity distribution", fill=(220, 220, 220), font=ax_font)
            return img
        except Exception as exc:
            logger.warning("Pixel histogram render failed: %s", exc)
            return Image.new("RGB", (640, 300), (30, 30, 35))

    def _refresh_pixel_histogram(self):
        if not self._has_current_panel():
            return
        try:
            img = self._render_pixel_histogram()
            self.pixhist_view.set_image(img, fit_if_needed=True)
        except Exception:
            pass

    def _on_tab_changed(self):
        try:
            idx = self.left_tabs.index(self.left_tabs.select())
            if idx == 2:  # Pixel Histogram tab
                self._refresh_pixel_histogram()
        except Exception:
            pass

    def _refresh_hist(self):
        if not self._has_current_panel():
            return
        pc = self.current_pc()
        panel = self.current_panel()
        lane = max(1, min(int(pc.lanes), int(self.lane_var.get())))
        self.lane_var.set(lane)
        # Update lane indicator with context info
        try:
            _ctx, ctx_lines = self._lane_context(lane)
            indicator = " · ".join(ctx_lines[1:]) if len(ctx_lines) > 1 else ""
            self.lane_indicator.configure(text=indicator)
        except Exception:
            pass
        sig = self._profiles.get(lane)
        if sig is None:
            return
        peaks = self._peaks.get(lane, [])
        try:
            lane_bounds = list(self._lane_bounds or compute_lane_bounds(panel, pc))
            lane_idx = max(1, min(len(lane_bounds), int(lane))) if lane_bounds else int(lane)
            if lane_bounds:
                x0, x1 = lane_bounds[lane_idx - 1]
            else:
                x0, x1 = lane_x_bounds(panel, pc, lane)
            lane_slice = panel.crop((x0, 0, x1, panel.height))
        except Exception:
            lane_slice = None
        peak_val_labels: Dict[int, str] = {}
        try:
            vals = list(self._detected_band_values_by_lane.get(int(lane), []) or [])
            unit = str((self._fit or {}).get("unit", ""))
            for p, v in zip(list(peaks or []), vals):
                if v is None:
                    continue
                peak_val_labels[int(p)] = format_size_pretty(v, unit)
        except Exception:
            peak_val_labels = {}
        global_hist_x_max = float(max(1.0, float(getattr(self, "_global_hist_x_max", 1.0) or 1.0)))
        if (not np.isfinite(global_hist_x_max)) or global_hist_x_max <= 0:
            try:
                global_hist_x_max = float(max(1.0, np.nanmax(np.asarray(sig, dtype=float))))
            except Exception:
                global_hist_x_max = 1.0
        img = render_lane_histogram(
            sig,
            peaks,
            threshold=float(pc.analysis_peak_threshold),
            lane_slice=lane_slice,
            title=f"Lane {lane} intensity profile",
            lane_meta_lines=list(self._lane_meta_lines.get(int(lane), [])),
            peak_value_labels=peak_val_labels,
            size_unit=str((self._fit or {}).get("unit", "")),
            x_scale_max=float(global_hist_x_max),
        )
        self.hist_view.set_image(img, fit_if_needed=True)

    def _on_overlay_right_click(self, event):
        """Right-click on overlay: show context menu to add/delete band."""
        if not self._has_current_panel():
            return
        try:
            canvas = self.overlay_view
            # Convert screen coords to image coords
            img_x, img_y = canvas.screen_to_image(event.x, event.y)
            if img_x is None or img_y is None:
                return
            img_x = int(img_x)
            img_y = int(img_y)
            # Only interact within the panel image area (not the plot to the right)
            panel = self.current_panel()
            if img_x >= panel.width or img_y >= panel.height or img_y < 0 or img_x < 0:
                return
            pc = self.current_pc()
            lane_bounds = list(self._lane_bounds or compute_lane_bounds(panel, pc))
            # Find which lane was clicked
            clicked_lane = None
            for lane_idx, (x0, x1) in enumerate(lane_bounds, start=1):
                if x0 <= img_x < x1:
                    clicked_lane = lane_idx
                    break
            if clicked_lane is None:
                return
            # Check if there's a nearby peak to delete (within 8px)
            existing_peaks = list(self._peaks.get(clicked_lane, []))
            nearby_peak = None
            for p in existing_peaks:
                if abs(p - img_y) <= 8:
                    nearby_peak = p
                    break
            menu = tk.Menu(self, tearoff=0)
            if nearby_peak is not None:
                menu.add_command(
                    label=f"Delete band at y={nearby_peak} (lane {clicked_lane})",
                    command=lambda: self._delete_band(clicked_lane, nearby_peak)
                )
            menu.add_command(
                label=f"Add band at y={img_y} (lane {clicked_lane})",
                command=lambda: self._add_band(clicked_lane, img_y)
            )
            menu.tk_popup(event.x_root, event.y_root)
        except Exception as exc:
            logger.warning("Overlay right-click failed: %s", exc)

    def _add_band(self, lane: int, y_px: int):
        """Manually add a peak at y_px in the given lane."""
        peaks = list(self._peaks.get(lane, []))
        if y_px not in peaks:
            peaks.append(int(y_px))
            peaks.sort()
        self._peaks[lane] = peaks
        self._recompute_band_values_for_lane(lane)
        self._refresh_overlay()
        self._refresh_hist()

    def _delete_band(self, lane: int, y_px: int):
        """Manually remove a peak at y_px from the given lane."""
        peaks = list(self._peaks.get(lane, []))
        peaks = [p for p in peaks if p != y_px]
        self._peaks[lane] = peaks
        self._recompute_band_values_for_lane(lane)
        self._refresh_overlay()
        self._refresh_hist()

    def _recompute_band_values_for_lane(self, lane: int):
        """Recompute size estimates for all bands in a lane after a manual edit."""
        try:
            peaks = self._peaks.get(lane, [])
            sig = self._profiles.get(lane)
            unit = str((self._fit or {}).get("unit", ""))
            vals: List[Optional[float]] = []
            sizes_str: List[str] = []
            for y in peaks:
                est = estimate_size_from_y(float(y), self._fit)
                vals.append(est)
                sizes_str.append(format_size_pretty(est, unit) if est is not None else "?")
            self._detected_band_values_by_lane[lane] = vals
            pc = self.current_pc()
            pc.analysis_band_values_by_lane[lane] = [v for v in vals]
            pc.analysis_band_sizes_by_lane[lane] = sizes_str
        except Exception as exc:
            logger.warning("_recompute_band_values_for_lane failed: %s", exc)

    def _compute_band_tracks(self, tolerance_px: int = 12) -> List[Dict[int, int]]:
        """
        Group peaks from different lanes that are within tolerance_px of each other.
        Returns a list of track dicts: {lane_idx: peak_y_px}.
        """
        if not self._peaks:
            return []
        pc = self.current_pc()
        lanes = list(range(1, max(1, int(pc.lanes)) + 1))
        # Collect all (lane, y) pairs sorted by y
        all_peaks: List[Tuple[int, int]] = []
        for lane in lanes:
            for y in self._peaks.get(lane, []):
                all_peaks.append((int(lane), int(y)))
        all_peaks.sort(key=lambda t: t[1])
        tracks: List[Dict[int, int]] = []
        used: set = set()
        for i, (lane_i, y_i) in enumerate(all_peaks):
            if i in used:
                continue
            track: Dict[int, int] = {lane_i: y_i}
            used.add(i)
            for j, (lane_j, y_j) in enumerate(all_peaks):
                if j in used or lane_j == lane_i:
                    continue
                # Check if y_j is close to the median y of the current track
                track_ys = list(track.values())
                median_y = float(sorted(track_ys)[len(track_ys) // 2])
                if abs(y_j - median_y) <= tolerance_px and lane_j not in track:
                    track[lane_j] = y_j
                    used.add(j)
            if len(track) >= 2:
                tracks.append(track)
        return tracks

    def save_report(self):
        if not self._profiles:
            messagebox.showinfo("Analysis", "No analysis data to export yet. Run analysis first.")
            return
        if _HAS_OPENPYXL:
            filetypes = [("Excel workbook", "*.xlsx"), ("CSV (UTF-8)", "*.csv"), ("Text", "*.txt")]
            default_ext = ".xlsx"
            initial = "band_analysis.xlsx"
        else:
            filetypes = [("CSV (UTF-8)", "*.csv"), ("Text", "*.txt")]
            default_ext = ".csv"
            initial = "band_analysis.csv"
        path = filedialog.asksaveasfilename(
            title="Export analysis data",
            defaultextension=default_ext,
            initialfile=initial,
            filetypes=filetypes,
        )
        if not path:
            return
        ext = pathlib.Path(path).suffix.lower()
        try:
            if ext == ".xlsx" and _HAS_OPENPYXL:
                self._save_xlsx(path)
            elif ext in (".csv", ".txt"):
                self._save_csv(path)
            else:
                # If openpyxl not available and .xlsx chosen, fall back to csv
                path = str(pathlib.Path(path).with_suffix(".csv"))
                self._save_csv(path)
            messagebox.showinfo("Analysis", f"Exported:\n{path}")
        except Exception as exc:
            logger.warning("Export failed: %s", exc)
            messagebox.showerror("Export failed", str(exc))

    def _build_export_rows(self) -> List[Dict[str, Any]]:
        """Build a list of row dicts for export."""
        rows: List[Dict[str, Any]] = []
        pc = self.current_pc()
        panel_idx = int(self.app.state.current_panel_index) + 1
        unit = str((self._fit or {}).get("unit", ""))
        for lane in range(1, max(1, int(pc.lanes)) + 1):
            peaks = self._peaks.get(lane, [])
            sig = self._profiles.get(lane)
            vals = list((self._detected_band_values_by_lane or {}).get(lane, []) or [])
            # Try to get a lane label from the first header row
            lane_label = ""
            try:
                if pc.header_rows:
                    hr = pc.header_rows[0]
                    lane_label = str((hr.values or [])[lane - 1]) if lane <= len(hr.values or []) else ""
            except Exception:
                pass
            if not peaks:
                rows.append({
                    "Panel": panel_idx,
                    "Lane": lane,
                    "Lane label": lane_label,
                    "Band": "",
                    f"Size ({unit})": "",
                    "Intensity": "",
                    "Peak Y (px)": "",
                })
                continue
            for band_idx, y in enumerate(peaks, start=1):
                intensity = float(sig[y]) if sig is not None and 0 <= y < len(sig) else ""
                size_val = vals[band_idx - 1] if band_idx - 1 < len(vals) else None
                size_str = "" if size_val is None else round(float(size_val), 2)
                rows.append({
                    "Panel": panel_idx,
                    "Lane": lane,
                    "Lane label": lane_label,
                    "Band": band_idx,
                    f"Size ({unit})": size_str,
                    "Intensity": round(float(intensity), 3) if intensity != "" else "",
                    "Peak Y (px)": int(y),
                })
        return rows

    def _save_xlsx(self, path: str) -> None:
        rows = self._build_export_rows()
        if not rows:
            raise ValueError("No data to export.")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Band Analysis"
        headers = list(rows[0].keys())
        # Header row styling
        header_fill = PatternFill("solid", fgColor="2E6EA0")
        header_font = XLFont(bold=True, color="FFFFFF")
        header_align = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border
        # Data rows
        alt_fill = PatternFill("solid", fgColor="EAF2FB")
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, h in enumerate(headers, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row[h])
                cell.alignment = Alignment(horizontal="center")
                cell.border = border
                if row_idx % 2 == 0:
                    cell.fill = alt_fill
        # Auto-fit column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max(10, min(30, max_len + 4))
        # Also write the text report on a second sheet
        ws2 = wb.create_sheet("Report (text)")
        for i, line in enumerate(self._report_text.splitlines(), start=1):
            ws2.cell(row=i, column=1, value=line)
        wb.save(path)

    def _save_csv(self, path: str) -> None:
        rows = self._build_export_rows()
        if not rows:
            # Fall back to text report
            pathlib.Path(path).write_text(self._report_text, encoding="utf-8-sig")
            return
        import csv
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
            writer.writeheader()
            writer.writerows(rows)


# ----------------------------
# Step 7: Review / Save / Next panel ->/ Next image
# ----------------------------
class ReviewFrame(PanelAwareFrame):
    def __init__(self, parent, app: GelAnnotatorApp):
        super().__init__(parent)
        self.app = app

        ttk.Label(self, text="Step 7 - Review and export", font=("Arial", 18, "bold")).pack(pady=8)

        main = ttk.Frame(self)
        main.pack(fill="both", expand=True)

        self.viewport = ZoomableCanvas(main, bg="#222222", show_tools=True)
        self.viewport.pack(side="left", fill="both", expand=True, padx=10, pady=10)

        self.right_scroll = VScrollPanel(main, width=self.app._scaled_sidebar_width(SIDEBAR_WIDTH_NARROW))
        self.right_scroll.pack(side="left", fill="y", padx=10, pady=10)
        right = self.right_scroll.inner

        ttk.Label(right, text="Go back to step:").pack(anchor="w")
        self.back_step = tk.StringVar(value="Layout")
        ttk.Combobox(
            right,
            textvariable=self.back_step,
            values=["Edit", "Panels", "Layout", "Annotations", "Marker", "Analysis"],
            state="readonly",
        ).pack(fill="x", pady=4)
        ttk.Button(right, text="Go", command=self.go_back).pack(fill="x", pady=4)

        ttk.Separator(right).pack(fill="x", pady=10)

        ttk.Button(right, text="💾  Save annotated image", style="Success.TButton", command=self.save).pack(fill="x", pady=4)
        ttk.Button(right, text="Next panel ->", command=self.next_panel).pack(fill="x", pady=4)
        ttk.Button(right, text="Start next image (keep settings)", command=self.next_image).pack(fill="x", pady=8)

        ttk.Separator(right).pack(fill="x", pady=6)
        ttk.Label(right, text="Project file", font=("Arial", 9, "bold")).pack(anchor="w", pady=(2, 0))
        ttk.Button(right, text="Save project (.gelproj)", command=self.app.save_project).pack(fill="x", pady=2)
        ttk.Button(right, text="Load project (.gelproj)", command=self.app.load_project).pack(fill="x", pady=2)

        ttk.Separator(right).pack(fill="x", pady=8)
        self.all_band_label_font_var = tk.IntVar(value=15)
        bl_row = ttk.Frame(right)
        bl_row.pack(fill="x", pady=2)
        ttk.Label(bl_row, text="Band label font size (all panels):").pack(side="left")
        ttk.Spinbox(bl_row, from_=6, to=96, textvariable=self.all_band_label_font_var, width=6).pack(side="right")
        ttk.Button(right, text="Apply band label size to all", command=self._apply_band_label_font_all).pack(fill="x", pady=2)

        ttk.Label(right, text="Tip: You can iterate by going back, then returning here.\nWhen saving, prefer PNG for figures.", justify="left").pack(anchor="w", pady=10)

        ttk.Separator(right).pack(fill="x", pady=8)
        self.edit_mode_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            right,
            text="Interactive text edit (BETA)",
            variable=self.edit_mode_var,
        ).pack(anchor="w", pady=(0, 4))
        ttk.Label(
            right,
            text="Click text in the preview, then edit selected element style.",
            justify="left",
            wraplength=260,
        ).pack(anchor="w")

        self.selected_elem_var = tk.StringVar(value="Selected: none")
        ttk.Label(right, textvariable=self.selected_elem_var, justify="left", wraplength=260).pack(anchor="w", pady=(6, 4))

        self.elem_font_size_var = tk.IntVar(value=14)
        self.elem_font_family_var = tk.StringVar(value="default")
        self.elem_color_var = tk.StringVar(value="#000000")
        self.elem_angle_var = tk.DoubleVar(value=0.0)
        self.elem_dx_var = tk.IntVar(value=0)
        self.elem_dy_var = tk.IntVar(value=0)
        self.elem_width_var = tk.IntVar(value=2)

        er1 = ttk.Frame(right); er1.pack(fill="x", pady=2)
        ttk.Label(er1, text="Font size").pack(side="left")
        ttk.Spinbox(er1, from_=6, to=96, textvariable=self.elem_font_size_var, width=6).pack(side="right")

        er2 = ttk.Frame(right); er2.pack(fill="x", pady=2)
        ttk.Label(er2, text="Font").pack(side="left")
        ttk.Combobox(er2, textvariable=self.elem_font_family_var, values=["default", "DejaVuSans"], state="readonly", width=14).pack(side="right")

        er3 = ttk.Frame(right); er3.pack(fill="x", pady=2)
        ttk.Label(er3, text="Color").pack(side="left")
        ttk.Entry(er3, textvariable=self.elem_color_var, width=12).pack(side="left", padx=4)
        ttk.Button(er3, text="Pick", command=self._choose_elem_color).pack(side="right")

        er4 = ttk.Frame(right); er4.pack(fill="x", pady=2)
        ttk.Label(er4, text="Angle").pack(side="left")
        ttk.Spinbox(er4, from_=-180, to=180, textvariable=self.elem_angle_var, width=6).pack(side="right")

        er5 = ttk.Frame(right); er5.pack(fill="x", pady=2)
        ttk.Label(er5, text="X offset").pack(side="left")
        ttk.Spinbox(er5, from_=-500, to=500, textvariable=self.elem_dx_var, width=6).pack(side="right")

        er6 = ttk.Frame(right); er6.pack(fill="x", pady=2)
        ttk.Label(er6, text="Y offset").pack(side="left")
        ttk.Spinbox(er6, from_=-500, to=500, textvariable=self.elem_dy_var, width=6).pack(side="right")

        er7 = ttk.Frame(right); er7.pack(fill="x", pady=2)
        ttk.Label(er7, text="Line width").pack(side="left")
        ttk.Spinbox(er7, from_=1, to=30, textvariable=self.elem_width_var, width=6).pack(side="right")

        ttk.Button(right, text="Apply to selected", command=self._apply_selected_element_style).pack(fill="x", pady=(6, 2))
        ttk.Button(right, text="Reset selected style", command=self._reset_selected_element_style).pack(fill="x", pady=2)
        ttk.Button(right, text="Clear styles (current panel)", command=self._clear_current_panel_overrides).pack(fill="x", pady=2)

        self._tk_img = None
        self._elements: List[Dict[str, Any]] = []
        self._selected_element: Optional[Dict[str, Any]] = None
        self._panel_render_cache: Dict[str, Image.Image] = {}
        self.viewport.canvas.bind("<Button-1>", self._on_preview_click, add="+")

    def on_show(self):
        if not self.app.state.panels:
            self.app.show_frame("SettingsFrame")
            return
        try:
            idx = int(self.app.state.current_panel_index)
            idx = max(0, min(idx, len(self.app.state.panel_configs) - 1))
            fs = int(getattr(self.app.state.panel_configs[idx], "final_band_label_font_size", 15) or 15)
            self.all_band_label_font_var.set(max(6, fs))
        except Exception:
            self.all_band_label_font_var.set(15)
        self.refresh_preview()

    def _panel_cache_key(self, idx: int) -> str:
        try:
            pc = self.app.state.panel_configs[idx]
            panel = self.app.state.panels[idx]
            pc_hash = hashlib.md5(repr(asdict(pc)).encode()).hexdigest()[:12]
            return f"{idx}:{id(panel)}:{pc_hash}"
        except Exception:
            return f"{idx}:nocache"

    def invalidate_panel_cache(self, idx: Optional[int] = None):
        if idx is None:
            self._panel_render_cache.clear()
        else:
            keys_to_remove = [k for k in self._panel_render_cache if k.startswith(f"{idx}:")]
            for k in keys_to_remove:
                del self._panel_render_cache[k]

    def refresh_preview(self):
        rendered: List[Image.Image] = []
        elements: List[Dict[str, Any]] = []
        spacing = PANEL_SPACING_PX
        y_off = 0
        for idx, (panel, pc) in enumerate(zip(self.app.state.panels, self.app.state.panel_configs)):
            key = self._panel_cache_key(idx)
            img = self._panel_render_cache.get(key)
            if img is None:
                img = render_panel(panel, pc, self.app.state.settings, preview=True)
                self._panel_render_cache[key] = img
            rendered.append(img)
            y_off += int(img.height) + int(spacing)
        # Also collect elements for interactive editing (skip cache for element collection)
        _final_with_elems, elements = render_full_project_with_elements(self.app.state, preview=True)
        self._elements = list(elements)
        if not rendered:
            self.viewport.set_image(Image.new("RGB", (800, 600), (255, 255, 255)), fit_if_needed=True)
            return
        W = max(img.width for img in rendered)
        H = sum(img.height for img in rendered) + spacing * (len(rendered) - 1)
        out = Image.new("RGB", (W, H), (255, 255, 255))
        y = 0
        for img in rendered:
            out.paste(img, (0, y))
            y += img.height + spacing
        self.viewport.set_image(out, fit_if_needed=True)

    def go_back(self):
        step = self.back_step.get()
        if step == "Edit":
            self.app.show_frame("EditFrame")
        elif step == "Panels":
            if self.app.state.multi_source_images:
                self.app.show_frame("ComposeFrame")
            else:
                self.app.show_frame("PanelSelectFrame")
        elif step == "Layout":
            self.app.show_frame("LayoutFrame")
        elif step == "Annotations":
            self.app.show_frame("AnnotationFrame")
        elif step == "Marker":
            self.app.show_frame("MarkerFrame")
        elif step == "Analysis":
            self.app.show_frame("AnalysisFrame")

    def save(self):
        img = render_full_project(self.app.state, preview=False)
        default_name = "annotated.png"
        path = filedialog.asksaveasfilename(
            title="Save annotated image",
            defaultextension=".png",
            initialfile=default_name,
            filetypes=[("PNG", "*.png"), ("JPEG", "*.jpg *.jpeg"), ("TIFF", "*.tif *.tiff")]
        )
        if not path:
            return
        try:
            ext = pathlib.Path(path).suffix.lower()
            if ext in (".jpg", ".jpeg"):
                img.save(path, quality=95)
            else:
                img.save(path)
            self.app.state.last_export_path = path
            messagebox.showinfo("Saved", f"Saved:\n{path}")
        except Exception as exc:
            logger.warning("Export failed: %s", exc)
            messagebox.showerror("Save failed", f"Could not save image:\n{exc}")

    def next_panel(self):
        # Advance to next panel if present; else stay.
        idx = self.app.state.current_panel_index
        if idx + 1 < len(self.app.state.panels):
            self.app.state.current_panel_index += 1
            self.app.show_frame("LayoutFrame")
        else:
            messagebox.showinfo("Panels", "This was the last panel.")

    def next_image(self):
        self.app.reset_for_next_image()
        self.app.show_frame("SettingsFrame")

    def _apply_band_label_font_all(self):
        try:
            fs = max(6, int(self.all_band_label_font_var.get()))
        except Exception:
            fs = 15
            self.all_band_label_font_var.set(15)
        for pc in self.app.state.panel_configs:
            try:
                pc.final_band_label_font_size = int(fs)
            except Exception:
                continue
        self.refresh_preview()

    def _choose_elem_color(self):
        picked = colorchooser.askcolor(color=self.elem_color_var.get(), title="Choose text color")
        if not picked or not picked[1]:
            return
        self.elem_color_var.set(_safe_hex_color(str(picked[1])))

    def _on_preview_click(self, event):
        if not bool(self.edit_mode_var.get()):
            return
        ix, iy = self.viewport.canvas_to_image(event.x, event.y)
        best: Optional[Dict[str, Any]] = None
        best_area = None
        for el in self._elements:
            bb = tuple(el.get("bbox", ()))
            if len(bb) != 4:
                continue
            x0, y0, x1, y1 = [float(v) for v in bb]
            if x0 <= ix <= x1 and y0 <= iy <= y1:
                area = max(1.0, (x1 - x0) * (y1 - y0))
                if best is None or (best_area is not None and area < best_area):
                    best = el
                    best_area = area
        self._selected_element = best
        if best is None:
            self.selected_elem_var.set("Selected: none")
            return
        key = str(best.get("key", ""))
        txt = str(best.get("text", ""))
        self.selected_elem_var.set(f"Selected: {key} ({best.get('etype', '')})\n{txt[:80]}")
        st = dict(best.get("style", {}) or {})
        try:
            self.elem_font_size_var.set(int(st.get("font_size", 14)))
        except Exception:
            pass
        self.elem_font_family_var.set(str(st.get("font_family", "default") or "default"))
        self.elem_color_var.set(_safe_hex_color(str(st.get("color", "#000000"))))
        try:
            self.elem_angle_var.set(float(st.get("angle", 0.0)))
        except Exception:
            pass
        try:
            self.elem_dx_var.set(int(round(float(st.get("dx", 0.0)))))
        except Exception:
            pass
        try:
            self.elem_dy_var.set(int(round(float(st.get("dy", 0.0)))))
        except Exception:
            pass
        try:
            self.elem_width_var.set(max(1, int(st.get("width", 2))))
        except Exception:
            pass

    def _apply_selected_element_style(self):
        el = self._selected_element
        if not el:
            messagebox.showinfo("Review edit", "Select an element in the preview first.")
            return
        try:
            panel_idx = int(el.get("panel_index", -1))
        except Exception:
            panel_idx = -1
        if panel_idx < 0 or panel_idx >= len(self.app.state.panel_configs):
            messagebox.showerror("Review edit", "Invalid selected element panel index.")
            return
        key = str(el.get("key", "")).strip()
        if not key:
            messagebox.showerror("Review edit", "Invalid selected element key.")
            return
        pc = self.app.state.panel_configs[panel_idx]
        ov = dict(getattr(pc, "review_text_overrides", {}) or {})
        fam = str(self.elem_font_family_var.get() or "default").strip()
        if fam.lower() not in {"default", "dejavusans"}:
            fam = "default"
        ov[key] = dict(
            font_size=max(6, int(self.elem_font_size_var.get())),
            font_family=fam,
            color=_safe_hex_color(str(self.elem_color_var.get())),
            angle=float(self.elem_angle_var.get()),
            dx=int(self.elem_dx_var.get()),
            dy=int(self.elem_dy_var.get()),
            width=max(1, int(self.elem_width_var.get())),
        )
        pc.review_text_overrides = ov
        self.refresh_preview()

    def _reset_selected_element_style(self):
        el = self._selected_element
        if not el:
            messagebox.showinfo("Review edit", "Select an element first.")
            return
        try:
            panel_idx = int(el.get("panel_index", -1))
        except Exception:
            panel_idx = -1
        if panel_idx < 0 or panel_idx >= len(self.app.state.panel_configs):
            return
        key = str(el.get("key", "")).strip()
        if not key:
            return
        pc = self.app.state.panel_configs[panel_idx]
        ov = dict(getattr(pc, "review_text_overrides", {}) or {})
        if key in ov:
            ov.pop(key, None)
            pc.review_text_overrides = ov
            self.refresh_preview()

    def _clear_current_panel_overrides(self):
        idx = int(self.app.state.current_panel_index)
        if idx < 0 or idx >= len(self.app.state.panel_configs):
            return
        pc = self.app.state.panel_configs[idx]
        pc.review_text_overrides = {}
        self.refresh_preview()


# ----------------------------
# Rendering (core figure composition)
# ----------------------------

def point_in_bbox(pt: Tuple[int, int], bbox: Tuple[int, int, int, int]) -> bool:
    x, y = pt
    x0, y0, x1, y1 = bbox
    return (x0 <= x <= x1) and (y0 <= y <= y1)

def _wrap_text_lines(draw: ImageDraw.ImageDraw, font: ImageFont.FreeTypeFont, text: str, box_w: int) -> List[str]:
    lines: List[str] = []
    for raw in text.splitlines():
        words = raw.split(" ")
        cur = ""
        for w in words:
            trial = (cur + " " + w).strip()
            if draw.textlength(trial, font=font) > box_w - 20:
                if cur:
                    lines.append(cur)
                cur = w
            else:
                cur = trial
        if cur:
            lines.append(cur)
    return lines


def _format_marker_size(s: float) -> str:
    return f"{int(s) if abs(s - int(s)) < 1e-6 else s:g}"

def marker_base_positions(marker: 'MarkerDefinition', calib: 'MarkerCalibration') -> Dict[float, int]:
    """Return mapping size->y (panel-relative px) before per-panel overrides/hiding."""
    # Use explicit y_positions from marker file if present
    if marker.y_positions and len(marker.y_positions) == len(marker.sizes):
        return {float(s): int(y) for s, y in zip(marker.sizes, marker.y_positions)}

    # Ensure we have a fit if possible
    a = calib.fit_a
    b = calib.fit_b
    if (a is None or b is None) and calib.picked and len(calib.picked) >= 2:
        ys = np.array([p[0] for p in calib.picked], dtype=float)
        sizes = np.array([p[1] for p in calib.picked], dtype=float)
        xs = np.log10(sizes)
        a2, b2 = np.polyfit(xs, ys, 1)
        a, b = float(a2), float(b2)

    if a is not None and b is not None:
        out: Dict[float, int] = {}
        for s in marker.sizes:
            try:
                ypix = a * math.log10(float(s)) + b
                out[float(s)] = int(round(ypix))
            except Exception:
                continue
        return out

    # Fall back to picked points only
    if calib.picked:
        return {float(s): int(y) for (y, s) in calib.picked}

    return {}

def marker_positions_for_panel(
    panel: Image.Image,
    pc: 'PanelConfig',
    settings: 'AppSettings',
    include_hidden: bool = False,
) -> List[Tuple[float, int]]:
    """Return marker tick list (size, y panel px) after overrides; hidden ticks optional."""
    if (not pc.include_marker) or (not pc.marker_calibration.marker_name):
        return []
    try:
        marker = load_marker(pc.marker_calibration.marker_name)
    except Exception:
        return []
    base_map = marker_base_positions(marker, pc.marker_calibration)
    overrides = pc.marker_tick_overrides or {}
    if (not base_map) and (not overrides):
        return []

    hidden = set(pc.marker_tick_hidden or [])
    out: List[Tuple[float, int]] = []
    for s in marker.sizes:
        fs = float(s)
        if (not include_hidden) and (fs in hidden):
            continue
        if fs in overrides:
            y = int(overrides[fs])
        elif fs in base_map:
            y = int(base_map[fs])
        else:
            continue
        out.append((fs, y))
    return out

def get_panel_gel_regions(panel: Image.Image, pc: PanelConfig) -> List[Tuple[int, int]]:
    """Return normalized, sorted gel regions in panel pixel coordinates."""
    W = int(panel.width)
    raw: List[Tuple[int, int]] = []
    has_custom = False

    for reg in list(getattr(pc, "gel_regions", []) or []):
        try:
            a, b = int(reg[0]), int(reg[1])
        except Exception:
            continue
        has_custom = True
        if b < a:
            a, b = b, a
        a = max(0, min(W - 1, a))
        b = max(1, min(W, b))
        if b - a >= 2:
            raw.append((a, b))

    if not raw and (pc.gel_left is not None) and (pc.gel_right is not None):
        has_custom = True
        a = int(pc.gel_left)
        b = int(pc.gel_right)
        if b < a:
            a, b = b, a
        a = max(0, min(W - 1, a))
        b = max(1, min(W, b))
        if b - a >= 2:
            raw.append((a, b))

    if not raw and not has_custom:
        return [(0, W)]
    if not raw:
        raw = [(0, W)]

    raw.sort(key=lambda t: (t[0], t[1]))
    merged: List[Tuple[int, int]] = []
    for a, b in raw:
        if not merged:
            merged.append((a, b))
            continue
        pa, pb = merged[-1]
        if a <= pb:  # merge overlaps
            merged[-1] = (pa, max(pb, b))
        else:
            merged.append((a, b))

    if has_custom:
        pc.gel_regions = [(int(a), int(b)) for (a, b) in merged]
        if pc.gel_regions:
            pc.gel_left = int(pc.gel_regions[0][0])
            pc.gel_right = int(pc.gel_regions[0][1])
        return list(pc.gel_regions)
    return merged


def _allocate_lanes_to_regions(total_lanes: int, widths: List[int]) -> List[int]:
    n = len(widths)
    if n <= 0:
        return []
    lanes = max(1, int(total_lanes))
    widths2 = [max(1, int(w)) for w in widths]

    if lanes < n:
        # Not enough lanes for all regions: assign one lane to the widest regions first.
        order = sorted(range(n), key=lambda i: (-widths2[i], i))
        counts = [0] * n
        for i in order[:lanes]:
            counts[i] = 1
        return counts

    total_w = float(sum(widths2))
    raw = [lanes * (w / total_w) for w in widths2]
    counts = [max(1, int(math.floor(v))) for v in raw]

    while sum(counts) > lanes:
        candidates = [i for i, c in enumerate(counts) if c > 1]
        if not candidates:
            break
        j = max(candidates, key=lambda i: (counts[i], raw[i] - math.floor(raw[i]), widths2[i]))
        counts[j] -= 1

    while sum(counts) < lanes:
        j = max(range(n), key=lambda i: (raw[i] - math.floor(raw[i]), widths2[i], -i))
        counts[j] += 1

    return counts


def compute_lane_bounds(panel: Image.Image, pc: PanelConfig) -> List[Tuple[int, int]]:
    W = int(panel.width)
    lanes = max(1, int(pc.lanes))
    has_custom = bool(pc.gel_regions) or ((pc.gel_left is not None) and (pc.gel_right is not None))
    regions = get_panel_gel_regions(panel, pc)
    widths = [max(1, int(r - l)) for (l, r) in regions]
    counts: List[int] = []
    raw_counts = list(getattr(pc, "gel_region_lane_counts", []) or [])
    if has_custom and len(raw_counts) == len(regions):
        ok = True
        for v in raw_counts:
            try:
                iv = int(v)
            except Exception:
                ok = False
                break
            if iv < 1:
                ok = False
                break
            counts.append(iv)
        if ok:
            lanes = max(1, int(sum(counts)))
    if not counts:
        counts = _allocate_lanes_to_regions(lanes, widths)
    if has_custom:
        pc.gel_region_lane_counts = [int(c) for c in counts]
        pc.lanes = max(1, int(sum(counts)))
        lanes = int(pc.lanes)
    else:
        pc.gel_region_lane_counts = []

    out: List[Tuple[int, int]] = []
    for (left, right), n in zip(regions, counts):
        if n <= 0:
            continue
        rw = max(1, int(right - left))
        for i in range(n):
            x0 = int(round(left + rw * (i / n)))
            x1 = int(round(left + rw * ((i + 1) / n)))
            x0 = max(0, min(W - 1, x0))
            x1 = max(x0 + 1, min(W, x1))
            out.append((x0, x1))

    if not out:
        out = [(0, max(1, W))]

    if len(out) > lanes:
        out = out[:lanes]
    while len(out) < lanes:
        lx0, lx1 = out[-1]
        mid = (lx0 + lx1) // 2
        if mid <= lx0 or mid >= lx1:
            nx0 = max(0, min(W - 1, lx1 - 1))
            nx1 = min(W, nx0 + 1)
            out.append((nx0, nx1))
            continue
        out[-1] = (lx0, mid)
        out.append((mid, lx1))
    return out


def lane_center_x(panel: Image.Image, pc: PanelConfig, lane_idx_1based: int) -> float:
    bounds = compute_lane_bounds(panel, pc)
    lane_idx = max(1, min(len(bounds), int(lane_idx_1based)))
    x0, x1 = bounds[lane_idx - 1]
    return float((x0 + x1) / 2.0)


def lane_x_bounds(panel: Image.Image, pc: PanelConfig, lane_idx_1based: int) -> Tuple[int, int]:
    bounds = compute_lane_bounds(panel, pc)
    lane_idx = max(1, min(len(bounds), int(lane_idx_1based)))
    return bounds[lane_idx - 1]

def _smooth_1d(sig: np.ndarray, window: int) -> np.ndarray:
    w = max(1, int(window))
    if w <= 1:
        return sig.copy()
    if w % 2 == 0:
        w += 1
    kernel = np.ones(w, dtype=float) / float(w)
    return np.convolve(sig, kernel, mode="same")

def _lane_mean_gray_profile(
    panel: Image.Image,
    pc: PanelConfig,
    lane_idx_1based: int,
    gray: Optional[np.ndarray] = None,
    lane_bounds: Optional[List[Tuple[int, int]]] = None,
) -> np.ndarray:
    if gray is None:
        gray = np.asarray(panel.convert("L"), dtype=float)
    if lane_bounds is not None and len(lane_bounds) > 0:
        lane_idx = max(1, min(len(lane_bounds), int(lane_idx_1based)))
        x0, x1 = lane_bounds[lane_idx - 1]
    else:
        x0, x1 = lane_x_bounds(panel, pc, lane_idx_1based)
    W = panel.width
    x0 = max(0, min(W - 1, int(x0)))
    x1 = max(x0 + 1, min(W, int(x1)))
    return gray[:, x0:x1].mean(axis=1).astype(float)

def _rolling_ball_background(signal: np.ndarray, radius: int = 50) -> np.ndarray:
    """Estimate background via 1-D rolling ball (morphological opening with sphere cross-section)."""
    sig = np.asarray(signal, dtype=float)
    n = int(sig.size)
    r = max(3, int(radius))
    # Build the ball profile: a parabolic arc cross-section
    x = np.arange(-r, r + 1, dtype=float)
    ball = r - np.sqrt(np.maximum(0.0, r ** 2 - x ** 2))
    # Erosion then dilation (morphological opening = background estimate)
    bg = np.full(n, np.inf, dtype=float)
    for i in range(n):
        lo = max(0, i - r)
        hi = min(n, i + r + 1)
        blo = lo - (i - r)
        sub = sig[lo:hi] + ball[blo: blo + (hi - lo)]
        bg[i] = float(np.min(sub))
    bg2 = np.full(n, -np.inf, dtype=float)
    for i in range(n):
        lo = max(0, i - r)
        hi = min(n, i + r + 1)
        blo = lo - (i - r)
        sub = bg[lo:hi] - ball[blo: blo + (hi - lo)]
        bg2[i] = float(np.max(sub))
    return np.clip(bg2, 0.0, None)


def _local_adjacent_background(
    panel: Image.Image,
    pc: PanelConfig,
    lane_idx_1based: int,
    gray: Optional[np.ndarray] = None,
    lane_bounds: Optional[List[Tuple[int, int]]] = None,
) -> np.ndarray:
    """Average the two immediately adjacent lane strips as local background."""
    neighbors: List[np.ndarray] = []
    for delta in (-1, 1):
        nbr = lane_idx_1based + delta
        if nbr < 1 or nbr > max(1, int(pc.lanes)):
            continue
        try:
            nbr_sig = _lane_mean_gray_profile(panel, pc, nbr, gray=gray, lane_bounds=lane_bounds)
            neighbors.append(nbr_sig)
        except Exception:
            pass
    if not neighbors:
        return np.zeros(panel.height, dtype=float)
    stacked = np.stack(neighbors, axis=0)
    return np.mean(stacked, axis=0).astype(float)


def lane_intensity_profile(
    panel: Image.Image,
    pc: PanelConfig,
    lane_idx_1based: int,
    half_width: int = 2,
    polarity: str = "dark",
    smooth_window: int = 10,
    background_signal: Optional[np.ndarray] = None,
    gray: Optional[np.ndarray] = None,
    lane_bounds: Optional[List[Tuple[int, int]]] = None,
    bg_mode: str = "lane",
    rolling_ball_radius: int = 50,
) -> np.ndarray:
    # Use the full lane slice width instead of only a center line.
    _ = half_width  # legacy compatibility; lane width is derived from lane bounds
    col = _lane_mean_gray_profile(panel, pc, lane_idx_1based, gray=gray, lane_bounds=lane_bounds)
    sig = (255.0 - col) if str(polarity).lower() == "dark" else col
    if str(bg_mode).strip() == "rolling_ball":
        try:
            raw_col = _lane_mean_gray_profile(panel, pc, lane_idx_1based, gray=gray, lane_bounds=lane_bounds)
            raw_sig = (255.0 - raw_col) if str(polarity).lower() == "dark" else raw_col
            ball_bg = _rolling_ball_background(raw_sig, radius=int(rolling_ball_radius))
            sig = np.maximum(raw_sig - ball_bg, 0.0)
        except Exception:
            pass
    elif str(bg_mode).strip() == "local_adjacent":
        try:
            nbr_bg = _local_adjacent_background(panel, pc, lane_idx_1based, gray=gray, lane_bounds=lane_bounds)
            nbr_sig = (255.0 - nbr_bg) if str(polarity).lower() == "dark" else nbr_bg
            sig = np.maximum(sig - nbr_sig, 0.0)
        except Exception:
            pass
    elif background_signal is not None:
        # Existing lane-subtraction logic
        try:
            bg = np.asarray(background_signal, dtype=float)
            if bg.shape == sig.shape:
                sig = sig - bg
            else:
                n = min(int(sig.size), int(bg.size))
                if n > 0:
                    sig2 = sig.copy()
                    sig2[:n] = sig2[:n] - bg[:n]
                    sig = sig2
            sig = np.maximum(sig, 0.0)
        except Exception:
            pass
    sig = _smooth_1d(sig.astype(float), int(smooth_window))
    return sig.astype(float)

def detect_bands_from_profile(
    signal: np.ndarray,
    threshold: float = 20.0,
    prominence: float = 7.0,
    min_distance: int = 10,
    prom_window: int = 20,
) -> List[int]:
    sig = np.asarray(signal, dtype=float)
    n = int(sig.size)
    if n < 3:
        return []
    th = float(threshold)
    prom_th = float(prominence)
    win = max(3, int(prom_window))
    cand: List[Tuple[int, float]] = []
    for i in range(1, n - 1):
        if sig[i] < th:
            continue
        if not (sig[i] > sig[i - 1] and sig[i] >= sig[i + 1]):
            continue
        l0 = max(0, i - win)
        r0 = min(n - 1, i + win)
        left_min = float(np.min(sig[l0 : i + 1]))
        right_min = float(np.min(sig[i : r0 + 1]))
        base = max(left_min, right_min)
        prom = float(sig[i] - base)
        if prom >= prom_th:
            cand.append((i, float(sig[i])))
    if not cand:
        return []

    # Non-maximum suppression by minimum distance, strongest peaks first.
    d = max(1, int(min_distance))
    chosen: List[int] = []
    for idx, _val in sorted(cand, key=lambda t: t[1], reverse=True):
        if all(abs(idx - c) >= d for c in chosen):
            chosen.append(int(idx))
    chosen.sort()
    return chosen

def _pchip_slopes(x: np.ndarray, y: np.ndarray) -> np.ndarray:
    """Shape-preserving slopes for monotone cubic Hermite interpolation."""
    n = int(len(x))
    if n <= 1:
        return np.zeros(n, dtype=float)
    h = np.diff(x)
    d = np.diff(y) / np.maximum(h, 1e-12)
    if n == 2:
        return np.array([d[0], d[0]], dtype=float)
    m = np.zeros(n, dtype=float)
    for k in range(1, n - 1):
        dk1 = float(d[k - 1])
        dk = float(d[k])
        if (dk1 == 0.0) or (dk == 0.0) or ((dk1 > 0) != (dk > 0)):
            m[k] = 0.0
            continue
        w1 = 2.0 * h[k] + h[k - 1]
        w2 = h[k] + 2.0 * h[k - 1]
        m[k] = (w1 + w2) / ((w1 / dk1) + (w2 / dk))

    # Endpoints (PCHIP one-sided estimates with monotonicity limiting)
    d0 = float(d[0]); d1 = float(d[1]); h0 = float(h[0]); h1 = float(h[1])
    m0 = ((2.0 * h0 + h1) * d0 - h0 * d1) / max(h0 + h1, 1e-12)
    if (m0 > 0) != (d0 > 0):
        m0 = 0.0
    elif ((d0 > 0) != (d1 > 0)) and (abs(m0) > abs(3.0 * d0)):
        m0 = 3.0 * d0
    m[0] = m0

    dn1 = float(d[-1]); dn2 = float(d[-2]); hn1 = float(h[-1]); hn2 = float(h[-2])
    mn = ((2.0 * hn1 + hn2) * dn1 - hn1 * dn2) / max(hn1 + hn2, 1e-12)
    if (mn > 0) != (dn1 > 0):
        mn = 0.0
    elif ((dn1 > 0) != (dn2 > 0)) and (abs(mn) > abs(3.0 * dn1)):
        mn = 3.0 * dn1
    m[-1] = mn
    return m


def _interp_monotone_eval(
    x_nodes: List[float],
    y_nodes: List[float],
    xq: float,
    method: str = "pchip",
    slopes: Optional[List[float]] = None,
) -> Optional[float]:
    try:
        xv = float(xq)
    except Exception:
        return None
    x = np.asarray(x_nodes, dtype=float)
    y = np.asarray(y_nodes, dtype=float)
    n = int(len(x))
    if n == 0:
        return None
    if n == 1:
        return float(y[0]) if np.isfinite(y[0]) else None
    if not np.isfinite(xv):
        return None

    # Extrapolate linearly using endpoint slopes.
    if xv <= x[0]:
        if method == "pchip" and slopes is not None and len(slopes) == n:
            return float(y[0] + (xv - x[0]) * float(slopes[0]))
        slope0 = float((y[1] - y[0]) / max(x[1] - x[0], 1e-12))
        return float(y[0] + (xv - x[0]) * slope0)
    if xv >= x[-1]:
        if method == "pchip" and slopes is not None and len(slopes) == n:
            return float(y[-1] + (xv - x[-1]) * float(slopes[-1]))
        slope1 = float((y[-1] - y[-2]) / max(x[-1] - x[-2], 1e-12))
        return float(y[-1] + (xv - x[-1]) * slope1)

    i = int(np.searchsorted(x, xv, side="right") - 1)
    i = max(0, min(n - 2, i))
    x0 = float(x[i]); x1 = float(x[i + 1])
    y0 = float(y[i]); y1 = float(y[i + 1])
    h = x1 - x0
    if h <= 1e-12:
        return float(y0)

    if method != "pchip" or slopes is None or len(slopes) != n:
        t = (xv - x0) / h
        return float(y0 + t * (y1 - y0))

    m = np.asarray(slopes, dtype=float)
    t = (xv - x0) / h
    t2 = t * t
    t3 = t2 * t
    h00 = 2.0 * t3 - 3.0 * t2 + 1.0
    h10 = t3 - 2.0 * t2 + t
    h01 = -2.0 * t3 + 3.0 * t2
    h11 = t3 - t2
    yv = h00 * y0 + h10 * h * float(m[i]) + h01 * y1 + h11 * h * float(m[i + 1])
    return float(yv)


def _prep_unique_nodes(x_raw: List[float], y_raw: List[float]) -> Tuple[List[float], List[float]]:
    """Sort by x and merge duplicate x values (average y)."""
    pairs: List[Tuple[float, float]] = []
    for xa, ya in zip(list(x_raw or []), list(y_raw or [])):
        try:
            xv = float(xa); yv = float(ya)
        except Exception:
            continue
        if np.isfinite(xv) and np.isfinite(yv):
            pairs.append((xv, yv))
    if not pairs:
        return [], []
    pairs.sort(key=lambda t: (t[0], t[1]))
    xs: List[float] = []
    ys: List[float] = []
    cur_x = pairs[0][0]
    buf: List[float] = [pairs[0][1]]
    for xv, yv in pairs[1:]:
        if abs(xv - cur_x) <= 1e-9:
            buf.append(yv)
            continue
        xs.append(float(cur_x))
        ys.append(float(sum(buf) / max(1, len(buf))))
        cur_x = xv
        buf = [yv]
    xs.append(float(cur_x))
    ys.append(float(sum(buf) / max(1, len(buf))))
    return xs, ys


def _fit_eval_x_from_y(y_px: float, fit: Optional[Dict[str, Any]]) -> Optional[float]:
    if not fit:
        return None
    try:
        yv = float(y_px)
    except Exception:
        return None
    fit_kind = str(fit.get("fit_kind", "") or "")
    if fit_kind == "poly_x_from_y":
        try:
            coeffs = [float(c) for c in list(fit.get("coeffs_y_to_x", []) or [])]
            deg = int(fit.get("degree", max(0, len(coeffs) - 1)))
            if coeffs and len(coeffs) == deg + 1:
                xv = float(np.polyval(np.asarray(coeffs, dtype=float), yv))
                if np.isfinite(xv):
                    return xv
        except Exception:
            return None
        return None
    method = str(fit.get("interp_method", "") or "")
    y_nodes = [float(v) for v in list(fit.get("interp_y_nodes", []) or [])]
    x_nodes = [float(v) for v in list(fit.get("interp_x_nodes", []) or [])]
    slopes = [float(v) for v in list(fit.get("interp_dxdy_slopes", []) or [])] if fit.get("interp_dxdy_slopes") else None
    if y_nodes and x_nodes and len(y_nodes) == len(x_nodes):
        val = _interp_monotone_eval(y_nodes, x_nodes, yv, method=method or "linear", slopes=slopes)
        if val is not None and np.isfinite(val):
            return float(val)
    return None


def _fit_eval_y_from_size(size_val: float, fit: Optional[Dict[str, Any]]) -> Optional[float]:
    if not fit:
        return None
    try:
        s = float(size_val)
    except Exception:
        return None
    if (not np.isfinite(s)) or s <= 0:
        return None
    xq = float(math.log10(max(1e-12, s)))
    fit_kind = str(fit.get("fit_kind", "") or "")
    if fit_kind == "poly_x_from_y":
        try:
            coeffs = [float(c) for c in list(fit.get("coeffs_y_to_x", []) or [])]
            deg = int(fit.get("degree", max(0, len(coeffs) - 1)))
        except Exception:
            coeffs = []
            deg = 0
        if coeffs and len(coeffs) == deg + 1:
            try:
                if deg == 1:
                    a = float(coeffs[0])
                    b = float(coeffs[1])
                    if abs(a) < 1e-12:
                        return None
                    yv = float((xq - b) / a)
                    return yv if np.isfinite(yv) else None
                if deg == 2:
                    # Fast closed-form inverse of x(y)=c2*y^2 + c1*y + c0.
                    c2 = float(coeffs[0]); c1 = float(coeffs[1]); c0 = float(coeffs[2])
                    if abs(c2) < 1e-12:
                        if abs(c1) < 1e-12:
                            return None
                        yv = float((xq - c0) / c1)
                        return yv if np.isfinite(yv) else None
                    disc = float(c1 * c1 - 4.0 * c2 * (c0 - xq))
                    if disc < 0.0:
                        return None
                    root = math.sqrt(max(0.0, disc))
                    den = 2.0 * c2
                    r1 = float((-c1 + root) / den)
                    r2 = float((-c1 - root) / den)
                    candidates = [r for r in (r1, r2) if np.isfinite(r)]
                    if not candidates:
                        return None
                    y_min = float(fit.get("y_min", float("nan")))
                    y_max = float(fit.get("y_max", float("nan")))
                    if np.isfinite(y_min) and np.isfinite(y_max):
                        lo = min(y_min, y_max)
                        hi = max(y_min, y_max)
                        in_range = [r for r in candidates if (lo - 1e-6) <= r <= (hi + 1e-6)]
                        if in_range:
                            mid = 0.5 * (lo + hi)
                            return float(min(in_range, key=lambda rr: abs(rr - mid)))
                        return float(min(candidates, key=lambda rr: min(abs(rr - lo), abs(rr - hi))))
                    return float(candidates[0])
                poly = np.asarray(coeffs, dtype=float).copy()
                poly[-1] -= xq
                roots = np.roots(poly)
                real_roots = [float(r.real) for r in roots if abs(float(r.imag)) < 1e-7 and np.isfinite(float(r.real))]
                if not real_roots:
                    return None
                y_min = float(fit.get("y_min", float("nan")))
                y_max = float(fit.get("y_max", float("nan")))
                if np.isfinite(y_min) and np.isfinite(y_max):
                    lo = min(y_min, y_max)
                    hi = max(y_min, y_max)
                    in_range = [r for r in real_roots if (lo - 1e-6) <= r <= (hi + 1e-6)]
                    if in_range:
                        y_sol = min(in_range, key=lambda rr: abs(rr - ((lo + hi) / 2.0)))
                    else:
                        y_sol = min(real_roots, key=lambda rr: min(abs(rr - lo), abs(rr - hi)))
                else:
                    y_sol = real_roots[0]
                return y_sol if np.isfinite(y_sol) else None
            except Exception:
                return None
    # Preferred path: direct monotonic interpolation y = g(log10(size)).
    method = str(fit.get("interp_method", "") or "")
    x_nodes = [float(v) for v in list(fit.get("interp_x_to_y_nodes_x", []) or [])]
    y_nodes = [float(v) for v in list(fit.get("interp_x_to_y_nodes_y", []) or [])]
    slopes = [float(v) for v in list(fit.get("interp_dydx_slopes", []) or [])] if fit.get("interp_dydx_slopes") else None
    if x_nodes and y_nodes and len(x_nodes) == len(y_nodes):
        yv = _interp_monotone_eval(x_nodes, y_nodes, xq, method=method or "linear", slopes=slopes)
        if yv is not None and np.isfinite(yv):
            return float(yv)
    # Legacy polynomial fallback.
    try:
        coeffs = [float(c) for c in list(fit.get("coeffs", []) or [])]
        deg = int(fit.get("degree", max(0, len(coeffs) - 1)))
        if coeffs and len(coeffs) == deg + 1:
            yv = float(np.polyval(np.asarray(coeffs, dtype=float), xq))
            if np.isfinite(yv):
                return yv
    except Exception:
        pass
    try:
        a = float(fit.get("a", 0.0))
        b = float(fit.get("b", 0.0))
        if abs(a) < 1e-12:
            return None
        yv = float(a * xq + b)
        if np.isfinite(yv):
            return yv
    except Exception:
        return None
    return None


def _fit_equation_text(fit: Optional[Dict[str, Any]]) -> str:
    if not fit:
        return "No marker fit available (need at least 2 marker points)."
    fit_kind = str(fit.get("fit_kind", ""))
    if fit_kind == "monotone_interp":
        unit = str(fit.get("unit", ""))
        method = str(fit.get("interp_method", "linear"))
        n_pts = int(fit.get("n_points", 0) or 0)
        return f"Calibration: monotonic {method.upper()} interpolation of log10(size) vs y (n={n_pts}) [{unit}]"
    if fit_kind == "poly_x_from_y":
        unit = str(fit.get("unit", ""))
        coeffs = [float(c) for c in list(fit.get("coeffs_y_to_x", []) or [])]
        deg = int(fit.get("degree", max(0, len(coeffs) - 1)))
        r2 = float(fit.get("r2", float("nan")))
        rmse = float(fit.get("rmse", float("nan")))
        if deg == 1 and len(coeffs) == 2:
            a, b = float(coeffs[0]), float(coeffs[1])
            return (
                f"log10(size) = {a:.8f}*y + {b:.6f};  size = 10^({a:.8f}*y + {b:.6f})   "
                f"[{unit}], R^2={r2:.4f}, RMSE(log10)={rmse:.4f}"
            )
        if deg == 2 and len(coeffs) == 3:
            a2, a1, a0 = float(coeffs[0]), float(coeffs[1]), float(coeffs[2])
            return (
                f"log10(size) = {a2:.10f}*y^2 + {a1:.8f}*y + {a0:.6f}   "
                f"[{unit}], R^2={r2:.4f}, RMSE(log10)={rmse:.4f}"
            )
        if coeffs:
            terms: List[str] = []
            p = deg
            for c in coeffs:
                if p > 1:
                    terms.append(f"{c:.8g}*y^{p}")
                elif p == 1:
                    terms.append(f"{c:.8g}*y")
                else:
                    terms.append(f"{c:.8g}")
                p -= 1
            return f"log10(size) = {' + '.join(terms)}   [{unit}], R^2={r2:.4f}, RMSE(log10)={rmse:.4f}"
    coeffs = [float(c) for c in list(fit.get("coeffs", []) or [])]
    deg = int(fit.get("degree", max(0, len(coeffs) - 1)))
    unit = str(fit.get("unit", ""))
    r2 = float(fit.get("r2", float("nan")))
    if deg == 1 and len(coeffs) == 2:
        a, b = float(coeffs[0]), float(coeffs[1])
        return f"y = {a:.6f}*log10(size) + {b:.3f};  size = 10^((y-{b:.3f})/{a:.6f})   [{unit}], R^2={r2:.4f}"
    if deg == 2 and len(coeffs) == 3:
        c2, c1, c0 = float(coeffs[0]), float(coeffs[1]), float(coeffs[2])
        return (
            f"y = {c2:.6f}*log10(size)^2 + {c1:.6f}*log10(size) + {c0:.3f}   "
            f"[{unit}], R^2={r2:.4f}; size solved numerically"
        )
    if coeffs:
        terms: List[str] = []
        p = deg
        for c in coeffs:
            if p > 1:
                terms.append(f"{c:.6f}*log10(size)^{p}")
            elif p == 1:
                terms.append(f"{c:.6f}*log10(size)")
            else:
                terms.append(f"{c:.3f}")
            p -= 1
        return f"y = {' + '.join(terms)}   [{unit}], R^2={r2:.4f}; size solved numerically"
    return "Marker fit available, but equation could not be formatted."


def fit_marker_curve_for_panel(
    panel: Image.Image,
    pc: PanelConfig,
    settings: AppSettings,
    marker_positions_hidden: Optional[List[Tuple[float, float]]] = None,
) -> Optional[Dict[str, Any]]:
    if not pc.marker_calibration.marker_name:
        return None
    try:
        marker = load_marker(pc.marker_calibration.marker_name)
    except Exception:
        return None
    if marker_positions_hidden is None:
        pos = marker_positions_for_panel(panel, pc, settings, include_hidden=True)
    else:
        pos = list(marker_positions_hidden)
    if len(pos) < 2:
        return None
    sizes: List[float] = []
    ys: List[float] = []
    for s, y in pos:
        fs = float(s)
        if fs <= 0:
            continue
        sizes.append(fs)
        ys.append(float(y) + float(pc.marker_y_offset))
    if len(sizes) < 2:
        return None

    fit_mode = str(getattr(pc, "analysis_fit_mode", "monotone_interp") or "monotone_interp").strip().lower()
    if fit_mode not in {"monotone_interp", "log_linear", "log_quadratic", "log_cubic"}:
        fit_mode = "monotone_interp"

    # Regression options often used in gel calibration: semilog linear/quadratic/cubic.
    # We fit in the prediction direction used later in analysis:
    #   x = log10(size) = f(y)
    if fit_mode in {"log_linear", "log_quadratic", "log_cubic"}:
        y_arr = np.asarray([float(v) for v in ys], dtype=float)
        x_arr = np.asarray([float(math.log10(float(s))) for s in sizes], dtype=float)
        ok = np.isfinite(y_arr) & np.isfinite(x_arr)
        y_arr = y_arr[ok]
        x_arr = x_arr[ok]
        if y_arr.size >= 2:
            req_deg = 1 if fit_mode == "log_linear" else (2 if fit_mode == "log_quadratic" else 3)
            # Degenerate y duplicates can reduce effective degree; degrade gracefully.
            n_unique_y = int(len(np.unique(np.round(y_arr.astype(float), 6))))
            deg = max(1, min(req_deg, int(y_arr.size) - 1, max(1, n_unique_y - 1)))
            try:
                # Improve numerical conditioning and fit balance:
                # - center/scale y before fitting (quadratic is sensitive on large pixel ranges)
                # - weight by local spacing in y so densely packed markers do not dominate
                y_sort_idx = np.argsort(y_arr)
                y_sorted = y_arr[y_sort_idx]
                x_sorted = x_arr[y_sort_idx]
                if y_sorted.size >= 2:
                    y_d = np.diff(y_sorted)
                    y_d = np.where(np.isfinite(y_d), np.abs(y_d), 0.0)
                    w_sorted = np.zeros_like(y_sorted, dtype=float)
                    if y_sorted.size == 2:
                        w_sorted[:] = max(1e-6, float(abs(y_sorted[1] - y_sorted[0])))
                    else:
                        w_sorted[0] = max(1e-6, float(y_d[0]))
                        w_sorted[-1] = max(1e-6, float(y_d[-1]))
                        for k in range(1, y_sorted.size - 1):
                            w_sorted[k] = max(1e-6, float(0.5 * (y_d[k - 1] + y_d[k])))
                    weights = np.empty_like(w_sorted)
                    weights[y_sort_idx] = w_sorted
                else:
                    weights = np.ones_like(y_arr, dtype=float)

                # Bias regression toward the high-MW end (large sizes / larger log10(size)),
                # where users care most about fit quality and marker points are often sparse.
                x_lo = float(np.min(x_arr))
                x_hi = float(np.max(x_arr))
                if (x_hi - x_lo) > 1e-9:
                    t_hi = np.clip((x_arr - x_lo) / (x_hi - x_lo), 0.0, 1.0)
                else:
                    t_hi = np.zeros_like(x_arr, dtype=float)
                hi_bias_strength = 2.0 if fit_mode == "log_linear" else (3.5 if fit_mode == "log_quadratic" else 5.0)
                hi_bias = 1.0 + hi_bias_strength * np.power(t_hi, 2.0)
                weights = np.asarray(np.maximum(weights, 1e-9), dtype=float) * hi_bias

                # Extra anchor on the largest marker(s) so the regression reaches the top end better.
                # This preserves smoothness while reducing drift above ~5 kDa / 5 kb.
                try:
                    top_idx = np.argsort(x_arr)[-max(1, min(2, int(x_arr.size))):]
                    if top_idx.size >= 1:
                        weights[top_idx[-1]] *= (2.5 if fit_mode == "log_linear" else (3.0 if fit_mode == "log_quadratic" else 4.0))
                    if top_idx.size >= 2:
                        weights[top_idx[-2]] *= (1.5 if fit_mode == "log_linear" else (2.0 if fit_mode == "log_quadratic" else 2.5))
                except Exception:
                    pass

                y_center = float(np.mean(y_arr))
                y_scale = float(np.std(y_arr))
                if (not np.isfinite(y_scale)) or y_scale < 1e-9:
                    y_scale = float(max(1.0, np.max(y_arr) - np.min(y_arr)))
                if y_scale < 1e-9:
                    y_scale = 1.0
                z_arr = (y_arr - y_center) / y_scale

                # numpy.polyfit's w applies to unsquared residuals; use sqrt(weight) for LS weighting.
                coeffs_z = np.polyfit(z_arr, x_arr, deg, w=np.sqrt(np.maximum(weights, 1e-9)))
                # Convert back to x(y) polynomial so downstream eval/inversion code remains unchanged.
                pz = np.poly1d(np.asarray(coeffs_z, dtype=float))
                z_of_y = np.poly1d([1.0 / y_scale, -y_center / y_scale])
                coeffs_y_to_x = np.asarray((pz(z_of_y)).c, dtype=float)
                x_hat = np.polyval(coeffs_y_to_x, y_arr)
                rss = float(np.sum((x_arr - x_hat) ** 2))
                ss_tot = float(np.sum((x_arr - np.mean(x_arr)) ** 2))
                r2 = 1.0 - rss / ss_tot if ss_tot > 1e-12 else 1.0
                rmse = float(math.sqrt(max(rss, 0.0) / max(1, int(y_arr.size))))
                return dict(
                    fit_kind="poly_x_from_y",
                    degree=int(deg),
                    coeffs_y_to_x=[float(c) for c in np.asarray(coeffs_y_to_x, dtype=float)],
                    r2=float(r2),
                    rmse=float(rmse),
                    unit=str(marker.unit),
                    marker_name=str(marker.name),
                    n_points=int(y_arr.size),
                    x_min=float(np.min(x_arr)),
                    x_max=float(np.max(x_arr)),
                    y_min=float(np.min(y_arr)),
                    y_max=float(np.max(y_arr)),
                    requested_mode=str(fit_mode),
                    fit_weighting="y_spacing",
                )
            except Exception:
                # Fallback to interpolation below.
                pass

    # Direct calibration used for prediction: x=log10(size) as a monotonic function of y.
    x_raw = [float(math.log10(float(s))) for s in sizes]
    y_raw = [float(v) for v in ys]

    y_nodes, x_nodes = _prep_unique_nodes(y_raw, x_raw)
    if len(y_nodes) < 2:
        return None

    # Build reverse mapping only for plotting y(size) in the overlay.
    x_to_y_nodes_x, x_to_y_nodes_y = _prep_unique_nodes(x_nodes, y_nodes)
    if len(x_to_y_nodes_x) < 2:
        return None

    interp_method = "pchip" if (len(y_nodes) >= 3 and len(x_to_y_nodes_x) >= 3) else "linear"
    dxdy_slopes: Optional[List[float]] = None
    dydx_slopes: Optional[List[float]] = None
    if interp_method == "pchip":
        try:
            dxdy_slopes = [float(v) for v in _pchip_slopes(np.asarray(y_nodes, dtype=float), np.asarray(x_nodes, dtype=float))]
            dydx_slopes = [float(v) for v in _pchip_slopes(np.asarray(x_to_y_nodes_x, dtype=float), np.asarray(x_to_y_nodes_y, dtype=float))]
        except Exception:
            interp_method = "linear"
            dxdy_slopes = None
            dydx_slopes = None

    # Quality metric in the prediction direction (x from y). Interpolation is exact on nodes after de-duplication.
    x_pred = []
    for yv in y_nodes:
        xv = _interp_monotone_eval(y_nodes, x_nodes, float(yv), method=interp_method, slopes=dxdy_slopes)
        x_pred.append(float(xv) if xv is not None else float("nan"))
    x_arr = np.asarray(x_nodes, dtype=float)
    xhat = np.asarray(x_pred, dtype=float)
    ok = np.isfinite(x_arr) & np.isfinite(xhat)
    if np.any(ok):
        rss = float(np.sum((x_arr[ok] - xhat[ok]) ** 2))
        ss_tot = float(np.sum((x_arr[ok] - np.mean(x_arr[ok])) ** 2))
        r2 = 1.0 - rss / ss_tot if ss_tot > 1e-12 else 1.0
        rmse = float(math.sqrt(max(rss, 0.0) / max(1, int(np.sum(ok)))))
    else:
        r2 = float("nan")
        rmse = float("nan")

    out: Dict[str, Any] = dict(
        fit_kind="monotone_interp",
        interp_method=str(interp_method),
        interp_y_nodes=[float(v) for v in y_nodes],          # independent y
        interp_x_nodes=[float(v) for v in x_nodes],          # dependent log10(size)
        interp_dxdy_slopes=([float(v) for v in dxdy_slopes] if dxdy_slopes is not None else []),
        interp_x_to_y_nodes_x=[float(v) for v in x_to_y_nodes_x],  # independent log10(size)
        interp_x_to_y_nodes_y=[float(v) for v in x_to_y_nodes_y],  # dependent y
        interp_dydx_slopes=([float(v) for v in dydx_slopes] if dydx_slopes is not None else []),
        r2=float(r2),
        rmse=float(rmse),
        unit=str(marker.unit),
        marker_name=str(marker.name),
        n_points=int(len(y_nodes)),
        x_min=float(min(x_to_y_nodes_x)),
        x_max=float(max(x_to_y_nodes_x)),
    )
    return out

def estimate_size_from_y(y_px: float, fit: Optional[Dict[str, Any]]) -> Optional[float]:
    if not fit:
        return None
    try:
        yv = float(y_px)
    except Exception:
        return None
    # Preferred path: direct prediction function x=log10(size) from y.
    x_from_y = _fit_eval_x_from_y(yv, fit)
    if x_from_y is not None and np.isfinite(x_from_y):
        try:
            out = float(10 ** float(x_from_y))
            return out if (np.isfinite(out) and out > 0) else None
        except Exception:
            return None

    coeffs = [float(c) for c in list(fit.get("coeffs", []) or [])]
    deg = int(fit.get("degree", max(0, len(coeffs) - 1)))

    if coeffs and len(coeffs) == deg + 1:
        if deg == 1:
            a = float(coeffs[0])
            b = float(coeffs[1])
            if abs(a) < 1e-9:
                return None
            try:
                out = float(10 ** ((yv - b) / a))
                return out if (np.isfinite(out) and out > 0) else None
            except Exception:
                return None
        # Solve poly(log10(size)) - y = 0 numerically for log10(size).
        try:
            poly = np.asarray(coeffs, dtype=float).copy()
            poly[-1] -= yv
            roots = np.roots(poly)
            real_roots = [float(r.real) for r in roots if abs(float(r.imag)) < 1e-7 and np.isfinite(float(r.real))]
            if not real_roots:
                return None
            x_min = float(fit.get("x_min", float("nan")))
            x_max = float(fit.get("x_max", float("nan")))
            if np.isfinite(x_min) and np.isfinite(x_max):
                lo = min(x_min, x_max)
                hi = max(x_min, x_max)
                in_range = [r for r in real_roots if (lo - 1e-6) <= r <= (hi + 1e-6)]
                if in_range:
                    x_sol = min(in_range, key=lambda rr: abs(rr - ((lo + hi) / 2.0)))
                else:
                    # Pick root closest to calibrated domain.
                    x_sol = min(real_roots, key=lambda rr: min(abs(rr - lo), abs(rr - hi)))
            else:
                x_sol = real_roots[0]
            out = float(10 ** float(x_sol))
            return out if (np.isfinite(out) and out > 0) else None
        except Exception:
            return None

    # Legacy fallback
    a = float(fit.get("a", 0.0))
    b = float(fit.get("b", 0.0))
    if abs(a) < 1e-9:
        return None
    try:
        out = float(10 ** ((yv - b) / a))
        return out if (np.isfinite(out) and out > 0) else None
    except Exception:
        return None

def format_size_pretty(size_val: Optional[float], unit: str = "") -> str:
    if size_val is None:
        return "n/a"
    try:
        v = float(size_val)
    except Exception:
        return "n/a"
    if not np.isfinite(v) or v <= 0:
        return "n/a"
    if v >= 1000:
        s = f"{int(round(v)):,}".replace(",", ".")
    elif v >= 100:
        s = f"{v:.1f}".rstrip("0").rstrip(".")
    elif v >= 10:
        s = f"{v:.2f}".rstrip("0").rstrip(".")
    else:
        s = f"{v:.3f}".rstrip("0").rstrip(".")
    unit = str(unit).strip()
    return f"{s} {unit}".strip()


def _parse_size_value(raw_val: Any) -> Optional[float]:
    if raw_val is None:
        return None
    if isinstance(raw_val, (int, float)):
        try:
            v = float(raw_val)
        except Exception:
            return None
        if np.isfinite(v) and v > 0:
            return v
        return None
    s = str(raw_val).strip()
    if not s:
        return None
    s = s.replace("kDa", "").replace("bp", "").strip()
    s = s.replace(" ", "")
    # Handle thousands separators formatted as 1.429 (for 1429).
    if re.fullmatch(r"\d{1,3}(?:\.\d{3})+", s):
        s = s.replace(".", "")
    else:
        s = s.replace(",", ".")
    try:
        v = float(s)
    except Exception:
        return None
    if np.isfinite(v) and v > 0:
        return v
    return None


def final_band_size_labels_for_panel(pc: PanelConfig) -> Dict[int, List[str]]:
    """Lane -> stacked label lines for optional final-image size labels."""
    if not bool(getattr(pc, "show_band_sizes_on_final", True)):
        return {}

    excluded: set[int] = set()
    for v in list(getattr(pc, "band_size_excluded_lanes", []) or []):
        try:
            excluded.add(int(v))
        except Exception:
            continue

    include_raw = getattr(pc, "band_size_included_bands_by_lane", {}) or {}
    raw_vals_any = getattr(pc, "analysis_band_values_by_lane", {}) or {}
    raw_txt = getattr(pc, "analysis_band_sizes_by_lane", {}) or {}
    out: Dict[int, List[str]] = {}

    lanes: set[int] = set()
    if isinstance(raw_vals_any, dict):
        for k in raw_vals_any.keys():
            try:
                lanes.add(int(k))
            except Exception:
                continue
    if isinstance(raw_txt, dict):
        for k in raw_txt.keys():
            try:
                lanes.add(int(k))
            except Exception:
                continue

    for lane in sorted(lanes):
        if lane in excluded:
            continue

        vals_any = raw_vals_any.get(lane, raw_vals_any.get(str(lane), None))
        vals_txt = raw_txt.get(lane, raw_txt.get(str(lane), None))
        if vals_any is None:
            vals_any = list(vals_txt or [])
        vals = list(vals_any or [])
        if not vals:
            continue

        include_list = include_raw.get(lane, include_raw.get(str(lane), list(range(1, len(vals) + 1))))
        include_set = set()
        for bi in list(include_list or []):
            try:
                idx = int(bi)
            except Exception:
                continue
            if 1 <= idx <= len(vals):
                include_set.add(idx)
        if not include_set:
            continue

        picked: List[float] = []
        for idx, raw_v in enumerate(vals, start=1):
            if idx not in include_set:
                continue
            v = _parse_size_value(raw_v)
            if v is not None:
                picked.append(float(v))
        if not picked:
            continue
        # Top bands first; for standard gels this corresponds to larger sizes first.
        picked.sort(reverse=True)
        # "next full number" = round up to integer.
        lines = [str(int(math.ceil(v))) for v in picked]
        out[int(lane)] = lines
    return out

def render_lane_histogram(
    signal: np.ndarray,
    peaks: List[int],
    threshold: float,
    lane_slice: Optional[Image.Image] = None,
    title: str = "Lane intensity",
    lane_meta_lines: Optional[List[str]] = None,
    peak_value_labels: Optional[Dict[int, str]] = None,
    size_unit: str = "",
    x_scale_max: Optional[float] = None,
) -> Image.Image:
    sig = np.asarray(signal, dtype=float)
    n = int(sig.size)
    W, H = 1240, 540
    mt, mb = 40, 92
    ph = max(10, H - mt - mb)
    lane_w = 140 if lane_slice is not None else 0
    left_pad = 18
    gap = 14 if lane_slice is not None else 0
    y_axis_w = 72
    ann_w = 250
    x0 = left_pad + lane_w + gap + y_axis_w
    x1 = W - ann_w - 18
    pw = max(10, x1 - x0)
    out = Image.new("RGB", (W, H), (255, 255, 255))
    draw = ImageDraw.Draw(out)

    # Optional lane slice (scaled to match Y extent so it aligns with profile)
    if lane_slice is not None:
        try:
            lane_img = lane_slice.resize((lane_w, ph), Image.Resampling.BICUBIC)
            out.paste(lane_img, (left_pad, mt))
            draw.rectangle((left_pad, mt, left_pad + lane_w, mt + ph), outline=(0, 0, 0), width=1)
            draw.text((left_pad, 8), "Lane slice", fill=(0, 0, 0))
        except Exception:
            pass

    # Axes: X=intensity (linear), Y=position(px)
    draw.line((x0, mt + ph, x1, mt + ph), fill=(0, 0, 0), width=2)   # x-axis
    draw.line((x0, mt, x0, mt + ph), fill=(0, 0, 0), width=2)         # y-axis

    try:
        sig_max = float(np.nanmax(sig)) if sig.size > 0 else 1.0
    except Exception:
        sig_max = 1.0
    if (not np.isfinite(sig_max)) or sig_max <= 0:
        sig_max = 1.0
    if x_scale_max is None:
        vmax = float(max(1.0, sig_max))
    else:
        try:
            vmax = float(x_scale_max)
            if (not np.isfinite(vmax)) or vmax <= 0:
                vmax = float(max(1.0, sig_max))
        except Exception:
            vmax = float(max(1.0, sig_max))
    pts: List[Tuple[float, float]] = []
    if n > 1:
        for i, v in enumerate(sig):
            try:
                vf = float(v)
            except Exception:
                continue
            if not np.isfinite(vf):
                continue
            x = x0 + (vf / vmax) * pw
            y = mt + (i / (n - 1)) * ph
            pts.append((x, y))
    elif n == 1:
        try:
            vf = float(sig[0])
            if np.isfinite(vf):
                pts.append((x0 + (vf / vmax) * pw, mt))
        except Exception:
            pass
    if len(pts) >= 2:
        draw.line(pts, fill=(30, 90, 180), width=2)

    # Threshold line (vertical in intensity-vs-position plot)
    th = float(max(0.0, threshold))
    th_x = x0 + (min(th, vmax) / vmax) * pw
    draw.line((th_x, mt, th_x, mt + ph), fill=(220, 120, 0), width=1)

    # Peak marks (horizontal at detected Y)
    for p in peaks:
        if n <= 1:
            continue
        y = mt + (float(p) / (n - 1)) * ph
        draw.line((x0, y, x1, y), fill=(200, 0, 0), width=1)

    # Grid + ticks (more labels, but kept outside the plot area)
    y_tick_count = 7 if n > 1 else 1
    y_ticks: List[int] = []
    if n > 1:
        y_ticks = sorted(set(int(round((n - 1) * i / max(1, y_tick_count - 1))) for i in range(y_tick_count)))
    else:
        y_ticks = [0]
    for yp in y_ticks:
        yy = mt if n <= 1 else (mt + (float(yp) / max(1, n - 1)) * ph)
        draw.line((x0 - 5, yy, x0, yy), fill=(0, 0, 0), width=1)
        draw.line((x0, yy, x1, yy), fill=(230, 230, 230), width=1)
        txt = str(int(yp))
        bb = draw.textbbox((0, 0), txt)
        draw.text((x0 - 8 - (bb[2] - bb[0]), yy - (bb[3] - bb[1]) / 2.0), txt, fill=(0, 0, 0))

    x_tick_vals = [0.0, 0.25 * vmax, 0.5 * vmax, 0.75 * vmax, vmax]
    for xv in x_tick_vals:
        xx = x0 + (float(xv) / max(vmax, 1e-9)) * pw
        draw.line((xx, mt + ph, xx, mt + ph + 5), fill=(0, 0, 0), width=1)
        draw.line((xx, mt, xx, mt + ph), fill=(240, 240, 240), width=1)
        txt = f"{xv:.0f}"
        bb = draw.textbbox((0, 0), txt)
        tw = bb[2] - bb[0]
        tx = max(left_pad, min(W - ann_w - 6 - tw, xx - tw / 2.0))
        draw.text((tx, mt + ph + 7), txt, fill=(0, 0, 0))

    # Labels/titles
    draw.text((x0, 8), title, fill=(0, 0, 0))
    draw.text((x0, mt + ph + 30), "Intensity (a.u.)", fill=(0, 0, 0))
    draw.text((x0 - 56, mt + 2), "Y (px)", fill=(0, 0, 0))

    # Keep threshold text out of the graph area.
    ann_x0 = x1 + 10
    draw.rectangle((ann_x0, mt, W - 10, mt + ph), outline=(220, 220, 220), width=1)
    draw.text((ann_x0 + 6, mt + 4), f"Threshold: {th:.1f}", fill=(220, 120, 0))
    draw.text((ann_x0 + 6, mt + 22), "Peak labels", fill=(0, 0, 0))

    # Peak labels (y px + optional fitted size) placed in the annotation column with collision avoidance.
    peak_lbl_map = {int(k): str(v) for k, v in (peak_value_labels or {}).items()}
    placed_label_ys: List[float] = []
    for p in list(peaks or []):
        if n <= 1:
            continue
        yy = mt + (float(p) / (n - 1)) * ph
        draw.line((x1 - 4, yy, x1 + 4, yy), fill=(180, 0, 0), width=1)
        base_txt = f"y={int(p)} px"
        size_txt = peak_lbl_map.get(int(p), "").strip()
        if size_txt:
            if str(size_unit).strip() and (str(size_unit).strip() not in size_txt):
                txt = f"{base_txt}, {size_txt} {str(size_unit).strip()}".strip()
            else:
                txt = f"{base_txt}, {size_txt}"
        else:
            txt = base_txt
        # Collision-avoidance in annotation column.
        ty = float(yy) - 7.0
        for _ in range(20):
            if all(abs(float(ty) - py) >= 14.0 for py in placed_label_ys):
                break
            ty += 14.0
        ty = max(float(mt + 40), min(float(mt + ph - 14), float(ty)))
        placed_label_ys.append(float(ty))
        draw.text((ann_x0 + 10, ty), txt, fill=(120, 0, 0))

    # Optional lane context text below lane slice (top-left area).
    meta = [str(x).strip() for x in (lane_meta_lines or []) if str(x).strip()]
    if meta:
        y0 = mt + ph + 50
        for i, line in enumerate(meta[:3]):
            draw.text((left_pad, y0 + i * 14), line, fill=(0, 0, 0))
    return out

def bracket_levels(groups: List[GroupLabel], lane_count: int) -> Tuple[List[Tuple[GroupLabel, int, int, int]], int]:
    """
    Assign a vertical level to each bracket so overlapping lane ranges stack.
    Returns a list of (group, start_lane, end_lane, level) and max level.
    """
    levels: List[List[Tuple[int, int]]] = []
    specs: List[Tuple[GroupLabel, int, int, int]] = []

    for g in groups:
        s = max(1, int(g.start_lane))
        e = min(int(lane_count), int(g.end_lane))
        if e < s:
            s, e = e, s
        if e < 1 or s > lane_count:
            continue

        level = 0
        while True:
            if level >= len(levels):
                levels.append([])
            overlap = any(not (e < a or s > b) for a, b in levels[level])
            if not overlap:
                levels[level].append((s, e))
                specs.append((g, s, e, level))
                break
            level += 1

    max_level = len(levels) - 1
    return specs, max_level


def rotated_text_size(text: str, font: ImageFont.ImageFont, angle_deg: float) -> Tuple[int, int]:
    if not text:
        return (0, 0)
    bb = font.getbbox(text)
    tw = max(1, int(bb[2] - bb[0]))
    th = max(1, int(bb[3] - bb[1]))
    if abs(float(angle_deg)) < 1e-6:
        return (tw, th)
    rad = math.radians(float(angle_deg))
    rw = abs(tw * math.cos(rad)) + abs(th * math.sin(rad))
    rh = abs(tw * math.sin(rad)) + abs(th * math.cos(rad))
    return (int(math.ceil(rw)), int(math.ceil(rh)))


def rotated_text_pivot_x_extents(
    text: str,
    font: ImageFont.ImageFont,
    angle_deg: float,
    pivot_mode: str = "bl",  # bl | br
) -> Tuple[float, float]:
    """
    Approximate x-extents of a rotated text box relative to its pivot.
    Used to reserve output width so angled header labels are not clipped.
    """
    if not text:
        return (0.0, 0.0)
    bb = font.getbbox(text)
    pad = 8.0
    tx = float(pad - bb[0])
    ty = float(pad - bb[1])
    tw = max(1.0, float(bb[2] - bb[0]))
    th = max(1.0, float(bb[3] - bb[1]))
    w = max(2.0, tw + 2.0 * pad)
    h = max(2.0, th + 2.0 * pad)
    if str(pivot_mode).lower() == "br":
        px = float(tx + bb[2])
    else:
        px = float(tx + bb[0])
    py = float(ty + bb[3])
    cx = w / 2.0
    cy = h / 2.0
    ang = math.radians(float(angle_deg))
    ca = math.cos(ang)
    sa = math.sin(ang)

    def _rot_about_center(x: float, y: float) -> Tuple[float, float]:
        dx = x - cx
        dy = y - cy
        xr = dx * ca - dy * sa + cx
        yr = dx * sa + dy * ca + cy
        return (xr, yr)

    # Use the actual text bbox corners (not the padded image corners).
    text_corners = [
        _rot_about_center(float(tx + bb[0]), float(ty + bb[1])),
        _rot_about_center(float(tx + bb[2]), float(ty + bb[1])),
        _rot_about_center(float(tx + bb[2]), float(ty + bb[3])),
        _rot_about_center(float(tx + bb[0]), float(ty + bb[3])),
    ]
    min_x = min(p[0] for p in text_corners)
    p_rx, _p_ry = _rot_about_center(px, py)
    # Return extents relative to the pivot point.
    ex0 = float(min_x - p_rx)
    ex1 = float(max(p[0] for p in text_corners) - p_rx)
    return (ex0, ex1)


def rotated_text_pivot_left_anchor_offset(
    text: str,
    font: ImageFont.ImageFont,
    angle_deg: float,
    pivot_mode: str = "bl",  # bl | br
) -> Tuple[float, float]:
    """
    Return the offset (dx, dy) from pivot to the rotated-box corner that defines the
    left-most x position. Used to anchor that left edge consistently to a lane center + baseline.
    """
    if not text:
        return (0.0, 0.0)
    bb = font.getbbox(text)
    pad = 8.0
    tx = float(pad - bb[0])
    ty = float(pad - bb[1])
    tw = max(1.0, float(bb[2] - bb[0]))
    th = max(1.0, float(bb[3] - bb[1]))
    w = max(2.0, tw + 2.0 * pad)
    h = max(2.0, th + 2.0 * pad)
    if str(pivot_mode).lower() == "br":
        px = float(tx + bb[2])
    else:
        px = float(tx + bb[0])
    py = float(ty + bb[3])
    cx = w / 2.0
    cy = h / 2.0
    ang = math.radians(float(angle_deg))
    ca = math.cos(ang)
    sa = math.sin(ang)

    def _rot_about_center(x: float, y: float) -> Tuple[float, float]:
        dx = x - cx
        dy = y - cy
        xr = dx * ca - dy * sa + cx
        yr = dx * sa + dy * ca + cy
        return (xr, yr)

    p_rx, p_ry = _rot_about_center(px, py)
    rel_corners: List[Tuple[float, float]] = []
    for x, y in (
        (float(tx + bb[0]), float(ty + bb[1])),
        (float(tx + bb[2]), float(ty + bb[1])),
        (float(tx + bb[2]), float(ty + bb[3])),
        (float(tx + bb[0]), float(ty + bb[3])),
    ):
        xr, yr = _rot_about_center(x, y)
        rel_corners.append((float(xr - p_rx), float(yr - p_ry)))

    xs = [p[0] for p in rel_corners]
    ys = [p[1] for p in rel_corners]
    # Anchor to the lower end of the left edge of the rotated *axis-aligned* text box.
    # That point is (min_x, max_y) in rotated bbox coordinates relative to the pivot.
    return (float(min(xs)), float(max(ys)))


def draw_rotated_text_center_top(
    out: Image.Image,
    text: str,
    center_x: float,
    top_y: float,
    angle_deg: float,
    font: ImageFont.ImageFont,
    fill: Tuple[int, int, int] = (0, 0, 0),
) -> None:
    if not text:
        return
    angle = float(angle_deg)
    if abs(angle) < 1e-6:
        draw = ImageDraw.Draw(out)
        bb = draw.textbbox((0, 0), text, font=font)
        tw = bb[2] - bb[0]
        draw.text((float(center_x) - tw / 2.0, float(top_y)), text, fill=fill, font=font)
        return

    bb = font.getbbox(text)
    tw = max(1, int(bb[2] - bb[0]))
    th = max(1, int(bb[3] - bb[1]))
    pad = 6
    txt = Image.new("RGBA", (tw + 2 * pad, th + 2 * pad), (0, 0, 0, 0))
    td = ImageDraw.Draw(txt)
    td.text((pad, pad), text, font=font, fill=(fill[0], fill[1], fill[2], 255))
    rot = txt.rotate(angle, expand=True, resample=Image.Resampling.BICUBIC)
    x = int(round(float(center_x) - rot.width / 2.0))
    y = int(round(float(top_y)))
    out.paste(rot, (x, y), rot)


def draw_rotated_text_pivot(
    out: Image.Image,
    text: str,
    pivot_x: float,
    pivot_y: float,
    angle_deg: float,
    font: ImageFont.ImageFont,
    pivot_mode: str = "bl",  # bl | br
    fill: Tuple[int, int, int] = (0, 0, 0),
) -> None:
    """
    Draw text rotated around a specific pivot point, then place that pivot at (pivot_x, pivot_y).
    Used for header values so angled labels "start" from the bottom above the lane center.
    """
    if not text:
        return
    angle = float(angle_deg)
    bb = font.getbbox(text)
    tw = max(1, int(bb[2] - bb[0]))
    th = max(1, int(bb[3] - bb[1]))
    pad = 8
    w = max(2, tw + 2 * pad)
    h = max(2, th + 2 * pad)

    txt = Image.new("RGBA", (w, h), (0, 0, 0, 0))
    td = ImageDraw.Draw(txt)
    tx = int(pad - bb[0])
    ty = int(pad - bb[1])
    td.text((tx, ty), text, font=font, fill=(fill[0], fill[1], fill[2], 255))

    if str(pivot_mode).lower() == "br":
        px = int(round(tx + bb[2]))
    else:
        px = int(round(tx + bb[0]))
    py = int(round(ty + bb[3]))
    px = max(0, min(w - 1, px))
    py = max(0, min(h - 1, py))

    rot = txt.rotate(angle, expand=True, resample=Image.Resampling.BICUBIC)

    # Compute pivot location in the expanded rotated image analytically.
    # This avoids pixel-quantized jumps that occur when tracking a 1px marker.
    cx = float(w) / 2.0
    cy = float(h) / 2.0
    ang = math.radians(float(angle))
    ca = math.cos(ang)
    sa = math.sin(ang)

    def _rot_about_center(x: float, y: float) -> Tuple[float, float]:
        dx = x - cx
        dy = y - cy
        xr = dx * ca - dy * sa + cx
        yr = dx * sa + dy * ca + cy
        return (xr, yr)

    corners = [
        _rot_about_center(0.0, 0.0),
        _rot_about_center(float(w), 0.0),
        _rot_about_center(float(w), float(h)),
        _rot_about_center(0.0, float(h)),
    ]
    min_x = min(p[0] for p in corners)
    min_y = min(p[1] for p in corners)
    p_rx, p_ry = _rot_about_center(float(px), float(py))
    rp_x = float(p_rx - min_x)
    rp_y = float(p_ry - min_y)
    x = int(round(float(pivot_x) - rp_x))
    y = int(round(float(pivot_y) - rp_y))
    out.paste(rot, (x, y), rot)


def draw_rotated_header_value(
    out: Image.Image,
    text: str,
    cell_center_x: float,
    row_top_y: float,
    row_h: float,
    baseline_h: float,
    angle_deg: float,
    font: ImageFont.ImageFont,
    fill: Tuple[int, int, int] = (0, 0, 0),
) -> None:
    """
    Header cell values:
    - 0 deg: centered above the lane cell (legacy)
    - +angle: anchor at bottom-left, with anchor centered over the cell
    - -angle: anchor at bottom-right, with anchor centered over the cell
    """
    if not text:
        return
    ang = float(angle_deg)
    if abs(ang) < 1e-6:
        draw_rotated_text_center_top(
            out,
            text=text,
            center_x=float(cell_center_x),
            top_y=float(row_top_y),
            angle_deg=0.0,
            font=font,
            fill=fill,
        )
        return
    # Bottom anchor a little above the row bottom to avoid clipping and match publication-style labels.
    # Anchor near the row bottom so the left-edge point sits close to the image.
    # The earlier drift issue was caused by incorrect anchor geometry, which is now fixed.
    anchor_y = float(row_top_y) + float(row_h) - 3.0
    # Use the left-most edge of the rotated label box as the centered anchor above the cell.
    pivot_mode = "bl"
    ex0, ey0 = rotated_text_pivot_left_anchor_offset(text, font, ang, pivot_mode=pivot_mode)
    pivot_x = float(cell_center_x) - float(ex0)
    pivot_y = float(anchor_y) - float(ey0)
    draw_rotated_text_pivot(
        out,
        text=text,
        pivot_x=pivot_x,
        pivot_y=pivot_y,
        angle_deg=ang,
        font=font,
        pivot_mode=pivot_mode,
        fill=fill,
    )


def compute_render_layout(panel: Image.Image, pc: PanelConfig, settings: AppSettings) -> Dict[str, Any]:
    """
    Compute the geometry that render_panel() uses, so interactive steps can place
    lane guides, headers, brackets, and marker annotations consistently.
    """
    font = get_font(settings)
    marker_font = get_font_with_size(settings, pc.marker_font_size if pc.marker_font_size else int(settings.font_size))
    default_font_size = max(8, int(settings.font_size))

    W, H = panel.size
    lanes = max(1, pc.lanes)
    _normalize_top_annotations(pc)

    line_h = font.getbbox("Ag")[3] - font.getbbox("Ag")[1] + 6
    header_rows_all = [hr for hr in pc.header_rows if (hr.name.strip() or any(v.strip() for v in hr.values))]
    top_header_rows: List[HeaderRow] = []
    bottom_header_rows: List[HeaderRow] = []
    for hr in header_rows_all:
        pos = str(getattr(hr, "position", "top")).strip().lower()
        if pos == "bottom":
            bottom_header_rows.append(hr)
        else:
            top_header_rows.append(hr)
    group_rows = [g for g in pc.group_labels if str(g.text).strip()]
    header_angle_deg = int(getattr(pc, "header_value_angle_deg", 0))
    header_heading_x_off = int(getattr(pc, "header_heading_x_offset", 0))
    header_heading_y_off = int(getattr(pc, "header_heading_y_offset", 0))
    header_values_x_off = int(getattr(pc, "header_values_x_offset", 0))
    header_values_y_off = int(getattr(pc, "header_values_y_offset", 0))
    bracket_label_x_off = int(getattr(pc, "bracket_label_x_offset", 0))
    bracket_label_y_off = int(getattr(pc, "bracket_label_y_offset", 0))
    bracket_line_x_off = int(getattr(pc, "bracket_line_x_offset", 0))
    bracket_line_y_off = int(getattr(pc, "bracket_line_y_offset", 0))

    header_by_id = {hr.id: hr for hr in top_header_rows}
    groups_by_height: Dict[str, List[GroupLabel]] = {}
    for g in group_rows:
        hgid = str(getattr(g, "height_group", "")).strip()
        if not hgid:
            continue
        groups_by_height.setdefault(hgid, []).append(g)
    top_items: List[Tuple[str, Any]] = []
    for oid in pc.top_annotation_order:
        if oid in header_by_id:
            top_items.append(("header", header_by_id[oid]))
            continue
        if oid.startswith("BG:"):
            hgid = oid[3:]
            if hgid in groups_by_height:
                top_items.append(("group_set", groups_by_height[hgid]))
                continue
        if oid in groups_by_height:
            top_items.append(("group_set", groups_by_height[oid]))

    header_row_h_by_id: Dict[str, int] = {}
    header_row_angle_by_id: Dict[str, int] = {}
    header_row_heading_font_size_by_id: Dict[str, int] = {}
    header_row_value_font_size_by_id: Dict[str, int] = {}
    header_row_heading_line_h_by_id: Dict[str, int] = {}
    header_row_value_line_h_by_id: Dict[str, int] = {}
    for hr in header_rows_all:
        row_angle = max(-80, min(80, int(getattr(hr, "angle_deg", header_angle_deg))))
        header_row_angle_by_id[hr.id] = int(row_angle)
        heading_fs = max(6, int(getattr(hr, "heading_font_size", 0) or default_font_size))
        value_fs = max(6, int(getattr(hr, "value_font_size", 0) or default_font_size))
        header_row_heading_font_size_by_id[hr.id] = int(heading_fs)
        header_row_value_font_size_by_id[hr.id] = int(value_fs)
        heading_font = get_font_with_size(settings, heading_fs)
        value_font = get_font_with_size(settings, value_fs)
        heading_line_h = int(heading_font.getbbox("Ag")[3] - heading_font.getbbox("Ag")[1] + 6)
        value_line_h = int(value_font.getbbox("Ag")[3] - value_font.getbbox("Ag")[1] + 6)
        header_row_heading_line_h_by_id[hr.id] = int(heading_line_h)
        header_row_value_line_h_by_id[hr.id] = int(value_line_h)
        row_h = int(max(heading_line_h, value_line_h))
        if abs(row_angle) > 0:
            max_h = int(max(heading_line_h, value_line_h))
            for v in hr.values:
                vv = str(v).strip()
                if not vv:
                    continue
                _rw, rh = rotated_text_size(vv, value_font, row_angle)
                max_h = max(max_h, int(rh) + 14)
            row_h = max_h
        header_row_h_by_id[hr.id] = int(row_h)
    header_row_h = max([int(line_h)] + [int(v) for v in header_row_h_by_id.values()])

    # extra vertical space so bracket text is clearly separated from the bracket line
    BRACKET_TEXT_TO_LINE_GAP = max(0, int(getattr(pc, "bracket_text_gap", 26)))
    bracket_row_h = line_h + BRACKET_TEXT_TO_LINE_GAP + 18
    top_block_h = 0
    for kind, item in top_items:
        if kind == "header":
            top_block_h += int(header_row_h_by_id.get(getattr(item, "id", ""), header_row_h))
        else:
            top_block_h += bracket_row_h

    y_offsets = [header_heading_y_off, header_values_y_off, bracket_label_y_off, bracket_line_y_off]
    offset_up = max(0, -min(y_offsets))
    offset_down = max(0, max(y_offsets))
    has_top_angle = any(abs(int(header_row_angle_by_id.get(hr.id, header_angle_deg))) > 0 for hr in top_header_rows)
    header_top_pad = 18 if (top_header_rows and has_top_angle) else 10
    top_margin = int(header_top_pad + offset_up + top_block_h + offset_down)

    x_offsets = [header_heading_x_off, header_values_x_off, bracket_label_x_off, bracket_line_x_off]
    offset_left = max(0, -min(x_offsets))
    offset_right = max(0, max(x_offsets))
    has_header = any(kind == "header" for kind, _ in top_items)
    left_name_margin = 10
    if has_header:
        max_name_w = 0
        for hr in header_rows_all:
            nm = str(getattr(hr, "name", "") or "").strip()
            if not nm:
                continue
            heading_font = get_font_with_size(
                settings,
                int(header_row_heading_font_size_by_id.get(hr.id, default_font_size)),
            )
            bb_nm = heading_font.getbbox(nm)
            max_name_w = max(max_name_w, int(bb_nm[2] - bb_nm[0]))
        left_name_margin = max(120, int(max_name_w + 12))
    left_marker_margin = 70 if (pc.include_marker and pc.marker_calibration.marker_name) else 10
    left_margin = int(max(left_name_margin, left_marker_margin) + offset_left)
    right_margin = int(10 + offset_right)
    bottom_margin = 10
    bottom_offsets = [header_heading_y_off, header_values_y_off]
    bottom_offset_up = max(0, -min(bottom_offsets))
    bottom_offset_down = max(0, max(bottom_offsets))
    bottom_header_block_h = 0
    if bottom_header_rows:
        bottom_h = sum(int(header_row_h_by_id.get(hr.id, header_row_h)) for hr in bottom_header_rows)
        bottom_header_block_h = int(bottom_offset_up + bottom_offset_down + 8 + bottom_h)
        bottom_margin = max(bottom_margin, bottom_header_block_h + 8)
    final_label_font_size = int(getattr(pc, "final_band_label_font_size", 15) or 15)
    final_label_font = get_font_with_size(settings, final_label_font_size if final_label_font_size > 0 else int(settings.font_size))
    final_label_line_h = max(12, int(final_label_font.getbbox("Ag")[3] - final_label_font.getbbox("Ag")[1] + 7))
    final_label_stagger_px = 0
    final_size_labels = final_band_size_labels_for_panel(pc)
    if final_size_labels:
        max_stack = max(len(v) for v in final_size_labels.values()) if final_size_labels else 1
        bottom_margin = max(
            bottom_margin,
            int(bottom_header_block_h + 6 + max(1, max_stack) * final_label_line_h + 6),
        )

    # Lane geometry in panel-local coords (supports multiple gel regions).
    lane_bounds = compute_lane_bounds(panel, pc)
    lane_lefts_local = [float(x0) for (x0, _x1) in lane_bounds]
    lane_rights_local = [float(x1) for (_x0, x1) in lane_bounds]
    lane_centers_local = [((lx + rx) / 2.0) for lx, rx in zip(lane_lefts_local, lane_rights_local)]

    # Extend horizontal margins if angled header values would overflow panel bounds.
    extra_label_left = 0.0
    extra_label_right = 0.0
    for hr in header_rows_all:
        row_angle = int(header_row_angle_by_id.get(getattr(hr, "id", ""), header_angle_deg))
        value_font = get_font_with_size(
            settings,
            int(header_row_value_font_size_by_id.get(getattr(hr, "id", ""), default_font_size)),
        )
        for i in range(min(len(lane_centers_local), max(0, lanes))):
            v = str(hr.values[i] if i < len(hr.values) else "").strip()
            if not v:
                continue
            cx = float(lane_centers_local[i]) + float(header_values_x_off)
            if abs(float(row_angle)) < 1e-6:
                bbv = value_font.getbbox(v)
                tw = float(bbv[2] - bbv[0])
                minx = cx - tw / 2.0
                maxx = cx + tw / 2.0
            else:
                pivot_mode = "bl"
                ex0, ex1 = rotated_text_pivot_x_extents(v, value_font, row_angle, pivot_mode=pivot_mode)
                # Header-value draw anchors the rotated label's left-most edge to the column center.
                # Therefore the rendered x-span is [cx, cx + (ex1 - ex0)].
                minx = float(cx)
                maxx = float(cx) + float(ex1 - ex0)
            if minx < 0.0:
                # Keep left margin stable to avoid the whole panel "jumping" while changing angle.
                # User requested automatic extension for right overflow; left overflow can still be
                # handled manually via header X offset if needed.
                extra_label_left = max(extra_label_left, 0.0)
            if maxx > float(W):
                extra_label_right = max(extra_label_right, maxx - float(W) + 2.0)

    left_margin = int(left_margin)
    right_margin = int(right_margin + math.ceil(extra_label_right))

    out_w = int(W + left_margin + right_margin)
    out_h = int(H + top_margin + bottom_margin)
    panel_origin = (int(left_margin), int(top_margin))

    # Lane geometry in output coords.
    lane_lefts = [panel_origin[0] + float(x0) for (x0, _x1) in lane_bounds]
    lane_rights = [panel_origin[0] + float(x1) for (_x0, x1) in lane_bounds]
    lane_centers = [((lx + rx) / 2.0) for lx, rx in zip(lane_lefts, lane_rights)]

    # Marker bbox (for interactive overlays in Step 6)
    marker_bbox = None
    marker_tick_x = None
    try:
        calib = pc.marker_calibration
        pos = marker_positions_for_panel(panel, pc, settings)
        if pos:
            lanes2 = max(1, pc.lanes)
            ladder_lane = max(1, min(lanes2, int(calib.ladder_lane) if calib.ladder_lane else 1))
            x_tick = lane_lefts[ladder_lane - 1] - 8 + int(pc.marker_x_offset)
            marker_tick_x = int(x_tick)
            tick_len = max(2, int(pc.marker_tick_length))
            label_gap = max(0, int(pc.marker_label_gap))

            max_tw = 0
            for s, _y in pos:
                label = _format_marker_size(float(s))
                try:
                    bb = marker_font.getbbox(label)
                    tw = int(bb[2] - bb[0])
                except Exception:
                    tw = int(max(6, 0.6 * len(label) * max(8, int(getattr(marker_font, "size", 10)))))
                max_tw = max(max_tw, int(tw))

            ys = [int(y) for (_s, y) in pos]
            y0 = panel_origin[1] + int(min(ys)) + int(pc.marker_y_offset)
            y1 = panel_origin[1] + int(max(ys)) + int(pc.marker_y_offset)
            x0 = int(x_tick - tick_len - label_gap - max_tw - 12)
            x1 = int(x_tick + 10)
            marker_bbox = (x0, int(y0 - 14), x1, int(y1 + 14))
    except Exception:
        marker_bbox = None
        marker_tick_x = None

    return dict(
        W=W,
        H=H,
        out_w=out_w,
        out_h=out_h,
        panel_origin=panel_origin,
        line_h=line_h,
        header_top_pad=header_top_pad,
        header_offset_up=offset_up,
        header_row_h=header_row_h,
        header_row_h_by_id=header_row_h_by_id,
        header_row_angle_by_id=header_row_angle_by_id,
        header_row_heading_font_size_by_id=header_row_heading_font_size_by_id,
        header_row_value_font_size_by_id=header_row_value_font_size_by_id,
        header_row_heading_line_h_by_id=header_row_heading_line_h_by_id,
        header_row_value_line_h_by_id=header_row_value_line_h_by_id,
        bracket_row_h=bracket_row_h,
        header_angle_deg=header_angle_deg,
        header_rows=top_header_rows,
        bottom_header_rows=bottom_header_rows,
        bottom_header_block_h=bottom_header_block_h,
        bottom_header_offset_up=bottom_offset_up,
        top_items=top_items,
        BRACKET_TEXT_TO_LINE_GAP=BRACKET_TEXT_TO_LINE_GAP,
        group_level_h=bracket_row_h,
        group_specs=[],
        group_max_level=0,
        lane_centers=lane_centers,
        lane_lefts=lane_lefts,
        lane_rights=lane_rights,
        marker_bbox=marker_bbox,
        marker_tick_x=marker_tick_x,
        final_size_labels=final_size_labels,
        final_label_font_size=final_label_font_size,
        final_label_line_h=final_label_line_h,
        final_label_stagger_px=final_label_stagger_px,
    )

def render_panel(
    panel: Image.Image,
    pc: PanelConfig,
    settings: AppSettings,
    preview: bool = False,
    collect_elements: Optional[List[Dict[str, Any]]] = None,
    panel_index: int = 0,
    panel_y_offset: int = 0,
) -> Image.Image:
    """
    Render a single panel with header rows, group brackets, highlights, marker sizes, and marker annotations.
    """
    layout = compute_render_layout(panel, pc, settings)
    font = get_font(settings)
    marker_font = get_font_with_size(settings, pc.marker_font_size if pc.marker_font_size else int(settings.font_size))
    default_font_size = max(8, int(settings.font_size))
    final_label_font_size = int(layout.get("final_label_font_size", 0) or 0)
    final_label_font = get_font_with_size(settings, final_label_font_size if final_label_font_size > 0 else int(settings.font_size))
    _font_cache: Dict[int, ImageFont.ImageFont] = {}

    def _row_font(sz: int) -> ImageFont.ImageFont:
        key = int(max(6, sz))
        f = _font_cache.get(key)
        if f is None:
            f = get_font_with_size(settings, key)
            _font_cache[key] = f
        return f

    W, H = layout["W"], layout["H"]
    out = Image.new("RGB", (layout["out_w"], layout["out_h"]), (255, 255, 255))
    out.paste(panel, layout["panel_origin"])

    draw = ImageDraw.Draw(out)
    review_overrides: Dict[str, Dict[str, Any]] = dict(getattr(pc, "review_text_overrides", {}) or {})

    def _style_for(
        key: str,
        base_size: int,
        base_color: str = "#000000",
        base_angle: float = 0.0,
        base_width: int = 2,
    ) -> Dict[str, Any]:
        ov = dict(review_overrides.get(str(key), {}) or {})
        try:
            size = max(6, int(ov.get("font_size", base_size)))
        except Exception:
            size = max(6, int(base_size))
        color = _safe_hex_color(str(ov.get("color", base_color)))
        try:
            ang = float(ov.get("angle", base_angle))
        except Exception:
            ang = float(base_angle)
        try:
            dx = float(ov.get("dx", 0.0))
        except Exception:
            dx = 0.0
        try:
            dy = float(ov.get("dy", 0.0))
        except Exception:
            dy = 0.0
        try:
            width = max(1, int(ov.get("width", base_width)))
        except Exception:
            width = max(1, int(base_width))
        fam = str(ov.get("font_family", "default") or "default").strip()
        if fam.lower() not in {"default", "dejavusans"}:
            fam = "default"
        return dict(font_size=size, color=color, angle=ang, dx=dx, dy=dy, width=width, font_family=fam)

    def _font_by_family(font_size: int, family: str) -> ImageFont.ImageFont:
        fs = max(6, int(font_size))
        fam = str(family or "default").strip().lower()
        if fam == "dejavusans":
            try:
                return ImageFont.truetype("DejaVuSans.ttf", size=fs)
            except Exception:
                pass
        return get_font_with_size(settings, fs)

    def _hex_rgb(hex_color: str) -> Tuple[int, int, int]:
        try:
            return _hex_to_rgb(_safe_hex_color(hex_color))
        except Exception:
            return (0, 0, 0)

    def _record_element(
        key: str,
        text: str,
        etype: str,
        bbox_local: Tuple[float, float, float, float],
        style: Dict[str, Any],
    ) -> None:
        if collect_elements is None:
            return
        x0, y0, x1, y1 = bbox_local
        collect_elements.append(
            dict(
                panel_index=int(panel_index),
                key=str(key),
                text=str(text),
                etype=str(etype),
                bbox=(
                    float(x0),
                    float(y0 + panel_y_offset),
                    float(x1),
                    float(y1 + panel_y_offset),
                ),
                style=dict(style),
            )
        )

    def _draw_text_top_left(
        key: str,
        text: str,
        etype: str,
        x: float,
        y: float,
        base_size: int,
        base_color: str = "#000000",
        base_angle: float = 0.0,
    ) -> Tuple[float, float, float, float]:
        s = _style_for(key, base_size=base_size, base_color=base_color, base_angle=base_angle)
        fnt = _font_by_family(int(s["font_size"]), str(s["font_family"]))
        fill = _hex_rgb(str(s["color"]))
        xx = float(x) + float(s["dx"])
        yy = float(y) + float(s["dy"])
        ang = float(s["angle"])
        if abs(ang) < 1e-6:
            draw.text((xx, yy), text, fill=fill, font=fnt)
            bb = draw.textbbox((xx, yy), text, font=fnt)
            bbox = (float(bb[0]), float(bb[1]), float(bb[2]), float(bb[3]))
            _record_element(key, text, etype, bbox, s)
            return bbox

        bb0 = fnt.getbbox(text)
        tw = max(1, int(bb0[2] - bb0[0]))
        th = max(1, int(bb0[3] - bb0[1]))
        pad = 6
        txt = Image.new("RGBA", (tw + 2 * pad, th + 2 * pad), (0, 0, 0, 0))
        td = ImageDraw.Draw(txt)
        td.text((pad - bb0[0], pad - bb0[1]), text, font=fnt, fill=(fill[0], fill[1], fill[2], 255))
        rot = txt.rotate(float(ang), expand=True, resample=Image.Resampling.BICUBIC)
        xi = int(round(xx))
        yi = int(round(yy))
        out.paste(rot, (xi, yi), rot)
        bbox = (float(xi), float(yi), float(xi + rot.width), float(yi + rot.height))
        _record_element(key, text, etype, bbox, s)
        return bbox

    def _rotate_points(points: List[Tuple[float, float]], angle_deg: float, cx: float, cy: float) -> List[Tuple[float, float]]:
        if abs(float(angle_deg)) < 1e-6:
            return list(points)
        a = math.radians(float(angle_deg))
        ca = math.cos(a)
        sa = math.sin(a)
        out_pts: List[Tuple[float, float]] = []
        for x, y in points:
            dx = float(x) - float(cx)
            dy = float(y) - float(cy)
            rx = float(cx) + dx * ca - dy * sa
            ry = float(cy) + dx * sa + dy * ca
            out_pts.append((rx, ry))
        return out_pts

    def _bbox_from_points(points: List[Tuple[float, float]], pad: float = 0.0) -> Tuple[float, float, float, float]:
        xs = [float(p[0]) for p in points]
        ys = [float(p[1]) for p in points]
        return (min(xs) - pad, min(ys) - pad, max(xs) + pad, max(ys) + pad)

    def _draw_line_element(
        key: str,
        etype: str,
        points: List[Tuple[float, float]],
        base_width: int = 2,
        base_color: str = "#000000",
    ) -> Tuple[float, float, float, float]:
        s = _style_for(key, base_size=base_width, base_color=base_color, base_angle=0.0, base_width=base_width)
        pts = [(float(x) + float(s["dx"]), float(y) + float(s["dy"])) for (x, y) in points]
        if pts:
            cx = sum(p[0] for p in pts) / max(1, len(pts))
            cy = sum(p[1] for p in pts) / max(1, len(pts))
            pts = _rotate_points(pts, float(s["angle"]), cx, cy)
        fill = _hex_rgb(str(s["color"]))
        wd = max(1, int(s["width"]))
        if len(pts) >= 2:
            draw.line(pts, fill=fill, width=wd)
        bbox = _bbox_from_points(pts, pad=max(2.0, float(wd)))
        _record_element(key, "", etype, bbox, s)
        return bbox

    def _draw_polygon_outline_element(
        key: str,
        etype: str,
        points: List[Tuple[float, float]],
        closed: bool = True,
        base_width: int = 2,
        base_color: str = "#000000",
    ) -> Tuple[float, float, float, float]:
        s = _style_for(key, base_size=base_width, base_color=base_color, base_angle=0.0, base_width=base_width)
        pts = [(float(x) + float(s["dx"]), float(y) + float(s["dy"])) for (x, y) in points]
        if pts:
            cx = sum(p[0] for p in pts) / max(1, len(pts))
            cy = sum(p[1] for p in pts) / max(1, len(pts))
            pts = _rotate_points(pts, float(s["angle"]), cx, cy)
        fill = _hex_rgb(str(s["color"]))
        wd = max(1, int(s["width"]))
        if closed and len(pts) >= 2:
            draw.line(pts + [pts[0]], fill=fill, width=wd)
        elif len(pts) >= 2:
            draw.line(pts, fill=fill, width=wd)
        bbox = _bbox_from_points(pts, pad=max(2.0, float(wd)))
        _record_element(key, "", etype, bbox, s)
        return bbox

    def _draw_polygon_filled_element(
        key: str,
        etype: str,
        points: List[Tuple[float, float]],
        base_color: str = "#000000",
    ) -> Tuple[float, float, float, float]:
        s = _style_for(key, base_size=1, base_color=base_color, base_angle=0.0, base_width=1)
        pts = [(float(x) + float(s["dx"]), float(y) + float(s["dy"])) for (x, y) in points]
        if pts:
            cx = sum(p[0] for p in pts) / max(1, len(pts))
            cy = sum(p[1] for p in pts) / max(1, len(pts))
            pts = _rotate_points(pts, float(s["angle"]), cx, cy)
        fill = _hex_rgb(str(s["color"]))
        if len(pts) >= 3:
            draw.polygon(pts, fill=fill, outline=fill)
        bbox = _bbox_from_points(pts, pad=2.0)
        _record_element(key, "", etype, bbox, s)
        return bbox

    line_h = layout["line_h"]
    header_top_pad = int(layout.get("header_top_pad", 10))
    header_offset_up = int(layout.get("header_offset_up", 0))
    header_row_h = int(layout.get("header_row_h", line_h))
    header_row_h_by_id: Dict[str, int] = dict(layout.get("header_row_h_by_id", {}) or {})
    header_row_angle_by_id: Dict[str, int] = dict(layout.get("header_row_angle_by_id", {}) or {})
    header_row_heading_font_size_by_id: Dict[str, int] = dict(layout.get("header_row_heading_font_size_by_id", {}) or {})
    header_row_value_font_size_by_id: Dict[str, int] = dict(layout.get("header_row_value_font_size_by_id", {}) or {})
    header_row_value_line_h_by_id: Dict[str, int] = dict(layout.get("header_row_value_line_h_by_id", {}) or {})
    bracket_row_h = int(layout.get("bracket_row_h", line_h + int(layout["BRACKET_TEXT_TO_LINE_GAP"]) + 18))
    header_angle_deg = int(layout.get("header_angle_deg", 0))
    top_items = list(layout.get("top_items", []))
    bottom_header_rows = list(layout.get("bottom_header_rows", []))
    bottom_header_block_h = int(layout.get("bottom_header_block_h", 0))
    bottom_header_offset_up = int(layout.get("bottom_header_offset_up", 0))
    lane_centers = layout["lane_centers"]
    lane_lefts = layout["lane_lefts"]
    lane_rights = layout["lane_rights"]
    panel_origin_x, panel_origin_y = layout["panel_origin"]
    header_heading_x_off = int(getattr(pc, "header_heading_x_offset", 0))
    header_heading_y_off = int(getattr(pc, "header_heading_y_offset", 0))
    header_values_x_off = int(getattr(pc, "header_values_x_offset", 0))
    header_values_y_off = int(getattr(pc, "header_values_y_offset", 0))
    bracket_label_x_off = int(getattr(pc, "bracket_label_x_offset", 0))
    bracket_label_y_off = int(getattr(pc, "bracket_label_y_offset", 0))
    bracket_line_x_off = int(getattr(pc, "bracket_line_x_offset", 0))
    bracket_line_y_off = int(getattr(pc, "bracket_line_y_offset", 0))

    # Draw top annotations in user-defined order (headers and brackets can be interleaved).
    y = max(6, header_top_pad - 6 + header_offset_up)
    BRACKET_TEXT_TO_LINE_GAP = int(layout["BRACKET_TEXT_TO_LINE_GAP"])
    lanes = len(lane_centers)
    for kind, item in top_items:
        if kind == "header":
            hr = item
            row_angle = int(header_row_angle_by_id.get(getattr(hr, "id", ""), header_angle_deg))
            row_h = int(header_row_h_by_id.get(getattr(hr, "id", ""), header_row_h))
            heading_font = _row_font(int(header_row_heading_font_size_by_id.get(getattr(hr, "id", ""), default_font_size)))
            value_font = _row_font(int(header_row_value_font_size_by_id.get(getattr(hr, "id", ""), default_font_size)))
            value_line_h = int(header_row_value_line_h_by_id.get(getattr(hr, "id", ""), line_h))
            if hr.name.strip():
                bb_name = draw.textbbox((0, 0), hr.name, font=heading_font)
                name_x = float(panel_origin_x + header_heading_x_off - bb_name[2])
                name_y = float(y + header_heading_y_off)
                if abs(int(row_angle)) > 0:
                    th_name = float(bb_name[3] - bb_name[1])
                    name_y = float(y) + (float(row_h) - th_name) / 2.0 - float(bb_name[1]) + float(header_heading_y_off)
                _draw_text_top_left(
                    key=f"hdr_name_top:{hr.id}",
                    text=str(hr.name),
                    etype="header_name",
                    x=float(name_x),
                    y=float(name_y),
                    base_size=int(header_row_heading_font_size_by_id.get(getattr(hr, "id", ""), default_font_size)),
                    base_color="#000000",
                    base_angle=0.0,
                )
            for i in range(lanes):
                v = hr.values[i] if i < len(hr.values) else ""
                if not v:
                    continue
                key = f"hdr_val_top:{hr.id}:{i}"
                s = _style_for(
                    key=key,
                    base_size=int(header_row_value_font_size_by_id.get(getattr(hr, "id", ""), default_font_size)),
                    base_color="#000000",
                    base_angle=float(row_angle),
                )
                value_font_use = _font_by_family(int(s["font_size"]), str(s["font_family"]))
                x = lane_centers[i] + header_values_x_off + float(s["dx"])
                draw_rotated_header_value(
                    out,
                    text=str(v),
                    cell_center_x=float(x),
                    row_top_y=float(y + header_values_y_off + float(s["dy"])),
                    row_h=float(row_h),
                    baseline_h=float(value_line_h),
                    angle_deg=float(s["angle"]),
                    font=value_font_use,
                    fill=_hex_rgb(str(s["color"])),
                )
                try:
                    if abs(float(s["angle"])) < 1e-6:
                        bbv = value_font_use.getbbox(str(v))
                        tw = float(bbv[2] - bbv[0])
                        th = float(bbv[3] - bbv[1])
                        bx0 = float(x) - tw / 2.0
                        by0 = float(y + header_values_y_off + float(s["dy"]))
                        bbox = (bx0, by0, bx0 + tw, by0 + th)
                    else:
                        rw, rh = rotated_text_size(str(v), value_font_use, float(s["angle"]))
                        bx0 = float(x)
                        by0 = float(y + header_values_y_off + float(s["dy"]) + row_h - rh)
                        bbox = (bx0, by0, bx0 + float(rw), by0 + float(rh))
                    _record_element(key, str(v), "header_value", bbox, s)
                except Exception:
                    pass
            y += row_h
            continue

        if kind == "group_set":
            groups = list(item or [])
            if not groups:
                y += bracket_row_h
                continue
            y_text_base = y + 2
            y_line_base = y_text_base + line_h + BRACKET_TEXT_TO_LINE_GAP
            y_text = y_text_base + bracket_label_y_off
            y_line = y_line_base + bracket_line_y_off

            # Stable order on the row for deterministic drawing.
            groups.sort(key=lambda gg: (min(int(gg.start_lane), int(gg.end_lane)), max(int(gg.start_lane), int(gg.end_lane)), str(gg.text)))
            for g in groups:
                s = max(1, min(lanes, int(g.start_lane)))
                e = max(1, min(lanes, int(g.end_lane)))
                if e < s:
                    s, e = e, s
                x0 = lane_lefts[s - 1] + bracket_line_x_off
                x1 = lane_rights[e - 1] + bracket_line_x_off
                t = str(g.text).strip()
                if t:
                    bbox = draw.textbbox((0, 0), t, font=font)
                    tw = bbox[2] - bbox[0]
                    _draw_text_top_left(
                        key=f"bracket_label:{g.id}",
                        text=t,
                        etype="bracket_label",
                        x=float((x0 + x1) / 2 - tw / 2 + bracket_label_x_off),
                        y=float(y_text),
                        base_size=default_font_size,
                        base_color="#000000",
                        base_angle=0.0,
                    )
                if g.bracket:
                    _draw_line_element(
                        key=f"bracket_line_main:{g.id}",
                        etype="bracket_line",
                        points=[(float(x0), float(y_line)), (float(x1), float(y_line))],
                        base_width=2,
                        base_color="#000000",
                    )
                    _draw_line_element(
                        key=f"bracket_line_left:{g.id}",
                        etype="bracket_line",
                        points=[(float(x0), float(y_line)), (float(x0), float(y_line + 6))],
                        base_width=2,
                        base_color="#000000",
                    )
                    _draw_line_element(
                        key=f"bracket_line_right:{g.id}",
                        etype="bracket_line",
                        points=[(float(x1), float(y_line)), (float(x1), float(y_line + 6))],
                        base_width=2,
                        base_color="#000000",
                    )
            y += bracket_row_h

    # Highlights (relative to panel, optional)
    if bool(getattr(pc, "highlight_enabled", False)):
        for h_idx, h in enumerate(pc.highlights):
            x0 = panel_origin_x + int(h.x0)
            y0 = panel_origin_y + int(h.y0)
            x1 = panel_origin_x + int(h.x1)
            y1 = panel_origin_y + int(h.y1)
            outline = h.color if getattr(h, "color", "") else "#DC0000"
            width = max(1, int(getattr(h, "width", 2)))
            kind = str(getattr(h, "kind", "box") or "box").lower()
            if kind == "arrow":
                _draw_line_element(
                    key=f"highlight_arrow_line:{h_idx}",
                    etype="highlight_arrow",
                    points=[(float(x0), float(y0)), (float(x1), float(y1))],
                    base_width=width,
                    base_color=str(outline),
                )
                dx = float(x1 - x0)
                dy = float(y1 - y0)
                if abs(dx) + abs(dy) > 1e-6:
                    ang = math.atan2(dy, dx)
                    head_len = max(8.0, 4.0 * float(width))
                    head_half = max(4.0, 2.5 * float(width))
                    bx = float(x1) - head_len * math.cos(ang)
                    by = float(y1) - head_len * math.sin(ang)
                    lx = bx + head_half * math.cos(ang + math.pi / 2.0)
                    ly = by + head_half * math.sin(ang + math.pi / 2.0)
                    rx = bx + head_half * math.cos(ang - math.pi / 2.0)
                    ry = by + head_half * math.sin(ang - math.pi / 2.0)
                    _draw_polygon_outline_element(
                        key=f"highlight_arrow_head:{h_idx}",
                        etype="highlight_arrow",
                        points=[(float(x1), float(y1)), (float(lx), float(ly)), (float(rx), float(ry))],
                        closed=True,
                        base_width=max(1, int(width)),
                        base_color=str(outline),
                    )
                    _draw_polygon_filled_element(
                        key=f"highlight_arrow_head_fill:{h_idx}",
                        etype="highlight_arrow",
                        points=[(float(x1), float(y1)), (float(lx), float(ly)), (float(rx), float(ry))],
                        base_color=str(outline),
                    )
            elif kind == "asterisk":
                cx = 0.5 * (float(x0) + float(x1))
                cy = 0.5 * (float(y0) + float(y1))
                rr = 0.5 * max(abs(float(x1) - float(x0)), abs(float(y1) - float(y0)))
                rr = max(4.0, float(rr), 2.0 * float(width))
                for seg_idx, ((sx0, sy0), (sx1, sy1)) in enumerate(_asterisk_segments(cx, cy, rr)):
                    _draw_line_element(
                        key=f"highlight_asterisk:{h_idx}:{seg_idx}",
                        etype="highlight_asterisk",
                        points=[(float(sx0), float(sy0)), (float(sx1), float(sy1))],
                        base_width=max(1, int(width)),
                        base_color=str(outline),
                    )
            else:
                _draw_polygon_outline_element(
                    key=f"highlight_box:{h_idx}",
                    etype="highlight_box",
                    points=[(float(x0), float(y0)), (float(x1), float(y0)), (float(x1), float(y1)), (float(x0), float(y1))],
                    closed=True,
                    base_width=width,
                    base_color=str(outline),
                )

    # Marker / ladder
    if pc.include_marker and pc.marker_calibration.marker_name:
        try:
            marker = load_marker(pc.marker_calibration.marker_name)
            pos = marker_positions_for_panel(panel, pc, settings)

            if pos:
                lanes = max(1, pc.lanes)
                ladder_lane = max(1, min(lanes, int(pc.marker_calibration.ladder_lane) if pc.marker_calibration.ladder_lane else 1))
                x_tick = lane_lefts[ladder_lane - 1] - 8 + int(pc.marker_x_offset)
                tick_len = max(2, int(pc.marker_tick_length))
                label_gap = max(0, int(pc.marker_label_gap))
                labels_top_y: List[float] = []
                label_right_anchor_x = float(x_tick - tick_len - label_gap)

                for marker_idx, (s, ypix) in enumerate(pos, start=1):
                    y_canvas = panel_origin_y + int(ypix) + int(pc.marker_y_offset)
                    # tick
                    _draw_line_element(
                        key=f"marker_tick:{marker_idx}",
                        etype="marker_tick",
                        points=[(float(x_tick - tick_len), float(y_canvas)), (float(x_tick), float(y_canvas))],
                        base_width=2,
                        base_color="#000000",
                    )
                    label = _format_marker_size(float(s))
                    bbox = draw.textbbox((0, 0), label, font=marker_font)
                    tw = bbox[2] - bbox[0]
                    th = bbox[3] - bbox[1]
                    label_x = x_tick - tick_len - label_gap - tw
                    label_y = int(round(float(y_canvas) - (float(th) / 2.0) - float(bbox[1])))
                    labels_top_y.append(float(label_y + bbox[1]))
                    _draw_text_top_left(
                        key=f"marker_label:{marker_idx}",
                        text=label,
                        etype="marker_label",
                        x=float(label_x),
                        y=float(label_y),
                        base_size=max(6, int(getattr(pc, "marker_font_size", settings.font_size) or settings.font_size)),
                        base_color="#000000",
                        base_angle=0.0,
                    )

            # Center unit label above the right edge of the number labels (e.g. above the end of "10000").
            if pos:
                unit_text = str(marker.unit).strip()
                if unit_text:
                    bb_u = draw.textbbox((0, 0), unit_text, font=marker_font)
                    uw = float(bb_u[2] - bb_u[0])
                    uh = float(bb_u[3] - bb_u[1])
                    unit_center_x = float(label_right_anchor_x)
                    unit_x = float(unit_center_x) - uw / 2.0 - float(bb_u[0])
                    if labels_top_y:
                        unit_y = min(labels_top_y) - uh - 3.0 - float(bb_u[1])
                    else:
                        unit_y = float(panel_origin_y + 6)
                    unit_x = max(0.0, float(unit_x))
                    _draw_text_top_left(
                        key="marker_unit",
                        text=unit_text,
                        etype="marker_unit",
                        x=float(unit_x),
                        y=float(unit_y),
                        base_size=max(6, int(getattr(pc, "marker_font_size", settings.font_size) or settings.font_size)),
                        base_color="#000000",
                        base_angle=0.0,
                    )
        except Exception:
            pass

    # Header rows that are moved to the bottom of the panel.
    if bottom_header_rows:
        y_bottom = panel_origin_y + H + 4 + bottom_header_offset_up
        for hr in bottom_header_rows:
            row_angle = int(header_row_angle_by_id.get(getattr(hr, "id", ""), header_angle_deg))
            row_h = int(header_row_h_by_id.get(getattr(hr, "id", ""), header_row_h))
            heading_font = _row_font(int(header_row_heading_font_size_by_id.get(getattr(hr, "id", ""), default_font_size)))
            value_font = _row_font(int(header_row_value_font_size_by_id.get(getattr(hr, "id", ""), default_font_size)))
            value_line_h = int(header_row_value_line_h_by_id.get(getattr(hr, "id", ""), line_h))
            if hr.name.strip():
                bb_name = draw.textbbox((0, 0), hr.name, font=heading_font)
                name_x = float(panel_origin_x + header_heading_x_off - bb_name[2])
                name_y = float(y_bottom + header_heading_y_off)
                if abs(int(row_angle)) > 0:
                    th_name = float(bb_name[3] - bb_name[1])
                    name_y = float(y_bottom) + (float(row_h) - th_name) / 2.0 - float(bb_name[1]) + float(header_heading_y_off)
                _draw_text_top_left(
                    key=f"hdr_name_bottom:{hr.id}",
                    text=str(hr.name),
                    etype="header_name",
                    x=float(name_x),
                    y=float(name_y),
                    base_size=int(header_row_heading_font_size_by_id.get(getattr(hr, "id", ""), default_font_size)),
                    base_color="#000000",
                    base_angle=0.0,
                )
            for i in range(lanes):
                v = hr.values[i] if i < len(hr.values) else ""
                if not v:
                    continue
                key = f"hdr_val_bottom:{hr.id}:{i}"
                s = _style_for(
                    key=key,
                    base_size=int(header_row_value_font_size_by_id.get(getattr(hr, "id", ""), default_font_size)),
                    base_color="#000000",
                    base_angle=float(row_angle),
                )
                value_font_use = _font_by_family(int(s["font_size"]), str(s["font_family"]))
                x = lane_centers[i] + header_values_x_off + float(s["dx"])
                draw_rotated_header_value(
                    out,
                    text=str(v),
                    cell_center_x=float(x),
                    row_top_y=float(y_bottom + header_values_y_off + float(s["dy"])),
                    row_h=float(row_h),
                    baseline_h=float(value_line_h),
                    angle_deg=float(s["angle"]),
                    font=value_font_use,
                    fill=_hex_rgb(str(s["color"])),
                )
                try:
                    if abs(float(s["angle"])) < 1e-6:
                        bbv = value_font_use.getbbox(str(v))
                        tw = float(bbv[2] - bbv[0])
                        th = float(bbv[3] - bbv[1])
                        bx0 = float(x) - tw / 2.0
                        by0 = float(y_bottom + header_values_y_off + float(s["dy"]))
                        bbox = (bx0, by0, bx0 + tw, by0 + th)
                    else:
                        rw, rh = rotated_text_size(str(v), value_font_use, float(s["angle"]))
                        bx0 = float(x)
                        by0 = float(y_bottom + header_values_y_off + float(s["dy"]) + row_h - rh)
                        bbox = (bx0, by0, bx0 + float(rw), by0 + float(rh))
                    _record_element(key, str(v), "header_value", bbox, s)
                except Exception:
                    pass
            y_bottom += row_h

    # Optional final-size labels under panel (from analysis).
    final_size_labels = dict(layout.get("final_size_labels", {}) or {})
    if final_size_labels:
        y_base = panel_origin_y + H + 4 + max(0, bottom_header_block_h) + 2
        line_h_lbl = max(10, int(layout.get("final_label_line_h", line_h)))
        panel_unit = str(getattr(pc, "analysis_band_size_unit", "") or "").strip()
        lane_items: List[Tuple[int, List[str]]] = []
        for lane, lines in sorted(final_size_labels.items(), key=lambda kv: int(kv[0])):
            try:
                li = int(lane)
            except Exception:
                continue
            if li < 1 or li > len(lane_centers):
                continue
            parts = [str(x).strip() for x in list(lines or []) if str(x).strip()]
            if not parts:
                continue
            lane_items.append((li, parts))
        if lane_items and panel_unit:
            last_lane, last_parts = lane_items[-1]
            if last_parts:
                last_parts[-1] = f"{last_parts[-1]} {panel_unit}"
            lane_items[-1] = (last_lane, last_parts)
        # Keep similarly sized bands horizontally aligned across lanes:
        # row j in each lane is drawn at the same y position (no stagger/de-overlap shifting).
        for li, parts in lane_items:
            x = lane_centers[li - 1]
            for j, t in enumerate(parts):
                yy = float(y_base + j * line_h_lbl)
                bb = draw.textbbox((0, 0), t, font=final_label_font)
                tw = bb[2] - bb[0]
                x0 = float(x - tw / 2.0)
                _draw_text_top_left(
                    key=f"final_size:{li}:{j}",
                    text=str(t),
                    etype="final_size_label",
                    x=float(x0),
                    y=float(yy),
                    base_size=int(final_label_font_size if final_label_font_size > 0 else settings.font_size),
                    base_color="#000000",
                    base_angle=0.0,
                )

    return out


def render_full_project(state: ProjectState, preview: bool = False) -> Image.Image:
    """
    Compose all panels vertically into a final figure.
    """
    rendered = []
    for panel, pc in zip(state.panels, state.panel_configs):
        rendered.append(render_panel(panel, pc, state.settings, preview=preview))

    if not rendered:
        return Image.new("RGB", (800, 600), (255, 255, 255))

    spacing = PANEL_SPACING_PX
    W = max(img.width for img in rendered)
    H = sum(img.height for img in rendered) + spacing * (len(rendered) - 1)
    out = Image.new("RGB", (W, H), (255, 255, 255))
    y = 0
    for img in rendered:
        out.paste(img, (0, y))
        y += img.height + spacing
    return out


def render_full_project_with_elements(
    state: ProjectState,
    preview: bool = False,
) -> Tuple[Image.Image, List[Dict[str, Any]]]:
    """
    Render full project image and return clickable text element metadata
    for Review-step interactive editing.
    """
    rendered: List[Image.Image] = []
    elements: List[Dict[str, Any]] = []
    spacing = PANEL_SPACING_PX
    y_off = 0
    for idx, (panel, pc) in enumerate(zip(state.panels, state.panel_configs)):
        img = render_panel(
            panel,
            pc,
            state.settings,
            preview=preview,
            collect_elements=elements,
            panel_index=int(idx),
            panel_y_offset=int(y_off),
        )
        rendered.append(img)
        y_off += int(img.height) + int(spacing)

    if not rendered:
        return Image.new("RGB", (800, 600), (255, 255, 255)), []

    W = max(img.width for img in rendered)
    H = sum(img.height for img in rendered) + spacing * (len(rendered) - 1)
    out = Image.new("RGB", (W, H), (255, 255, 255))
    y = 0
    for img in rendered:
        out.paste(img, (0, y))
        y += img.height + spacing
    return out, elements


def main():
    app = GelAnnotatorApp()
    app.mainloop()

if __name__ == "__main__":
    main()










